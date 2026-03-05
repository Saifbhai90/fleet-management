from flask import render_template, redirect, url_for, flash, request, Response, jsonify, send_from_directory, session
from app import app, db
from models import (
    Company, Project, Vehicle, Driver, ParkingStation, District, EmployeePost, Employee,
    project_district, vehicle_district, ProjectTransfer, VehicleTransfer, DriverTransfer, DriverStatusChange,
    DriverAttendance,
    VehicleDailyTask, EmergencyTaskRecord, VehicleMileageRecord, RedTask, VehicleMoveWithoutTask, PenaltyRecord,
    Party, Product, FuelExpense, ProductBalance, OilExpense, OilExpenseItem, OilExpenseAttachment,
    MaintenanceExpense, MaintenanceExpenseItem, MaintenanceExpenseAttachment,
    Notification,
)
from forms import (
    CompanyForm, ProjectForm, VehicleForm, VehicleImportForm, DriverForm, DriverImportForm, ParkingForm, DistrictForm,
    AssignProjectToCompanyForm, EditProjectAssignmentForm,
    AssignProjectToDistrictForm, AssignVehicleToDistrictForm, AssignVehicleToParkingForm,
    AssignDriverToVehicleForm, ProjectTransferForm, VehicleTransferForm, EditVehicleTransferForm, DriverTransferForm, DriverJobLeftForm, DriverRejoinForm,
    DriverAttendanceFilterForm, DriverAttendanceReportForm, ATTENDANCE_STATUS_CHOICES,
    TaskReportForm, TaskReportFilterForm, EmergencyTaskUploadForm, VehicleMileageUploadForm, ParkingImportForm,
    TaskReportUploadBothForm, RedTaskFilterForm, RedTaskForm, VehicleMoveWithoutTaskFilterForm, VehicleMoveWithoutTaskForm, PenaltyRecordForm, PenaltyRecordFilterForm,
    FuelExpenseFilterForm, FuelExpenseForm,
    OilExpenseFilterForm, OilExpenseForm,
    MaintenanceExpenseFilterForm, MaintenanceExpenseForm,
    PartyForm,
    ProductForm,
    EmployeePostForm,
    EmployeeForm,
    LoginForm,
)
from datetime import datetime, date, time
import csv
from io import StringIO
from sqlalchemy import func, text, inspect
from sqlalchemy.exc import OperationalError, IntegrityError
from utils import generate_csv_response, parse_date
import re
import os
from werkzeug.utils import secure_filename

@app.before_request
def require_login():
    """Redirect to login if user not logged in. Only login and static are public."""
    # endpoint can be None for some requests
    endpoint = request.endpoint or ''
    if endpoint.startswith('static'):
        return
    if endpoint in ('login',):
        return
    if not session.get('user'):
        return redirect(url_for('login'))


# ────────────────────────────────────────────────
# API: CNIC / License duplicate check (for driver form)
# ────────────────────────────────────────────────
def _cnic_digits(cnic):
    if not cnic:
        return ''
    return re.sub(r'[-\s]', '', str(cnic).strip())

@app.route('/api/check-cnic')
def api_check_cnic():
    """Returns { exists: bool, message: str }. Call with ?cnic=xxx&exclude_driver_id=1 (optional, for edit)."""
    cnic = request.args.get('cnic', '').strip()
    digits = _cnic_digits(cnic)
    if not cnic or len(digits) != 13:
        return jsonify({'exists': False, 'message': ''})
    exclude = request.args.get('exclude_driver_id', type=int)
    drivers = Driver.query.filter(Driver.cnic_no.isnot(None), Driver.cnic_no != '').all()
    for d in drivers:
        if exclude and d.id == exclude:
            continue
        if _cnic_digits(d.cnic_no) == digits:
            return jsonify({'exists': True, 'message': f'CNIC already registered for driver: {d.name} ({d.driver_id})'})
    return jsonify({'exists': False, 'message': ''})

@app.route('/api/check-license')
def api_check_license():
    """Returns { exists: bool, message: str }. Call with ?license=xxx&exclude_driver_id=1 (optional)."""
    license_no = request.args.get('license', '').strip()
    if not license_no:
        return jsonify({'exists': False, 'message': ''})
    exclude = request.args.get('exclude_driver_id', type=int)
    q = Driver.query.filter(Driver.license_no.ilike(license_no))
    if exclude:
        q = q.filter(Driver.id != exclude)
    other = q.first()
    if other:
        return jsonify({'exists': True, 'message': f'License number already registered for driver: {other.name} ({other.driver_id})'})
    return jsonify({'exists': False, 'message': ''})

# ────────────────────────────────────────────────
# Serve uploaded files (vehicles/drivers documents and photos)
# ────────────────────────────────────────────────
@app.route('/uploads/<path:filename>')
def uploaded_file(filename):
    """Serve files from UPLOAD_FOLDER. Path must be under uploads (no traversal)."""
    base = os.path.abspath(app.config['UPLOAD_FOLDER'])
    path = os.path.abspath(os.path.join(base, filename))
    if not path.startswith(base) or not os.path.isfile(path):
        return '', 404
    return send_from_directory(base, filename)

# ────────────────────────────────────────────────
# Dashboard / Home
# ────────────────────────────────────────────────
@app.route('/')
@app.route('/dashboard')
def dashboard():
    total_companies = Company.query.count()
    total_projects = Project.query.count()
    total_vehicles = Vehicle.query.count()
    total_drivers = Driver.query.count()
    total_parking = ParkingStation.query.count()
    total_districts = District.query.count()
    recent_projects = Project.query.order_by(Project.created_at.desc()).limit(5).all()

    # Notifications: unread first, recent
    notifications = Notification.query.filter(Notification.read_at.is_(None)).order_by(Notification.created_at.desc()).limit(20).all()
    # Optionally seed a few from data (e.g. expiry) - only if none exist
    if not notifications and total_drivers:
        from datetime import timedelta
        today = date.today()
        end = today + timedelta(days=60)
        expiring_count = 0
        for d in Driver.query.filter(Driver.status == 'Active').all():
            if (d.license_expiry_date and today <= d.license_expiry_date <= end) or (d.cnic_expiry_date and today <= d.cnic_expiry_date <= end):
                expiring_count += 1
        if expiring_count > 0:
            n = Notification(
                title='Document expiry',
                message=f'{expiring_count} driver(s) have license or CNIC expiring in the next 60 days.',
                link=url_for('report_expiry'),
                link_text='View expiry report',
                notification_type='warning'
            )
            db.session.add(n)
            db.session.commit()
            notifications = [n]
    # Optional: one "parking full" notification if any station is at capacity (only when we have no unread notifications)
    if not notifications and total_parking:
        for s in ParkingStation.query.all():
            if s.capacity and s.capacity > 0:
                occ = Vehicle.query.filter_by(parking_station_id=s.id).count()
                if occ >= s.capacity:
                    n2 = Notification(
                        title='Parking full',
                        message=f'"{s.name}" is at full capacity ({occ}/{s.capacity}).',
                        link=url_for('report_parking_utilization'),
                        link_text='View parking',
                        notification_type='danger'
                    )
                    db.session.add(n2)
                    db.session.commit()
                    notifications = Notification.query.filter(Notification.read_at.is_(None)).order_by(Notification.created_at.desc()).limit(20).all()
                    break

    # Optional: one "attendance missing" notification if many active drivers have no recent attendance (only when we have no unread notifications)
    if not notifications and total_drivers:
        from datetime import timedelta
        today = date.today()
        start = today - timedelta(days=7)
        active_drivers = Driver.query.filter(Driver.status == 'Active').all()
        missing_count = 0
        for d in active_drivers:
            has_recent = DriverAttendance.query.filter(
                DriverAttendance.driver_id == d.id,
                DriverAttendance.attendance_date >= start
            ).first()
            if not has_recent:
                missing_count += 1
        if missing_count > 0:
            n3 = Notification(
                title='Attendance missing',
                message=f'{missing_count} active driver(s) have no attendance in the last 7 days.',
                link=url_for('driver_attendance_list'),
                link_text='View attendance',
                notification_type='info'
            )
            db.session.add(n3)
            db.session.commit()
            notifications = Notification.query.filter(Notification.read_at.is_(None)).order_by(Notification.created_at.desc()).limit(20).all()

    return render_template('dashboard.html',
                           total_companies=total_companies,
                           total_projects=total_projects,
                           total_vehicles=total_vehicles,
                           total_drivers=total_drivers,
                           total_parking=total_parking,
                           total_districts=total_districts,
                           recent_projects=recent_projects,
                           notifications=notifications)


@app.route('/notification/<int:pk>/read', methods=['POST'])
def notification_mark_read(pk):
    n = Notification.query.get_or_404(pk)
    n.read_at = datetime.utcnow()
    db.session.commit()
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({'ok': True})
    next_url = request.args.get('next') or url_for('dashboard')
    return redirect(next_url)


# ────────────────────────────────────────────────
# Companies
# ────────────────────────────────────────────────
@app.route('/companies/')
def companies():
    search = request.args.get('search', '')
    query = Company.query
    if search:
        query = query.filter(Company.name.ilike(f'%{search}%'))
    companies_list = query.order_by(Company.name).all()
    return render_template('companies.html', companies=companies_list, search=search)


@app.route('/company/add', methods=['GET', 'POST'])
@app.route('/company/edit/<int:id>', methods=['GET', 'POST'])
def company_form(id=None):
    company = Company.query.get_or_404(id) if id else None
    form = CompanyForm(obj=company)
    if form.validate_on_submit():
        try:
            if not company:
                company = Company()
            form.populate_obj(company)
            if not id:
                db.session.add(company)
            db.session.commit()
            flash('Company saved successfully!', 'success')
            return redirect(url_for('companies'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error saving company: {str(e)}', 'danger')
    return render_template('company_form.html', form=form, title='Company', back_url=url_for('companies'))


@app.route('/company/delete/<int:id>', methods=['POST'])
def delete_company(id):
    company = Company.query.get_or_404(id)
    if company.projects:
        flash(f'Cannot delete "{company.name}". It has {len(company.projects)} project(s). Delete projects first.', 'danger')
        return redirect(url_for('companies'))
    try:
        db.session.delete(company)
        db.session.commit()
        flash('Company deleted successfully.', 'success')
    except Exception as e:
        flash(f'Error deleting company: {str(e)}', 'danger')
    return redirect(url_for('companies'))


# ────────────────────────────────────────────────
# Projects
# ────────────────────────────────────────────────
@app.route('/projects/')
def projects_list():
    search = request.args.get('search', '').strip()
    query = Project.query
    if search:
        query = query.filter(Project.name.ilike(f'%{search}%'))
    projects = query.order_by(Project.name).all()
    return render_template('projects_list.html', projects=projects, search=search)


@app.route('/projects/export')
def projects_export():
    """Export projects list (with optional search) to CSV."""
    search = request.args.get('search', '').strip()
    query = Project.query
    if search:
        query = query.filter(Project.name.ilike(f'%{search}%'))
    projects = query.order_by(Project.name).all()
    headers = ['ID', 'Project Name', 'Start Date', 'Status']
    rows = []
    for p in projects:
        rows.append([
            p.id,
            p.name,
            p.start_date.strftime('%Y-%m-%d') if getattr(p, 'start_date', None) else '',
            p.status
        ])
    filename = 'projects.csv' if not search else f'projects_search_{search}.csv'
    return generate_csv_response(headers, rows, filename=filename)


@app.route('/project/<int:id>')
def project_detail(id):
    project = Project.query.get_or_404(id)
    v_search = request.args.get('v_search', '')
    d_search = request.args.get('d_search', '')
    p_search = request.args.get('p_search', '')
    vehicles = project.vehicles
    if v_search:
        vehicles = [v for v in vehicles if v_search.lower() in v.vehicle_no.lower() or v_search.lower() in v.model.lower() or v_search.lower() in v.vehicle_type.lower()]
    drivers = project.drivers
    if d_search:
        drivers = [d for d in drivers if d_search.lower() in d.name.lower() or d_search.lower() in (d.cnic_no or '').lower() or d_search.lower() in (d.license_no or '').lower()]
    parkings = project.parking_stations
    if p_search:
        parkings = [p for p in parkings if p_search.lower() in p.name.lower() or p_search.lower() in (p.district or '').lower() or p_search.lower() in (p.address_location or '').lower()]
    return render_template('project_detail.html',
                           project=project,
                           vehicles=vehicles,
                           drivers=drivers,
                           parkings=parkings,
                           v_search=v_search,
                           d_search=d_search,
                           p_search=p_search)


@app.route('/company/<int:company_id>/projects/')
def company_projects(company_id):
    company = Company.query.get_or_404(company_id)
    search = request.args.get('search', '')
    query = Project.query.filter_by(company_id=company_id)
    if search:
        query = query.filter(Project.name.ilike(f'%{search}%'))
    projects = query.all()
    return render_template('projects.html', company=company, projects=projects, search=search)


@app.route('/project/add', methods=['GET', 'POST'])
@app.route('/project/edit/<int:id>', methods=['GET', 'POST'])
def project_form(id=None):
    project = Project.query.get_or_404(id) if id else None
    form = ProjectForm(obj=project)
    if form.validate_on_submit():
        try:
            if not project:
                project = Project()
            form.populate_obj(project)
            if form.status.data == 'Inactive' and not form.inactive_date.data:
                flash('Inactive Date is required when status is Inactive.', 'danger')
                return render_template('project_form.html', form=form, title='Project', back_url=url_for('projects_list'))
            if not id:
                db.session.add(project)
            db.session.commit()
            flash('Project saved successfully!', 'success')
            return redirect(url_for('projects_list'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error saving project: {str(e)}', 'danger')
    back_url = url_for('projects_list')
    return render_template('project_form.html', form=form, title='Project', back_url=back_url)


@app.route('/project/delete/<int:id>', methods=['POST'])
def delete_project(id):
    project = Project.query.get_or_404(id)
    linked_districts = db.session.query(project_district).filter_by(project_id=project.id).count()
    if linked_districts > 0:
        flash(f'Cannot delete "{project.name}". It is linked to {linked_districts} district(s). Remove district assignments first.', 'danger')
        return redirect(url_for('assign_project_to_company'))
    try:
        db.session.delete(project)
        db.session.commit()
        flash(f'Project "{project.name}" deleted successfully.', 'success')
    except Exception as e:
        flash(f'Error deleting project: {str(e)}', 'danger')
    return redirect(url_for('projects_list'))


@app.route('/project/toggle_status/<int:id>', methods=['POST'])
def toggle_project_status(id):
    project = Project.query.get_or_404(id)
    if project.status == 'Active':
        project.status = 'Inactive'
        flash('Project marked as Inactive. Update Inactive Date from Edit if needed.', 'info')
    else:
        project.status = 'Active'
        project.inactive_date = None
        flash('Project marked as Active.', 'success')
    db.session.commit()
    return redirect(url_for('projects_list'))


# ────────────────────────────────────────────────
# Vehicles
# ────────────────────────────────────────────────
@app.route('/vehicles/')
def vehicles_list():
    search = request.args.get('search', '').strip()
    query = Vehicle.query
    if search:
        like = f'%{search}%'
        query = query.filter(
            Vehicle.vehicle_no.ilike(like) |
            Vehicle.model.ilike(like) |
            Vehicle.vehicle_type.ilike(like)
        )
    # Sort by ID as requested
    vehicles = query.order_by(Vehicle.id).all()
    return render_template('vehicles_list.html', vehicles=vehicles, search=search)


@app.route('/whats-new')
def whats_new():
    """Simple 'What's New' page so users can see recent changes."""
    return render_template('whats_new.html')


@app.route('/vehicles/export')
def vehicles_export():
    """Export vehicles list (with optional search) to CSV."""
    search = request.args.get('search', '').strip()
    query = Vehicle.query
    if search:
        like = f'%{search}%'
        query = query.filter(
            Vehicle.vehicle_no.ilike(like) |
            Vehicle.model.ilike(like) |
            Vehicle.vehicle_type.ilike(like)
        )
    vehicles = query.order_by(Vehicle.id).all()
    headers = ['ID', 'Vehicle No', 'Model', 'Type', 'Driver Capacity', 'Phone', 'Active Date']
    rows = []
    for v in vehicles:
        rows.append([
            v.id,
            v.vehicle_no,
            v.model,
            v.vehicle_type,
            getattr(v, 'driver_capacity', None),
            v.phone_no,
            v.active_date.strftime('%Y-%m-%d') if getattr(v, 'active_date', None) else ''
        ])
    filename = 'vehicles.csv' if not search else f'vehicles_search_{search}.csv'
    return generate_csv_response(headers, rows, filename=filename)


@app.route('/vehicles/import', methods=['GET', 'POST'])
def vehicles_import():
    """
    Import vehicles from Excel/CSV.
    """
    form = VehicleImportForm()
    if request.method == 'POST' and form.validate_on_submit():
        f = form.file.data
        if not f:
            flash('Please select an Excel/CSV file.', 'warning')
            return redirect(url_for('vehicles_import'))
        try:
            import pandas as pd
            filename = f.filename or ''
            ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
            if ext in ('xlsx', 'xls'):
                df = pd.read_excel(f)
            elif ext == 'csv':
                df = pd.read_csv(f)
            else:
                flash('Unsupported file type. Please upload .xlsx, .xls or .csv.', 'danger')
                return redirect(url_for('vehicles_import'))

            expected_cols = ['vehicle_no', 'model', 'vehicle_type', 'engine_no', 'chassis_no', 'driver_capacity', 'phone_no', 'active_date', 'remarks']
            missing = [c for c in expected_cols if c not in df.columns]
            if missing:
                flash(f'Missing required columns in file: {", ".join(missing)}', 'danger')
                return redirect(url_for('vehicles_import'))

            created = 0
            from utils import parse_date
            for _, row in df.iterrows():
                vno = str(row.get('vehicle_no') or '').strip()
                if not vno:
                    continue
                existing = Vehicle.query.filter(Vehicle.vehicle_no == vno).first()
                if existing:
                    continue

                cap_val = row.get('driver_capacity')
                try:
                    driver_capacity = int(cap_val) if pd.notna(cap_val) else None
                except Exception:
                    driver_capacity = None

                active_raw = row.get('active_date')
                active_date = None
                if pd.notna(active_raw):
                    if isinstance(active_raw, str):
                        active_date = parse_date(active_raw)
                    else:
                        try:
                            active_date = pd.to_datetime(active_raw).date()
                        except Exception:
                            active_date = None

                v = Vehicle(
                    vehicle_no=vno,
                    model=str(row.get('model') or '').strip(),
                    engine_no=str(row.get('engine_no') or '').strip() or None,
                    chassis_no=str(row.get('chassis_no') or '').strip() or None,
                    vehicle_type=str(row.get('vehicle_type') or '').strip() or None,
                    driver_capacity=driver_capacity,
                    phone_no=str(row.get('phone_no') or '').strip() or None,
                    active_date=active_date,
                    remarks=str(row.get('remarks') or '').strip() or None,
                )
                db.session.add(v)
                created += 1
            db.session.commit()
            if created:
                flash(f'{created} vehicle(s) imported successfully.', 'success')
            else:
                flash('File processed but no new vehicles were imported (duplicates or empty rows).', 'info')
            return redirect(url_for('vehicles_list'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error importing vehicles: {str(e)}', 'danger')
    return render_template('vehicles_import.html', form=form)


@app.route('/vehicles/import/template')
def vehicles_import_template():
    """
    Downloadable template for vehicle import (open in Excel and fill).
    """
    headers = ['vehicle_no', 'model', 'vehicle_type', 'engine_no', 'chassis_no', 'driver_capacity', 'phone_no', 'active_date', 'remarks']
    rows = [
        ['LEA-1234', 'Suzuki Cultus', 'Ambulance', 'ENG123', 'CHS123', 1, '0300-1112233', '01-01-2024', 'Example row 1'],
        ['LEA-5678', 'Toyota Hiace', 'Passanger', 'ENG456', 'CHS456', 2, '0300-2223344', '05-02-2024', 'Example row 2'],
    ]
    return generate_csv_response(headers, rows, filename='vehicles_import_template.csv')


@app.route('/vehicles/print')
def vehicles_print():
    """Print-friendly view of vehicles (browser print to PDF)."""
    search = request.args.get('search', '').strip()
    query = Vehicle.query
    if search:
        like = f'%{search}%'
        query = query.filter(
            Vehicle.vehicle_no.ilike(like) |
            Vehicle.model.ilike(like) |
            Vehicle.vehicle_type.ilike(like)
        )
    vehicles = query.order_by(Vehicle.id).all()
    return render_template('vehicles_print.html', vehicles=vehicles, search=search)


@app.route('/vehicle/add', methods=['GET', 'POST'])
@app.route('/vehicle/edit/<int:id>', methods=['GET', 'POST'])
def vehicle_form(id=None):
    vehicle = Vehicle.query.get_or_404(id) if id else None
    form = VehicleForm(obj=vehicle)
    if form.validate_on_submit():
        try:
            if not vehicle:
                vehicle = Vehicle()
            # Normalize vehicle number (trim spaces)
            vehicle_no_normalized = (form.vehicle_no.data or '').strip()
            form.populate_obj(vehicle)
            vehicle.vehicle_no = vehicle_no_normalized

            # Duplicate vehicle_no check (case-insensitive, ignore current record)
            existing_q = Vehicle.query.filter(
                func.lower(Vehicle.vehicle_no) == vehicle_no_normalized.lower()
            )
            if id:
                existing_q = existing_q.filter(Vehicle.id != vehicle.id)
            if existing_q.first():
                flash('This Vehicle Number is already registered', 'danger')
                target = url_for('vehicle_form', id=id) if id else url_for('vehicle_form')
                return redirect(target)

            if not id:
                db.session.add(vehicle)
            db.session.commit()
            doc_file = request.files.get('document')
            if doc_file and doc_file.filename:
                ext = (os.path.splitext(secure_filename(doc_file.filename))[1] or '.pdf').lower()
                if ext != '.pdf':
                    ext = '.pdf'
                subdir = os.path.join(app.config['UPLOAD_FOLDER'], 'vehicles', str(vehicle.id))
                os.makedirs(subdir, exist_ok=True)
                fname = secure_filename('document') + ext
                filepath = os.path.join(subdir, fname)
                doc_file.save(filepath)
                vehicle.document_path = os.path.join('vehicles', str(vehicle.id), fname).replace('\\', '/')
                db.session.commit()
            flash('Vehicle saved successfully!', 'success')
            return redirect(url_for('vehicles_list'))
        except IntegrityError:
            db.session.rollback()
            flash('This Vehicle Number is already registered', 'danger')
            target = url_for('vehicle_form', id=id) if id else url_for('vehicle_form')
            return redirect(target)
        except Exception as e:
            db.session.rollback()
            flash(f'Error saving vehicle: {str(e)}', 'danger')
    back_url = url_for('vehicles_list')
    return render_template('vehicle_form.html', form=form, vehicle=vehicle, title='Vehicle', back_url=back_url)


@app.route('/vehicle/delete/<int:id>', methods=['POST'])
def delete_vehicle(id):
    vehicle = Vehicle.query.get_or_404(id)
    try:
        db.session.delete(vehicle)
        db.session.commit()
        flash('Vehicle deleted.', 'success')
    except Exception as e:
        flash(f'Error deleting vehicle: {str(e)}', 'danger')
    return redirect(url_for('vehicles_list'))


# ────────────────────────────────────────────────
# Drivers
# ────────────────────────────────────────────────
@app.route('/drivers')
def drivers_list():
    search = request.args.get('search', '').strip()
    query = Driver.query
    if search:
        like = f"%{search}%"
        query = query.filter(
            Driver.name.ilike(like) |
            Driver.driver_id.ilike(like) |
            Driver.cnic_no.ilike(like) |
            Driver.license_no.ilike(like) |
            Driver.phone1.ilike(like) |
            Driver.phone2.ilike(like) |
            Driver.driver_district.ilike(like)
        )
    # Sort by ID as requested
    drivers = query.order_by(Driver.id.asc()).all()
    today = date.today()
    return render_template('drivers_list.html', drivers=drivers, search=search, today=today)


@app.route('/drivers/export')
def drivers_export():
    """Export drivers list (with optional search) to CSV."""
    search = request.args.get('search', '').strip()
    query = Driver.query
    if search:
        like = f"%{search}%"
        query = query.filter(
            Driver.name.ilike(like) |
            Driver.driver_id.ilike(like) |
            Driver.cnic_no.ilike(like) |
            Driver.license_no.ilike(like) |
            Driver.phone1.ilike(like) |
            Driver.phone2.ilike(like) |
            Driver.driver_district.ilike(like)
        )
    drivers = query.order_by(Driver.id.asc()).all()
    headers = ['Driver ID', 'Name', 'CNIC', 'License', 'Phone', 'District', 'Status']
    rows = []
    for d in drivers:
        rows.append([
            d.driver_id,
            d.name,
            d.cnic_no,
            d.license_no,
            d.phone1,
            d.driver_district,
            d.status
        ])
    filename = 'drivers.csv' if not search else f'drivers_search_{search}.csv'
    return generate_csv_response(headers, rows, filename=filename)


@app.route('/drivers/import', methods=['GET', 'POST'])
def drivers_import():
    """
    Import drivers from Excel/CSV (basic fields).
    """
    form = DriverImportForm()
    if request.method == 'POST' and form.validate_on_submit():
        f = form.file.data
        if not f:
            flash('Please select an Excel/CSV file.', 'warning')
            return redirect(url_for('drivers_import'))
        try:
            import pandas as pd
            filename = f.filename or ''
            ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
            if ext in ('xlsx', 'xls'):
                df = pd.read_excel(f)
            elif ext == 'csv':
                df = pd.read_csv(f)
            else:
                flash('Unsupported file type. Please upload .xlsx, .xls or .csv.', 'danger')
                return redirect(url_for('drivers_import'))

            expected_cols = [
                'driver_id', 'name', 'father_name', 'phone1', 'driver_district',
                'cnic_no', 'license_no', 'status'
            ]
            missing = [c for c in expected_cols if c not in df.columns]
            if missing:
                flash(f'Missing required columns in file: {", ".join(missing)}', 'danger')
                return redirect(url_for('drivers_import'))

            created = 0
            for _, row in df.iterrows():
                did = str(row.get('driver_id') or '').strip()
                if not did:
                    continue
                existing = Driver.query.filter(Driver.driver_id == did).first()
                if existing:
                    continue
                d = Driver(
                    driver_id=did,
                    name=str(row.get('name') or '').strip(),
                    father_name=str(row.get('father_name') or '').strip() or None,
                    phone1=str(row.get('phone1') or '').strip() or None,
                    driver_district=str(row.get('driver_district') or '').strip() or None,
                    cnic_no=str(row.get('cnic_no') or '').strip() or None,
                    license_no=str(row.get('license_no') or '').strip() or None,
                    status=str(row.get('status') or '').strip() or 'Active',
                )
                db.session.add(d)
                created += 1
            db.session.commit()
            if created:
                flash(f'{created} driver(s) imported successfully.', 'success')
            else:
                flash('File processed but no new drivers were imported (duplicates or empty rows).', 'info')
            return redirect(url_for('drivers_list'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error importing drivers: {str(e)}', 'danger')
    return render_template('drivers_import.html', form=form)


@app.route('/drivers/import/template')
def drivers_import_template():
    """
    Downloadable template for driver import (open in Excel and fill).
    """
    headers = ['driver_id', 'name', 'father_name', 'phone1', 'driver_district', 'cnic_no', 'license_no', 'status']
    rows = [
        ['DRV-2024-1001', 'Ali Ahmad', 'Ahmad Khan', '0300-1112233', 'Lahore', '32304-1111111-5', 'LTV-12345', 'Active'],
        ['DRV-2024-1002', 'Bilal Hussain', 'Hussain Raza', '0300-2223344', 'Lahore', '32304-2222222-5', 'HTV-56789', 'Active'],
    ]
    return generate_csv_response(headers, rows, filename='drivers_import_template.csv')


@app.route('/drivers/print')
def drivers_print():
    """Print-friendly view of drivers (browser print to PDF)."""
    search = request.args.get('search', '').strip()
    query = Driver.query
    if search:
        like = f"%{search}%"
        query = query.filter(
            Driver.name.ilike(like) |
            Driver.driver_id.ilike(like) |
            Driver.cnic_no.ilike(like) |
            Driver.license_no.ilike(like) |
            Driver.phone1.ilike(like) |
            Driver.phone2.ilike(like) |
            Driver.driver_district.ilike(like)
        )
    drivers = query.order_by(Driver.id.asc()).all()
    return render_template('drivers_print.html', drivers=drivers, search=search)


@app.route('/driver/add', methods=['GET', 'POST'])
@app.route('/driver/edit/<int:id>', methods=['GET', 'POST'])
def driver_form(id=None):
    # Common Employee Post master choices for both add/edit
    posts = EmployeePost.query.order_by(EmployeePost.full_name).all()
    post_choices = [('', '-- Select Post --')] + [
        (p.full_name, f"{p.full_name} ({p.short_name})") for p in posts
    ]

    if id:
        driver = Driver.query.get_or_404(id)
        form = DriverForm(obj=driver)
        form.post.choices = post_choices
        if driver.application_date:
            if isinstance(driver.application_date, str):
                form.application_date.data = parse_date(driver.application_date)
            else:
                form.application_date.data = driver.application_date
        title = f"Edit Driver - {driver.name}"
    else:
        driver = Driver()
        form = DriverForm()
        form.post.choices = post_choices
        title = "Add New Driver"
    if form.validate_on_submit():
        try:
            form.populate_obj(driver)
            if not id:
                db.session.add(driver)
            db.session.commit()
            subdir = os.path.join(app.config['UPLOAD_FOLDER'], 'drivers', str(driver.id))
            os.makedirs(subdir, exist_ok=True)
            photo_file = request.files.get('photo')
            if photo_file and photo_file.filename:
                ext = os.path.splitext(secure_filename(photo_file.filename))[1] or '.jpg'
                ext = ext.lower() if ext.lower() in ('.jpg', '.jpeg', '.png', '.gif', '.webp') else '.jpg'
                fname = 'photo' + ext
                filepath = os.path.join(subdir, fname)
                photo_file.save(filepath)
                driver.photo_path = os.path.join('drivers', str(driver.id), fname).replace('\\', '/')
            doc_file = request.files.get('document')
            if doc_file and doc_file.filename:
                ext = (os.path.splitext(secure_filename(doc_file.filename))[1] or '.pdf').lower()
                if ext != '.pdf':
                    ext = '.pdf'
                fname = secure_filename('document') + ext
                filepath = os.path.join(subdir, fname)
                doc_file.save(filepath)
                driver.document_path = os.path.join('drivers', str(driver.id), fname).replace('\\', '/')
            if photo_file and photo_file.filename or doc_file and doc_file.filename:
                db.session.commit()
            flash(f"Driver '{driver.name}' successfully {'updated' if id else 'added'}!", 'success')
            return redirect(url_for('drivers_list'))
        except Exception as e:
            db.session.rollback()
            flash(f"Error saving driver: {str(e)}", 'danger')
    return render_template('driver_form.html', form=form, title=title, driver=driver)


# ────────────────────────────────────────────────
# Employees (non-driver staff)
# ────────────────────────────────────────────────
@app.route('/employees')
def employees_list():
    search = request.args.get('search', '').strip()
    query = Employee.query
    if search:
        like = f"%{search}%"
        query = query.filter(
            Employee.name.ilike(like) |
            Employee.code.ilike(like) |
            Employee.department.ilike(like) |
            Employee.cnic_no.ilike(like)
        )
    employees = query.order_by(Employee.id.asc()).all()
    return render_template('employees_list.html', employees=employees, search=search)


@app.route('/employee/add', methods=['GET', 'POST'])
@app.route('/employee/edit/<int:id>', methods=['GET', 'POST'])
def employee_form(id=None):
    # Build Employee Post dropdown (id-based, shows full + short name)
    posts = EmployeePost.query.order_by(EmployeePost.full_name).all()
    post_choices = [(0, '-- Select Post --')] + [
        (p.id, f"{p.full_name} ({p.short_name})") for p in posts
    ]

    # Department history (distinct existing departments)
    dept_rows = (
        db.session.query(Employee.department)
        .filter(Employee.department.isnot(None), Employee.department != '')
        .distinct()
        .order_by(Employee.department)
        .all()
    )
    departments = [row[0] for row in dept_rows]

    if id:
        emp = Employee.query.get_or_404(id)
        form = EmployeeForm(obj=emp)
        form.post_id.choices = post_choices
        if emp.joining_date:
            if isinstance(emp.joining_date, str):
                form.joining_date.data = parse_date(emp.joining_date)
            else:
                form.joining_date.data = emp.joining_date
        title = f"Edit Employee - {emp.name}"
    else:
        emp = Employee()
        form = EmployeeForm()
        form.post_id.choices = post_choices
        title = "Add New Employee"

    if form.validate_on_submit():
        try:
            # Auto-generate code if user left it blank
            code_val = (form.code.data or "").strip()
            if not code_val:
                year = datetime.now().year
                last_emp = Employee.query.order_by(Employee.id.desc()).first()
                next_num = (last_emp.id + 1) if last_emp else 1
                form.code.data = f"EMP-{year}-{next_num:04d}"

            form.populate_obj(emp)
            if form.post_id.data == 0:
                emp.post_id = None
            if not id:
                db.session.add(emp)
            db.session.commit()
            flash('Employee saved successfully!', 'success')
            return redirect(url_for('employees_list'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error saving employee: {str(e)}', 'danger')

    return render_template('employee_form.html', form=form, title=title, employee=emp, departments=departments)


@app.route('/employee/delete/<int:id>', methods=['POST'])
def employee_delete(id):
    emp = Employee.query.get_or_404(id)
    try:
        db.session.delete(emp)
        db.session.commit()
        flash('Employee deleted.', 'success')
    except Exception as e:
        flash(f'Error deleting employee: {str(e)}', 'danger')
    return redirect(url_for('employees_list'))


@app.route('/employees/export')
def employees_export():
    """Export employees list (with optional search) to CSV."""
    search = request.args.get('search', '').strip()
    query = Employee.query
    if search:
        like = f"%{search}%"
        query = query.filter(
            Employee.name.ilike(like) |
            Employee.code.ilike(like) |
            Employee.department.ilike(like) |
            Employee.cnic_no.ilike(like)
        )
    employees = query.order_by(Employee.id.asc()).all()
    headers = ['ID', 'Code', 'Name', 'CNIC', 'Post', 'Department', 'Phone', 'Joining Date', 'Status']
    rows = []
    for e in employees:
        rows.append([
            e.id,
            e.code,
            e.name,
            e.cnic_no or '',
            e.post.full_name if e.post else '',
            e.department or '',
            e.phone1 or e.phone2 or '',
            e.joining_date.strftime('%Y-%m-%d') if e.joining_date else '',
            e.status or '',
        ])
    filename = 'employees.csv' if not search else f'employees_search_{search}.csv'
    return generate_csv_response(headers, rows, filename=filename)


@app.route('/employees/print')
def employees_print():
    """Print-friendly view of employees (browser print to PDF)."""
    search = request.args.get('search', '').strip()
    query = Employee.query
    if search:
        like = f"%{search}%"
        query = query.filter(
            Employee.name.ilike(like) |
            Employee.code.ilike(like) |
            Employee.department.ilike(like) |
            Employee.cnic_no.ilike(like)
        )
    employees = query.order_by(Employee.id.asc()).all()
    return render_template('employees_print.html', employees=employees, search=search)


# ────────────────────────────────────────────────
# Simple Login (admin / admin)
# ────────────────────────────────────────────────
@app.route('/login', methods=['GET', 'POST'])
def login():
    form = LoginForm()
    if form.validate_on_submit():
        username = (form.username.data or '').strip()
        password = (form.password.data or '').strip()
        if username == 'admin' and password == 'admin':
            session['user'] = 'admin'
            flash('Login successful. Welcome admin!', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid user ID or password.', 'danger')
    return render_template('login.html', form=form)


@app.route('/logout')
def logout():
    session.pop('user', None)
    flash('You have been logged out.', 'info')
    return redirect(url_for('login'))


@app.route('/driver/delete/<int:id>', methods=['POST'])
def delete_driver(id):
    driver = Driver.query.get_or_404(id)
    try:
        db.session.delete(driver)
        db.session.commit()
        flash('Driver deleted.', 'success')
    except Exception as e:
        flash(f'Error deleting driver: {str(e)}', 'danger')
    return redirect(url_for('drivers_list'))


# ────────────────────────────────────────────────
# Parking Stations
# ────────────────────────────────────────────────
@app.route('/parking/')
def parking_list():
    search = request.args.get('search', '')
    query = ParkingStation.query
    if search:
        query = query.filter(
            ParkingStation.name.ilike(f'%{search}%') |
            ParkingStation.district.ilike(f'%{search}%') |
            ParkingStation.address_location.ilike(f'%{search}%')
        )
    # Sort by ID as requested
    parkings = query.order_by(ParkingStation.id).all()
    return render_template('parking_list.html', parkings=parkings, search=search)


@app.route('/parking/import', methods=['GET', 'POST'])
def parking_import():
    """
    Import parking locations from Excel/CSV.
    User can also download a template.
    """
    form = ParkingImportForm()
    if request.method == 'POST' and form.validate_on_submit():
        f = form.file.data
        if not f:
            flash('Please select an Excel/CSV file.', 'warning')
            return redirect(url_for('parking_import'))
        try:
            import pandas as pd
            filename = f.filename or ''
            ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
            if ext in ('xlsx', 'xls'):
                df = pd.read_excel(f)
            elif ext == 'csv':
                df = pd.read_csv(f)
            else:
                flash('Unsupported file type. Please upload .xlsx, .xls or .csv.', 'danger')
                return redirect(url_for('parking_import'))

            expected_cols = ['name', 'district', 'tehsil', 'mouza', 'uc_name', 'capacity', 'address_location', 'remarks']
            missing = [c for c in expected_cols if c not in df.columns]
            if missing:
                flash(f'Missing required columns in file: {", ".join(missing)}', 'danger')
                return redirect(url_for('parking_import'))

            created = 0
            for _, row in df.iterrows():
                name = str(row.get('name') or '').strip()
                if not name:
                    continue
                # Skip if already exists with same name and district
                existing = ParkingStation.query.filter(
                    ParkingStation.name == name,
                    ParkingStation.district == (str(row.get('district') or '').strip() or None)
                ).first()
                if existing:
                    continue

                capacity_val = row.get('capacity')
                try:
                    capacity = int(capacity_val) if pd.notna(capacity_val) else 0
                except Exception:
                    capacity = 0

                p = ParkingStation(
                    name=name,
                    district=str(row.get('district') or '').strip() or None,
                    tehsil=str(row.get('tehsil') or '').strip() or None,
                    mouza=str(row.get('mouza') or '').strip() or None,
                    uc_name=str(row.get('uc_name') or '').strip() or None,
                    capacity=capacity or 0,
                    address_location=str(row.get('address_location') or '').strip() or None,
                    remarks=str(row.get('remarks') or '').strip() or None,
                )
                db.session.add(p)
                created += 1
            db.session.commit()
            if created:
                flash(f'{created} parking station(s) imported successfully.', 'success')
            else:
                flash('File processed but no new parking stations were imported (duplicates or empty rows).', 'info')
            return redirect(url_for('parking_list'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error importing parking locations: {str(e)}', 'danger')
    return render_template('parking_import.html', form=form)


@app.route('/parking/import/template')
def parking_import_template():
    """
    Downloadable template for parking import (open in Excel and fill).
    """
    headers = ['name', 'district', 'tehsil', 'mouza', 'uc_name', 'capacity', 'address_location', 'remarks']
    rows = [
        ['Central Parking A', 'Lahore', 'Model Town', 'Mouza 1', 'UC-1', 20, 'Near Main Road, Model Town', 'Example row 1'],
        ['Central Parking B', 'Lahore', 'Johar Town', 'Mouza 2', 'UC-2', 15, 'Near Hospital, Johar Town', 'Example row 2'],
    ]
    return generate_csv_response(headers, rows, filename='parking_import_template.csv')


@app.route('/parking/export')
def parking_export():
    """Export parking list (with optional search) to CSV/Excel."""
    search = request.args.get('search', '').strip()
    query = ParkingStation.query
    if search:
        query = query.filter(
            ParkingStation.name.ilike(f'%{search}%') |
            ParkingStation.district.ilike(f'%{search}%') |
            ParkingStation.address_location.ilike(f'%{search}%')
        )
    parkings = query.order_by(ParkingStation.id).all()
    headers = ['ID', 'Station Name', 'District', 'Tehsil', 'Capacity', 'Address/Location']
    rows = []
    for p in parkings:
        rows.append([
            p.id,
            p.name,
            p.district or '',
            p.tehsil or '',
            p.capacity,
            p.address_location or '',
        ])
    return generate_csv_response(headers, rows, filename='parking_stations.csv')


@app.route('/parking/print')
def parking_print():
    """Print-friendly view of parking stations (browser print to PDF)."""
    search = request.args.get('search', '').strip()
    query = ParkingStation.query
    if search:
        query = query.filter(
            ParkingStation.name.ilike(f'%{search}%') |
            ParkingStation.district.ilike(f'%{search}%') |
            ParkingStation.address_location.ilike(f'%{search}%')
        )
    parkings = query.order_by(ParkingStation.id).all()
    return render_template('parking_print.html', parkings=parkings, search=search)


@app.route('/parking/add', methods=['GET', 'POST'])
@app.route('/parking/edit/<int:id>', methods=['GET', 'POST'])
def parking_form(id=None):
    parking = ParkingStation.query.get_or_404(id) if id else None
    form = ParkingForm(obj=parking)
    if form.validate_on_submit():
        try:
            if not parking:
                parking = ParkingStation()
            # Normalize name (trim spaces)
            name_normalized = (form.name.data or '').strip()
            form.populate_obj(parking)
            parking.name = name_normalized

            # Duplicate name check (case-insensitive, ignore current record)
            existing_q = ParkingStation.query.filter(
                func.lower(ParkingStation.name) == name_normalized.lower()
            )
            if id:
                existing_q = existing_q.filter(ParkingStation.id != parking.id)
            if existing_q.first():
                flash('Parking Station with this name already exists. Please use a different name.', 'danger')
                return render_template('parking_form.html', form=form, title='Parking Station', back_url=url_for('parking_list'))

            if not id:
                db.session.add(parking)
            db.session.commit()
            flash('Parking Station saved successfully!', 'success')
            return redirect(url_for('parking_list'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error saving parking station: {str(e)}', 'danger')
    back_url = url_for('parking_list')
    return render_template('parking_form.html', form=form, title='Parking Station', back_url=back_url)


@app.route('/parking/delete/<int:id>', methods=['POST'])
def delete_parking(id):
    parking = ParkingStation.query.get_or_404(id)
    try:
        db.session.delete(parking)
        db.session.commit()
        flash('Parking Station deleted.', 'success')
    except Exception as e:
        flash(f'Error deleting parking station: {str(e)}', 'danger')
    return redirect(url_for('parking_list'))


# ────────────────────────────────────────────────
# Districts
# ────────────────────────────────────────────────
@app.route('/districts/')
def districts_list():
    search = request.args.get('search', '')
    query = District.query
    if search:
        query = query.filter(
            District.name.ilike(f'%{search}%') |
            District.province.ilike(f'%{search}%')
        )
    # Sort by ID as requested
    districts = query.order_by(District.id).all()
    return render_template('districts_list.html', districts=districts, search=search)


@app.route('/districts/export')
def districts_export():
    """Export districts list (with optional search) to CSV/Excel."""
    search = request.args.get('search', '').strip()
    query = District.query
    if search:
        query = query.filter(
            District.name.ilike(f'%{search}%') |
            District.province.ilike(f'%{search}%')
        )
    districts = query.order_by(District.id).all()
    headers = ['ID', 'District Name', 'Province/Region', 'Remarks', 'Created']
    rows = []
    for d in districts:
        rows.append([
            d.id,
            d.name,
            d.province or '',
            (d.remarks or '')[:200] if getattr(d, 'remarks', None) else '',
            d.created_at.strftime('%Y-%m-%d') if getattr(d, 'created_at', None) else '',
        ])
    return generate_csv_response(headers, rows, filename='districts.csv')


@app.route('/districts/print')
def districts_print():
    """Print-friendly view of districts list (browser print to PDF)."""
    search = request.args.get('search', '').strip()
    query = District.query
    if search:
        query = query.filter(
            District.name.ilike(f'%{search}%') |
            District.province.ilike(f'%{search}%')
        )
    districts = query.order_by(District.id).all()
    return render_template('districts_print.html', districts=districts, search=search)


@app.route('/district/add', methods=['GET', 'POST'])
@app.route('/district/edit/<int:id>', methods=['GET', 'POST'])
def district_form(id=None):
    district = District.query.get_or_404(id) if id else None
    form = DistrictForm(obj=district)
    if form.validate_on_submit():
        try:
            if not district:
                district = District()
            form.populate_obj(district)
            if not id:
                db.session.add(district)
            db.session.commit()
            flash('District saved successfully!', 'success')
            return redirect(url_for('districts_list'))
        except IntegrityError:
            db.session.rollback()
            flash('This district name already exists', 'danger')
        except Exception as e:
            db.session.rollback()
            flash(f'Error saving district: {str(e)}', 'danger')
    back_url = url_for('districts_list')
    return render_template('district_form.html', form=form, title='District', back_url=back_url)


@app.route('/district/delete/<int:id>', methods=['POST'])
def delete_district(id):
    district = District.query.get_or_404(id)
    try:
        db.session.delete(district)
        db.session.commit()
        flash('District deleted.', 'success')
    except Exception as e:
        flash(f'Error deleting district: {str(e)}', 'danger')
    return redirect(url_for('districts_list'))


@app.route('/project/<int:id>/export_vehicles')
def export_vehicles(id):
    project = Project.query.get_or_404(id)
    headers = ['ID', 'Vehicle No#', 'Model', 'Engine No', 'Chassis No', 'Vehicle Type', 'Phone No', 'Active Date']
    rows = []
    for v in project.vehicles:
        rows.append([
            v.id, v.vehicle_no, v.model, v.engine_no, v.chassis_no,
            v.vehicle_type, v.phone_no,
            v.active_date.strftime('%Y-%m-%d') if v.active_date else ''
        ])
    return generate_csv_response(headers, rows, f"vehicles_project_{id}.csv")


# ────────────────────────────────────────────────
# Assignment: Project → Company
# ────────────────────────────────────────────────
@app.route('/assign_project_to_company')
def assign_project_to_company():
    assigned_projects = Project.query.filter(Project.company_id.isnot(None)).order_by(Project.name).all()
    return render_template('assign_project_to_company.html', assigned_projects=assigned_projects)


@app.route('/assign_project_to_company/new', methods=['GET', 'POST'])
def assign_project_to_company_new():
    form = AssignProjectToCompanyForm()
    form.company_id.choices = [(c.id, c.name) for c in Company.query.order_by(Company.name).all()]
    form.project_id.choices = [(p.id, p.name) for p in Project.query.filter(Project.company_id.is_(None)).order_by(Project.name).all()]
    if form.validate_on_submit():
        try:
            company = Company.query.get_or_404(form.company_id.data)
            project = Project.query.get_or_404(form.project_id.data)
            if project.company_id:
                flash(f'Project "{project.name}" already assigned.', 'warning')
            else:
                project.company_id = company.id
                project.assign_date = form.assign_date.data
                project.assign_remarks = form.assign_remarks.data
                db.session.commit()
                flash(f'Project "{project.name}" assigned to "{company.name}".', 'success')
                return redirect(url_for('assign_project_to_company'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'danger')
    return render_template('assign_project_to_company_new.html', form=form)


@app.route('/assign_project_to_company/edit/<int:project_id>', methods=['GET', 'POST'])
def assign_project_to_company_edit(project_id):
    old_project = Project.query.get_or_404(project_id)
    form = EditProjectAssignmentForm()
    form.company_id.choices = [(c.id, c.name) for c in Company.query.order_by(Company.name).all()]
    available_projects = Project.query.filter((Project.company_id == None) | (Project.id == project_id)).all()
    form.project_id.choices = [(p.id, p.name) for p in available_projects]
    
    if request.method == 'GET':
        form.company_id.data = old_project.company_id
        form.project_id.data = old_project.id
        form.assign_date.data = old_project.assign_date
        form.assign_remarks.data = old_project.assign_remarks
    
    if form.validate_on_submit():
        try:
            new_project_id = int(form.project_id.data)
            if new_project_id != old_project.id:
                new_project = Project.query.get(new_project_id)
                new_project.company_id = form.company_id.data
                new_project.assign_date = form.assign_date.data
                new_project.assign_remarks = form.assign_remarks.data
                old_project.company_id = None
                old_project.assign_date = None
                old_project.assign_remarks = None
            else:
                old_project.company_id = form.company_id.data
                old_project.assign_date = form.assign_date.data
                old_project.assign_remarks = form.assign_remarks.data
            
            db.session.commit()
            flash('Assignment updated successfully.', 'success')
            return redirect(url_for('assign_project_to_company'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'danger')
    return render_template('assign_project_to_company_edit.html', form=form, project=old_project)


@app.route('/assign_project_to_company/desassign/<int:project_id>', methods=['POST'])
def desassign_project_from_company(project_id):
    project = Project.query.get_or_404(project_id)
    if not project.company_id:
        flash("This project is not assigned to any company.", 'info')
        return redirect(url_for('assign_project_to_company'))
    try:
        company_name = project.company.name if project.company else 'N/A'
        project.company_id = None
        project.assign_date = None
        project.assign_remarks = None
        db.session.commit()
        flash(f'Project "{project.name}" successfully desassigned from "{company_name}".', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error desassigning project: {str(e)}', 'danger')
    return redirect(url_for('assign_project_to_company'))


# ────────────────────────────────────────────────
# Assignment: Project → District
# ────────────────────────────────────────────────
@app.route('/assign_project_to_district')
def assign_project_to_district():
    results = db.session.query(
        Project, District, project_district.c.assign_date, project_district.c.remarks
    ).join(project_district, Project.id == project_district.c.project_id)\
     .join(District, District.id == project_district.c.district_id)\
     .order_by(Project.name).all()
    
    assigned_structured = []
    for proj, dist, a_date, a_remarks in results:
        link_data = {'assign_date': a_date, 'remarks': a_remarks}
        assigned_structured.append((link_data, proj, dist))
    
    return render_template('assign_project_to_district.html', assigned=assigned_structured)


@app.route('/assign_project_to_district/new', methods=['GET', 'POST'])
def assign_project_to_district_new():
    form = AssignProjectToDistrictForm()
    form.project_id.choices = [(p.id, p.name) for p in Project.query.order_by(Project.name).all()]
    form.district_id.choices = [(d.id, d.name) for d in District.query.order_by(District.name).all()]
    if form.validate_on_submit():
        try:
            project = Project.query.get_or_404(form.project_id.data)
            district = District.query.get_or_404(form.district_id.data)
            exists = db.session.query(project_district).filter_by(
                project_id=project.id, district_id=district.id
            ).first()
            if exists:
                flash(f'District "{district.name}" already assigned to "{project.name}".', 'warning')
            else:
                db.session.execute(
                    project_district.insert().values(
                        project_id=project.id,
                        district_id=district.id,
                        assign_date=form.assign_date.data,
                        remarks=form.remarks.data
                    )
                )
                db.session.commit()
                flash(f'District "{district.name}" assigned to "{project.name}".', 'success')
                return redirect(url_for('assign_project_to_district'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'danger')
    return render_template('assign_project_to_district_new.html', form=form)


@app.route('/assign_project_to_district/edit/<int:project_id>/<int:district_id>', methods=['GET', 'POST'])
def assign_project_to_district_edit(project_id, district_id):
    link = db.session.query(project_district).filter_by(project_id=project_id, district_id=district_id).first()
    if not link:
        flash("Assignment not found.", 'danger')
        return redirect(url_for('assign_project_to_district'))
    project = Project.query.get_or_404(project_id)
    district = District.query.get_or_404(district_id)
    form = AssignProjectToDistrictForm()
    form.project_id.choices = [(p.id, p.name) for p in Project.query.all()]
    form.district_id.choices = [(d.id, d.name) for d in District.query.all()]
    if request.method == 'GET':
        form.project_id.data = project_id
        form.district_id.data = district_id
        form.assign_date.data = link.assign_date
        form.remarks.data = link.remarks
    if form.validate_on_submit():
        try:
            db.session.execute(
                project_district.update().where(
                    project_district.c.project_id == project_id,
                    project_district.c.district_id == district_id
                ).values(
                    assign_date=form.assign_date.data,
                    remarks=form.remarks.data
                )
            )
            db.session.commit()
            flash("Assignment updated successfully!", "success")
            return redirect(url_for('assign_project_to_district'))
        except Exception as e:
            db.session.rollback()
            flash(f"Error: {str(e)}", "danger")
    return render_template('assign_project_to_district_edit.html', form=form, project=project, district=district)


@app.route('/assign_project_to_district/desassign/<int:project_id>/<int:district_id>', methods=['POST'])
def desassign_district_from_project(project_id, district_id):
    try:
        db.session.execute(
            project_district.delete().where(
                project_district.c.project_id == project_id,
                project_district.c.district_id == district_id
            )
        )
        db.session.commit()
        flash('District successfully desassigned from project.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error desassigning: {str(e)}', 'danger')
    return redirect(url_for('assign_project_to_district'))


# ────────────────────────────────────────────────
# Assignment: Vehicle → District
# ────────────────────────────────────────────────
@app.route('/assign_vehicle_to_district')
def assign_vehicle_to_district():
    search = request.args.get('search', '').strip()
    query = Vehicle.query.filter(Vehicle.district_id.isnot(None))
    if search:
        query = query.join(Project).join(District).filter(
            (Vehicle.vehicle_no.ilike(f'%{search}%')) |
            (Project.name.ilike(f'%{search}%')) |
            (District.name.ilike(f'%{search}%'))
        )
    assigned = query.all()
    return render_template('assign_vehicle_to_district_list.html', assigned_vehicles=assigned, search=search)


@app.route('/assign_vehicle_to_district/new', methods=['GET', 'POST'])
def assign_vehicle_to_district_new():
    form = AssignVehicleToDistrictForm()
    form.project_id.choices = [(p.id, p.name) for p in Project.query.order_by(Project.name).all()]
    form.vehicle_id.choices = [(v.id, f"{v.vehicle_no} ({v.model})") for v in Vehicle.query.all()]
    if request.method == 'POST':
        p_id = request.form.get('project_id', type=int)
        if p_id:
            project = Project.query.get(p_id)
            form.district_id.choices = [(d.id, d.name) for d in project.districts] if project else []
        else:
            form.district_id.choices = []

    if form.validate_on_submit():
        try:
            vehicle = Vehicle.query.get(form.vehicle_id.data)
            if vehicle:
                vehicle.project_id = form.project_id.data
                vehicle.district_id = form.district_id.data
                vehicle.assign_to_district_date = form.assign_date.data
                vehicle.assignment_remarks = form.remarks.data
                db.session.commit()
                flash(f"Vehicle {vehicle.vehicle_no} assigned successfully!", "success")
                return redirect(url_for('assign_vehicle_to_district'))
        except Exception as e:
            db.session.rollback()
            flash(f"Error: {str(e)}", "danger")
    return render_template('assign_vehicle_to_district_new.html', form=form)


@app.route('/get_districts_by_project/<int:project_id>')
def get_districts_by_project(project_id):
    districts = District.query.join(project_district) \
                     .filter(project_district.c.project_id == project_id) \
                     .order_by(District.name).all()
    return jsonify([{"id": d.id, "name": d.name} for d in districts])


@app.route('/assign_vehicle_to_district/desassign/<int:vehicle_id>', methods=['POST'])
def desassign_vehicle_from_district(vehicle_id):
    vehicle = Vehicle.query.get_or_404(vehicle_id)
    vehicle.district_id = None
    vehicle.assign_to_district_date = None
    vehicle.assignment_remarks = None
    db.session.commit()
    flash("Vehicle desassigned successfully.", "info")
    return redirect(url_for('assign_vehicle_to_district'))

@app.route('/company/report/<int:id>')
def company_report(id):
    company = Company.query.get_or_404(id)
    today = date.today().strftime('%d %b, %Y')
    return render_template('company_report.html', company=company, current_date=today)


@app.route('/assign_vehicle_to_district/edit/<int:vehicle_id>', methods=['GET', 'POST'])
def assign_vehicle_to_district_edit(vehicle_id):
    vehicle = Vehicle.query.get_or_404(vehicle_id)
    form = AssignVehicleToDistrictForm()
    form.project_id.choices = [(p.id, p.name) for p in Project.query.order_by(Project.name).all()]
    form.vehicle_id.choices = [(v.id, f"{v.vehicle_no} ({v.model})") for v in Vehicle.query.all()]
    
    current_p_id = request.form.get('project_id', type=int) or vehicle.project_id
    if current_p_id:
        project = Project.query.get(current_p_id)
        form.district_id.choices = [(d.id, d.name) for d in project.districts]
    else:
        form.district_id.choices = []

    if request.method == 'GET':
        form.project_id.data = vehicle.project_id
        form.district_id.data = vehicle.district_id
        form.vehicle_id.data = vehicle.id
        form.assign_date.data = vehicle.assign_to_district_date
        form.remarks.data = vehicle.assignment_remarks

    if form.validate_on_submit():
        try:
            if vehicle.id != form.vehicle_id.data:
                vehicle.project_id = None
                vehicle.district_id = None
                vehicle = Vehicle.query.get(form.vehicle_id.data)
            
            vehicle.project_id = form.project_id.data
            vehicle.district_id = form.district_id.data
            vehicle.assign_to_district_date = form.assign_date.data
            vehicle.assignment_remarks = form.remarks.data
            
            db.session.commit()
            flash("Assignment updated successfully!", "success")
            return redirect(url_for('assign_vehicle_to_district'))
        except Exception as e:
            db.session.rollback()
            flash(f"Error: {str(e)}", "danger")

    return render_template('assign_vehicle_to_district_edit.html', form=form, vehicle=vehicle)


@app.route('/get_parking_by_project/<int:project_id>')
def get_parking_by_project(project_id):
    project = Project.query.get(project_id)
    if not project:
        return jsonify([])
    stations = [{"id": p.id, "name": p.name} for p in project.parking_stations]
    return jsonify(stations)

@app.route('/assign_vehicle_to_parking/new', methods=['GET', 'POST'])
def assign_vehicle_to_parking_new():
    form = AssignVehicleToParkingForm()
    form.project_id.choices = [(0, "Select Project")] + [(p.id, p.name) for p in Project.query.all()]
    
    if request.method == 'POST':
        form.district_id.choices = [(int(request.form.get('district_id')), '')] if request.form.get('district_id') else []
        form.vehicle_id.choices = [(int(request.form.get('vehicle_id')), '')] if request.form.get('vehicle_id') else []
        form.parking_station_id.choices = [(int(request.form.get('parking_station_id')), '')] if request.form.get('parking_station_id') else []

    if form.validate_on_submit():
        try:
            vehicle = Vehicle.query.get(form.vehicle_id.data)
            vehicle.parking_station_id = form.parking_station_id.data
            vehicle.parking_assign_date = form.assign_date.data 
            vehicle.parking_remarks = form.remarks.data
            db.session.commit()
            flash("Vehicle assigned to parking successfully!", "success")
            return redirect(url_for('assign_vehicle_to_parking_list'))
        except Exception as e:
            db.session.rollback()
            flash(f"Error: {str(e)}", "danger")

    return render_template('assign_vehicle_to_parking_new.html', form=form)

@app.route('/get_vehicles_by_district/<int:project_id>/<int:district_id>')
def get_vehicles_by_district(project_id, district_id):
    vehicles = Vehicle.query.filter(
        Vehicle.project_id == project_id,
        Vehicle.district_id == district_id
    ).order_by(Vehicle.vehicle_no).all()
    return jsonify([{"id": v.id, "no": v.vehicle_no} for v in vehicles])

@app.route('/get_parking_by_district/<int:district_id>')
def get_parking_by_district(district_id):
    district = District.query.get(district_id)
    if not district:
        return jsonify([])
    stations = ParkingStation.query.filter_by(district=district.name).all()
    output = []
    for s in stations:
        occupied = Vehicle.query.filter_by(parking_station_id=s.id).count()
        available = s.capacity - occupied
        output.append({
            "id": s.id, 
            "name": f"{s.name} (Available: {available}/{s.capacity})",
            "is_full": available <= 0
        })
    return jsonify(output)


@app.route('/assign_vehicle_to_parking')
def assign_vehicle_to_parking_list():
    search = request.args.get('search', '').strip()
    query = Vehicle.query.filter(Vehicle.parking_station_id.isnot(None))
    if search:
        query = query.join(Project).join(ParkingStation).filter(
            (Vehicle.vehicle_no.ilike(f'%{search}%')) |
            (Project.name.ilike(f'%{search}%')) |
            (ParkingStation.name.ilike(f'%{search}%'))
        )
    parked_vehicles = query.all()
    return render_template('assign_vehicle_to_parking_list.html', parked_vehicles=parked_vehicles, search=search)

@app.route('/assign_vehicle_to_parking/desassign/<int:vehicle_id>', methods=['POST'])
def desassign_vehicle_from_parking(vehicle_id):
    vehicle = Vehicle.query.get_or_404(vehicle_id)
    station_name = vehicle.parking_station.name if vehicle.parking_station else "Parking"
    vehicle.parking_station_id = None
    db.session.commit()
    flash(f"Vehicle {vehicle.vehicle_no} removed from {station_name}.", "info")
    return redirect(url_for('assign_vehicle_to_parking_list'))

@app.route('/assign_vehicle_to_parking/edit/<int:vehicle_id>', methods=['GET', 'POST'])
def assign_vehicle_to_parking_edit(vehicle_id):
    vehicle = Vehicle.query.get_or_404(vehicle_id)
    form = AssignVehicleToParkingForm()
    
    form.project_id.choices = [(p.id, p.name) for p in Project.query.order_by(Project.name).all()]
    form.vehicle_id.choices = [(v.id, f"{v.vehicle_no} ({v.model})") for v in Vehicle.query.all()]
    
    p_id = request.form.get('project_id', type=int) or vehicle.project_id
    d_id = request.form.get('district_id', type=int) or vehicle.district_id

    if p_id:
        proj = Project.query.get(p_id)
        form.district_id.choices = [(d.id, d.name) for d in proj.districts.all()]
    else:
        form.district_id.choices = []

    if d_id:
        dist_obj = District.query.get(d_id)
        stations = ParkingStation.query.filter_by(district=dist_obj.name).all()
        form.parking_station_id.choices = [(s.id, s.name) for s in stations]
    else:
        form.parking_station_id.choices = []

    if request.method == 'GET':
        form.project_id.data = vehicle.project_id
        form.district_id.data = vehicle.district_id
        form.vehicle_id.data = vehicle.id
        form.parking_station_id.data = vehicle.parking_station_id
        form.assign_date.data = vehicle.parking_assign_date
        form.remarks.data = vehicle.parking_remarks 

    if form.validate_on_submit():
        try:
            new_ps_id = form.parking_station_id.data
            if new_ps_id != vehicle.parking_station_id:
                parking = ParkingStation.query.get(new_ps_id)
                occupied = Vehicle.query.filter_by(parking_station_id=new_ps_id).count()
                if occupied >= parking.capacity:
                    flash(f"Error: {parking.name} is full!", "danger")
                    return render_template('assign_vehicle_to_parking_edit.html', form=form, vehicle=vehicle)
            
            if vehicle.id != form.vehicle_id.data:
                vehicle.parking_station_id = None
                vehicle = Vehicle.query.get(form.vehicle_id.data)

            vehicle.parking_station_id = new_ps_id
            vehicle.parking_assign_date = form.assign_date.data
            vehicle.parking_remarks = form.remarks.data
            db.session.commit()
            flash("Parking assignment updated successfully!", "success")
            return redirect(url_for('assign_vehicle_to_parking_list'))
        except Exception as e:
            db.session.rollback()
            flash(f"Error: {str(e)}", "danger")

    return render_template('assign_vehicle_to_parking_edit.html', form=form, vehicle=vehicle)

@app.route('/get_unassigned_drivers')
def get_unassigned_drivers():
    drivers = Driver.query.filter(Driver.vehicle_id == None).all()
    return jsonify([{"id": d.id, "name": f"{d.name} ({d.driver_id})"} for d in drivers])


@app.route('/assign_driver_to_vehicle/new', methods=['GET', 'POST'])
def assign_driver_to_vehicle_new():
    form = AssignDriverToVehicleForm()
    form.project_id.choices = [(0, '-- Select Project --')] + [
        (p.id, p.name) for p in Project.query.order_by(Project.name).all()
    ]

    selected_project_id = None
    selected_district_id = None

    if request.method == 'POST':
        selected_project_id = form.project_id.data
        selected_district_id = form.district_id.data

        if selected_project_id and selected_project_id != 0:
            project = Project.query.get(selected_project_id)
            if project:
                form.district_id.choices = [(d.id, d.name) for d in project.districts]

                if selected_district_id and selected_district_id != 0:
                    vehicles = Vehicle.query.filter_by(
                        project_id=selected_project_id,
                        district_id=selected_district_id
                    ).all()
                    form.vehicle_id.choices = [
                        (v.id, f"{v.vehicle_no} – {v.model or 'N/A'}") for v in vehicles
                    ]

    unassigned_drivers = Driver.query.filter(Driver.vehicle_id.is_(None)).order_by(Driver.name).all()
    form.driver_id.choices = [(0, '-- Select Driver --')] + [
        (d.id, f"{d.name} ({d.driver_id})") for d in unassigned_drivers
    ]

    if form.validate_on_submit():
        vehicle = Vehicle.query.get(form.vehicle_id.data)
        driver = Driver.query.get(form.driver_id.data)

        if not vehicle or not driver:
            flash("Selected vehicle or driver not found.", "danger")
            return render_template('assign_driver_to_vehicle_new.html', form=form)

        current_count = Driver.query.filter_by(vehicle_id=vehicle.id).count()
        if current_count >= (vehicle.driver_capacity or 1):
            flash(f"Vehicle capacity ({vehicle.driver_capacity or 1}) already reached!", "danger")
            return render_template('assign_driver_to_vehicle_new.html', form=form)

        shift_taken = Driver.query.filter_by(vehicle_id=vehicle.id, shift=form.shift.data).first()
        if shift_taken:
            flash(f"{form.shift.data} shift already assigned to this vehicle!", "danger")
            return render_template('assign_driver_to_vehicle_new.html', form=form)

        try:
            driver.vehicle_id = vehicle.id
            driver.shift = form.shift.data
            driver.project_id = form.project_id.data
            driver.assign_date = form.assign_date.data
            driver.assign_remarks = form.remarks.data
            db.session.commit()
            flash(f"Driver {driver.name} assigned to {vehicle.vehicle_no} ({form.shift.data}) successfully!", "success")
            return redirect(url_for('assign_driver_to_vehicle_list'))
        except Exception as e:
            db.session.rollback()
            flash(f"Error saving assignment: {str(e)}", "danger")

    if not form.district_id.choices: form.district_id.choices = [(0, '-- Select District --')]
    if not form.vehicle_id.choices: form.vehicle_id.choices = [(0, '-- Select Vehicle --')]

    return render_template('assign_driver_to_vehicle_new.html', form=form)

@app.route('/get_vehicle_capacity_info/<int:vehicle_id>')
def get_vehicle_capacity_info(vehicle_id):
    vehicle = Vehicle.query.get(vehicle_id)
    if not vehicle:
        return jsonify({"error": "Not found"}), 404
    assigned_drivers = Driver.query.filter_by(vehicle_id=vehicle_id).all()
    occupied_shifts = [d.shift for d in assigned_drivers]
    return jsonify({
        "capacity": vehicle.driver_capacity or 1,
        "occupied_shifts": occupied_shifts,
        "available_morning": "Morning" not in occupied_shifts,
        "available_night": "Night" not in occupied_shifts
    })

@app.route('/assign_driver_to_vehicle')
def assign_driver_to_vehicle_list():
    search = request.args.get('search', '').strip()
    query = db.session.query(Driver, Vehicle).join(Vehicle, Driver.vehicle_id == Vehicle.id)
    if search:
        query = query.outerjoin(Project, Driver.project_id == Project.id).filter(
            (Driver.name.ilike(f'%{search}%')) |
            (Driver.driver_id.ilike(f'%{search}%')) |
            (Vehicle.vehicle_no.ilike(f'%{search}%')) |
            (Project.name.ilike(f'%{search}%'))
        )
    assigned_drivers = query.all()
    return render_template('assign_driver_to_vehicle_list.html', assigned_drivers=assigned_drivers, search=search)

@app.route('/assign_driver_to_vehicle/desassign/<int:driver_id>', methods=['POST'])
def desassign_driver_from_vehicle(driver_id):
    driver = Driver.query.get_or_404(driver_id)
    vehicle = Vehicle.query.get(driver.vehicle_id) if driver.vehicle_id else None
    vehicle_no = vehicle.vehicle_no if vehicle else "Vehicle"
    
    driver.vehicle_id = None
    driver.shift = None
    db.session.commit()
    
    flash(f"Driver '{driver.name}' successfully removed from Vehicle '{vehicle_no}'.", "info")
    return redirect(url_for('assign_driver_to_vehicle_list'))

@app.route('/assign_driver_to_vehicle/edit/<int:driver_id>', methods=['GET', 'POST'])
def assign_driver_to_vehicle_edit(driver_id):
    driver = Driver.query.get_or_404(driver_id)
    current_vehicle = Vehicle.query.get(driver.vehicle_id) if driver.vehicle_id else None

    form = AssignDriverToVehicleForm()
    form.project_id.choices = [(0, '-- Select Project --')] + [
        (p.id, p.name) for p in Project.query.order_by(Project.name).all()
    ]

    pid = form.project_id.data if request.method == 'POST' else driver.project_id
    did = form.district_id.data if request.method == 'POST' else (current_vehicle.district_id if current_vehicle else None)

    form.district_id.choices = [(0, '-- Select District --')]
    form.vehicle_id.choices = [(0, '-- Select Vehicle --')]

    if pid and pid != 0:
        proj = Project.query.get(pid)
        if proj:
            form.district_id.choices += [(d.id, d.name) for d in proj.districts]
            if did and did != 0:
                vehicles = Vehicle.query.filter_by(project_id=pid, district_id=did).all()
                form.vehicle_id.choices += [(v.id, f"{v.vehicle_no} – {v.model or 'N/A'}") for v in vehicles]

    current_driver_choice = (driver.id, f"{driver.name} ({driver.driver_id}) – Current")
    unassigned = Driver.query.filter(Driver.vehicle_id.is_(None), Driver.id != driver.id).order_by(Driver.name).all()
    unassigned_choices = [(d.id, f"{d.name} ({d.driver_id}) – Unassigned") for d in unassigned]
    form.driver_id.choices = [current_driver_choice] + unassigned_choices

    if request.method == 'GET':
        form.project_id.data    = driver.project_id
        form.district_id.data   = current_vehicle.district_id if current_vehicle else None
        form.vehicle_id.data    = driver.vehicle_id
        form.driver_id.data     = driver.id
        form.shift.data         = driver.shift
        form.assign_date.data   = driver.assign_date or date.today()
        form.remarks.data       = driver.assign_remarks or ''

    if form.validate_on_submit():
        vehicle = Vehicle.query.get(form.vehicle_id.data)
        if not vehicle:
            flash("Vehicle not found", "danger")
            return render_template('assign_driver_to_vehicle_edit.html', form=form, driver=driver)

        if vehicle.id != driver.vehicle_id:
            count = Driver.query.filter_by(vehicle_id=vehicle.id).count()
            if count >= (vehicle.driver_capacity or 1):
                flash("Target vehicle capacity reached", "danger")
                return render_template('assign_driver_to_vehicle_edit.html', form=form, driver=driver)

        conflict = Driver.query.filter(
            Driver.vehicle_id == vehicle.id, Driver.shift == form.shift.data, Driver.id != driver.id
        ).first()
        if conflict:
            flash(f"{form.shift.data} shift already taken", "danger")
            return render_template('assign_driver_to_vehicle_edit.html', form=form, driver=driver)

        try:
            new_driver_id = form.driver_id.data
            if new_driver_id != driver.id:
                driver.vehicle_id = None
                driver.shift = None
                driver.assign_date = None
                driver.assign_remarks = None

                new_driver = Driver.query.get(new_driver_id)
                if new_driver:
                    new_driver.vehicle_id = vehicle.id
                    new_driver.shift = form.shift.data
                    new_driver.project_id = form.project_id.data
                    new_driver.assign_date = form.assign_date.data
                    new_driver.assign_remarks = form.remarks.data
            else:
                driver.vehicle_id = vehicle.id
                driver.shift = form.shift.data
                driver.project_id = form.project_id.data
                driver.assign_date = form.assign_date.data
                driver.assign_remarks = form.remarks.data

            db.session.commit()
            flash("Assignment updated successfully", "success")
            return redirect(url_for('assign_driver_to_vehicle_list'))
        except Exception as e:
            db.session.rollback()
            flash(f"Error: {str(e)}", "danger")

    return render_template('assign_driver_to_vehicle_edit.html', form=form, driver=driver)


# Transfer Section - Project
@app.route('/project-transfers')
def project_transfers():
    transfers = ProjectTransfer.query.order_by(ProjectTransfer.transfer_date.desc()).all()
    return render_template('project_transfers.html', transfers=transfers)

@app.route('/project-transfer/new', methods=['GET', 'POST'])
def project_transfer_new():
    form = ProjectTransferForm()
    form.project_id.choices = [
        (p.id, f"{p.name} (Current: {p.company.name if p.company else 'Unassigned'})")
        for p in Project.query.filter(Project.company_id.isnot(None)).order_by(Project.name).all()
    ]
    form.new_company_id.choices = [(c.id, c.name) for c in Company.query.order_by(Company.name).all()]

    if form.validate_on_submit():
        try:
            project = Project.query.get_or_404(form.project_id.data)
            new_company = Company.query.get_or_404(form.new_company_id.data)
            if project.company_id == new_company.id:
                flash(f'Project already belongs to "{new_company.name}".', 'info')
                return redirect(url_for('project_transfer_new'))
            old_company = project.company
            old_company_name = old_company.name if old_company else 'Unassigned'
            
            transfer = ProjectTransfer(
                project_id=project.id,
                old_company_id=project.company_id,
                new_company_id=new_company.id,
                transfer_date=form.transfer_date.data,
                remarks=form.remarks.data
            )
            project.company_id = new_company.id
            project.assign_date = form.transfer_date.data
            project.assign_remarks = form.remarks.data
            db.session.add(transfer)
            db.session.commit()
            flash(f'Project "{project.name}" transferred from "{old_company_name}" to "{new_company.name}".', 'success')
            return redirect(url_for('project_transfers'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error during transfer: {str(e)}', 'danger')

    return render_template('project_transfer_new.html', form=form)

@app.route('/project-transfer/edit/<int:id>', methods=['GET', 'POST'])
def project_transfer_edit(id):
    transfer = ProjectTransfer.query.get_or_404(id)
    form = ProjectTransferForm(obj=transfer)
    form.project_id.choices = [(p.id, p.name) for p in Project.query.all()]
    form.new_company_id.choices = [(c.id, c.name) for c in Company.query.all()]
    if form.validate_on_submit():
        transfer.project_id = form.project_id.data
        transfer.new_company_id = form.new_company_id.data
        transfer.transfer_date = form.transfer_date.data
        transfer.remarks = form.remarks.data
        db.session.commit()
        flash('Transfer updated successfully!', 'success')
        return redirect(url_for('project_transfers'))
    return render_template('project_transfer_edit.html', form=form, transfer=transfer)

@app.route('/project-transfer/delete/<int:id>', methods=['POST'])
def project_transfer_delete(id):
    transfer_record = ProjectTransfer.query.get_or_404(id)
    try:
        db.session.delete(transfer_record)
        db.session.commit()
        flash('Transfer record permanently deleted.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting record: {str(e)}', 'danger')
    return redirect(url_for('project_transfers'))

# ==========================================
# VEHICLE TRANSFER ROUTES
# ==========================================

@app.route('/vehicle-transfers')
def vehicle_transfers():
    transfers = VehicleTransfer.query.order_by(VehicleTransfer.transfer_date.desc()).all()
    return render_template('vehicle_transfers.html', transfers=transfers)

@app.route('/vehicle-transfer/new', methods=['GET', 'POST'])
def vehicle_transfer_new():
    form = VehicleTransferForm()
    all_projects = [(p.id, p.name) for p in Project.query.order_by(Project.name).all()]
    form.from_project_id.choices = [(0, '-- Select From Project --')] + all_projects
    form.new_project_id.choices = [(0, '-- Select New Project --')] + all_projects

    if request.method == 'POST':
        form.from_district_id.choices = [(int(request.form.get('from_district_id') or 0), '')]
        form.vehicle_id.choices = [(int(request.form.get('vehicle_id') or 0), '')]
        form.new_district_id.choices = [(int(request.form.get('new_district_id') or 0), '')]
        form.new_parking_id.choices = [(int(request.form.get('new_parking_id') or 0), '')]

    if form.validate_on_submit():
        if form.vehicle_id.data == 0 or form.new_project_id.data == 0 or form.new_district_id.data == 0:
            flash("Vehicle, New Project, and New District are required fields.", "danger")
            return redirect(url_for('vehicle_transfer_new'))

        try:
            vehicle = Vehicle.query.get_or_404(form.vehicle_id.data)
            old_p_id = vehicle.project_id
            old_d_id = vehicle.district_id
            old_park_id = vehicle.parking_station_id
            
            new_p_id = form.new_project_id.data
            new_d_id = form.new_district_id.data
            new_park_id = form.new_parking_id.data if form.new_parking_id.data != 0 else None

            if old_p_id == new_p_id and old_d_id == new_d_id and old_park_id == new_park_id:
                flash("Cannot transfer! The vehicle is already at this exact Location.", "danger")
                return redirect(url_for('vehicle_transfer_new'))

            if new_park_id:
                parking = ParkingStation.query.get(new_park_id)
                occupied = Vehicle.query.filter_by(parking_station_id=new_park_id).count()
                if occupied >= parking.capacity:
                    flash(f"Transfer Failed! '{parking.name}' is FULL.", "danger")
                    return redirect(url_for('vehicle_transfer_new'))

            drivers_freed = 0
            if old_p_id != new_p_id or old_d_id != new_d_id:
                attached_drivers = Driver.query.filter_by(vehicle_id=vehicle.id).all()
                for d in attached_drivers:
                    d.vehicle_id = None
                    d.shift = None
                    drivers_freed += 1
            
            transfer = VehicleTransfer(
                vehicle_id=vehicle.id,
                old_project_id=old_p_id, old_district_id=old_d_id, old_parking_id=old_park_id,
                new_project_id=new_p_id, new_district_id=new_d_id, new_parking_id=new_park_id,
                transfer_date=form.transfer_date.data,
                remarks=form.remarks.data
            )
            
            vehicle.project_id = new_p_id
            vehicle.district_id = new_d_id
            vehicle.parking_station_id = new_park_id

            db.session.add(transfer)
            db.session.commit()

            msg = f"Vehicle {vehicle.vehicle_no} transferred successfully."
            if drivers_freed > 0:
                msg += f" {drivers_freed} driver(s) were unassigned."
                flash(msg, 'warning')
            else:
                flash(msg, 'success')

            return redirect(url_for('vehicle_transfers'))

        except Exception as e:
            db.session.rollback()
            flash(f"Error: {str(e)}", "danger")

    if not form.from_district_id.choices: form.from_district_id.choices = [(0, '-- Select District --')]
    if not form.vehicle_id.choices: form.vehicle_id.choices = [(0, '-- Select Vehicle --')]
    if not form.new_district_id.choices: form.new_district_id.choices = [(0, '-- Select District --')]
    if not form.new_parking_id.choices: form.new_parking_id.choices = [(0, '-- Select Parking --')]

    return render_template('vehicle_transfer_new.html', form=form)

@app.route('/vehicle-transfer/edit/<int:id>', methods=['GET', 'POST'])
def vehicle_transfer_edit(id):
    transfer = VehicleTransfer.query.get_or_404(id)
    vehicle = transfer.vehicle
    form = EditVehicleTransferForm()

    form.new_project_id.choices = [(0, '-- Select Project --')] + [(p.id, p.name) for p in Project.query.order_by(Project.name).all()]

    pid = form.new_project_id.data if request.method == 'POST' else transfer.new_project_id
    did = form.new_district_id.data if request.method == 'POST' else transfer.new_district_id
    
    if pid and pid != 0:
        proj = Project.query.get(pid)
        form.new_district_id.choices = [(0, '-- Select District --')] + [(d.id, d.name) for d in proj.districts]
    else:
        form.new_district_id.choices = [(0, '-- Select District --')]

    if did and did != 0:
        dist = District.query.get(did)
        stations = ParkingStation.query.filter_by(district=dist.name).all()
        form.new_parking_id.choices = [(0, '-- No Parking --')] + [(s.id, s.name) for s in stations]
    else:
        form.new_parking_id.choices = [(0, '-- No Parking --')]

    if request.method == 'GET':
        form.new_project_id.data = transfer.new_project_id
        form.new_district_id.data = transfer.new_district_id
        form.new_parking_id.data = transfer.new_parking_id if transfer.new_parking_id else 0
        form.transfer_date.data = transfer.transfer_date
        form.remarks.data = transfer.remarks

    if form.validate_on_submit():
        new_park_id = form.new_parking_id.data if form.new_parking_id.data != 0 else None
        
        if new_park_id and new_park_id != transfer.new_parking_id:
            parking = ParkingStation.query.get(new_park_id)
            occupied = Vehicle.query.filter(Vehicle.parking_station_id == new_park_id, Vehicle.id != vehicle.id).count()
            if occupied >= parking.capacity:
                flash(f"Cannot update! '{parking.name}' is FULL.", "danger")
                return render_template('vehicle_transfer_edit.html', form=form, transfer=transfer)

        try:
            transfer.new_project_id = form.new_project_id.data
            transfer.new_district_id = form.new_district_id.data
            transfer.new_parking_id = new_park_id
            transfer.transfer_date = form.transfer_date.data
            transfer.remarks = form.remarks.data
            
            vehicle.project_id = form.new_project_id.data
            vehicle.district_id = form.new_district_id.data
            vehicle.parking_station_id = new_park_id

            db.session.commit()
            flash("Transfer record updated successfully.", "success")
            return redirect(url_for('vehicle_transfers'))
        except Exception as e:
            db.session.rollback()
            flash(f"Error: {str(e)}", "danger")

    return render_template('vehicle_transfer_edit.html', form=form, transfer=transfer)

@app.route('/vehicle-transfer/delete/<int:id>', methods=['POST'])
def vehicle_transfer_delete(id):
    transfer_record = VehicleTransfer.query.get_or_404(id)
    try:
        db.session.delete(transfer_record)
        db.session.commit()
        flash('Transfer record permanently deleted.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting record: {str(e)}', 'danger')
    return redirect(url_for('vehicle_transfers'))

@app.route('/get_vehicle_current_info/<int:vehicle_id>')
def get_vehicle_current_info(vehicle_id):
    v = Vehicle.query.get(vehicle_id)
    if not v:
        return jsonify({"parking_name": "Not Found"})
    if v.parking_station:
        return jsonify({"parking_name": v.parking_station.name})
    else:
        return jsonify({"parking_name": "No Parking Assigned (Free)"})

# ==========================================
# DRIVER TRANSFER ROUTES & APIs
# ==========================================

@app.route('/get_assigned_drivers/<int:project_id>')
def get_assigned_drivers(project_id):
    if project_id == 0:
        drivers = Driver.query.filter(Driver.vehicle_id.isnot(None)).all()
    else:
        drivers = Driver.query.filter(Driver.vehicle_id.isnot(None), Driver.project_id == project_id).all()
    return jsonify([{"id": d.id, "name": f"{d.name} ({d.driver_id})"} for d in drivers])

@app.route('/get_driver_current_info/<int:driver_id>')
def get_driver_current_info(driver_id):
    d = Driver.query.get(driver_id)
    if not d or not d.vehicle: return jsonify({"info": "Not Assigned"})
    return jsonify({"info": f"{d.vehicle.vehicle_no} ({d.shift} Shift)"})

@app.route('/get_available_shifts/<int:vehicle_id>')
def get_available_shifts(vehicle_id):
    v = Vehicle.query.get(vehicle_id)
    if not v: return jsonify([])
    
    cap = v.driver_capacity or 1
    driver_id_to_exclude = request.args.get('exclude_driver_id', type=int)
    
    query = Driver.query.filter_by(vehicle_id=v.id)
    if driver_id_to_exclude:
        query = query.filter(Driver.id != driver_id_to_exclude)
        
    existing_shifts = [d.shift for d in query.all()]
    
    shifts = []
    if cap == 1:
        if "Morning" not in existing_shifts: shifts.append({"id": "Morning", "name": "Morning"})
    else:
        if "Morning" not in existing_shifts: shifts.append({"id": "Morning", "name": "Morning"})
        if "Night" not in existing_shifts: shifts.append({"id": "Night", "name": "Night"})
        
    return jsonify(shifts)

@app.route('/driver-transfers')
def driver_transfers():
    transfers = DriverTransfer.query.order_by(DriverTransfer.transfer_date.desc()).all()
    return render_template('driver_transfers.html', transfers=transfers)

@app.route('/driver-transfer/new', methods=['GET', 'POST'])
def driver_transfer_new():
    form = DriverTransferForm()

    form.from_project_id.choices = [(0, '-- Select Project --')] + [
        (p.id, p.name) for p in Project.query.order_by(Project.name).all()
    ]
    form.new_project_id.choices = form.from_project_id.choices[:] 

    form.new_shift.choices = [
        ('', '-- Select Shift --'),
        ('Morning', 'Morning'),
        ('Night', 'Night')
    ]

    form.from_district_id.choices = [(0, '-- Select Project first --')]
    form.from_vehicle_id.choices  = [(0, '-- Select District first --')]
    form.driver_id.choices        = [(0, '-- Select Vehicle first --')]
    form.new_district_id.choices  = [(0, '-- Select Project first --')]
    form.new_vehicle_id.choices   = [(0, '-- Select District first --')]

    if request.method == 'POST':
        # Dynamically populate choices based on selections (for validation & UX)
        if form.from_project_id.data and form.from_project_id.data != 0:
            districts = District.query.join(project_district).filter(
                project_district.c.project_id == form.from_project_id.data
            ).order_by(District.name).all()
            form.from_district_id.choices = [(0, '-- Select District --')] + [(d.id, d.name) for d in districts]

            if form.from_district_id.data and form.from_district_id.data != 0:
                vehicles = Vehicle.query.filter_by(
                    project_id=form.from_project_id.data, district_id=form.from_district_id.data
                ).order_by(Vehicle.vehicle_no).all()
                form.from_vehicle_id.choices = [(0, '-- Select Vehicle --')] + [(v.id, v.vehicle_no) for v in vehicles]

                if form.from_vehicle_id.data and form.from_vehicle_id.data != 0:
                    drivers = Driver.query.filter_by(vehicle_id=form.from_vehicle_id.data).order_by(Driver.name).all()
                    form.driver_id.choices = [(0, '-- Select Driver --')] + [(d.id, f"{d.name} ({d.driver_id})") for d in drivers]

        if form.new_project_id.data and form.new_project_id.data != 0:
            districts = District.query.join(project_district).filter(
                project_district.c.project_id == form.new_project_id.data
            ).order_by(District.name).all()
            form.new_district_id.choices = [(0, '-- Select District --')] + [(d.id, d.name) for d in districts]

            if form.new_district_id.data and form.new_district_id.data != 0:
                vehicles = Vehicle.query.filter_by(
                    project_id=form.new_project_id.data, district_id=form.new_district_id.data
                ).order_by(Vehicle.vehicle_no).all()
                form.new_vehicle_id.choices = [(0, '-- Select Vehicle --')] + [(v.id, v.vehicle_no) for v in vehicles]

    if form.validate_on_submit():
        try:
            driver = Driver.query.get_or_404(form.driver_id.data)
            new_vehicle = Vehicle.query.get_or_404(form.new_vehicle_id.data)

            # Capacity check
            existing_count = Driver.query.filter(
                Driver.vehicle_id == new_vehicle.id,
                Driver.id != driver.id
            ).count()

            if existing_count >= (new_vehicle.driver_capacity or 1):
                flash(f"Vehicle {new_vehicle.vehicle_no} is full (capacity reached).", "danger")
                return render_template('driver_transfer_new.html', form=form)

            # Shift already taken check
            shift_taken = Driver.query.filter(
                Driver.vehicle_id == new_vehicle.id,
                Driver.shift == form.new_shift.data,
                Driver.id != driver.id
            ).first()

            if shift_taken:
                flash(f"Shift '{form.new_shift.data}' already assigned in vehicle {new_vehicle.vehicle_no}.", "danger")
                return render_template('driver_transfer_new.html', form=form)

            # Create transfer record with OLD and NEW values
            transfer = DriverTransfer(
                driver_id=driver.id,
                old_project_id=driver.project_id,
                old_vehicle_id=driver.vehicle_id,
                old_shift=driver.shift,
                old_district_id=driver.district_id,                    # ← Yeh zaroori hai revert ke liye
                new_project_id=form.new_project_id.data,
                new_vehicle_id=new_vehicle.id,
                new_shift=form.new_shift.data,
                new_district_id=form.new_district_id.data,             # ← Yeh bhi zaroori
                transfer_date=form.transfer_date.data,
                remarks=form.remarks.data
            )
            db.session.add(transfer)

            # Update driver's current assignment
            driver.project_id = form.new_project_id.data
            driver.district_id = form.new_district_id.data
            driver.vehicle_id = new_vehicle.id
            driver.shift = form.new_shift.data

            # Optional: agar driver_district string field bhi update karna chahte ho
            # new_dist = District.query.get(form.new_district_id.data)
            # driver.driver_district = new_dist.name if new_dist else driver.driver_district

            db.session.commit()

            flash(f"Driver {driver.name} transferred successfully to {new_vehicle.vehicle_no} ({form.new_shift.data} shift).", "success")
            return redirect(url_for('driver_transfers'))

        except Exception as e:
            db.session.rollback()
            flash(f"Error during transfer: {str(e)}", "danger")

    return render_template('driver_transfer_new.html', form=form)

@app.route('/driver-transfer/delete/<int:id>', methods=['POST'])
def driver_transfer_delete(id):
    transfer = DriverTransfer.query.get_or_404(id)
    driver = transfer.driver

    if not driver:
        flash("Associated driver record not found.", "danger")
        return redirect(url_for('driver_transfers'))

    # Optional: Sirf sabse latest transfer hi delete hone de (safety ke liye)
    latest_transfer = DriverTransfer.query.filter_by(driver_id=driver.id)\
                                           .order_by(DriverTransfer.created_at.desc())\
                                           .first()

    if transfer.id != latest_transfer.id:
        flash("Only the most recent transfer can be deleted to avoid breaking history.", "warning")
        return redirect(url_for('driver_transfers'))

    try:
        # Revert driver to OLD values from this transfer
        driver.project_id = transfer.old_project_id
        driver.vehicle_id = transfer.old_vehicle_id
        driver.shift = transfer.old_shift
        driver.district_id = transfer.old_district_id

        # Optional: string district field revert (agar use kar rahe ho)
        # if transfer.old_district:
        #     driver.driver_district = transfer.old_district.name
        # else:
        #     driver.driver_district = None

        db.session.delete(transfer)
        db.session.commit()

        flash(f"Transfer record deleted and driver {driver.name} reverted to previous assignment.", "success")

    except Exception as e:
        db.session.rollback()
        flash(f"Error while reverting transfer: {str(e)}", "danger")

    return redirect(url_for('driver_transfers'))

# API: Ek makhsoos gaari (Vehicle) par assign drivers nikalna
@app.route('/get_drivers_by_vehicle/<int:vehicle_id>')
def get_drivers_by_vehicle(vehicle_id):
    drivers = Driver.query.filter_by(vehicle_id=vehicle_id).all()
    return jsonify([{
        "id": d.id,
        "name": f"{d.name} ({d.driver_id or 'No ID'}) - {d.shift or 'No Shift'}"
    } for d in drivers])


@app.route('/driver/job-left/new', methods=['GET', 'POST'])
def driver_job_left_new():
    form = DriverJobLeftForm()
    
    # Page load pe project choices set karo (dropdown populate ke liye)
    projects = Project.query.order_by(Project.name).all()
    form.project_id.choices = [(0, '-- Select Project --')] + [(p.id, p.name) for p in projects]
    
    # Default empty choices cascading fields ke liye (validation crash na ho)
    form.district_id.choices = [(0, '-- Select District --')]
    form.vehicle_id.choices = [(0, '-- Select Vehicle --')]
    form.driver_id.choices = [(0, '-- Select Driver --')]
    
    if form.is_submitted():  # Yeh check POST request pe chalega (submit hone pe)
        # Step 1: Selected values lo
        selected_project = form.project_id.data
        selected_district = form.district_id.data
        selected_vehicle = form.vehicle_id.data
        
        # Step 2: Choices ko re-build karo taake validation pass ho
        # District choices re-create based on selected project
        if selected_project and selected_project != 0:
            districts = District.query.join(project_district)\
                               .filter(project_district.c.project_id == selected_project)\
                               .order_by(District.name).all()
            form.district_id.choices = [(0, '-- Select District --')] + [(d.id, d.name) for d in districts]
        
        # Vehicle choices re-create based on project + district
        if selected_project != 0 and selected_district != 0:
            vehicles = Vehicle.query.filter(
                Vehicle.project_id == selected_project,
                Vehicle.district_id == selected_district
            ).order_by(Vehicle.vehicle_no).all()
            form.vehicle_id.choices = [(0, '-- Select Vehicle --')] + [(v.id, v.vehicle_no) for v in vehicles]
        
        # Driver choices re-create based on vehicle
        if selected_vehicle != 0:
            drivers = Driver.query.filter_by(vehicle_id=selected_vehicle).all()
            form.driver_id.choices = [(0, '-- Select Driver --')] + \
                                    [(d.id, f"{d.name} ({d.driver_id or 'No ID'}) - {d.shift or 'No Shift'}") for d in drivers]
        
        # Ab form ko dobara validate karo (choices updated hain)
        if form.validate():
            driver = Driver.query.get(form.driver_id.data)
            
            if not driver:
                flash("Invalid driver selected.", "danger")
            elif driver.vehicle_id != form.vehicle_id.data:
                flash("This driver is not assigned to the selected vehicle.", "danger")
            else:
                reason = form.reason.data
                if reason == 'Other':
                    reason = form.other_reason.data.strip() or 'Other (unspecified)'
                
                # Save to database
                status_change = DriverStatusChange(
                    driver_id=driver.id,
                    action_type='left',
                    reason=reason,
                    change_date=form.leave_date.data,
                    remarks=form.remarks.data,
                    left_project_id=form.project_id.data if form.project_id.data != 0 else None,
                    left_district_id=form.district_id.data if form.district_id.data != 0 else None,
                    left_vehicle_id=form.vehicle_id.data if form.vehicle_id.data != 0 else None,
                    left_shift=driver.shift
                )
                db.session.add(status_change)
                
                # Clear driver's current assignment
                driver.vehicle_id = None
                driver.shift = None
                driver.district_id = None  # agar Driver model mein hai
                driver.project_id = None  # <-- Ye bhi add karein, taake project se bhi hat jaye
                driver.status = 'Left'    # <-- Ye naya add karein

                db.session.commit()
                
                flash(f"Driver {driver.name} successfully marked as Job Left!", "success")
                return redirect(url_for('drivers_list'))
        else:
            flash("Please correct the errors below.", "danger")
    
    return render_template('driver_job_left_new.html', form=form)

# --- 1. REJOIN MAIN ROUTE ---
@app.route('/driver/rejoin/new', methods=['GET', 'POST'])
def driver_rejoin_new():
    form = DriverRejoinForm()
    projects = Project.query.order_by(Project.name).all()
    
    # 1. Base Choices (Initial)
    form.project_id.choices = [(0, '-- Select Project --')] + [(p.id, p.name) for p in projects]
    form.district_id.choices = [(0, '-- Select District --')]
    form.vehicle_id.choices = [(0, '-- Select Vehicle --')]
    form.driver_id.choices = [(0, '-- Select Driver --')]
    form.shift.choices = [('', 'Select'), ('Morning', 'Morning'), ('Night', 'Night')]

    # 2. Re-populate choices BEFORE validation if it's a POST request
    if request.method == 'POST':
        # Hum posted data se choices dubara bana rahe hain taake validation pass ho
        if form.project_id.data and form.project_id.data != 0:
            p_id = form.project_id.data
            districts = District.query.join(project_district).filter(project_district.c.project_id == p_id).all()
            form.district_id.choices = [(d.id, d.name) for d in districts]
            
            d_id = request.form.get('district_id', type=int)
            if d_id:
                vehicles = Vehicle.query.filter_by(district_id=d_id).all()
                form.vehicle_id.choices = [(v.id, v.vehicle_no) for v in vehicles]
                
                v_id = request.form.get('vehicle_id', type=int)
                if v_id:
                    # 'Left' drivers load karna taake validation pass ho sake
                    left_drivers = Driver.query.filter_by(status='Left').all()
                    form.driver_id.choices = [(d.id, d.name) for d in left_drivers]

    # 3. Validation and Saving
    if form.validate_on_submit():
        try:
            driver = Driver.query.get(form.driver_id.data)
            if not driver:
                flash("Error: Selected driver not found.", "danger")
                return redirect(url_for('driver_rejoin_new'))

            # Updates
            driver.status = 'Active'
            driver.project_id = form.project_id.data
            driver.district_id = form.district_id.data
            driver.vehicle_id = form.vehicle_id.data
            driver.shift = form.shift.data

            rejoin_log = DriverStatusChange(
                driver_id=driver.id,
                action_type='rejoin',
                change_date=form.rejoin_date.data,
                remarks=form.remarks.data,
                new_project_id=form.project_id.data,
                new_district_id=form.district_id.data,
                new_vehicle_id=form.vehicle_id.data,
                new_shift=form.shift.data
            )
            db.session.add(rejoin_log)
            db.session.commit()
            
            flash(f"Success! Driver {driver.name} is now ACTIVE.", "success")
            return redirect(url_for('driver_rejoin_list'))
            
        except Exception as e:
            db.session.rollback()
            print(f"DATABASE ERROR: {str(e)}") # Terminal mein check karein
            flash("Database error occurred.", "danger")
    else:
        # AGAR VALIDATION FAIL HO TO TERMINAL MEIN DEKHO KYUN FAIL HUI
        if request.method == 'POST':
            print("Form Validation Errors:", form.errors)
            flash("Form validation failed. Check terminal for details.", "warning")

    return render_template('driver_rejoin_new.html', form=form, projects=projects)

@app.route('/driver/job-left/list')
def driver_job_left_list():
    # Saare job left records order by date descending
    left_records = DriverStatusChange.query.filter_by(action_type='left') \
                                 .order_by(DriverStatusChange.change_date.desc()) \
                                 .all()
    
    return render_template('driver_job_left_list.html', 
                           records=left_records,
                           title="Driver Job Left History")

@app.route('/driver-job-left/view/<int:id>')
def driver_job_left_view(id):
    record = DriverStatusChange.query.get_or_404(id)
    return render_template('driver_job_left_view.html', record=record)

# 2. EDIT ROUTE
@app.route('/driver-job-left/edit/<int:id>', methods=['GET', 'POST'])
def driver_job_left_edit(id):
    record = DriverStatusChange.query.get_or_404(id)
    
    # Form initialize karein aur populate karein
    form = DriverJobLeftForm()
    
    # Dropdowns ko populate karein
    form.project_id.choices = [(p.id, p.name) for p in Project.query.all()]
    form.district_id.choices = [(d.id, d.name) for d in District.query.all()]
    form.vehicle_id.choices = [(v.id, v.vehicle_no) for v in Vehicle.query.all()]
    form.driver_id.choices = [(d.id, d.name) for d in Driver.query.all()]

    if request.method == 'GET':
        # Pre-fill data from DB to Form
        form.project_id.data = record.left_project_id
        form.district_id.data = record.left_district_id
        form.vehicle_id.data = record.left_vehicle_id
        form.driver_id.data = record.driver_id
        form.reason.data = record.reason
        form.leave_date.data = record.change_date
        form.remarks.data = record.remarks

    if form.validate_on_submit():
        record.reason = form.reason.data
        if form.reason.data == 'Other':
            record.reason = form.other_reason.data
        record.change_date = form.leave_date.data
        record.remarks = form.remarks.data
        record.driver_id = form.driver_id.data
        record.left_project_id = form.project_id.data
        record.left_district_id = form.district_id.data
        record.left_vehicle_id = form.vehicle_id.data
        
        db.session.commit()
        flash('Record updated successfully!', 'success')
        return redirect(url_for('driver_job_left_view', id=record.id))
        
    return render_template('driver_job_left_edit.html', record=record, form=form)

# 3. DELETE ROUTE
@app.route('/driver-job-left/delete/<int:id>', methods=['POST'])
def driver_job_left_delete(id):
    record = DriverStatusChange.query.get_or_404(id)
    
    try:
        # Optionally Revert Driver Status back to Active
        driver = record.driver
        if driver:
            driver.status = 'Active'
            # Agar assignments wapas deni hain:
            # driver.project_id = record.left_project_id
            # driver.district_id = record.left_district_id
            # driver.vehicle_id = record.left_vehicle_id
            # driver.shift = record.left_shift
            
        db.session.delete(record)
        db.session.commit()
        flash('Job left record has been deleted and driver activated.', 'warning')
    except Exception as e:
        db.session.rollback()
        flash('Error deleting record.', 'danger')
        
    return redirect(url_for('driver_job_left_list'))

# routes.py mein ye function add karein
@app.route('/driver/rejoin/list')
def driver_rejoin_list():
    # Sirf 'rejoin' action waale records uthayein
    rejoin_records = DriverStatusChange.query.filter_by(action_type='rejoin') \
                                     .order_by(DriverStatusChange.change_date.desc()) \
                                     .all()
    
    return render_template('driver_rejoin_list.html', 
                           records=rejoin_records, 
                           title="Driver Rejoin History")

@app.route('/api/get_left_drivers_by_vehicle/<int:vid>')
def get_left_drivers_by_vehicle(vid):
    # Query: Wo drivers jinka status 'Left' hai 
    # Aur unka DriverStatusChange table mein aakhri record is Vehicle ID ka ho
    from sqlalchemy import func
    
    drivers = db.session.query(Driver).join(DriverStatusChange).filter(
        func.lower(Driver.status) == 'left',
        DriverStatusChange.left_vehicle_id == vid,
        DriverStatusChange.action_type == 'left'
    ).all()

    results = []
    for d in drivers:
        results.append({
            'id': d.id,
            'name': d.name,
            'driver_id': d.driver_id
        })
    return jsonify(results)

@app.route('/api/check_vehicle_shifts/<int:vid>')
def check_vehicle_shifts(vid):
    vehicle = Vehicle.query.get_or_404(vid)
    active_drivers = Driver.query.filter_by(vehicle_id=vid, status='Active').all()
    occupied_shifts = [d.shift for d in active_drivers if d.shift]
    
    return jsonify({
        'capacity': vehicle.driver_capacity or 1,
        'occupied_shifts': occupied_shifts
    })

@app.route('/driver_rejoin_view/<int:id>')
def driver_rejoin_view(id):
    """
    Yeh function driver ke rejoin hone ki report dikhayega.
    'id' woh unique ID hai jo rejoin hone ke waqt history table mein save hui thi.
    """
    # 1. Database se record dhoondo, agar nahi milta toh 404 error dikhao
    # 'DriverRejoin' ki jagah apne Model ka naam likhein agar mukhtalif hai
    record = DriverStatusChange.query.get_or_404(id)
    
    # 2. Record ko 'driver_rejoin_view.html' template par bhej do
    return render_template('driver_rejoin_view.html', record=record)


# ────────────────────────────────────────────────
# Driver Attendance
# ────────────────────────────────────────────────
@app.route('/driver-attendance/')
def driver_attendance_list():
    form = DriverAttendanceFilterForm()
    form.project_id.choices = [(0, '-- All Projects --')] + [(p.id, p.name) for p in Project.query.order_by(Project.name).all()]
    view_date = date.today()
    project_id = request.args.get('project_id', type=int)
    district_id = request.args.get('district_id', type=int)
    search = (request.args.get('search') or '').strip()
    if request.args.get('date'):
        view_date = parse_date(request.args.get('date')) or view_date
    # Pre-fill form from query params
    form.attendance_date.data = view_date
    form.project_id.data = project_id if project_id else 0
    # District choices: only when project selected, from project_district
    if project_id and project_id != 0:
        districts = District.query.join(project_district).filter(project_district.c.project_id == project_id).order_by(District.name).all()
        form.district_id.choices = [(0, '-- All Districts --')] + [(d.id, d.name) for d in districts]
    else:
        form.district_id.choices = [(0, '-- All Districts --')]
    form.district_id.data = district_id if district_id else 0
    query = DriverAttendance.query.filter_by(attendance_date=view_date)
    if project_id:
        query = query.filter(DriverAttendance.project_id == project_id)
    attendance_list = query.order_by(DriverAttendance.driver_id).all()
    drivers_query = Driver.query.filter_by(status='Active')
    if project_id:
        drivers_query = drivers_query.filter(Driver.project_id == project_id)
    if district_id:
        drivers_query = drivers_query.filter(Driver.district_id == district_id)
    if search:
        q = f'%{search}%'
        drivers_query = drivers_query.outerjoin(Vehicle, Driver.vehicle_id == Vehicle.id).filter(
            db.or_(
                Driver.name.ilike(q),
                Driver.driver_id.ilike(q),
                Vehicle.vehicle_no.ilike(q),
            )
        )
    drivers = drivers_query.order_by(Driver.name).all()
    by_driver = {a.driver_id: a for a in attendance_list}
    return render_template('driver_attendance_list.html', form=form, view_date=view_date, drivers=drivers, by_driver=by_driver, project_id=project_id, district_id=district_id, search=search)


@app.route('/driver-attendance/mark', methods=['GET', 'POST'])
def driver_attendance_mark():
    form = DriverAttendanceFilterForm()
    form.project_id.choices = [(0, '-- All Projects --')] + [(p.id, p.name) for p in Project.query.order_by(Project.name).all()]
    view_date = date.today()
    project_id = request.args.get('project_id', type=int) or (request.form.get('project_id', type=int))
    district_id = request.args.get('district_id', type=int) or (request.form.get('district_id', type=int))
    search = (request.args.get('search') or request.form.get('search') or '').strip()
    if request.args.get('date'):
        view_date = parse_date(request.args.get('date')) or view_date
    if request.method == 'POST' and request.form.get('attendance_date'):
        view_date = parse_date(request.form.get('attendance_date')) or view_date
        proj_id = request.form.get('project_id', type=int)
        if proj_id == 0:
            proj_id = None
        project_id = proj_id
        dist_id = request.form.get('district_id', type=int)
        if dist_id == 0:
            dist_id = None
        district_id = dist_id
    if request.method == 'GET' and form.validate_on_submit():
        view_date = form.attendance_date.data
        project_id = form.project_id.data if form.project_id.data else None
        if project_id == 0:
            project_id = None
        district_id = form.district_id.data if form.district_id.data else None
        if district_id == 0:
            district_id = None
        return redirect(url_for('driver_attendance_mark', date=view_date.strftime('%d-%m-%Y'), project_id=project_id or '', district_id=district_id or '', search=search))
    if request.method == 'GET':
        form.attendance_date.data = view_date
        form.project_id.data = project_id if project_id else 0
    if project_id and project_id != 0:
        districts = District.query.join(project_district).filter(project_district.c.project_id == project_id).order_by(District.name).all()
        form.district_id.choices = [(0, '-- All Districts --')] + [(d.id, d.name) for d in districts]
    else:
        form.district_id.choices = [(0, '-- All Districts --')]
    if request.method == 'GET':
        form.district_id.data = district_id if district_id else 0
    drivers_query = Driver.query.filter_by(status='Active')
    if project_id:
        drivers_query = drivers_query.filter(Driver.project_id == project_id)
    if district_id:
        drivers_query = drivers_query.filter(Driver.district_id == district_id)
    if search:
        q = f'%{search}%'
        drivers_query = drivers_query.outerjoin(Vehicle, Driver.vehicle_id == Vehicle.id).filter(
            db.or_(
                Driver.name.ilike(q),
                Driver.driver_id.ilike(q),
                Vehicle.vehicle_no.ilike(q),
            )
        )
    drivers = drivers_query.order_by(Driver.name).all()
    existing = {a.driver_id: a for a in DriverAttendance.query.filter_by(attendance_date=view_date).all()}
    if request.method == 'POST' and request.form.get('save_attendance'):
        for d in drivers:
            did = d.id
            status = request.form.get(f'driver_{did}_status', 'Absent').strip() or 'Absent'
            check_in_s = request.form.get(f'driver_{did}_check_in', '')
            check_out_s = request.form.get(f'driver_{did}_check_out', '')
            remarks = request.form.get(f'driver_{did}_remarks', '')
            check_in_t = None
            check_out_t = None
            if check_in_s:
                try:
                    parts = check_in_s.strip().split(':')
                    if len(parts) >= 2:
                        check_in_t = time(int(parts[0]), int(parts[1]), int(parts[2]) if len(parts) > 2 else 0)
                except (ValueError, IndexError):
                    pass
            if check_out_s:
                try:
                    parts = check_out_s.strip().split(':')
                    if len(parts) >= 2:
                        check_out_t = time(int(parts[0]), int(parts[1]), int(parts[2]) if len(parts) > 2 else 0)
                except (ValueError, IndexError):
                    pass
            rec = existing.get(did)
            if rec:
                rec.status = status
                rec.check_in = check_in_t
                rec.check_out = check_out_t
                rec.remarks = remarks or None
                rec.project_id = project_id
                rec.updated_at = datetime.utcnow()
            else:
                rec = DriverAttendance(
                    driver_id=did, attendance_date=view_date, status=status,
                    check_in=check_in_t, check_out=check_out_t, remarks=remarks or None,
                    project_id=project_id
                )
                db.session.add(rec)
        try:
            db.session.commit()
            flash('Attendance saved successfully.', 'success')
            return redirect(url_for('driver_attendance_list', date=view_date.strftime('%d-%m-%Y'), project_id=project_id or '', district_id=district_id or '', search=search))
        except Exception as e:
            db.session.rollback()
            flash(f'Error saving attendance: {str(e)}', 'danger')
    return render_template('driver_attendance_mark.html', form=form, view_date=view_date, drivers=drivers, existing=existing, project_id=project_id, district_id=district_id, search=search, status_choices=ATTENDANCE_STATUS_CHOICES)


@app.route('/driver-attendance/report', methods=['GET', 'POST'])
def driver_attendance_report():
    form = DriverAttendanceReportForm()
    form.project_id.choices = [(0, '-- All Projects --')] + [(p.id, p.name) for p in Project.query.order_by(Project.name).all()]
    today = date.today()
    form.month.data = form.month.data or today.month
    form.year.data = form.year.data or today.year
    report = []
    if request.method == 'POST' and form.validate_on_submit():
        month = form.month.data
        year = form.year.data
        project_id = form.project_id.data if form.project_id.data else 0
        if project_id == 0:
            project_id = None
        district_id = form.district_id.data if form.district_id.data else 0
        if district_id == 0:
            district_id = None
        search = (form.search.data or '').strip()
        drivers_query = Driver.query.filter_by(status='Active')
        if project_id:
            drivers_query = drivers_query.filter(Driver.project_id == project_id)
        if district_id:
            drivers_query = drivers_query.filter(Driver.district_id == district_id)
        if search:
            q = f'%{search}%'
            drivers_query = drivers_query.outerjoin(Vehicle, Driver.vehicle_id == Vehicle.id).filter(
                db.or_(
                    Driver.name.ilike(q),
                    Driver.driver_id.ilike(q),
                    Vehicle.vehicle_no.ilike(q),
                )
            )
        drivers = drivers_query.order_by(Driver.name).all()
        from calendar import monthrange
        _, ndays = monthrange(year, month)
        start_d = date(year, month, 1)
        end_d = date(year, month, ndays)
        for d in drivers:
            rows = DriverAttendance.query.filter(
                DriverAttendance.driver_id == d.id,
                DriverAttendance.attendance_date >= start_d,
                DriverAttendance.attendance_date <= end_d
            ).all()
            by_status = {}
            for r in rows:
                by_status[r.status] = by_status.get(r.status, 0) + 1
            report.append({
                'driver': d,
                'present': by_status.get('Present', 0),
                'absent': by_status.get('Absent', 0),
                'leave': by_status.get('Leave', 0),
                'late': by_status.get('Late', 0),
                'half_day': by_status.get('Half-Day', 0),
                'off': by_status.get('Off', 0),
                'total_marked': len(rows),
                'days_in_month': ndays,
            })
    if request.method == 'POST' and form.validate_on_submit():
        project_id = form.project_id.data if form.project_id.data else 0
        if project_id and project_id != 0:
            districts = District.query.join(project_district).filter(project_district.c.project_id == project_id).order_by(District.name).all()
            form.district_id.choices = [(0, '-- All Districts --')] + [(d.id, d.name) for d in districts]
        else:
            form.district_id.choices = [(0, '-- All Districts --')]
    else:
        form.district_id.choices = [(0, '-- All Districts --')]
    return render_template('driver_attendance_report.html', form=form, report=report)


# ─── Task Report ─────────────────────────────────
@app.route('/get_projects_by_district/<int:district_id>')
def get_projects_by_district(district_id):
    if not district_id:
        return jsonify([])
    projects = Project.query.join(project_district).filter(project_district.c.district_id == district_id).order_by(Project.name).all()
    return jsonify([{'id': p.id, 'name': p.name} for p in projects])


@app.route('/get_vehicles_by_project_district')
def get_vehicles_by_project_district():
    project_id = request.args.get('project_id', type=int)
    district_id = request.args.get('district_id', type=int)
    if not project_id:
        return jsonify([])
    q = Vehicle.query.filter(Vehicle.project_id == project_id)
    if district_id:
        q = q.filter(Vehicle.district_id == district_id)
    vehicles = q.order_by(Vehicle.vehicle_no).all()
    return jsonify([{'id': v.id, 'vehicle_no': v.vehicle_no, 'vehicle_type': v.vehicle_type or ''} for v in vehicles])


@app.route('/task-report')
def task_report_list():
    form = TaskReportFilterForm()
    form.district_id.choices = [(0, '-- All Districts --')] + [(d.id, d.name) for d in District.query.order_by(District.name).all()]
    form.project_id.choices = [(0, '-- All Projects --')] + [(p.id, p.name) for p in Project.query.order_by(Project.name).all()]
    today = date.today()
    from_date = today
    to_date = today
    district_id = request.args.get('district_id', type=int) or 0
    project_id = request.args.get('project_id', type=int) or 0
    if request.method == 'POST':
        from_date = parse_date(request.form.get('from_date')) or today
        to_date = parse_date(request.form.get('to_date')) or today
        district_id = request.form.get('district_id', type=int) or 0
        project_id = request.form.get('project_id', type=int) or 0
        if from_date and to_date and from_date > to_date:
            from_date, to_date = to_date, from_date
        return redirect(url_for('task_report_list', from_date=from_date.strftime('%d-%m-%Y') if from_date else '', to_date=to_date.strftime('%d-%m-%Y') if to_date else '', district_id=district_id, project_id=project_id))
    from_str = request.args.get('from_date', '')
    to_str = request.args.get('to_date', '')
    if from_str:
        from_date = parse_date(from_str) or from_date
    if to_str:
        to_date = parse_date(to_str) or to_date
    if from_date and to_date and from_date > to_date:
        from_date, to_date = to_date, from_date
    form.from_date.data = from_date
    form.to_date.data = to_date
    form.district_id.data = district_id
    form.project_id.data = project_id
    query = VehicleDailyTask.query.filter(
        VehicleDailyTask.task_date >= from_date,
        VehicleDailyTask.task_date <= to_date
    )
    if district_id:
        query = query.filter(VehicleDailyTask.district_id == district_id)
    if project_id:
        query = query.filter(VehicleDailyTask.project_id == project_id)
    tasks = query.order_by(VehicleDailyTask.task_date.desc(), VehicleDailyTask.id).all()
    rows = []
    for t in tasks:
        v = t.vehicle
        task_d = t.task_date
        prev = VehicleDailyTask.query.filter(
            VehicleDailyTask.vehicle_id == t.vehicle_id,
            VehicleDailyTask.task_date < task_d
        ).order_by(VehicleDailyTask.task_date.desc()).first()
        start_reading = float(prev.close_reading) if prev else 0
        close_reading = float(t.close_reading)
        kms_driven = close_reading - start_reading
        if kms_driven < 0:
            kms_driven = 0
        emg_rec = EmergencyTaskRecord.query.filter_by(task_date=task_d, vehicle_no=v.vehicle_no).first()
        emg_tasks = emg_rec.emg_tasks_count if emg_rec else 0
        mileage_rec = VehicleMileageRecord.query.filter_by(task_date=task_d, vehicle_no=v.vehicle_no).first()
        tracker_km = float(mileage_rec.tracker_km) if mileage_rec else 0
        kms_diff = kms_driven - tracker_km
        pct_diff = round((kms_diff / kms_driven) * 100, 1) if kms_driven and kms_driven != 0 else None
        rows.append({
            'task': t, 'vehicle': v, 'task_date': task_d,
            'start_reading': start_reading, 'close_reading': close_reading,
            'kms_driven': round(kms_driven, 2), 'tasks_count': t.tasks_count,
            'emg_tasks': emg_tasks, 'tracker_km': round(tracker_km, 2),
            'kms_diff': round(kms_diff, 2), 'pct_diff': pct_diff,
        })
    total_kms = sum(r['kms_driven'] for r in rows)
    total_tracker = sum(r['tracker_km'] for r in rows)
    total_diff = total_kms - total_tracker
    total_pct = round((total_diff / total_kms * 100), 1) if total_kms else None
    total_tasks = sum(r['tasks_count'] for r in rows)
    return render_template('task_report_list.html', form=form, rows=rows, from_date=from_date, to_date=to_date,
                           total_kms=total_kms, total_tracker=total_tracker, total_diff=total_diff, total_pct=total_pct, total_tasks=total_tasks)


def _logbook_vehicle_aggregate(vehicle_id, from_date, to_date):
    """For one vehicle, return start_reading, close_reading, total_kms, total_task for the date range."""
    tasks = VehicleDailyTask.query.filter(
        VehicleDailyTask.vehicle_id == vehicle_id,
        VehicleDailyTask.task_date >= from_date,
        VehicleDailyTask.task_date <= to_date
    ).order_by(VehicleDailyTask.task_date.asc()).all()
    if not tasks:
        return None
    first = tasks[0]
    last = tasks[-1]
    prev = VehicleDailyTask.query.filter(
        VehicleDailyTask.vehicle_id == vehicle_id,
        VehicleDailyTask.task_date < from_date
    ).order_by(VehicleDailyTask.task_date.desc()).first()
    start_reading = float(prev.close_reading) if prev else 0
    close_reading = float(last.close_reading)
    total_kms = close_reading - start_reading
    if total_kms < 0:
        total_kms = 0
    total_task = sum(t.tasks_count for t in tasks)
    return {
        'start_reading': start_reading,
        'close_reading': close_reading,
        'total_kms': round(total_kms, 2),
        'total_task': total_task,
    }


@app.route('/task-report/logbook-cover', methods=['GET', 'POST'])
def task_report_logbook_cover():
    form = TaskReportFilterForm()
    form.district_id.choices = [(0, '-- Select District --')] + [(d.id, d.name) for d in District.query.order_by(District.name).all()]
    form.project_id.choices = [(0, '-- Select Project --')] + [(p.id, p.name) for p in Project.query.order_by(Project.name).all()]
    today = date.today()
    from_date = today
    to_date = today
    district_id = request.args.get('district_id', type=int) or 0
    project_id = request.args.get('project_id', type=int) or 0
    if request.method == 'POST':
        from_date = parse_date(request.form.get('from_date')) or today
        to_date = parse_date(request.form.get('to_date')) or today
        district_id = request.form.get('district_id', type=int) or 0
        project_id = request.form.get('project_id', type=int) or 0
        if from_date and to_date and from_date > to_date:
            from_date, to_date = to_date, from_date
        return redirect(url_for('task_report_logbook_cover', from_date=from_date.strftime('%d-%m-%Y') if from_date else '', to_date=to_date.strftime('%d-%m-%Y') if to_date else '', district_id=district_id, project_id=project_id))
    from_str = request.args.get('from_date', '')
    to_str = request.args.get('to_date', '')
    if from_str:
        from_date = parse_date(from_str) or from_date
    if to_str:
        to_date = parse_date(to_str) or to_date
    if from_date and to_date and from_date > to_date:
        from_date, to_date = to_date, from_date
    form.from_date.data = from_date
    form.to_date.data = to_date
    form.district_id.data = district_id
    form.project_id.data = project_id
    if district_id:
        projects = Project.query.join(project_district).filter(project_district.c.district_id == district_id).order_by(Project.name).all()
        form.project_id.choices = [(0, '-- Select Project --')] + [(p.id, p.name) for p in projects]
    rows = []
    if project_id:
        q = Vehicle.query.filter(Vehicle.project_id == project_id)
        if district_id:
            q = q.filter(Vehicle.district_id == district_id)
        vehicles = q.order_by(Vehicle.vehicle_no).all()
        project = Project.query.get(project_id)
        district = District.query.get(district_id) if district_id else None
        for v in vehicles:
            agg = _logbook_vehicle_aggregate(v.id, from_date, to_date)
            district_name = v.district.name if v.district else (district.name if district else '-')
            tehsil_name = v.parking_station.tehsil if v.parking_station and v.parking_station.tehsil else '-'
            if agg:
                rows.append({
                    'vehicle': v, 'vehicle_id': v.id, 'from_date': from_date, 'to_date': to_date,
                    'district_name': district_name, 'tehsil_name': tehsil_name,
                    'start_reading': agg['start_reading'], 'close_reading': agg['close_reading'],
                    'total_kms': agg['total_kms'], 'total_task': agg['total_task'],
                    'project': project, 'project_name': project.name if project else '',
                })
            else:
                rows.append({
                    'vehicle': v, 'vehicle_id': v.id, 'from_date': from_date, 'to_date': to_date,
                    'district_name': district_name, 'tehsil_name': tehsil_name,
                    'start_reading': None, 'close_reading': None, 'total_kms': None, 'total_task': None,
                    'project': project, 'project_name': project.name if project else '',
                })
    return render_template('logbook_cover_list.html', form=form, rows=rows, from_date=from_date, to_date=to_date)


@app.route('/task-report/logbook-view')
def task_report_logbook_view():
    vehicle_id = request.args.get('vehicle_id', type=int)
    from_str = request.args.get('from_date', '')
    to_str = request.args.get('to_date', '')
    if not vehicle_id or not from_str or not to_str:
        flash('Missing vehicle or date range.', 'warning')
        return redirect(url_for('task_report_logbook_cover'))
    from_date = parse_date(from_str)
    to_date = parse_date(to_str)
    if not from_date or not to_date:
        flash('Invalid date range.', 'warning')
        return redirect(url_for('task_report_logbook_cover'))
    vehicle = Vehicle.query.get_or_404(vehicle_id)
    agg = _logbook_vehicle_aggregate(vehicle_id, from_date, to_date)
    if not agg:
        flash('No task data for this vehicle in the selected period.', 'warning')
        return redirect(url_for('task_report_logbook_cover'))
    project = vehicle.project
    project_name = project.name if project else ''
    district_name = vehicle.district.name if vehicle.district else '-'
    tehsil_name = vehicle.parking_station.tehsil if vehicle.parking_station and vehicle.parking_station.tehsil else '-'
    data = {
        'vehicle_no': vehicle.vehicle_no,
        'from_date': from_date,
        'to_date': to_date,
        'district_name': district_name,
        'tehsil_name': tehsil_name,
        'total_task': agg['total_task'],
        'start_reading': agg['start_reading'],
        'close_reading': agg['close_reading'],
        'total_kms': agg['total_kms'],
        'project_name': project_name,
    }
    if project_name == 'RAS-1034':
        return render_template('logbook_cover_ras.html', **data)
    if project_name == 'COW-1034':
        return render_template('logbook_cover_cow.html', **data)
    return render_template('logbook_cover_ras.html', **data)


@app.route('/task-report/logbook-view-all')
def task_report_logbook_view_all():
    from_str = request.args.get('from_date', '')
    to_str = request.args.get('to_date', '')
    district_id = request.args.get('district_id', type=int) or 0
    project_id = request.args.get('project_id', type=int) or 0
    if not from_str or not to_str or not project_id:
        flash('From Date, To Date aur Project select karein.', 'warning')
        return redirect(url_for('task_report_logbook_cover'))
    from_date = parse_date(from_str)
    to_date = parse_date(to_str)
    if not from_date or not to_date:
        flash('Invalid date range.', 'warning')
        return redirect(url_for('task_report_logbook_cover'))
    q = Vehicle.query.filter(Vehicle.project_id == project_id)
    if district_id:
        q = q.filter(Vehicle.district_id == district_id)
    vehicles = q.order_by(Vehicle.vehicle_no).all()
    project = Project.query.get(project_id)
    project_name = project.name if project else ''
    district = District.query.get(district_id) if district_id else None
    rows_with_data = []
    for v in vehicles:
        agg = _logbook_vehicle_aggregate(v.id, from_date, to_date)
        if not agg:
            continue
        district_name = v.district.name if v.district else (district.name if district else '-')
        tehsil_name = v.parking_station.tehsil if v.parking_station and v.parking_station.tehsil else '-'
        rows_with_data.append({
            'vehicle_no': v.vehicle_no,
            'from_date': from_date,
            'to_date': to_date,
            'district_name': district_name,
            'tehsil_name': tehsil_name,
            'total_task': agg['total_task'],
            'start_reading': agg['start_reading'],
            'close_reading': agg['close_reading'],
            'total_kms': agg['total_kms'],
        })
    if not rows_with_data:
        flash('Is period mein kisi vehicle ke paas task data nahi mila.', 'warning')
        return redirect(url_for('task_report_logbook_cover', from_date=from_date.strftime('%d-%m-%Y'), to_date=to_date.strftime('%d-%m-%Y'), district_id=district_id, project_id=project_id))
    return render_template('logbook_cover_all.html', rows=rows_with_data, project_name=project_name)


@app.route('/task-report/new', methods=['GET', 'POST'])
def task_report_new():
    districts = District.query.order_by(District.name).all()
    view_date = parse_date(request.args.get('date') or request.form.get('task_date')) or date.today()
    district_id = request.args.get('district_id', type=int) or request.form.get('district_id', type=int) or 0
    project_id = request.args.get('project_id', type=int) or request.form.get('project_id', type=int) or 0

    if request.method == 'POST' and request.form.get('save_batch'):
        task_date = parse_date(request.form.get('task_date')) or view_date
        district_id = request.form.get('district_id', type=int) or 0
        project_id = request.form.get('project_id', type=int) or 0
        if not project_id:
            flash('Select District and Project first.', 'warning')
            return redirect(url_for('task_report_new'))
        q = Vehicle.query.filter(Vehicle.project_id == project_id)
        if district_id:
            q = q.filter(Vehicle.district_id == district_id)
        vehicles = q.order_by(Vehicle.vehicle_no).all()
        missing = []
        to_save = []
        for v in vehicles:
            close_val = request.form.get('vehicle_%s_close_reading' % v.id)
            tasks_val = request.form.get('vehicle_%s_tasks_count' % v.id)
            try:
                close_reading = float(close_val) if close_val not in (None, '') else None
            except (TypeError, ValueError):
                close_reading = None
            tasks_count = int(float(tasks_val)) if tasks_val not in (None, '') else 1
            if close_reading is None:
                missing.append(v.vehicle_no)
            else:
                to_save.append((v, close_reading, tasks_count))
        if missing:
            flash('Sab vehicles ke liye Close Reading zaroori hai. Missing: ' + ', '.join(missing), 'danger')
            view_date = task_date
            rows = _build_vehicle_rows(vehicles, task_date, request.form)
            projects = Project.query.join(project_district).filter(project_district.c.district_id == district_id).order_by(Project.name).all() if district_id else []
            return render_template('task_report_new.html', rows=rows, view_date=view_date, district_id=district_id, project_id=project_id, districts=districts, projects=projects)
        for v, close_reading, tasks_count in to_save:
            existing = VehicleDailyTask.query.filter_by(vehicle_id=v.id, task_date=task_date).first()
            if existing:
                existing.close_reading = close_reading
                existing.tasks_count = tasks_count
            else:
                db.session.add(VehicleDailyTask(
                    vehicle_id=v.id, project_id=project_id or None, district_id=district_id or None,
                    task_date=task_date, close_reading=close_reading, tasks_count=tasks_count,
                ))
        try:
            db.session.commit()
            flash('Task entries saved successfully.', 'success')
            return redirect(url_for('task_report_list', date=task_date.strftime('%d-%m-%Y')))
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'danger')
            view_date = task_date
            rows = _build_vehicle_rows(vehicles, task_date, request.form)
            projects = Project.query.join(project_district).filter(project_district.c.district_id == district_id).order_by(Project.name).all() if district_id else []
            return render_template('task_report_new.html', rows=rows, view_date=view_date, district_id=district_id, project_id=project_id, districts=districts, projects=projects)

    rows = []
    projects = []
    if district_id:
        projects = Project.query.join(project_district).filter(project_district.c.district_id == district_id).order_by(Project.name).all()
    if district_id and project_id:
        q = Vehicle.query.filter(Vehicle.project_id == project_id)
        if district_id:
            q = q.filter(Vehicle.district_id == district_id)
        vehicles = q.order_by(Vehicle.vehicle_no).all()
        rows = _build_vehicle_rows(vehicles, view_date, request.form)
    return render_template('task_report_new.html', rows=rows, view_date=view_date, district_id=district_id, project_id=project_id, districts=districts, projects=projects)


def _build_vehicle_rows(vehicles, task_date, form=None):
    form = form or {}
    rows = []
    for v in vehicles:
        prev = VehicleDailyTask.query.filter(
            VehicleDailyTask.vehicle_id == v.id,
            VehicleDailyTask.task_date < task_date
        ).order_by(VehicleDailyTask.task_date.desc()).first()
        start_reading = float(prev.close_reading) if prev else 0
        emg_rec = EmergencyTaskRecord.query.filter_by(task_date=task_date, vehicle_no=v.vehicle_no).first()
        emg_tasks = emg_rec.emg_tasks_count if emg_rec else 0
        mileage_rec = VehicleMileageRecord.query.filter_by(task_date=task_date, vehicle_no=v.vehicle_no).first()
        tracker_km = float(mileage_rec.tracker_km) if mileage_rec else 0
        existing = VehicleDailyTask.query.filter_by(vehicle_id=v.id, task_date=task_date).first()
        existing_close = float(existing.close_reading) if existing and existing.close_reading is not None else None
        existing_tasks = existing.tasks_count if existing else 1
        if form:
            key_close = 'vehicle_%s_close_reading' % v.id
            key_tasks = 'vehicle_%s_tasks_count' % v.id
            if key_close in form and form[key_close] not in (None, ''):
                try:
                    existing_close = float(form[key_close])
                except (TypeError, ValueError):
                    pass
            if key_tasks in form and form[key_tasks] not in (None, ''):
                try:
                    existing_tasks = int(float(form[key_tasks]))
                except (TypeError, ValueError):
                    pass
        rows.append({
            'vehicle': v,
            'start_reading': start_reading,
            'emg_tasks': emg_tasks,
            'tracker_km': round(tracker_km, 2),
            'close_reading': existing_close,
            'tasks_count': existing_tasks,
        })
    return rows


@app.route('/task-report/upload/emergency', methods=['GET', 'POST'])
def task_report_upload_emergency():
    form = EmergencyTaskUploadForm()
    if request.method == 'POST' and form.validate_on_submit():
        f = form.file.data
        if not f:
            flash('Please select an Excel file.', 'warning')
            return redirect(url_for('task_report_upload_emergency'))
        task_date = form.task_date.data
        try:
            import openpyxl
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(min_row=1, values_only=True))
            wb.close()
            if not rows:
                flash('Excel file is empty.', 'warning')
                return redirect(url_for('task_report_upload_emergency'))
            headers = [str(c).strip().lower() if c else '' for c in rows[0]]
            vehicle_col = emg_col = 0
            for i, h in enumerate(headers):
                if 'vehicle' in h and ('no' in h or 'number' in h or '#' in h):
                    vehicle_col = i
                if 'emg' in h or 'emergency' in h:
                    emg_col = i
            if emg_col == 0 and len(headers) > 1:
                emg_col = 1
            count = 0
            for row in rows[1:]:
                if not row or row[vehicle_col] is None:
                    continue
                v_no = str(row[vehicle_col]).strip() if row[vehicle_col] else ''
                if not v_no:
                    continue
                try:
                    emg_val = int(float(row[emg_col])) if len(row) > emg_col and row[emg_col] is not None else 0
                except (TypeError, ValueError):
                    emg_val = 0
                existing = EmergencyTaskRecord.query.filter_by(task_date=task_date, vehicle_no=v_no).first()
                if existing:
                    existing.emg_tasks_count = emg_val
                    existing.upload_date = date.today()
                else:
                    db.session.add(EmergencyTaskRecord(task_date=task_date, vehicle_no=v_no, emg_tasks_count=emg_val, upload_date=date.today()))
                count += 1
            db.session.commit()
            flash(f'EmergencyTaskReport uploaded: {count} record(s).', 'success')
            return redirect(url_for('task_report_list', date=task_date.strftime('%d-%m-%Y')))
        except Exception as e:
            db.session.rollback()
            flash(f'Error parsing Excel: {str(e)}', 'danger')
    return render_template('task_report_upload_emergency.html', form=form)


@app.route('/task-report/upload/mileage', methods=['GET', 'POST'])
def task_report_upload_mileage():
    form = VehicleMileageUploadForm()
    if request.method == 'POST' and form.validate_on_submit():
        f = form.file.data
        if not f:
            flash('Please select an Excel file.', 'warning')
            return redirect(url_for('task_report_upload_mileage'))
        task_date = form.task_date.data
        try:
            import openpyxl
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(min_row=1, values_only=True))
            wb.close()
            if not rows:
                flash('Excel file is empty.', 'warning')
                return redirect(url_for('task_report_upload_mileage'))
            headers = [str(c).strip().lower() if c else '' for c in rows[0]]
            vehicle_col = km_col = 0
            for i, h in enumerate(headers):
                if 'vehicle' in h and ('no' in h or 'number' in h or '#' in h):
                    vehicle_col = i
                if 'km' in h or 'mileage' in h or 'tracker' in h:
                    km_col = i
            if km_col == 0 and len(headers) > 1:
                km_col = 1
            count = 0
            for row in rows[1:]:
                if not row or row[vehicle_col] is None:
                    continue
                v_no = str(row[vehicle_col]).strip() if row[vehicle_col] else ''
                if not v_no:
                    continue
                try:
                    km_val = float(row[km_col]) if len(row) > km_col and row[km_col] is not None else 0
                except (TypeError, ValueError):
                    km_val = 0
                existing = VehicleMileageRecord.query.filter_by(task_date=task_date, vehicle_no=v_no).first()
                if existing:
                    existing.tracker_km = km_val
                    existing.upload_date = date.today()
                else:
                    db.session.add(VehicleMileageRecord(task_date=task_date, vehicle_no=v_no, tracker_km=km_val, upload_date=date.today()))
                count += 1
            db.session.commit()
            flash(f'Vehicle Mileage report uploaded: {count} record(s).', 'success')
            return redirect(url_for('task_report_list', date=task_date.strftime('%d-%m-%Y')))
        except Exception as e:
            db.session.rollback()
            flash(f'Error parsing Excel: {str(e)}', 'danger')
    return render_template('task_report_upload_mileage.html', form=form)


def _parse_emergency_excel(f, task_date):
    import openpyxl
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    wb.close()
    if not rows:
        return 0
    headers = [str(c).strip().lower() if c else '' for c in rows[0]]
    vehicle_col = emg_col = 0
    for i, h in enumerate(headers):
        if 'vehicle' in h and ('no' in h or 'number' in h or '#' in h):
            vehicle_col = i
        if 'emg' in h or 'emergency' in h:
            emg_col = i
    if emg_col == 0 and len(headers) > 1:
        emg_col = 1
    count = 0
    for row in rows[1:]:
        if not row or row[vehicle_col] is None:
            continue
        v_no = str(row[vehicle_col]).strip() if row[vehicle_col] else ''
        if not v_no:
            continue
        try:
            emg_val = int(float(row[emg_col])) if len(row) > emg_col and row[emg_col] is not None else 0
        except (TypeError, ValueError):
            emg_val = 0
        existing = EmergencyTaskRecord.query.filter_by(task_date=task_date, vehicle_no=v_no).first()
        if existing:
            existing.emg_tasks_count = emg_val
            existing.upload_date = date.today()
        else:
            db.session.add(EmergencyTaskRecord(task_date=task_date, vehicle_no=v_no, emg_tasks_count=emg_val, upload_date=date.today()))
        count += 1
    return count


def _parse_mileage_excel(f, task_date):
    import openpyxl
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    wb.close()
    if not rows:
        return 0
    headers = [str(c).strip().lower() if c else '' for c in rows[0]]
    vehicle_col = km_col = 0
    for i, h in enumerate(headers):
        if 'vehicle' in h and ('no' in h or 'number' in h or '#' in h):
            vehicle_col = i
        if 'km' in h or 'mileage' in h or 'tracker' in h:
            km_col = i
    if km_col == 0 and len(headers) > 1:
        km_col = 1
    count = 0
    for row in rows[1:]:
        if not row or row[vehicle_col] is None:
            continue
        v_no = str(row[vehicle_col]).strip() if row[vehicle_col] else ''
        if not v_no:
            continue
        try:
            km_val = float(row[km_col]) if len(row) > km_col and row[km_col] is not None else 0
        except (TypeError, ValueError):
            km_val = 0
        existing = VehicleMileageRecord.query.filter_by(task_date=task_date, vehicle_no=v_no).first()
        if existing:
            existing.tracker_km = km_val
            existing.upload_date = date.today()
        else:
            db.session.add(VehicleMileageRecord(task_date=task_date, vehicle_no=v_no, tracker_km=km_val, upload_date=date.today()))
        count += 1
    return count


@app.route('/task-report/upload', methods=['GET', 'POST'])
def task_report_upload():
    """Single form: upload both EmergencyTaskReport and Vehicle Mileage Excel."""
    form = TaskReportUploadBothForm()
    if request.method == 'POST' and form.validate_on_submit():
        task_date = form.task_date.data
        fe = form.file_emergency.data
        fm = form.file_mileage.data
        if not fe and not fm:
            flash('Please select at least one Excel file.', 'warning')
            return redirect(url_for('task_report_upload'))
        try:
            c1 = c2 = 0
            if fe:
                c1 = _parse_emergency_excel(fe, task_date)
            if fm:
                c2 = _parse_mileage_excel(fm, task_date)
            db.session.commit()
            msg = []
            if c1:
                msg.append(f'EmergencyTaskReport: {c1} record(s)')
            if c2:
                msg.append(f'Mileage report: {c2} record(s)')
            flash('Uploaded. ' + '; '.join(msg) if msg else 'No data imported.', 'success')
            return redirect(url_for('task_report_list', date=task_date.strftime('%d-%m-%Y')))
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'danger')
    return render_template('task_report_upload.html', form=form)


# ────────────────────────────────────────────────
# Red Task Report
# ────────────────────────────────────────────────
@app.route('/red-task')
def red_task_list():
    form = RedTaskFilterForm()
    form.district_id.choices = [(0, '-- All Districts --')] + [(d.id, d.name) for d in District.query.order_by(District.name).all()]
    today = date.today()
    from_date = today
    to_date = today
    district_id = request.args.get('district_id', type=int) or 0
    project_id = request.args.get('project_id', type=int) or 0
    if request.method == 'POST':
        from_date = parse_date(request.form.get('from_date')) or today
        to_date = parse_date(request.form.get('to_date')) or today
        district_id = request.form.get('district_id', type=int) or 0
        project_id = request.form.get('project_id', type=int) or 0
        if from_date and to_date and from_date > to_date:
            from_date, to_date = to_date, from_date
        return redirect(url_for('red_task_list', from_date=from_date.strftime('%d-%m-%Y') if from_date else '', to_date=to_date.strftime('%d-%m-%Y') if to_date else '', district_id=district_id, project_id=project_id))
    from_str = request.args.get('from_date', '')
    to_str = request.args.get('to_date', '')
    if from_str:
        from_date = parse_date(from_str) or from_date
    if to_str:
        to_date = parse_date(to_str) or to_date
    if from_date and to_date and from_date > to_date:
        from_date, to_date = to_date, from_date
    form.from_date.data = from_date
    form.to_date.data = to_date
    form.district_id.data = district_id
    if district_id:
        projects = Project.query.join(project_district).filter(project_district.c.district_id == district_id).order_by(Project.name).all()
        form.project_id.choices = [(0, '-- All Projects --')] + [(p.id, p.name) for p in projects]
    else:
        form.project_id.choices = [(0, '-- All Projects --')]
    form.project_id.data = project_id
    query = RedTask.query.filter(RedTask.task_date >= from_date, RedTask.task_date <= to_date)
    if district_id:
        query = query.filter(RedTask.district_id == district_id)
    if project_id:
        query = query.filter(RedTask.project_id == project_id)
    rows = query.order_by(RedTask.task_date.desc(), RedTask.id.desc()).all()
    return render_template('red_task_list.html', form=form, rows=rows, from_date=from_date, to_date=to_date)


@app.route('/red-task/new', methods=['GET', 'POST'])
def red_task_new():
    form = RedTaskForm()
    form.district_id.choices = [(0, '-- Select District --')] + [(d.id, d.name) for d in District.query.order_by(District.name).all()]
    form.project_id.choices = [(0, '-- Select Project --')]
    form.vehicle_id.choices = [(0, '-- Select Vehicle --')] + [(v.id, v.vehicle_no) for v in Vehicle.query.order_by(Vehicle.vehicle_no).all()]
    if request.method == 'POST' and form.validate_on_submit():
        task_date = form.task_date.data
        district_id = form.district_id.data or None
        if district_id == 0:
            district_id = None
        project_id = form.project_id.data or None
        if project_id == 0:
            project_id = None
        vehicle_id = form.vehicle_id.data or None
        if vehicle_id == 0:
            vehicle_id = None
        rec = RedTask(task_date=task_date, task_id=form.task_id.data.strip() or None, district_id=district_id, project_id=project_id, vehicle_id=vehicle_id,
            reason=form.reason.data.strip() or None, driver_name=form.driver_name.data.strip() or None, call_to_dto=form.call_to_dto.data or None,
            dto_investigation=form.dto_investigation.data.strip() or None, action=form.action.data or None)
        db.session.add(rec)
        db.session.commit()
        flash('Red Task entry saved.', 'success')
        return redirect(url_for('red_task_list'))
    return render_template('red_task_form.html', form=form, title='Add Red Task')


@app.route('/red-task/<int:pk>/edit', methods=['GET', 'POST'])
def red_task_edit(pk):
    rec = RedTask.query.get_or_404(pk)
    form = RedTaskForm(obj=rec)
    form.district_id.choices = [(0, '-- Select District --')] + [(d.id, d.name) for d in District.query.order_by(District.name).all()]
    form.vehicle_id.choices = [(0, '-- Select Vehicle --')] + [(v.id, v.vehicle_no) for v in Vehicle.query.order_by(Vehicle.vehicle_no).all()]
    if rec.district_id:
        projects = Project.query.join(project_district).filter(project_district.c.district_id == rec.district_id).order_by(Project.name).all()
        form.project_id.choices = [(0, '-- Select Project --')] + [(p.id, p.name) for p in projects]
    else:
        form.project_id.choices = [(0, '-- Select Project --')]
    if request.method == 'GET':
        form.task_date.data = rec.task_date
        form.task_id.data = rec.task_id or ''
        form.district_id.data = rec.district_id or 0
        form.project_id.data = rec.project_id or 0
        form.vehicle_id.data = rec.vehicle_id or 0
        form.reason.data = rec.reason or ''
        form.driver_name.data = rec.driver_name or ''
        form.call_to_dto.data = rec.call_to_dto or ''
        form.dto_investigation.data = rec.dto_investigation or ''
        form.action.data = rec.action or ''
    if request.method == 'POST' and form.validate_on_submit():
        task_date = form.task_date.data
        district_id = form.district_id.data or None
        if district_id == 0:
            district_id = None
        project_id = form.project_id.data or None
        if project_id == 0:
            project_id = None
        vehicle_id = form.vehicle_id.data or None
        if vehicle_id == 0:
            vehicle_id = None
        rec.task_date = task_date
        rec.task_id = form.task_id.data.strip() or None
        rec.district_id = district_id
        rec.project_id = project_id
        rec.vehicle_id = vehicle_id
        rec.reason = form.reason.data.strip() or None
        rec.driver_name = form.driver_name.data.strip() or None
        rec.call_to_dto = form.call_to_dto.data or None
        rec.dto_investigation = form.dto_investigation.data.strip() or None
        rec.action = form.action.data or None
        db.session.commit()
        flash('Red Task updated.', 'success')
        return redirect(url_for('red_task_list'))
    return render_template('red_task_form.html', form=form, title='Edit Red Task', rec=rec)


# ────────────────────────────────────────────────
# Vehicle Move without Task Report
# ────────────────────────────────────────────────
@app.route('/vehicle-move-without-task')
def without_task_list():
    form = VehicleMoveWithoutTaskFilterForm()
    form.district_id.choices = [(0, '-- All Districts --')] + [(d.id, d.name) for d in District.query.order_by(District.name).all()]
    today = date.today()
    from_date = today
    to_date = today
    district_id = request.args.get('district_id', type=int) or 0
    project_id = request.args.get('project_id', type=int) or 0
    if request.method == 'POST':
        from_date = parse_date(request.form.get('from_date')) or today
        to_date = parse_date(request.form.get('to_date')) or today
        district_id = request.form.get('district_id', type=int) or 0
        project_id = request.form.get('project_id', type=int) or 0
        if from_date and to_date and from_date > to_date:
            from_date, to_date = to_date, from_date
        return redirect(url_for('without_task_list', from_date=from_date.strftime('%d-%m-%Y') if from_date else '', to_date=to_date.strftime('%d-%m-%Y') if to_date else '', district_id=district_id, project_id=project_id))
    from_str = request.args.get('from_date', '')
    to_str = request.args.get('to_date', '')
    if from_str:
        from_date = parse_date(from_str) or from_date
    if to_str:
        to_date = parse_date(to_str) or to_date
    if from_date and to_date and from_date > to_date:
        from_date, to_date = to_date, from_date
    form.from_date.data = from_date
    form.to_date.data = to_date
    form.district_id.data = district_id
    if district_id:
        projects = Project.query.join(project_district).filter(project_district.c.district_id == district_id).order_by(Project.name).all()
        form.project_id.choices = [(0, '-- All Projects --')] + [(p.id, p.name) for p in projects]
    else:
        form.project_id.choices = [(0, '-- All Projects --')]
    form.project_id.data = project_id
    query = VehicleMoveWithoutTask.query.filter(
        VehicleMoveWithoutTask.move_date >= from_date,
        VehicleMoveWithoutTask.move_date <= to_date
    )
    if district_id:
        query = query.filter(VehicleMoveWithoutTask.district_id == district_id)
    if project_id:
        query = query.filter(VehicleMoveWithoutTask.project_id == project_id)
    rows = query.order_by(VehicleMoveWithoutTask.move_date.desc(), VehicleMoveWithoutTask.id.desc()).all()
    return render_template('without_task_list.html', form=form, rows=rows, from_date=from_date, to_date=to_date)


@app.route('/vehicle-move-without-task/new', methods=['GET', 'POST'])
def without_task_new():
    form = VehicleMoveWithoutTaskForm()
    form.district_id.choices = [(0, '-- Select District --')] + [(d.id, d.name) for d in District.query.order_by(District.name).all()]
    form.project_id.choices = [(0, '-- Select Project --')]
    form.vehicle_id.choices = [(0, '-- Select Vehicle --')] + [(v.id, v.vehicle_no) for v in Vehicle.query.order_by(Vehicle.vehicle_no).all()]
    if request.method == 'POST' and form.validate_on_submit():
        move_date = form.move_date.data
        district_id = form.district_id.data or None
        if district_id == 0:
            district_id = None
        project_id = form.project_id.data or None
        if project_id == 0:
            project_id = None
        vehicle_id = form.vehicle_id.data or None
        if vehicle_id == 0:
            vehicle_id = None
        rec = VehicleMoveWithoutTask(
            move_date=move_date, district_id=district_id, project_id=project_id, vehicle_id=vehicle_id,
            km_in=form.km_in.data, km_out=form.km_out.data, d_km=form.d_km.data,
            logbook_task=form.logbook_task.data or 0, emg_task=form.emg_task.data or 0, t_km=form.t_km.data,
            remarks=form.remarks.data.strip() or None, fine=form.fine.data.strip() or None
        )
        db.session.add(rec)
        db.session.commit()
        flash('Vehicle Move without Task entry saved.', 'success')
        return redirect(url_for('without_task_list'))
    return render_template('without_task_form.html', form=form, title='Add Vehicle Move without Task')


@app.route('/vehicle-move-without-task/<int:pk>/edit', methods=['GET', 'POST'])
def without_task_edit(pk):
    rec = VehicleMoveWithoutTask.query.get_or_404(pk)
    form = VehicleMoveWithoutTaskForm(obj=rec)
    form.district_id.choices = [(0, '-- Select District --')] + [(d.id, d.name) for d in District.query.order_by(District.name).all()]
    form.vehicle_id.choices = [(0, '-- Select Vehicle --')] + [(v.id, v.vehicle_no) for v in Vehicle.query.order_by(Vehicle.vehicle_no).all()]
    if rec.district_id:
        projects = Project.query.join(project_district).filter(project_district.c.district_id == rec.district_id).order_by(Project.name).all()
        form.project_id.choices = [(0, '-- Select Project --')] + [(p.id, p.name) for p in projects]
    else:
        form.project_id.choices = [(0, '-- Select Project --')]
    if request.method == 'GET':
        form.move_date.data = rec.move_date
        form.district_id.data = rec.district_id or 0
        form.project_id.data = rec.project_id or 0
        form.vehicle_id.data = rec.vehicle_id or 0
        form.km_in.data = rec.km_in
        form.km_out.data = rec.km_out
        form.d_km.data = rec.d_km
        form.logbook_task.data = rec.logbook_task or 0
        form.emg_task.data = rec.emg_task or 0
        form.t_km.data = rec.t_km
        form.remarks.data = rec.remarks or ''
        form.fine.data = rec.fine or ''
    if request.method == 'POST' and form.validate_on_submit():
        move_date = form.move_date.data
        district_id = form.district_id.data or None
        if district_id == 0:
            district_id = None
        project_id = form.project_id.data or None
        if project_id == 0:
            project_id = None
        vehicle_id = form.vehicle_id.data or None
        if vehicle_id == 0:
            vehicle_id = None
        rec.move_date = move_date
        rec.district_id = district_id
        rec.project_id = project_id
        rec.vehicle_id = vehicle_id
        rec.km_in = form.km_in.data
        rec.km_out = form.km_out.data
        rec.d_km = form.d_km.data
        rec.logbook_task = form.logbook_task.data or 0
        rec.emg_task = form.emg_task.data or 0
        rec.t_km = form.t_km.data
        rec.remarks = form.remarks.data.strip() or None
        rec.fine = form.fine.data.strip() or None
        db.session.commit()
        flash('Vehicle Move without Task updated.', 'success')
        return redirect(url_for('without_task_list'))
    return render_template('without_task_form.html', form=form, title='Edit Vehicle Move without Task', rec=rec)


# ────────────────────────────────────────────────
# Penalty Record (Driver Status)
# ────────────────────────────────────────────────
@app.route('/penalty-record')
def penalty_record_list():
    form = PenaltyRecordFilterForm()
    form.district_id.choices = [(0, '-- All Districts --')] + [(d.id, d.name) for d in District.query.order_by(District.name).all()]
    today = date.today()
    from_date = today
    to_date = today
    district_id = request.args.get('district_id', type=int) or 0
    project_id = request.args.get('project_id', type=int) or 0
    if request.method == 'POST':
        from_date = parse_date(request.form.get('from_date')) or today
        to_date = parse_date(request.form.get('to_date')) or today
        district_id = request.form.get('district_id', type=int) or 0
        project_id = request.form.get('project_id', type=int) or 0
        if from_date and to_date and from_date > to_date:
            from_date, to_date = to_date, from_date
        return redirect(url_for('penalty_record_list', from_date=from_date.strftime('%d-%m-%Y') if from_date else '', to_date=to_date.strftime('%d-%m-%Y') if to_date else '', district_id=district_id, project_id=project_id))
    from_str = request.args.get('from_date', '')
    to_str = request.args.get('to_date', '')
    if from_str:
        from_date = parse_date(from_str) or from_date
    if to_str:
        to_date = parse_date(to_str) or to_date
    if from_date and to_date and from_date > to_date:
        from_date, to_date = to_date, from_date
    form.from_date.data = from_date
    form.to_date.data = to_date
    form.district_id.data = district_id
    if district_id:
        projects = Project.query.join(project_district).filter(project_district.c.district_id == district_id).order_by(Project.name).all()
        form.project_id.choices = [(0, '-- All Projects --')] + [(p.id, p.name) for p in projects]
    else:
        form.project_id.choices = [(0, '-- All Projects --')]
    form.project_id.data = project_id
    query = PenaltyRecord.query.filter(
        PenaltyRecord.record_date >= from_date,
        PenaltyRecord.record_date <= to_date
    )
    if district_id:
        query = query.filter(PenaltyRecord.district_id == district_id)
    if project_id:
        query = query.filter(PenaltyRecord.project_id == project_id)
    rows = query.order_by(PenaltyRecord.record_date.desc(), PenaltyRecord.id.desc()).all()
    return render_template('penalty_record_list.html', form=form, rows=rows, from_date=from_date, to_date=to_date)


@app.route('/penalty-record/new', methods=['GET', 'POST'])
def penalty_record_new():
    form = PenaltyRecordForm()
    form.district_id.choices = [(0, '-- Select District --')] + [(d.id, d.name) for d in District.query.order_by(District.name).all()]
    form.project_id.choices = [(0, '-- Select Project --')]
    form.vehicle_id.choices = [(0, '-- Select Vehicle --')] + [(v.id, v.vehicle_no) for v in Vehicle.query.order_by(Vehicle.vehicle_no).all()]
    form.driver_id.choices = [(0, '-- Select Driver --')] + [(d.id, d.name) for d in Driver.query.order_by(Driver.name).all()]
    if request.method == 'POST' and form.validate_on_submit():
        record_date = form.record_date.data
        district_id = form.district_id.data or None
        if district_id == 0:
            district_id = None
        project_id = form.project_id.data or None
        if project_id == 0:
            project_id = None
        vehicle_id = form.vehicle_id.data or None
        if vehicle_id == 0:
            vehicle_id = None
        driver_id = form.driver_id.data or None
        if driver_id == 0:
            driver_id = None
        rec = PenaltyRecord(
            record_date=record_date, district_id=district_id, project_id=project_id, vehicle_id=vehicle_id, driver_id=driver_id,
            fine=form.fine.data.strip() or None, remarks=form.remarks.data.strip() or None
        )
        db.session.add(rec)
        db.session.commit()
        flash('Penalty record saved.', 'success')
        return redirect(url_for('penalty_record_list'))
    return render_template('penalty_record_form.html', form=form, title='Add Penalty Record')


@app.route('/penalty-record/<int:pk>/edit', methods=['GET', 'POST'])
def penalty_record_edit(pk):
    rec = PenaltyRecord.query.get_or_404(pk)
    form = PenaltyRecordForm(obj=rec)
    form.district_id.choices = [(0, '-- Select District --')] + [(d.id, d.name) for d in District.query.order_by(District.name).all()]
    form.vehicle_id.choices = [(0, '-- Select Vehicle --')] + [(v.id, v.vehicle_no) for v in Vehicle.query.order_by(Vehicle.vehicle_no).all()]
    form.driver_id.choices = [(0, '-- Select Driver --')] + [(d.id, d.name) for d in Driver.query.order_by(Driver.name).all()]
    if rec.district_id:
        projects = Project.query.join(project_district).filter(project_district.c.district_id == rec.district_id).order_by(Project.name).all()
        form.project_id.choices = [(0, '-- Select Project --')] + [(p.id, p.name) for p in projects]
    else:
        form.project_id.choices = [(0, '-- Select Project --')]
    if request.method == 'GET':
        form.record_date.data = rec.record_date
        form.district_id.data = rec.district_id or 0
        form.project_id.data = rec.project_id or 0
        form.vehicle_id.data = rec.vehicle_id or 0
        form.driver_id.data = rec.driver_id or 0
        form.fine.data = rec.fine or ''
        form.remarks.data = rec.remarks or ''
    if request.method == 'POST' and form.validate_on_submit():
        record_date = form.record_date.data
        district_id = form.district_id.data or None
        if district_id == 0:
            district_id = None
        project_id = form.project_id.data or None
        if project_id == 0:
            project_id = None
        vehicle_id = form.vehicle_id.data or None
        if vehicle_id == 0:
            vehicle_id = None
        driver_id = form.driver_id.data or None
        if driver_id == 0:
            driver_id = None
        rec.record_date = record_date
        rec.district_id = district_id
        rec.project_id = project_id
        rec.vehicle_id = vehicle_id
        rec.driver_id = driver_id
        rec.fine = form.fine.data.strip() or None
        rec.remarks = form.remarks.data.strip() or None
        db.session.commit()
        flash('Penalty record updated.', 'success')
        return redirect(url_for('penalty_record_list'))
    return render_template('penalty_record_form.html', form=form, title='Edit Penalty Record', rec=rec)


# ────────────────────────────────────────────────
# Party Name (Pump / Workshop / Spare parts shop)
# ────────────────────────────────────────────────
@app.route('/driver-posts')
def driver_post_list():
    search = request.args.get('search', '').strip()
    query = EmployeePost.query
    if search:
        query = query.filter(EmployeePost.full_name.ilike(f'%{search}%'))
    posts = query.order_by(EmployeePost.full_name).all()
    return render_template('driver_post_list.html', posts=posts, search=search)


@app.route('/driver-post/add', methods=['GET', 'POST'])
@app.route('/driver-post/edit/<int:id>', methods=['GET', 'POST'])
def driver_post_form(id=None):
    post = EmployeePost.query.get_or_404(id) if id else None
    form = EmployeePostForm(obj=post)
    if form.validate_on_submit():
        if not post:
            post = EmployeePost()
        post.short_name = form.short_name.data.strip()
        post.full_name = form.full_name.data.strip()
        post.remarks = form.remarks.data.strip() if form.remarks.data else None
        if not id:
            db.session.add(post)
        db.session.commit()
        flash('Post saved.', 'success')
        return redirect(url_for('driver_post_list'))
    return render_template('driver_post_form.html', form=form, post=post)


@app.route('/driver-post/delete/<int:id>', methods=['POST'])
def driver_post_delete(id):
    post = EmployeePost.query.get_or_404(id)
    db.session.delete(post)
    db.session.commit()
    flash('Post deleted.', 'success')
    return redirect(url_for('driver_post_list'))


@app.route('/parties')
def party_list():
    search = request.args.get('search', '').strip()
    party_type = request.args.get('type', '').strip()
    query = Party.query
    if party_type:
        query = query.filter(Party.party_type == party_type)
    if search:
        query = query.filter(Party.name.ilike(f'%{search}%'))
    parties = query.order_by(Party.party_type, Party.name).all()
    return render_template('party_list.html', parties=parties, search=search, party_type=party_type)


@app.route('/party/add', methods=['GET', 'POST'])
@app.route('/party/edit/<int:id>', methods=['GET', 'POST'])
def party_form(id=None):
    party = Party.query.get_or_404(id) if id else None
    form = PartyForm(obj=party)
    if request.method == 'GET' and request.args.get('type'):
        form.party_type.data = request.args.get('type')
    if form.validate_on_submit():
        if not party:
            party = Party()
        form.populate_obj(party)
        if not id:
            db.session.add(party)
        db.session.commit()
        flash('Party saved.', 'success')
        next_url = request.form.get('next') or request.args.get('next') or url_for('party_list')
        return redirect(next_url)
    return render_template('party_form.html', form=form, party=party, title='Edit Party' if id else 'Create Party Name')


@app.route('/party/delete/<int:id>', methods=['POST'])
def party_delete(id):
    party = Party.query.get_or_404(id)
    db.session.delete(party)
    db.session.commit()
    flash('Party deleted.', 'success')
    return redirect(url_for('party_list'))


# ────────────────────────────────────────────────
# Products Name (used in Fueling / Oil / Maintenance)
# ────────────────────────────────────────────────
def _ensure_product_used_in_forms_column():
    """Add used_in_forms column to product table if missing (e.g. old DB before migration)."""
    try:
        insp = inspect(db.engine)
        if 'product' not in insp.get_table_names():
            return
        cols = [c['name'] for c in insp.get_columns('product')]
        if 'used_in_forms' in cols:
            return
        db.session.execute(text('ALTER TABLE product ADD COLUMN used_in_forms VARCHAR(100)'))
        db.session.commit()
    except Exception:
        db.session.rollback()


@app.route('/products')
def product_list():
    _ensure_product_used_in_forms_column()
    search = request.args.get('search', '').strip()
    query = Product.query
    if search:
        query = query.filter(Product.name.ilike(f'%{search}%'))
    products = query.order_by(Product.name).all()
    return render_template('product_list.html', products=products, search=search)


@app.route('/product/add', methods=['GET', 'POST'])
@app.route('/product/edit/<int:id>', methods=['GET', 'POST'])
def product_form(id=None):
    _ensure_product_used_in_forms_column()
    product = Product.query.get_or_404(id) if id else None
    form = ProductForm(obj=product)
    if request.method == 'GET' and product and product.used_in_forms:
        form.used_in_forms.data = [x.strip() for x in product.used_in_forms.split(',') if x.strip()]
    if form.validate_on_submit():
        if not product:
            product = Product()
        form.populate_obj(product)
        product.used_in_forms = ','.join(form.used_in_forms.data) if form.used_in_forms.data else None
        if not id:
            db.session.add(product)
        db.session.commit()
        flash('Product saved.', 'success')
        next_url = request.form.get('next') or request.args.get('next') or url_for('product_list')
        return redirect(next_url)
    return render_template('product_form.html', form=form, product=product, title='Edit Product' if id else 'Create Product Name')


@app.route('/product/delete/<int:id>', methods=['POST'])
def product_delete(id):
    product = Product.query.get_or_404(id)
    db.session.delete(product)
    db.session.commit()
    flash('Product deleted.', 'success')
    return redirect(url_for('product_list'))


@app.route('/api/parties')
def api_parties():
    """List parties for dropdown. Optional: type=Pump, search=xxx."""
    party_type = request.args.get('type', '').strip()
    search = request.args.get('search', '').strip()
    query = Party.query
    if party_type:
        query = query.filter(Party.party_type == party_type)
    if search:
        query = query.filter(Party.name.ilike(f'%{search}%'))
    parties = query.order_by(Party.name).limit(100).all()
    return jsonify([{'id': p.id, 'name': p.name, 'party_type': p.party_type} for p in parties])


# ────────────────────────────────────────────────
# Fuel Expense
# ────────────────────────────────────────────────
def _fuel_expense_task_readings(vehicle_id, task_date):
    """Return (km_out_day_start, km_in_day_close) from VehicleDailyTask for given vehicle and date. None if no task."""
    task = VehicleDailyTask.query.filter_by(vehicle_id=vehicle_id, task_date=task_date).first()
    if not task:
        return None, None
    prev = VehicleDailyTask.query.filter(
        VehicleDailyTask.vehicle_id == vehicle_id,
        VehicleDailyTask.task_date < task_date
    ).order_by(VehicleDailyTask.task_date.desc()).first()
    km_out = float(prev.close_reading) if prev else None
    km_in = float(task.close_reading)
    return km_out, km_in


@app.route('/api/fuel-expense/last-reading')
def api_fuel_expense_last_reading():
    """Return last fueling entry's current_reading for vehicle_id (for Previous Reading)."""
    vehicle_id = request.args.get('vehicle_id', type=int)
    if not vehicle_id:
        return jsonify({'previous_reading': None})
    last_entry = FuelExpense.query.filter_by(vehicle_id=vehicle_id).order_by(FuelExpense.fueling_date.desc(), FuelExpense.id.desc()).first()
    return jsonify({'previous_reading': float(last_entry.current_reading) if last_entry and last_entry.current_reading else None})


@app.route('/api/fuel-expense/task-readings')
def api_fuel_expense_task_readings():
    """Return km_out (Day Start) and km_in (Day Close) from task report for vehicle_id and fueling_date."""
    vehicle_id = request.args.get('vehicle_id', type=int)
    date_str = request.args.get('fueling_date', '')
    if not vehicle_id or not date_str:
        return jsonify({'km_out_task': None, 'km_in_task': None})
    task_date = parse_date(date_str)
    if not task_date:
        return jsonify({'km_out_task': None, 'km_in_task': None})
    km_out, km_in = _fuel_expense_task_readings(vehicle_id, task_date)
    return jsonify({'km_out_task': km_out, 'km_in_task': km_in})


@app.route('/api/fuel-expense/suggested-price')
def api_fuel_expense_suggested_price():
    """Return latest fuel_price for the given fuel_type (Diesel or Super) from last expense entry."""
    fuel_type = request.args.get('fuel_type', '').strip()
    if fuel_type not in ('Diesel', 'Super'):
        return jsonify({'fuel_price': None})
    last_entry = FuelExpense.query.filter(
        FuelExpense.fuel_type == fuel_type,
        FuelExpense.fuel_price.isnot(None)
    ).order_by(FuelExpense.fueling_date.desc(), FuelExpense.id.desc()).first()
    return jsonify({'fuel_price': float(last_entry.fuel_price) if last_entry and last_entry.fuel_price else None})


@app.route('/api/fuel-expense/price-hint')
def api_fuel_expense_price_hint():
    """Return last 2 entries for current pump + fuel_type, and 2 other pumps' latest entry (same fuel_type) for price reference."""
    fuel_type = request.args.get('fuel_type', '').strip()
    fuel_pump_id = request.args.get('fuel_pump_id', type=int)
    if fuel_type not in ('Diesel', 'Super'):
        return jsonify({'current_pump': [], 'other_pumps': []})
    current_pump = []
    if fuel_pump_id:
        rows = FuelExpense.query.filter(
            FuelExpense.fuel_pump_id == fuel_pump_id,
            FuelExpense.fuel_type == fuel_type,
            FuelExpense.fuel_price.isnot(None)
        ).order_by(FuelExpense.fueling_date.desc(), FuelExpense.id.desc()).limit(2).all()
        pump_name = Party.query.get(fuel_pump_id).name if Party.query.get(fuel_pump_id) else ''
        for r in rows:
            current_pump.append({
                'fuel_price': float(r.fuel_price),
                'date': r.fueling_date.strftime('%d-%m-%Y') if r.fueling_date else '',
                'pump_name': pump_name,
            })
    other_pumps = []
    q_other = db.session.query(FuelExpense.fuel_pump_id).filter(
        FuelExpense.fuel_type == fuel_type,
        FuelExpense.fuel_price.isnot(None),
        FuelExpense.fuel_pump_id.isnot(None),
    )
    if fuel_pump_id:
        q_other = q_other.filter(FuelExpense.fuel_pump_id != fuel_pump_id)
    other_pump_ids = [x[0] for x in q_other.distinct().limit(2).all() if x[0]]
    for pid in other_pump_ids:
        last_row = FuelExpense.query.filter(
            FuelExpense.fuel_pump_id == pid,
            FuelExpense.fuel_type == fuel_type,
            FuelExpense.fuel_price.isnot(None)
        ).order_by(FuelExpense.fueling_date.desc(), FuelExpense.id.desc()).first()
        if last_row:
            p = Party.query.get(pid)
            other_pumps.append({
                'pump_name': p.name if p else '',
                'fuel_price': float(last_row.fuel_price),
                'date': last_row.fueling_date.strftime('%d-%m-%Y') if last_row.fueling_date else '',
            })
    return jsonify({'current_pump': current_pump, 'other_pumps': other_pumps})


@app.route('/expenses/fuel')
def fuel_expense_list():
    form = FuelExpenseFilterForm()
    form.district_id.choices = [(0, '-- Select District --')] + [(d.id, d.name) for d in District.query.order_by(District.name).all()]
    form.project_id.choices = [(0, '-- Select Project --')]
    form.vehicle_id.choices = [(0, '-- All Vehicles --')]
    today = date.today()
    from_date = request.args.get('from_date')
    to_date = request.args.get('to_date')
    district_id = request.args.get('district_id', type=int) or 0
    project_id = request.args.get('project_id', type=int) or 0
    vehicle_id = request.args.get('vehicle_id', type=int) or 0
    if request.method == 'POST':
        from_date = request.form.get('from_date')
        to_date = request.form.get('to_date')
        district_id = request.form.get('district_id', type=int) or 0
        project_id = request.form.get('project_id', type=int) or 0
        vehicle_id = request.form.get('vehicle_id', type=int) or 0
        return redirect(url_for('fuel_expense_list', from_date=from_date or '', to_date=to_date or '', district_id=district_id, project_id=project_id, vehicle_id=vehicle_id))
    from_d = parse_date(from_date) if from_date else today
    to_d = parse_date(to_date) if to_date else today
    if from_d and to_d and from_d > to_d:
        from_d, to_d = to_d, from_d
    form.from_date.data = from_d
    form.to_date.data = to_d
    form.district_id.data = district_id
    form.project_id.data = project_id
    form.vehicle_id.data = vehicle_id
    if district_id:
        projects = Project.query.join(project_district).filter(project_district.c.district_id == district_id).order_by(Project.name).all()
        form.project_id.choices = [(0, '-- Select Project --')] + [(p.id, p.name) for p in projects]
    if project_id:
        q = Vehicle.query.filter(Vehicle.project_id == project_id)
        if district_id:
            q = q.filter(Vehicle.district_id == district_id)
        vehicles = q.order_by(Vehicle.vehicle_no).all()
        form.vehicle_id.choices = [(0, '-- All Vehicles --')] + [(v.id, v.vehicle_no) for v in vehicles]
    query = FuelExpense.query.filter(
        FuelExpense.fueling_date >= from_d,
        FuelExpense.fueling_date <= to_d
    )
    if district_id:
        query = query.filter(FuelExpense.district_id == district_id)
    if project_id:
        query = query.filter(FuelExpense.project_id == project_id)
    if vehicle_id:
        query = query.filter(FuelExpense.vehicle_id == vehicle_id)
    rows = query.order_by(FuelExpense.fueling_date.desc(), FuelExpense.id.desc()).all()
    totals = {}
    if rows:
        total_km = sum(float(r.km or 0) for r in rows)
        total_liters = sum(float(r.liters or 0) for r in rows)
        total_amount = sum(float(r.amount or 0) for r in rows)
        first_prev = rows[-1].previous_reading
        last_curr = rows[0].current_reading
        avg_mpg = round(total_km / total_liters, 2) if total_liters else None
        totals = {'total_km': total_km, 'total_liters': total_liters, 'total_amount': total_amount,
                  'first_previous_reading': float(first_prev) if first_prev else None,
                  'last_current_reading': float(last_curr) if last_curr else None,
                  'avg_mpg': avg_mpg}
    return render_template('fuel_expense_list.html', form=form, rows=rows, from_date=from_d, to_date=to_d, totals=totals)


@app.route('/expenses/fuel/add', methods=['GET', 'POST'])
def fuel_expense_add():
    form = FuelExpenseForm()
    form.district_id.choices = [(0, '-- Select District --')] + [(d.id, d.name) for d in District.query.order_by(District.name).all()]
    form.project_id.choices = [(0, '-- Select Project --')]
    form.vehicle_id.choices = [(0, '-- Select Vehicle --')]
    pumps = Party.query.filter_by(party_type='Pump').order_by(Party.name).all()
    form.fuel_pump_id.choices = [(0, '-- Select Pump --')] + [(p.id, p.name) for p in pumps]
    if request.method == 'GET':
        district_id = request.args.get('district_id', type=int)
        project_id = request.args.get('project_id', type=int)
        if district_id:
            projects = Project.query.join(project_district).filter(project_district.c.district_id == district_id).order_by(Project.name).all()
            form.project_id.choices = [(0, '-- Select Project --')] + [(p.id, p.name) for p in projects]
        if project_id:
            q = Vehicle.query.filter(Vehicle.project_id == project_id)
            if district_id:
                q = q.filter(Vehicle.district_id == district_id)
            vehicles = q.order_by(Vehicle.vehicle_no).all()
            form.vehicle_id.choices = [(0, '-- Select Vehicle --')] + [(v.id, v.vehicle_no) for v in vehicles]
        form.district_id.data = district_id or 0
        form.project_id.data = project_id or 0
    if request.method == 'POST' and form.validate_on_submit():
        vehicle_id = form.vehicle_id.data
        if vehicle_id == 0:
            flash('Please select a vehicle.', 'danger')
            return render_template('fuel_expense_form.html', form=form, title='Add Fuel Expense')
        vehicle = Vehicle.query.get_or_404(vehicle_id)
        district_id = form.district_id.data or None
        if district_id == 0:
            district_id = None
        project_id = form.project_id.data or None
        if project_id == 0:
            project_id = vehicle.project_id
        fueling_date = form.fueling_date.data
        card_swipe_date = form.card_swipe_date.data
        payment_type = (form.payment_type.data or '').strip() or None
        slip_no = (form.slip_no.data or '').strip() or None
        fuel_type = (form.fuel_type.data or '').strip() or None
        fuel_pump_id = form.fuel_pump_id.data or None
        if fuel_pump_id == 0:
            fuel_pump_id = None
        previous_reading = form.previous_reading.data
        current_reading = form.current_reading.data
        if previous_reading is None:
            last_entry = FuelExpense.query.filter_by(vehicle_id=vehicle_id).order_by(FuelExpense.fueling_date.desc(), FuelExpense.id.desc()).first()
            previous_reading = float(last_entry.current_reading) if last_entry and last_entry.current_reading else 0
        prev_f = float(previous_reading)
        curr_f = float(current_reading)
        km = curr_f - prev_f if curr_f >= prev_f else 0
        amount = form.amount.data
        fuel_price = form.fuel_price.data
        amount_f = float(amount) if amount else 0
        fuel_price_f = float(fuel_price) if fuel_price else 0
        liters = round(amount_f / fuel_price_f, 2) if fuel_price_f else None
        mpg = round(km / float(liters), 2) if liters and km else None
        km_out_task, km_in_task = _fuel_expense_task_readings(vehicle_id, fueling_date)
        if km_out_task is not None and km_in_task is not None:
            matched = (abs(prev_f - km_out_task) < 0.01 and abs(curr_f - km_in_task) < 0.01)
            meter_reading_matched = 'Yes' if matched else 'No'
        else:
            meter_reading_matched = 'No'
        rec = FuelExpense(
            district_id=district_id, project_id=project_id, vehicle_id=vehicle_id,
            fueling_date=fueling_date, card_swipe_date=card_swipe_date,
            payment_type=payment_type, slip_no=slip_no, fuel_type=fuel_type, fuel_pump_id=fuel_pump_id,
            previous_reading=previous_reading, current_reading=current_reading,
            km=km, fuel_price=fuel_price, liters=liters, mpg=mpg, amount=amount,
            km_out_task=km_out_task, km_in_task=km_in_task, meter_reading_matched=meter_reading_matched
        )
        db.session.add(rec)
        db.session.commit()
        flash('Fuel expense saved.', 'success')
        return redirect(url_for('fuel_expense_list'))
    return render_template('fuel_expense_form.html', form=form, title='Add Fuel Expense')


@app.route('/expenses/fuel/<int:pk>/edit', methods=['GET', 'POST'])
def fuel_expense_edit(pk):
    rec = FuelExpense.query.get_or_404(pk)
    form = FuelExpenseForm(obj=rec)
    form.district_id.choices = [(0, '-- Select District --')] + [(d.id, d.name) for d in District.query.order_by(District.name).all()]
    if rec.district_id:
        projects = Project.query.join(project_district).filter(project_district.c.district_id == rec.district_id).order_by(Project.name).all()
        form.project_id.choices = [(0, '-- Select Project --')] + [(p.id, p.name) for p in projects]
    else:
        form.project_id.choices = [(0, '-- Select Project --')]
    q = Vehicle.query
    if rec.project_id:
        q = q.filter(Vehicle.project_id == rec.project_id)
    if rec.district_id:
        q = q.filter(Vehicle.district_id == rec.district_id)
    vehicles = q.order_by(Vehicle.vehicle_no).all()
    form.vehicle_id.choices = [(0, '-- Select Vehicle --')] + [(v.id, v.vehicle_no) for v in vehicles]
    pumps = Party.query.filter_by(party_type='Pump').order_by(Party.name).all()
    form.fuel_pump_id.choices = [(0, '-- Select Pump --')] + [(p.id, p.name) for p in pumps]
    if request.method == 'GET':
        form.district_id.data = rec.district_id or 0
        form.project_id.data = rec.project_id or 0
        form.vehicle_id.data = rec.vehicle_id
        form.fueling_date.data = rec.fueling_date
        form.card_swipe_date.data = rec.card_swipe_date
        form.payment_type.data = rec.payment_type or ''
        form.slip_no.data = rec.slip_no or ''
        form.fuel_type.data = rec.fuel_type or ''
        form.fuel_pump_id.data = rec.fuel_pump_id or 0
        form.previous_reading.data = rec.previous_reading
        form.current_reading.data = rec.current_reading
        form.amount.data = rec.amount
        form.fuel_price.data = rec.fuel_price
    if request.method == 'POST' and form.validate_on_submit():
        vehicle_id = form.vehicle_id.data
        if vehicle_id == 0:
            flash('Please select a vehicle.', 'danger')
            return render_template('fuel_expense_form.html', form=form, title='Edit Fuel Expense', rec=rec)
        district_id = form.district_id.data or None
        if district_id == 0:
            district_id = None
        project_id = form.project_id.data or None
        if project_id == 0:
            project_id = None
        fueling_date = form.fueling_date.data
        card_swipe_date = form.card_swipe_date.data
        payment_type = (form.payment_type.data or '').strip() or None
        slip_no = (form.slip_no.data or '').strip() or None
        fuel_type = (form.fuel_type.data or '').strip() or None
        fuel_pump_id = form.fuel_pump_id.data or None
        if fuel_pump_id == 0:
            fuel_pump_id = None
        previous_reading = form.previous_reading.data
        current_reading = form.current_reading.data
        if previous_reading is None:
            last_entry = FuelExpense.query.filter(
                FuelExpense.vehicle_id == vehicle_id,
                FuelExpense.id != pk
            ).order_by(FuelExpense.fueling_date.desc(), FuelExpense.id.desc()).first()
            previous_reading = float(last_entry.current_reading) if last_entry and last_entry.current_reading else 0
        prev_f = float(previous_reading)
        curr_f = float(current_reading)
        km = curr_f - prev_f if curr_f >= prev_f else 0
        amount = form.amount.data
        fuel_price = form.fuel_price.data
        amount_f = float(amount) if amount else 0
        fuel_price_f = float(fuel_price) if fuel_price else 0
        liters = round(amount_f / fuel_price_f, 2) if fuel_price_f else None
        mpg = round(km / float(liters), 2) if liters and km else None
        km_out_task, km_in_task = _fuel_expense_task_readings(vehicle_id, fueling_date)
        if km_out_task is not None and km_in_task is not None:
            matched = (abs(prev_f - km_out_task) < 0.01 and abs(curr_f - km_in_task) < 0.01)
            meter_reading_matched = 'Yes' if matched else 'No'
        else:
            meter_reading_matched = 'No'
        rec.district_id = district_id
        rec.project_id = project_id
        rec.vehicle_id = vehicle_id
        rec.fueling_date = fueling_date
        rec.card_swipe_date = card_swipe_date
        rec.payment_type = payment_type
        rec.slip_no = slip_no
        rec.fuel_type = fuel_type
        rec.fuel_pump_id = fuel_pump_id
        rec.previous_reading = previous_reading
        rec.current_reading = current_reading
        rec.km = km
        rec.fuel_price = fuel_price
        rec.liters = liters
        rec.mpg = mpg
        rec.amount = amount
        rec.km_out_task = km_out_task
        rec.km_in_task = km_in_task
        rec.meter_reading_matched = meter_reading_matched
        db.session.commit()
        flash('Fuel expense updated.', 'success')
        return redirect(url_for('fuel_expense_list'))
    return render_template('fuel_expense_form.html', form=form, title='Edit Fuel Expense', rec=rec)


@app.route('/expenses/fuel/<int:pk>/delete', methods=['POST'])
def fuel_expense_delete(pk):
    rec = FuelExpense.query.get_or_404(pk)
    db.session.delete(rec)
    db.session.commit()
    flash('Fuel expense deleted.', 'success')
    return redirect(url_for('fuel_expense_list'))


# ────────────────────────────────────────────────
# Oil Expense
# ────────────────────────────────────────────────
@app.route('/api/oil-expense/last-reading')
def api_oil_expense_last_reading():
    vehicle_id = request.args.get('vehicle_id', type=int)
    if not vehicle_id:
        return jsonify({})
    last_entry = OilExpense.query.filter_by(vehicle_id=vehicle_id).order_by(
        OilExpense.expense_date.desc(), OilExpense.id.desc()
    ).first()
    if not last_entry or last_entry.current_reading is None:
        return jsonify({})
    return jsonify({'previous_reading': float(last_entry.current_reading)})


@app.route('/api/oil-expense/products-for-oil')
def api_oil_expense_products_for_oil():
    products = Product.query.filter(
        db.or_(
            Product.used_in_forms.is_(None),
            Product.used_in_forms == '',
            Product.used_in_forms.like('%Oil%')
        )
    ).order_by(Product.name).all()
    return jsonify([{'id': p.id, 'name': p.name} for p in products])


@app.route('/api/oil-expense/product-balance/<int:product_id>')
def api_oil_expense_product_balance(product_id):
    bal = ProductBalance.query.filter_by(product_id=product_id).first()
    qty = float(bal.balance_qty) if bal and bal.balance_qty is not None else 0
    return jsonify({'product_id': product_id, 'balance_qty': qty})


def _get_or_create_product_balance(product_id):
    bal = ProductBalance.query.filter_by(product_id=product_id).first()
    if not bal:
        bal = ProductBalance(product_id=product_id, balance_qty=0)
        db.session.add(bal)
    return bal


def _apply_oil_expense_items_balance(items, reverse=False):
    for item in items:
        bal = _get_or_create_product_balance(item.product_id)
        purchase_qty = float(item.purchase_qty or 0)
        used_qty = float(item.used_qty or 0)
        delta = (purchase_qty - used_qty) if not reverse else (used_qty - purchase_qty)
        if delta:
            bal.balance_qty = (bal.balance_qty or 0) + delta
    db.session.flush()


@app.route('/oil-expenses')
def oil_expense_list():
    form = OilExpenseFilterForm()
    form.district_id.choices = [(0, '-- Select District --')] + [(d.id, d.name) for d in District.query.order_by(District.name).all()]
    form.project_id.choices = [(0, '-- Select Project --')] + [(p.id, p.name) for p in Project.query.order_by(Project.name).all()]
    form.vehicle_id.choices = [(0, '-- All Vehicles --')] + [(v.id, v.vehicle_no) for v in Vehicle.query.order_by(Vehicle.vehicle_no).all()]
    from_date = request.args.get('from_date', '').strip()
    to_date = request.args.get('to_date', '').strip()
    district_id = request.args.get('district_id', type=int) or 0
    project_id = request.args.get('project_id', type=int) or 0
    vehicle_id = request.args.get('vehicle_id', type=int) or 0
    if from_date or to_date or district_id or project_id or vehicle_id:
        from_d = parse_date(from_date) if from_date else date.today().replace(day=1)
        to_d = parse_date(to_date) if to_date else date.today()
        if not from_d:
            from_d = date.today().replace(day=1)
        if not to_d:
            to_d = date.today()
        query = OilExpense.query.filter(
            OilExpense.expense_date >= from_d,
            OilExpense.expense_date <= to_d
        )
        if district_id:
            query = query.filter(OilExpense.district_id == district_id)
        if project_id:
            query = query.filter(OilExpense.project_id == project_id)
        if vehicle_id:
            query = query.filter(OilExpense.vehicle_id == vehicle_id)
        rows = query.order_by(OilExpense.expense_date.desc(), OilExpense.id.desc()).all()
    else:
        from_d = to_d = None
        rows = []
    # Attach item totals per row for list display
    rows_with_totals = []
    for r in rows:
        total_purchase = sum(float(it.purchase_qty or 0) for it in r.items)
        total_used = sum(float(it.used_qty or 0) for it in r.items)
        total_balance = total_purchase - total_used
        total_amount = sum(float(it.amount or 0) for it in r.items)
        rows_with_totals.append({
            'rec': r,
            'total_purchase_qty': total_purchase,
            'total_used_qty': total_used,
            'total_balance_qty': total_balance,
            'total_amount': total_amount
        })
    totals = {'count': len(rows)}
    return render_template('oil_expense_list.html', form=form, rows=rows_with_totals, from_date=from_d, to_date=to_d, totals=totals)


@app.route('/oil-expense/add', methods=['GET', 'POST'])
@app.route('/oil-expense/edit/<int:pk>', methods=['GET', 'POST'])
def oil_expense_form(pk=None):
    rec = OilExpense.query.get_or_404(pk) if pk else None
    form = OilExpenseForm(obj=rec)
    products_for_oil = Product.query.filter(
        db.or_(
            Product.used_in_forms.is_(None),
            Product.used_in_forms == '',
            Product.used_in_forms.like('%Oil%')
        )
    ).order_by(Product.name).all()
    form.district_id.choices = [(0, '-- Select District --')] + [(d.id, d.name) for d in District.query.order_by(District.name).all()]
    form.project_id.choices = [(0, '-- Select Project --')] + [(p.id, p.name) for p in Project.query.order_by(Project.name).all()]
    form.vehicle_id.choices = [(v.id, v.vehicle_no) for v in Vehicle.query.order_by(Vehicle.vehicle_no).all()]
    if not form.vehicle_id.choices:
        form.vehicle_id.choices = [(0, '-- No Vehicle --')]
    else:
        form.vehicle_id.choices.insert(0, (0, '-- Select Vehicle --'))

    if request.method == 'GET' and rec:
        if rec.district_id:
            form.district_id.data = rec.district_id
        if rec.project_id:
            form.project_id.data = rec.project_id
        if rec.vehicle_id:
            form.vehicle_id.data = rec.vehicle_id

    if form.validate_on_submit():
        vehicle_id = form.vehicle_id.data
        if not vehicle_id:
            flash('Select vehicle.', 'danger')
            return render_template('oil_expense_form.html', form=form, rec=rec, title='Edit Oil Expense' if rec else 'Add Oil Expense', products_for_oil=products_for_oil)
        expense_date = form.expense_date.data
        card_swipe_date = form.card_swipe_date.data
        prev_reading = form.previous_reading.data
        curr_reading = form.current_reading.data
        km = None
        if prev_reading is not None and curr_reading is not None:
            try:
                km = float(curr_reading) - float(prev_reading)
            except (TypeError, ValueError):
                pass
        remarks = form.remarks.data

        product_ids = request.form.getlist('product_id')
        payment_types = request.form.getlist('payment_type')
        purchase_qtys = request.form.getlist('purchase_qty')
        used_qtys = request.form.getlist('used_qty')
        prices = request.form.getlist('price')
        items_data = []
        n = max(len(product_ids or [0]), len(payment_types or [0]), len(purchase_qtys or [0]), len(used_qtys or [0]), len(prices or [0]))
        for i in range(n):
            pid = product_ids[i] if i < len(product_ids or []) else None
            try:
                pid = int(pid) if pid else None
            except (TypeError, ValueError):
                pid = None
            if not pid:
                continue
            pt = payment_types[i] if i < len(payment_types or []) else ''
            try:
                purchase_qty = float(purchase_qtys[i]) if i < len(purchase_qtys or []) and purchase_qtys[i] else 0
            except (TypeError, ValueError):
                purchase_qty = 0
            try:
                used_qty = float(used_qtys[i]) if i < len(used_qtys or []) and used_qtys[i] else 0
            except (TypeError, ValueError):
                used_qty = 0
            try:
                price = float(prices[i]) if i < len(prices or []) and prices[i] else 0
            except (TypeError, ValueError):
                price = 0
            amount = (purchase_qty * price) if price else None
            items_data.append({
                'product_id': pid, 'payment_type': pt,
                'purchase_qty': purchase_qty, 'used_qty': used_qty,
                'price': price, 'amount': amount
            })

        if rec:
            old_items = list(rec.items.all())
            _apply_oil_expense_items_balance(old_items, reverse=True)
            rec.items.delete()
        else:
            rec = OilExpense(
                district_id=form.district_id.data or None,
                project_id=form.project_id.data or None,
                vehicle_id=vehicle_id,
                expense_date=expense_date,
                card_swipe_date=card_swipe_date,
                previous_reading=prev_reading,
                current_reading=curr_reading,
                km=km,
                remarks=remarks
            )
            db.session.add(rec)
        db.session.flush()

        if rec.id:
            rec.district_id = form.district_id.data or None
            rec.project_id = form.project_id.data or None
            rec.vehicle_id = vehicle_id
            rec.expense_date = expense_date
            rec.card_swipe_date = card_swipe_date
            rec.previous_reading = prev_reading
            rec.current_reading = curr_reading
            rec.km = km
            rec.remarks = remarks

        for idx, it in enumerate(items_data):
            item = OilExpenseItem(
                oil_expense_id=rec.id,
                product_id=it['product_id'],
                payment_type=it['payment_type'] or None,
                purchase_qty=it['purchase_qty'],
                used_qty=it['used_qty'],
                qty=it['purchase_qty'],
                price=it['price'],
                amount=it['amount'],
                sort_order=idx
            )
            db.session.add(item)
        db.session.flush()
        _apply_oil_expense_items_balance(rec.items.all(), reverse=False)
        db.session.commit()

        allowed_image = {'image/jpeg', 'image/png', 'image/gif', 'image/webp'}
        allowed_video = {'video/mp4', 'video/webm', 'video/quicktime'}
        files = request.files.getlist('attachments')
        if files:
            subdir = os.path.join(app.config['UPLOAD_FOLDER'], 'oil_expense', str(rec.id))
            os.makedirs(subdir, exist_ok=True)
            for f in files:
                if not f or not f.filename:
                    continue
                fn = secure_filename(f.filename)
                if not fn:
                    continue
                ext = os.path.splitext(fn)[1].lower()
                content_type = f.content_type or ''
                if content_type in allowed_image or ext in {'.jpg', '.jpeg', '.png', '.gif', '.webp'}:
                    file_type = 'image'
                elif content_type in allowed_video or ext in {'.mp4', '.webm', '.mov'}:
                    file_type = 'video'
                else:
                    continue
                base, ext = os.path.splitext(fn)
                unique = f"{base}_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}{ext}"
                path = os.path.join(subdir, unique)
                f.save(path)
                rel_path = os.path.join('oil_expense', str(rec.id), unique)
                att = OilExpenseAttachment(oil_expense_id=rec.id, file_path=rel_path, file_type=file_type, original_name=fn)
                db.session.add(att)
            db.session.commit()

        flash('Oil expense saved.', 'success')
        return redirect(url_for('oil_expense_list'))
    return render_template('oil_expense_form.html', form=form, rec=rec, title='Edit Oil Expense' if rec else 'Add Oil Expense', products_for_oil=products_for_oil)


@app.route('/oil-expense/delete/<int:pk>', methods=['POST'])
def oil_expense_delete(pk):
    rec = OilExpense.query.get_or_404(pk)
    items = list(rec.items.all())
    _apply_oil_expense_items_balance(items, reverse=True)
    for att in rec.attachments:
        full_path = os.path.join(app.config['UPLOAD_FOLDER'], att.file_path)
        if os.path.isfile(full_path):
            try:
                os.remove(full_path)
            except OSError:
                pass
    db.session.delete(rec)
    db.session.commit()
    flash('Oil expense deleted.', 'success')
    return redirect(url_for('oil_expense_list'))


# ────────────────────────────────────────────────
# Maintenance Expense
# ────────────────────────────────────────────────
@app.route('/api/maintenance-expense/last-reading')
def api_maintenance_expense_last_reading():
    vehicle_id = request.args.get('vehicle_id', type=int)
    if not vehicle_id:
        return jsonify({})
    last_entry = MaintenanceExpense.query.filter_by(vehicle_id=vehicle_id).order_by(
        MaintenanceExpense.expense_date.desc(), MaintenanceExpense.id.desc()
    ).first()
    if not last_entry or last_entry.current_reading is None:
        return jsonify({})
    return jsonify({'previous_reading': float(last_entry.current_reading)})


@app.route('/api/maintenance-expense/products-for-maintenance')
def api_maintenance_expense_products():
    products = Product.query.filter(
        db.or_(
            Product.used_in_forms.is_(None),
            Product.used_in_forms == '',
            Product.used_in_forms.like('%Maintenance%')
        )
    ).order_by(Product.name).all()
    return jsonify([{'id': p.id, 'name': p.name} for p in products])


@app.route('/maintenance-expenses')
def maintenance_expense_list():
    form = MaintenanceExpenseFilterForm()
    form.district_id.choices = [(0, '-- Select District --')] + [(d.id, d.name) for d in District.query.order_by(District.name).all()]
    form.project_id.choices = [(0, '-- Select Project --')] + [(p.id, p.name) for p in Project.query.order_by(Project.name).all()]
    form.vehicle_id.choices = [(0, '-- All Vehicles --')] + [(v.id, v.vehicle_no) for v in Vehicle.query.order_by(Vehicle.vehicle_no).all()]
    from_date = request.args.get('from_date', '').strip()
    to_date = request.args.get('to_date', '').strip()
    district_id = request.args.get('district_id', type=int) or 0
    project_id = request.args.get('project_id', type=int) or 0
    vehicle_id = request.args.get('vehicle_id', type=int) or 0
    if from_date or to_date or district_id or project_id or vehicle_id:
        from_d = parse_date(from_date) if from_date else date.today().replace(day=1)
        to_d = parse_date(to_date) if to_date else date.today()
        if not from_d:
            from_d = date.today().replace(day=1)
        if not to_d:
            to_d = date.today()
        query = MaintenanceExpense.query.filter(
            MaintenanceExpense.expense_date >= from_d,
            MaintenanceExpense.expense_date <= to_d
        )
        if district_id:
            query = query.filter(MaintenanceExpense.district_id == district_id)
        if project_id:
            query = query.filter(MaintenanceExpense.project_id == project_id)
        if vehicle_id:
            query = query.filter(MaintenanceExpense.vehicle_id == vehicle_id)
        rows = query.order_by(MaintenanceExpense.expense_date.desc(), MaintenanceExpense.id.desc()).all()
    else:
        from_d = to_d = None
        rows = []
    rows_with_totals = []
    for r in rows:
        total_qty = sum(float(it.qty or 0) for it in r.items)
        total_amount = sum(float(it.amount or 0) for it in r.items)
        rows_with_totals.append({'rec': r, 'total_qty': total_qty, 'total_amount': total_amount})
    totals = {'count': len(rows)}
    return render_template('maintenance_expense_list.html', form=form, rows=rows_with_totals, from_date=from_d, to_date=to_d, totals=totals)


@app.route('/maintenance-expense/add', methods=['GET', 'POST'])
@app.route('/maintenance-expense/edit/<int:pk>', methods=['GET', 'POST'])
def maintenance_expense_form(pk=None):
    rec = MaintenanceExpense.query.get_or_404(pk) if pk else None
    form = MaintenanceExpenseForm(obj=rec)
    products_for_maintenance = Product.query.filter(
        db.or_(
            Product.used_in_forms.is_(None),
            Product.used_in_forms == '',
            Product.used_in_forms.like('%Maintenance%')
        )
    ).order_by(Product.name).all()
    form.district_id.choices = [(0, '-- Select District --')] + [(d.id, d.name) for d in District.query.order_by(District.name).all()]
    form.project_id.choices = [(0, '-- Select Project --')] + [(p.id, p.name) for p in Project.query.order_by(Project.name).all()]
    form.vehicle_id.choices = [(v.id, v.vehicle_no) for v in Vehicle.query.order_by(Vehicle.vehicle_no).all()]
    if not form.vehicle_id.choices:
        form.vehicle_id.choices = [(0, '-- No Vehicle --')]
    else:
        form.vehicle_id.choices.insert(0, (0, '-- Select Vehicle --'))
    if request.method == 'GET' and rec:
        if rec.district_id:
            form.district_id.data = rec.district_id
        if rec.project_id:
            form.project_id.data = rec.project_id
        if rec.vehicle_id:
            form.vehicle_id.data = rec.vehicle_id

    if form.validate_on_submit():
        vehicle_id = form.vehicle_id.data
        if not vehicle_id:
            flash('Select vehicle.', 'danger')
            return render_template('maintenance_expense_form.html', form=form, rec=rec, title='Edit Maintenance' if rec else 'Add Maintenance', products_for_maintenance=products_for_maintenance)
        expense_date = form.expense_date.data
        prev_reading = form.previous_reading.data
        curr_reading = form.current_reading.data
        km = None
        if prev_reading is not None and curr_reading is not None:
            try:
                km = float(curr_reading) - float(prev_reading)
            except (TypeError, ValueError):
                pass
        remarks = form.remarks.data

        product_ids = request.form.getlist('product_id')
        qtys = request.form.getlist('qty')
        prices = request.form.getlist('price')
        items_data = []
        n = max(len(product_ids or [0]), len(qtys or [0]), len(prices or [0]))
        for i in range(n):
            pid = product_ids[i] if i < len(product_ids or []) else None
            try:
                pid = int(pid) if pid else None
            except (TypeError, ValueError):
                pid = None
            if not pid:
                continue
            try:
                qty = float(qtys[i]) if i < len(qtys or []) and qtys[i] else 0
            except (TypeError, ValueError):
                qty = 0
            try:
                price = float(prices[i]) if i < len(prices or []) and prices[i] else 0
            except (TypeError, ValueError):
                price = 0
            amount = (qty * price) if price else None
            items_data.append({'product_id': pid, 'qty': qty, 'price': price, 'amount': amount})

        if rec:
            rec.items.delete()
        else:
            rec = MaintenanceExpense(
                district_id=form.district_id.data or None,
                project_id=form.project_id.data or None,
                vehicle_id=vehicle_id,
                expense_date=expense_date,
                previous_reading=prev_reading,
                current_reading=curr_reading,
                km=km,
                remarks=remarks
            )
            db.session.add(rec)
        db.session.flush()
        if rec.id:
            rec.district_id = form.district_id.data or None
            rec.project_id = form.project_id.data or None
            rec.vehicle_id = vehicle_id
            rec.expense_date = expense_date
            rec.previous_reading = prev_reading
            rec.current_reading = curr_reading
            rec.km = km
            rec.remarks = remarks
        for idx, it in enumerate(items_data):
            item = MaintenanceExpenseItem(
                maintenance_expense_id=rec.id,
                product_id=it['product_id'],
                qty=it['qty'],
                price=it['price'],
                amount=it['amount'],
                sort_order=idx
            )
            db.session.add(item)
        db.session.commit()

        allowed_image = {'image/jpeg', 'image/png', 'image/gif', 'image/webp'}
        allowed_video = {'video/mp4', 'video/webm', 'video/quicktime'}
        files = request.files.getlist('attachments')
        if files:
            subdir = os.path.join(app.config['UPLOAD_FOLDER'], 'maintenance_expense', str(rec.id))
            os.makedirs(subdir, exist_ok=True)
            for f in files:
                if not f or not f.filename:
                    continue
                fn = secure_filename(f.filename)
                if not fn:
                    continue
                ext = os.path.splitext(fn)[1].lower()
                content_type = f.content_type or ''
                if content_type in allowed_image or ext in {'.jpg', '.jpeg', '.png', '.gif', '.webp'}:
                    file_type = 'image'
                elif content_type in allowed_video or ext in {'.mp4', '.webm', '.mov'}:
                    file_type = 'video'
                else:
                    continue
                base, ext = os.path.splitext(fn)
                unique = f"{base}_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}{ext}"
                path = os.path.join(subdir, unique)
                f.save(path)
                rel_path = os.path.join('maintenance_expense', str(rec.id), unique)
                att = MaintenanceExpenseAttachment(maintenance_expense_id=rec.id, file_path=rel_path, file_type=file_type, original_name=fn)
                db.session.add(att)
            db.session.commit()
        flash('Maintenance expense saved.', 'success')
        return redirect(url_for('maintenance_expense_list'))
    return render_template('maintenance_expense_form.html', form=form, rec=rec, title='Edit Maintenance' if rec else 'Add Maintenance', products_for_maintenance=products_for_maintenance)


@app.route('/maintenance-expense/delete/<int:pk>', methods=['POST'])
def maintenance_expense_delete(pk):
    rec = MaintenanceExpense.query.get_or_404(pk)
    for att in rec.attachments:
        full_path = os.path.join(app.config['UPLOAD_FOLDER'], att.file_path)
        if os.path.isfile(full_path):
            try:
                os.remove(full_path)
            except OSError:
                pass
    db.session.delete(rec)
    db.session.commit()
    flash('Maintenance expense deleted.', 'success')
    return redirect(url_for('maintenance_expense_list'))


# ────────────────────────────────────────────────
# Reports Index & Multiple Report Types
# ────────────────────────────────────────────────
@app.route('/reports/')
def reports_index():
    return render_template('reports_index.html')


@app.route('/reports/ai', methods=['GET', 'POST'])
def report_ai():
    """AI-style custom report: user describes what they want; we show a temp view. Close = discard."""
    result_html = None
    report_title = None
    if request.method == 'POST':
        desc = (request.form.get('description') or '').strip().lower()
        if not desc:
            flash('Please describe the report you need.', 'warning')
            return redirect(url_for('report_ai'))
        try:
            if any(w in desc for w in ['driver', 'drivers', 'personnel']):
                drivers = Driver.query.order_by(Driver.name).all()
                report_title = 'Drivers List'
                rows = [{'name': d.name, 'driver_id': d.driver_id, 'cnic': d.cnic_no, 'license': d.license_no, 'phone': d.phone1, 'status': d.status} for d in drivers]
                result_html = _render_ai_report_table(['Name', 'Driver ID', 'CNIC', 'License', 'Phone', 'Status'], rows)
            elif any(w in desc for w in ['vehicle', 'vehicles', 'fleet']):
                vehicles = Vehicle.query.order_by(Vehicle.vehicle_no).all()
                report_title = 'Vehicles List'
                rows = [{'v_no': v.vehicle_no, 'model': v.model, 'type': v.vehicle_type, 'phone': v.phone_no} for v in vehicles]
                result_html = _render_ai_report_table(['Vehicle No', 'Model', 'Type', 'Phone'], rows)
            elif any(w in desc for w in ['project', 'projects']):
                projects = Project.query.order_by(Project.name).all()
                report_title = 'Project Summary'
                rows = [{'name': p.name, 'vehicles': len(p.vehicles), 'drivers': len(p.drivers), 'status': p.status} for p in projects]
                result_html = _render_ai_report_table(['Project', 'Vehicles', 'Drivers', 'Status'], rows)
            elif any(w in desc for w in ['district', 'districts']):
                districts = District.query.order_by(District.name).all()
                report_title = 'District Summary'
                rows = []
                for d in districts:
                    vc = Vehicle.query.filter_by(district_id=d.id).count()
                    dc = Driver.query.filter_by(district_id=d.id).count()
                    rows.append({'name': d.name, 'vehicles': vc, 'drivers': dc})
                result_html = _render_ai_report_table(['District', 'Vehicles', 'Drivers'], rows)
            elif any(w in desc for w in ['expiry', 'license', 'cnic', 'expiring']):
                from datetime import timedelta
                today = date.today()
                end = today + timedelta(days=60)
                drivers = Driver.query.filter(Driver.status == 'Active').all()
                expiring = []
                for d in drivers:
                    if (d.license_expiry_date and today <= d.license_expiry_date <= end) or (d.cnic_expiry_date and today <= d.cnic_expiry_date <= end):
                        expiring.append({'name': d.name, 'license_expiry': d.license_expiry_date, 'cnic_expiry': d.cnic_expiry_date})
                report_title = 'License / CNIC Expiry (next 60 days)'
                rows = [{'name': r['name'], 'license_expiry': str(r['license_expiry']) if r['license_expiry'] else '-', 'cnic_expiry': str(r['cnic_expiry']) if r['cnic_expiry'] else '-'} for r in expiring]
                result_html = _render_ai_report_table(['Driver', 'License Expiry', 'CNIC Expiry'], rows)
            elif any(w in desc for w in ['company', 'companies']):
                companies = Company.query.order_by(Company.name).all()
                report_title = 'Companies'
                rows = [{'name': c.name, 'mobile': c.mobile or '-', 'email': c.email or '-'} for c in companies]
                result_html = _render_ai_report_table(['Company', 'Mobile', 'Email'], rows)
            elif any(w in desc for w in ['parking', 'utilization', 'station']):
                stations = ParkingStation.query.order_by(ParkingStation.name).all()
                report_title = 'Parking Utilization'
                rows = []
                for s in stations:
                    occ = Vehicle.query.filter_by(parking_station_id=s.id).count()
                    rows.append({'name': s.name, 'capacity': s.capacity, 'occupied': occ, 'available': s.capacity - occ})
                result_html = _render_ai_report_table(['Station', 'Capacity', 'Occupied', 'Available'], rows)
            elif any(w in desc for w in ['product', 'products', 'item']):
                products = Product.query.order_by(Product.name).all()
                report_title = 'Products (Master)'
                rows = [{'name': p.name, 'used_in': getattr(p, 'used_in_forms', None) or '-'} for p in products]
                result_html = _render_ai_report_table(['Product', 'Used in forms'], rows)
            elif any(w in desc for w in ['party', 'parties']):
                parties = Party.query.order_by(Party.name).all()
                report_title = 'Parties'
                rows = [{'name': p.name, 'contact': (getattr(p, 'contact', None) or '-')[:40]} for p in parties]
                result_html = _render_ai_report_table(['Party', 'Contact'], rows)
            elif any(w in desc for w in ['fuel', 'diesel', 'petrol']):
                # Fuel expenses: last 30 days
                from datetime import timedelta
                today = date.today()
                start = today - timedelta(days=30)
                q = FuelExpense.query.filter(FuelExpense.fueling_date >= start).order_by(FuelExpense.fueling_date.desc())
                expenses = q.limit(500).all()
                report_title = 'Fuel Expenses (last 30 days)'
                rows = []
                for f in expenses:
                    rows.append({
                        'date': f.fueling_date.strftime('%Y-%m-%d') if f.fueling_date else '',
                        'vehicle': f.vehicle.vehicle_no if f.vehicle else '',
                        'project': f.project.name if f.project else '',
                        'fuel_type': f.fuel_type or '',
                        'liters': f.liters,
                        'amount': f.amount,
                        'km': f.km,
                        'mpg': f.mpg,
                    })
                result_html = _render_ai_report_table(
                    ['Date', 'Vehicle', 'Project', 'Fuel Type', 'Liters', 'Amount', 'KM', 'KM per liter'],
                    rows
                )
            elif any(w in desc for w in ['maintenance', 'repair', 'service']):
                # Maintenance expenses: last 90 days
                from datetime import timedelta
                today = date.today()
                start = today - timedelta(days=90)
                q = MaintenanceExpense.query.filter(MaintenanceExpense.expense_date >= start).order_by(MaintenanceExpense.expense_date.desc())
                recs = q.limit(500).all()
                report_title = 'Maintenance Expenses (last 90 days)'
                rows = []
                for r in recs:
                    rows.append({
                        'date': r.expense_date.strftime('%Y-%m-%d') if r.expense_date else '',
                        'vehicle': r.vehicle.vehicle_no if r.vehicle else '',
                        'project': r.project.name if r.project else '',
                        'district': r.district.name if r.district else '',
                        'remarks': (r.remarks or '')[:80],
                    })
                result_html = _render_ai_report_table(
                    ['Date', 'Vehicle', 'Project', 'District', 'Remarks'],
                    rows
                )
            elif any(w in desc for w in ['penalty', 'penalties', 'fine', 'fines']):
                # Penalty records: last 60 days
                from datetime import timedelta
                today = date.today()
                start = today - timedelta(days=60)
                q = PenaltyRecord.query.filter(PenaltyRecord.record_date >= start).order_by(PenaltyRecord.record_date.desc())
                recs = q.limit(500).all()
                report_title = 'Penalties (last 60 days)'
                rows = []
                for r in recs:
                    rows.append({
                        'date': r.record_date.strftime('%Y-%m-%d') if r.record_date else '',
                        'driver': r.driver.name if r.driver else '',
                        'vehicle': r.vehicle.vehicle_no if r.vehicle else '',
                        'project': r.project.name if r.project else '',
                        'fine': r.fine or '',
                        'remarks': (r.remarks or '')[:80],
                    })
                result_html = _render_ai_report_table(
                    ['Date', 'Driver', 'Vehicle', 'Project', 'Fine', 'Remarks'],
                    rows
                )
            elif any(w in desc for w in ['attendance', 'absent', 'present']):
                # Driver attendance summary: last 30 days
                from datetime import timedelta
                today = date.today()
                start = today - timedelta(days=30)
                q = DriverAttendance.query.filter(DriverAttendance.attendance_date >= start).order_by(DriverAttendance.attendance_date.desc())
                recs = q.limit(500).all()
                report_title = 'Driver Attendance (last 30 days)'
                rows = []
                for r in recs:
                    rows.append({
                        'date': r.attendance_date.strftime('%Y-%m-%d') if r.attendance_date else '',
                        'driver': r.driver.name if r.driver else '',
                        'project': r.project.name if r.project else '',
                        'status': r.status,
                        'remarks': (r.remarks or '')[:80],
                    })
                result_html = _render_ai_report_table(
                    ['Date', 'Driver', 'Project', 'Status', 'Remarks'],
                    rows
                )
            else:
                report_title = 'Suggested reports'
                result_html = (
                    '<p class="text-muted">Try: "list of drivers", "vehicles", "project summary", "district summary", '
                    '"license expiry", "companies", "parking utilization", "products", "parties", '
                    '"fuel expenses", "maintenance expenses", "penalties", or "attendance summary".</p>'
                )
        except Exception as e:
            app.logger.exception(e)
            report_title = 'Error'
            result_html = f'<p class="text-danger">Report could not be generated. Try another keyword (e.g. drivers, vehicles, projects).</p>'
    return render_template('report_ai.html', result_html=result_html, report_title=report_title)


def _render_ai_report_table(headers, rows):
    if not rows:
        return '<p class="text-muted">No data found.</p>'
    keys = list(rows[0].keys()) if rows else []
    h = '<thead class="table-light"><tr>' + ''.join(f'<th>{x}</th>' for x in headers) + '</tr></thead>'
    body = '<tbody>'
    for r in rows:
        body += '<tr>' + ''.join(f'<td>{r.get(k, "")}</td>' for k in keys) + '</tr>'
    body += '</tbody>'
    return '<table class="table table-bordered table-sm">' + h + body + '</table>'


@app.route('/reports/project-summary')
def report_project_summary():
    projects = Project.query.order_by(Project.name).all()
    data = []
    for p in projects:
        data.append({
            'project': p,
            'vehicle_count': len(p.vehicles),
            'driver_count': len(p.drivers),
            'parking_count': len(p.parking_stations),
            'district_count': p.districts.count(),
        })
    return render_template('report_project_summary.html', data=data)


@app.route('/reports/district-summary')
def report_district_summary():
    districts = District.query.order_by(District.name).all()
    data = []
    for d in districts:
        vehicle_count = Vehicle.query.filter_by(district_id=d.id).count()
        driver_count = Driver.query.filter_by(district_id=d.id).count()
        project_ids = db.session.query(project_district.c.project_id).filter(project_district.c.district_id == d.id).distinct().all()
        project_count = len(project_ids)
        data.append({'district': d, 'vehicle_count': vehicle_count, 'driver_count': driver_count, 'project_count': project_count})
    return render_template('report_district_summary.html', data=data)


@app.route('/reports/vehicle-summary')
def report_vehicle_summary():
    vehicles = Vehicle.query.order_by(Vehicle.vehicle_no).all()
    return render_template('report_vehicle_summary.html', vehicles=vehicles)


@app.route('/reports/driver-profile/<int:driver_id>')
def report_driver_profile(driver_id):
    driver = Driver.query.get_or_404(driver_id)
    transfers = DriverTransfer.query.filter_by(driver_id=driver_id).order_by(DriverTransfer.transfer_date.desc()).all()
    status_changes = DriverStatusChange.query.filter_by(driver_id=driver_id).order_by(DriverStatusChange.change_date.desc()).all()
    return render_template('report_driver_profile.html', driver=driver, transfers=transfers, status_changes=status_changes)


@app.route('/reports/vehicle-profile/<int:vehicle_id>')
def report_vehicle_profile(vehicle_id):
    vehicle = Vehicle.query.get_or_404(vehicle_id)
    transfers = VehicleTransfer.query.filter_by(vehicle_id=vehicle_id).order_by(VehicleTransfer.transfer_date.desc()).all()
    driver_history = DriverTransfer.query.filter(
        (DriverTransfer.old_vehicle_id == vehicle_id) | (DriverTransfer.new_vehicle_id == vehicle_id)
    ).order_by(DriverTransfer.transfer_date.desc()).all()
    return render_template('report_vehicle_profile.html', vehicle=vehicle, transfers=transfers, driver_history=driver_history)


@app.route('/reports/expiry')
def report_expiry():
    from datetime import timedelta
    today = date.today()
    days = request.args.get('days', type=int) or 60
    end = today + timedelta(days=days)
    drivers = Driver.query.filter(Driver.status == 'Active').all()
    expiring = []
    for d in drivers:
        row = {'driver': d, 'license_expiry': d.license_expiry_date, 'cnic_expiry': d.cnic_expiry_date}
        if d.license_expiry_date and today <= d.license_expiry_date <= end:
            row['license_soon'] = True
        else:
            row['license_soon'] = False
        if d.cnic_expiry_date and today <= d.cnic_expiry_date <= end:
            row['cnic_soon'] = True
        else:
            row['cnic_soon'] = False
        if row['license_soon'] or row['cnic_soon']:
            expiring.append(row)
    return render_template('report_expiry.html', expiring=expiring, days=days)


@app.route('/reports/parking-utilization')
def report_parking_utilization():
    stations = ParkingStation.query.order_by(ParkingStation.name).all()
    data = []
    for s in stations:
        occupied = Vehicle.query.filter_by(parking_station_id=s.id).count()
        data.append({'station': s, 'occupied': occupied, 'available': s.capacity - occupied})
    return render_template('report_parking_utilization.html', data=data)