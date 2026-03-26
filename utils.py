# utils.py
import csv
import re
from io import StringIO, BytesIO
from datetime import datetime, date, time as dt_time, timedelta
from flask import Response
from typing import List, Tuple, Union, Dict, Any, Optional

try:
    from zoneinfo import ZoneInfo
except ImportError:
    from backports.zoneinfo import ZoneInfo

_PKT = ZoneInfo('Asia/Karachi')


def pk_now() -> datetime:
    """Current naive datetime in Pakistan timezone (Asia/Karachi)."""
    from datetime import timezone
    return datetime.now(timezone.utc).astimezone(_PKT).replace(tzinfo=None)


def pk_date() -> date:
    """Today's date in Pakistan timezone."""
    return pk_now().date()


def pk_time() -> dt_time:
    """Current time in Pakistan timezone."""
    return pk_now().time()


# ---------- Date: dd-mm-yyyy display & parse ----------
def format_date_ddmmyyyy(d: Optional[date]) -> str:
    """Format date as dd-mm-yyyy. Returns '' if None."""
    if d is None:
        return ''
    if isinstance(d, str):
        try:
            parsed = parse_date(d)
            return parsed.strftime('%d-%m-%Y') if parsed else ''
        except Exception:
            return d
    return d.strftime('%d-%m-%Y')


def parse_date(s: Optional[str]) -> Optional[date]:
    """Parse date string. Accepts dd-mm-yyyy or yyyy-mm-dd."""
    if not s or not str(s).strip():
        return None
    s = str(s).strip()
    for fmt in ('%d-%m-%Y', '%Y-%m-%d', '%d/%m/%Y', '%Y/%m/%d'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


# ---------- CNIC: 32304-1111111-5 format ----------
def format_cnic(cnic: Optional[str]) -> str:
    """Format CNIC as xxxxx-xxxxxxx-x. Accepts with or without dashes."""
    if not cnic or not str(cnic).strip():
        return ''
    s = re.sub(r'[-\s]', '', str(cnic).strip())
    if len(s) == 13 and s.isdigit():
        return f'{s[:5]}-{s[5:12]}-{s[12]}'
    return str(cnic)


# ---------- Phone/Mobile: 03xx-xxxxxxx ----------
def format_phone(phone: Optional[str]) -> str:
    """Format phone as 03xx-xxxxxxx (e.g. 0300-1110810)."""
    if not phone or not str(phone).strip():
        return ''
    s = re.sub(r'[-\s]', '', str(phone).strip())
    if len(s) >= 11 and s.startswith('03'):
        return f'{s[:4]}-{s[4:11]}'
    if len(s) == 10 and s.startswith('03'):
        return f'{s[:4]}-{s[4:10]}'
    return str(phone)


def generate_csv_response(
    headers: List[str],
    rows: List[Union[List[Any], Tuple[Any, ...], Dict[str, Any]]],
    filename: str = "export.csv",
    bom: bool = True,
    encoding: str = 'utf-8-sig'  # UTF-8 with BOM for Excel
) -> Response:
    """
    Generate CSV file and return Flask Response for browser download.
    
    Features:
    - Supports list/tuple/dict rows
    - UTF-8 with BOM for Excel compatibility (Urdu/special characters)
    - Safe handling of None/empty values
    - Custom filename and encoding
    
    Args:
        headers: List of column names e.g. ['ID', 'Vehicle No#', 'Model']
        rows: List of data rows (list, tuple or dict)
        filename: Downloaded file name (default: export.csv)
        bom: Whether to add UTF-8 BOM (Excel ke liye recommended)
        encoding: Output encoding (utf-8-sig for BOM)
    
    Returns:
        Flask Response object ready to return from route
    """
    output = StringIO()
    
    # Add BOM if requested (Excel mein Urdu/Arabic/special chars sahi dikhne ke liye)
    if bom:
        output.write('\ufeff')
    
    writer = csv.writer(output, quoting=csv.QUOTE_MINIMAL)
    
    # Write headers
    writer.writerow(headers)
    
    # Write data rows
    for row in rows:
        if isinstance(row, dict):
            # Dict row → header order mein values lo
            writer.writerow([row.get(key, '') for key in headers])
        else:
            # List/Tuple row → direct write (length mismatch pe empty string)
            row_data = list(row) + [''] * (len(headers) - len(row))  # pad if short
            writer.writerow(row_data[:len(headers)])  # truncate if too long
    
    output.seek(0)
    
    return Response(
        output,
        mimetype=f"text/csv; charset={encoding}",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Content-Type": f"text/csv; charset={encoding}"
        }
    )


# Optional: Agar sirf string output chahiye (API/test ke liye)
def generate_csv_string(
    headers: List[str],
    rows: List[Union[List[Any], Tuple[Any, ...], Dict[str, Any]]],
    bom: bool = True
) -> str:
    """Same as above but returns CSV string instead of Response"""
    output = StringIO()
    if bom:
        output.write('\ufeff')
    writer = csv.writer(output, quoting=csv.QUOTE_MINIMAL)
    writer.writerow(headers)
    for row in rows:
        if isinstance(row, dict):
            writer.writerow([row.get(key, '') for key in headers])
        else:
            row_data = list(row) + [''] * (len(headers) - len(row))
            writer.writerow(row_data[:len(headers)])
    output.seek(0)
    return output.getvalue()


def generate_excel_template(
    headers: List[str],
    rows: List[Union[List[Any], Tuple[Any, ...]]],
    required_columns: Optional[List[str]] = None,
    filename: str = "template.xlsx"
) -> Response:
    """
    Excel (.xlsx) template generate karta hai jisme:
    - 1st row = headers
    - Neeche sample rows
    - Required columns ki heading light color + bold hoti hai
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Template"

    # Header row
    ws.append(headers)

    # Sample data rows
    for r in rows:
        ws.append(list(r))

    required_columns = required_columns or []

    # Style: required headers ko highlight + bold
    required_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")  # light yellow
    required_font = Font(bold=True)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        if header in required_columns:
            cell.fill = required_fill
            cell.font = required_font

    # Basic auto-width (best-effort)
    for col_idx, header in enumerate(headers, start=1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        max_len = max(len(str(header)), 10)
        ws.column_dimensions[col_letter].width = max_len + 2

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        },
    )