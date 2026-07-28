# utils.py
import csv
import re
from io import StringIO, BytesIO
from datetime import datetime, date, time as dt_time, timedelta
from flask import Response
from typing import List, Tuple, Union, Dict, Any, Optional

try:
    from zoneinfo import ZoneInfo
except ImportError:
    from backports.zoneinfo import ZoneInfo

_PKT = ZoneInfo('Asia/Karachi')


def pk_now() -> datetime:
    """Current naive datetime in Pakistan timezone (Asia/Karachi)."""
    from datetime import timezone
    return datetime.now(timezone.utc).astimezone(_PKT).replace(tzinfo=None)


def pk_date() -> date:
    """Today's date in Pakistan timezone."""
    return pk_now().date()


def pk_time() -> dt_time:
    """Current time in Pakistan timezone."""
    return pk_now().time()


# ---------- Date: dd-mm-yyyy display & parse ----------
def format_time_ampm(t: Optional[Union[dt_time, datetime]]) -> str:
    """Format time as hh:mm AM/PM (e.g. 09:30 AM). Returns '' if None."""
    if t is None:
        return ''
    if isinstance(t, datetime):
        t = t.time()
    return t.strftime('%I:%M %p').lstrip('0') or t.strftime('%I:%M %p')


def format_reading(value: Optional[Union[int, float]]) -> str:
    """Format odometer reading: show integer when fractional part is zero, else show decimals."""
    if value is None:
        return ''
    try:
        v = float(value)
    except (TypeError, ValueError):
        return str(value)
    if abs(v - round(v)) < 1e-9:
        return str(int(round(v)))
    return f'{v:.2f}'.rstrip('0').rstrip('.')


def format_date_ddmmyyyy(d: Optional[date]) -> str:
    """Format date as dd-mm-yyyy. Returns '' if None."""
    if d is None:
        return ''
    if isinstance(d, str):
        try:
            parsed = parse_date(d)
            return parsed.strftime('%d-%m-%Y') if parsed else ''
        except Exception:
            return d
    return d.strftime('%d-%m-%Y')


def format_ufone_date_short(s: Any) -> str:
    """Ufone task date -> dd-mm-yy. Accepts ISO, MM/DD/YYYY (Ufone default),
    'DD Mon YYYY HH:MM:SS', dd-mm-yyyy, etc. Mirrors the JS formatTaskDate() so
    server-rendered rows always show the date (no reliance on JS timing)."""
    if not s:
        return ''
    s = str(s).strip()
    if not s or s.startswith('01 Jan 1900') or s.startswith('1900') or s.startswith('01/01/1900'):
        return ''
    # ISO: 2026-07-22 or 2026-07-22T00:55:10
    m = re.match(r'^(\d{4})-(\d{1,2})-(\d{1,2})', s)
    if m:
        return f"{int(m.group(3)):02d}-{int(m.group(2)):02d}-{m.group(1)[-2:]}"
    # MM/DD/YYYY (Ufone default — slash = US format) — also MM/DD/YYYY HH:MM
    m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{2,4})', s)
    if m:
        yy = m.group(3)[-2:] if len(m.group(3)) == 4 else m.group(3).zfill(2)
        return f"{int(m.group(2)):02d}-{int(m.group(1)):02d}-{yy}"
    # '22 Jul 2026 00:55:10' or '22-Jul-2026'
    m = re.match(r'^(\d{1,2})[\s-]([A-Za-z]{3,})[\s-](\d{2,4})', s)
    if m:
        months = {'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
                  'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12}
        mon = months.get(m.group(2).lower())
        if mon:
            yy = m.group(3)[-2:] if len(m.group(3)) == 4 else m.group(3).zfill(2)
            return f"{int(m.group(1)):02d}-{mon:02d}-{yy}"
    # dd-mm-yyyy or dd-mm-yy (dash = day-first)
    m = re.match(r'^(\d{1,2})-(\d{1,2})-(\d{2,4})', s)
    if m:
        yy = m.group(3)[-2:] if len(m.group(3)) == 4 else m.group(3).zfill(2)
        return f"{int(m.group(1)):02d}-{int(m.group(2)):02d}-{yy}"
    return s


def format_ufone_datetime_short(s: Any, time_part: Any = None) -> str:
    """Date + time for dashboard table: dd-mm-yy HH:MM[:SS] when time exists."""
    date_part = format_ufone_date_short(s)
    if not date_part:
        return ''
    raw = str(s or '').strip()
    tm = None
    if time_part not in (None, ''):
        tm = str(time_part).strip()
    if not tm:
        m = re.search(r'(\d{1,2}:\d{2}(?::\d{2})?)', raw)
        if m:
            tm = m.group(1)
        else:
            # ISO T separator
            m = re.search(r'T(\d{2}:\d{2}(?::\d{2})?)', raw)
            if m:
                tm = m.group(1)
    if tm:
        # Drop trailing .fff if any
        tm = tm.split('.')[0]
        return f'{date_part} {tm}'
    return date_part


def parse_date(s: Optional[str]) -> Optional[date]:
    """Parse date string. Accepts dd-mm-yyyy or yyyy-mm-dd."""
    if not s or not str(s).strip():
        return None
    s = str(s).strip()
    for fmt in ('%d-%m-%Y', '%Y-%m-%d', '%d/%m/%Y', '%Y/%m/%d'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


# ---------- CNIC: 32304-1111111-5 format ----------
def format_cnic(cnic: Optional[str]) -> str:
    """Format CNIC as xxxxx-xxxxxxx-x. Accepts with or without dashes."""
    if not cnic or not str(cnic).strip():
        return ''
    s = re.sub(r'[-\s]', '', str(cnic).strip())
    if len(s) == 13 and s.isdigit():
        return f'{s[:5]}-{s[5:12]}-{s[12]}'
    return str(cnic)


# ---------- Phone/Mobile: 03xx-xxxxxxx ----------
def format_phone(phone: Optional[str]) -> str:
    """Format phone as 03xx-xxxxxxx (e.g. 0300-1110810)."""
    if not phone or not str(phone).strip():
        return ''
    s = re.sub(r'[-\s]', '', str(phone).strip())
    if len(s) >= 11 and s.startswith('03'):
        return f'{s[:4]}-{s[4:11]}'
    if len(s) == 10 and s.startswith('03'):
        return f'{s[:4]}-{s[4:10]}'
    return str(phone)


def generate_csv_response(
    headers: List[str],
    rows: List[Union[List[Any], Tuple[Any, ...], Dict[str, Any]]],
    filename: str = "export.csv",
    bom: bool = True,
    encoding: str = 'utf-8-sig'  # UTF-8 with BOM for Excel
) -> Response:
    """
    Generate CSV file and return Flask Response for browser download.
    
    Features:
    - Supports list/tuple/dict rows
    - UTF-8 with BOM for Excel compatibility (Urdu/special characters)
    - Safe handling of None/empty values
    - Custom filename and encoding
    
    Args:
        headers: List of column names e.g. ['ID', 'Vehicle No#', 'Model']
        rows: List of data rows (list, tuple or dict)
        filename: Downloaded file name (default: export.csv)
        bom: Whether to add UTF-8 BOM (Excel ke liye recommended)
        encoding: Output encoding (utf-8-sig for BOM)
    
    Returns:
        Flask Response object ready to return from route
    """
    output = StringIO()
    
    # Add BOM if requested (Excel mein Urdu/Arabic/special chars sahi dikhne ke liye)
    if bom:
        output.write('\ufeff')
    
    writer = csv.writer(output, quoting=csv.QUOTE_MINIMAL)
    
    # Write headers
    writer.writerow(headers)
    
    # Write data rows
    for row in rows:
        if isinstance(row, dict):
            # Dict row → header order mein values lo
            writer.writerow([row.get(key, '') for key in headers])
        else:
            # List/Tuple row → direct write (length mismatch pe empty string)
            row_data = list(row) + [''] * (len(headers) - len(row))  # pad if short
            writer.writerow(row_data[:len(headers)])  # truncate if too long
    
    output.seek(0)
    
    return Response(
        output,
        mimetype=f"text/csv; charset={encoding}",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Content-Type": f"text/csv; charset={encoding}"
        }
    )


# Optional: Agar sirf string output chahiye (API/test ke liye)
def generate_csv_string(
    headers: List[str],
    rows: List[Union[List[Any], Tuple[Any, ...], Dict[str, Any]]],
    bom: bool = True
) -> str:
    """Same as above but returns CSV string instead of Response"""
    output = StringIO()
    if bom:
        output.write('\ufeff')
    writer = csv.writer(output, quoting=csv.QUOTE_MINIMAL)
    writer.writerow(headers)
    for row in rows:
        if isinstance(row, dict):
            writer.writerow([row.get(key, '') for key in headers])
        else:
            row_data = list(row) + [''] * (len(headers) - len(row))
            writer.writerow(row_data[:len(headers)])
    output.seek(0)
    return output.getvalue()


def generate_excel_template(
    headers: List[str],
    rows: List[Union[List[Any], Tuple[Any, ...]]],
    required_columns: Optional[List[str]] = None,
    filename: str = "template.xlsx"
) -> Response:
    """
    Excel (.xlsx) template generate karta hai jisme:
    - 1st row = headers
    - Neeche sample rows
    - Required columns ki heading light color + bold hoti hai
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Template"

    # Header row
    ws.append(headers)

    # Sample data rows
    for r in rows:
        ws.append(list(r))

    required_columns = required_columns or []

    # Style: required headers ko highlight + bold
    required_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")  # light yellow
    required_font = Font(bold=True)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        if header in required_columns:
            cell.fill = required_fill
            cell.font = required_font

    # Basic auto-width (best-effort)
    for col_idx, header in enumerate(headers, start=1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        max_len = max(len(str(header)), 10)
        ws.column_dimensions[col_letter].width = max_len + 2

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        },
    )


# ---------- Public share links (e.g. Driver Profile, 24h) ----------
_SALT_DRIVER_PROFILE = "driver-profile-share-v1"
_MAX_AGE_DRIVER_PROFILE = 86400  # 24 hours
_SALT_NOTIFICATION_POPUP = "notification-popup-v1"
# Read-only Android notification popup token. Keep it time-limited, but long
# enough that a user can open a notification later in the day without hitting
# an "expired" popup after normal delivery/open delays.
_MAX_AGE_NOTIFICATION_POPUP = 86400  # 24 hours


def make_driver_profile_share_token(secret_key: str, driver_id: int) -> str:
    """Signed token for a time-limited public driver profile URL (itsdangerous)."""
    from itsdangerous import URLSafeTimedSerializer

    s = URLSafeTimedSerializer(secret_key, salt=_SALT_DRIVER_PROFILE)
    return s.dumps({"d": int(driver_id)})


def load_driver_profile_share_token(secret_key: str, token: str) -> Optional[int]:
    """
    Return driver_id if token is valid and not older than 24 hours, else None.
    """
    from itsdangerous import URLSafeTimedSerializer, BadSignature, SignatureExpired

    s = URLSafeTimedSerializer(secret_key, salt=_SALT_DRIVER_PROFILE)
    try:
        data = s.loads(token, max_age=_MAX_AGE_DRIVER_PROFILE)
        d = data.get("d")
        return int(d) if d is not None else None
    except (BadSignature, SignatureExpired, TypeError, ValueError, Exception):
        return None


def make_notification_popup_token(secret_key: str, payload: dict) -> str:
    """Signed token for a short-lived public notification popup."""
    from itsdangerous import URLSafeTimedSerializer

    data = {
        "title": str(payload.get("title") or "").strip()[:200],
        "message": str(payload.get("message") or "").strip()[:4000],
        "type": str(payload.get("type") or "info").strip()[:50],
        "source": str(payload.get("source") or "generic").strip()[:80],
        "save_enabled": bool(payload.get("save_enabled")),
        "created_at": str(payload.get("created_at") or "").strip()[:80],
        "original_link": str(payload.get("original_link") or "").strip()[:500],
    }
    s = URLSafeTimedSerializer(secret_key, salt=_SALT_NOTIFICATION_POPUP)
    return s.dumps(data)


def load_notification_popup_token(secret_key: str, token: str) -> Optional[dict]:
    """Return signed popup payload if valid and recent, else None."""
    from itsdangerous import URLSafeTimedSerializer, BadSignature, SignatureExpired

    s = URLSafeTimedSerializer(secret_key, salt=_SALT_NOTIFICATION_POPUP)
    try:
        data = s.loads(token, max_age=_MAX_AGE_NOTIFICATION_POPUP)
        if not isinstance(data, dict):
            return None
        return {
            "title": str(data.get("title") or "").strip()[:200],
            "message": str(data.get("message") or "").strip()[:4000],
            "type": str(data.get("type") or "info").strip()[:50],
            "source": str(data.get("source") or "generic").strip()[:80],
            "save_enabled": bool(data.get("save_enabled")),
            "created_at": str(data.get("created_at") or "").strip()[:80],
            "original_link": str(data.get("original_link") or "").strip()[:500],
        }
    except (BadSignature, SignatureExpired, TypeError, ValueError, Exception):
        return None


def user_profile_avatar_path(user) -> Optional[str]:
    """
    Profile photo for sidebar / account: linked driver's photo_path when
    login username matches driver CNIC (same logic as account profile page).
    """
    if user is None:
        return None
    try:
        from models import Driver
        from sqlalchemy import func
    except Exception:
        return None
    uname = (getattr(user, 'username', None) or '').strip()
    if not uname:
        return None
    variants = [uname]
    digits = re.sub(r'\D', '', uname)
    if len(digits) == 13:
        variants.append(digits[:5] + '-' + digits[5:12] + '-' + digits[12:])
    for c in variants:
        drv = Driver.query.filter(func.lower(Driver.cnic_no) == c.lower()).first()
        if drv and getattr(drv, 'photo_path', None):
            return drv.photo_path
    return None


# Ufone / PortalXS append operational tags to RegNo, e.g.
# "GBF-25-579 COW", "GBD-24-395-COW", "GBF-25-061 USG"
_UFONE_REG_TAG_RE = re.compile(
    r'[\s\-]+(COW|USG\+P|USG|RAS|MNHC|EMS|NHP)\s*$',
    re.IGNORECASE,
)


def strip_ufone_reg_tag(reg: Optional[str]) -> str:
    """Strip trailing COW/USG/RAS tags (space or hyphen) from a vehicle reg."""
    s = (reg or '').strip()
    if not s:
        return ''
    s = _UFONE_REG_TAG_RE.sub('', s).strip()
    if ' ' in s:
        s = s.split()[0].strip()
    return s


def normalize_vehicle_reg_key(reg: Optional[str]) -> str:
    """Alphanumeric key after stripping Ufone tags (for matching)."""
    return re.sub(r'[^A-Za-z0-9]', '', strip_ufone_reg_tag(reg).upper())


def emg_amb_reg_matches_vehicle(vehicle_no: Optional[str]):
    """SQLAlchemy OR filter: EmergencyTaskRecord.amb_reg_no ↔ fleet vehicle_no.

    Matches exact, base (no tag), and tagged variants:
    GBF-25-579 / GBF-25-579 COW / GBD-24-395-COW.
    """
    from models import EmergencyTaskRecord
    from sqlalchemy import or_

    raw = (vehicle_no or '').strip()
    if not raw:
        return EmergencyTaskRecord.id < 0
    base = strip_ufone_reg_tag(raw) or raw
    return or_(
        EmergencyTaskRecord.amb_reg_no == raw,
        EmergencyTaskRecord.amb_reg_no == base,
        EmergencyTaskRecord.amb_reg_no.ilike(base + ' %'),
        EmergencyTaskRecord.amb_reg_no.ilike(base + '-%'),
    )