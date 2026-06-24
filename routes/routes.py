# Force Rebuild — all syntax verified clean, pushing to unblock Render deploy queue
from flask import render_template, redirect, url_for, flash, request, Response, jsonify, send_from_directory, session, send_file, abort, make_response, after_this_request, current_app
from app import app, db, csrf
from vehicle_sort_utils import vehicle_order_by, sort_vehicles_in_memory
from models import (
    Company, Project, Vehicle, Driver, ParkingStation, District, EmployeePost, Employee, EmployeeDocument,
    project_district, employee_project, employee_district, vehicle_district, ProjectTransfer, VehicleTransfer, DriverTransfer, DriverStatusChange,
    DriverAttendance,
    LeaveRequest,
    VehicleDailyTask, EmergencyTaskRecord, VehicleMileageRecord, VehicleActivityRecord, RedTask, VehicleMoveWithoutTask, PenaltyRecord, UnexecutedTaskRecord,
    Party, Product, FuelExpense, FuelExpenseAttachment, ProductBalance, OilExpense, OilExpenseItem, OilExpenseAttachment,
    MaintenanceWorkOrder, MaintenanceWorkOrderAttachment, MaintenanceExpense, MaintenanceExpenseItem, MaintenanceExpenseAttachment,
    WorkspaceProduct,
    WorkspaceParty, WorkspaceAccount, WorkspaceJournalEntry, WorkspaceJournalEntryLine, WorkspaceVehicleReadingSetup, WorkspaceVehicleMaintenanceBaseline, WorkspaceExpense, ExpenseDeleteCleanupJob,
    Notification, NotificationRead,
    User, Role, Permission, role_permissions,
    LoginLog, ActivityLog, ClientActivityLog, ClientDiagnosticLog,
    Reminder,
    AttendanceTimeControl, AttendanceTimeOverride,
    PhysicalBook, BookAssignment,
    AttendanceSettings,
    DeviceFCMToken,
    AppRelease,
    LoginAttempt,
    SystemSetting,
    EmployeeAssignment,
)
from forms import (
    CompanyForm, ProjectForm, VehicleForm, VehicleImportForm, DriverForm, DriverImportForm, EmployeeImportForm, ParkingForm, DistrictForm,
    AssignProjectToCompanyForm, EditProjectAssignmentForm,
    AssignProjectToDistrictForm, AssignVehicleToDistrictForm, AssignVehicleToParkingForm,
    AssignDriverToVehicleForm, ProjectTransferForm, VehicleTransferForm, EditVehicleTransferForm, DriverTransferForm, DriverJobLeftForm, DriverRejoinForm,
    DriverAttendanceFilterForm, DriverAttendanceReportForm, ATTENDANCE_STATUS_CHOICES,
    ATTENDANCE_LIST_STATUS_FILTER_OPTIONS,
    TaskReportForm, TaskReportFilterForm, EmergencyTaskUploadForm, VehicleMileageUploadForm, ParkingImportForm, ProductImportForm,
    TaskReportUploadBothForm, RedTaskFilterForm, RedTaskForm, VehicleMoveWithoutTaskFilterForm, VehicleMoveWithoutTaskForm, PenaltyRecordForm, PenaltyRecordFilterForm,
    FuelExpenseFilterForm, FuelExpenseForm,
    OilExpenseFilterForm, OilExpenseForm,
    MaintenanceExpenseFilterForm, MaintenanceExpenseForm,
    PartyForm,
    ProductForm,
    EmployeePostForm,
    EmployeeForm,
    EmployeeFormStep1,
    EmployeeFormStep2,
    EmployeeFormStep3,
    EmployeeAssignmentForm,
    EmployeeDocumentForm,
    LoginForm,
    UserForm,
    RoleForm,
    NotificationForm,
    ReminderForm,
    ChangePasswordForm,
    SetNewPasswordForm,
    AttendanceTimeControlForm,
    AttendanceTimeOverrideForm,
    EmployeeAssignForm,
    EmployeeDeassignForm,
    EmployeeLeftForm,
    EmployeeRejoinForm,
)
from datetime import datetime, date, time, timezone
from datetime import timedelta
from decimal import Decimal
import base64
import csv
from io import StringIO, BytesIO
import io
import xlsxwriter
from sqlalchemy import func, text, inspect, or_, cast, and_, false, delete, insert, select
from sqlalchemy import String as SAString
from sqlalchemy.exc import OperationalError, IntegrityError, DataError
from sqlalchemy.orm import joinedload
from utils import (
    generate_csv_response, parse_date, generate_excel_template, format_cnic, format_phone, format_date_ddmmyyyy, format_reading,
    format_time_ampm,
    pk_now, pk_date, pk_time,
    make_driver_profile_share_token, load_driver_profile_share_token,
)
from auth_utils import get_required_permission, user_has_permission, user_can_access, check_password, is_endpoint_allowed_for_any_authed
from flask_wtf.csrf import CSRFError
from werkzeug.exceptions import HTTPException
from werkzeug.security import generate_password_hash
import re
import os
import json
import uuid
import math
import tempfile
import zipfile
import time as _time_mod
import threading
import mimetypes
import shutil
from urllib.request import Request, urlopen
from werkzeug.utils import secure_filename
from r2_storage import upload_image_file, upload_image_bytes
from freeze_utils import (
    get_freeze_config,
    save_freeze_config,
    get_freeze_form_catalog,
    is_freeze_protected_request,
    extract_effective_date,
    evaluate_freeze,
)
from finance_utils import (
    ensure_workspace_base_accounts,
    ensure_workspace_counterparty_account,
    workspace_create_journal_entry,
    workspace_reverse_journal_entry,
)


def _multi_word_filter(search_str, *columns):
    """Multi-word AND search: each space-separated token must appear in at least one column.
    
    Example: "COW 395" → (any_col ILIKE '%COW%') AND (any_col ILIKE '%395%')
    This lets users narrow results by typing multiple keywords from different columns.
    Returns a SQLAlchemy clause or None if no tokens found.
    """
    tokens = [t for t in search_str.split() if t]
    if not tokens:
        return None
    return and_(*(or_(*(col.ilike(f'%{tok}%') for col in columns)) for tok in tokens))


class SimplePagination:
    """Lightweight pagination wrapper for Python lists (no SQLAlchemy required).
    API-compatible with Flask-SQLAlchemy Pagination so templates share the same HTML."""
    def __init__(self, items, page, per_page):
        self.per_page = per_page
        self.total = len(items)
        self.pages = max(1, (self.total + per_page - 1) // per_page)
        self.page = max(1, min(page, self.pages))
        start = (self.page - 1) * per_page
        self.items = items[start: start + per_page]
        self.has_prev = self.page > 1
        self.has_next = self.page < self.pages
        self.prev_num = self.page - 1
        self.next_num = self.page + 1

    def iter_pages(self, left_edge=1, right_edge=1, left_current=2, right_current=2):
        last = 0
        for num in range(1, self.pages + 1):
            if (num <= left_edge or num > self.pages - right_edge or
                    abs(num - self.page) <= max(left_current, right_current)):
                if last + 1 != num:
                    yield None
                yield num
                last = num


def _attendance_list_resolve_per_page(request, total_count):
    """Attendance list pagination size.

    Desktop browser default: 20 rows (fast load).

    Mobile shell (Capacitor / Android System WebView): when ``per_page`` is omitted,
    return full filtered result set so the list + photo share see every row (no silent 20 cap).

    Explicit ``?per_page=all`` / ``0`` / negative → show all. ``?native=1`` forces full list.
    """
    raw = request.args.get('per_page')
    ua = (request.headers.get('User-Agent') or '').lower()

    def _full_size():
        return max(int(total_count or 0), 1)

    hint = (request.args.get('native') or request.args.get('app') or '').strip().lower()
    if hint in ('1', 'true', 'yes'):
        return _full_size()

    if raw is not None and str(raw).strip() != '':
        s = str(raw).strip().lower()
        if s in ('all', 'max', 'full'):
            return _full_size()
        try:
            n = int(raw)
            if n <= 0:
                return _full_size()
            return max(1, min(n, 10000))
        except (TypeError, ValueError):
            return 20

    # Same WebView as Fleet Manager Android app (see README_CAPACITOR)
    if 'capacitor' in ua or ('wv' in ua and 'android' in ua):
        return _full_size()

    return 20


_VEHICLE_FAMILY_SETTING_KEY = 'vehicle_family_options'
_DEFAULT_VEHICLE_FAMILIES = [
    'Suzuki Bolan',
    'Toyota Hilux Vigo',
    'Changan Karvaan',
]
_VEHICLE_FAMILY_OIL_LIMITS_KEY = 'vehicle_family_oil_change_limits'
_VEHICLE_FAMILY_OIL_NEAR_PERCENT_KEY = 'vehicle_family_oil_change_near_percent'
_MAINTENANCE_JOB_CATEGORY_KEY = 'maintenance_job_categories'
_FUEL_MARKET_SCAN_KEY = 'fuel_market_scan_status'


def _get_vehicle_family_options():
    """Load vehicle family dropdown options from system settings."""
    raw = (SystemSetting.get(_VEHICLE_FAMILY_SETTING_KEY, '') or '').strip()
    options = []
    if raw:
        try:
            parsed = json.loads(raw)
            if isinstance(parsed, list):
                options = [str(x).strip() for x in parsed if str(x).strip()]
        except Exception:
            # Backward compatibility if old value was stored as delimited text.
            options = [p.strip() for p in raw.split('|') if p.strip()]
    if not options:
        options = list(_DEFAULT_VEHICLE_FAMILIES)
    for name in _DEFAULT_VEHICLE_FAMILIES:
        if name not in options:
            options.append(name)
    return options


def _save_vehicle_family_options(options):
    clean = []
    seen = set()
    for name in options:
        val = (name or '').strip()
        if not val:
            continue
        key = val.lower()
        if key in seen:
            continue
        seen.add(key)
        clean.append(val)
    SystemSetting.set(_VEHICLE_FAMILY_SETTING_KEY, json.dumps(clean, ensure_ascii=True))
    return clean


def _get_maintenance_job_categories():
    """Load maintenance job categories from system settings.

    Format:
      [{ "name": "Engine Oil Service", "interval_mode": "interval_km", "interval_value": 5000 }]
    """
    raw = (SystemSetting.get(_MAINTENANCE_JOB_CATEGORY_KEY, '') or '').strip()
    options = []
    if raw:
        try:
            parsed = json.loads(raw)
            if isinstance(parsed, list):
                for entry in parsed:
                    if not isinstance(entry, dict):
                        continue
                    name = (entry.get('name') or '').strip()
                    mode = (entry.get('interval_mode') or '').strip().lower()
                    try:
                        interval_value = int(float(entry.get('interval_value')))
                    except Exception:
                        interval_value = 0
                    if not name or mode not in ('interval_km', 'interval_day') or interval_value <= 0:
                        continue
                    options.append({'name': name, 'interval_mode': mode, 'interval_value': interval_value})
        except Exception:
            options = []
    if not options:
        options = [
            {'name': 'Engine Oil Service', 'interval_mode': 'interval_km', 'interval_value': 5000},
            {'name': 'General Inspection', 'interval_mode': 'interval_day', 'interval_value': 180},
        ]
    return options


def _save_maintenance_job_categories(options):
    clean = []
    seen = set()
    for entry in (options or []):
        if not isinstance(entry, dict):
            continue
        name = (entry.get('name') or '').strip()
        mode = (entry.get('interval_mode') or '').strip().lower()
        try:
            interval_value = int(float(entry.get('interval_value')))
        except Exception:
            interval_value = 0
        if not name or mode not in ('interval_km', 'interval_day') or interval_value <= 0:
            continue
        key = f"{name.lower()}::{mode}"
        if key in seen:
            continue
        seen.add(key)
        clean.append({'name': name, 'interval_mode': mode, 'interval_value': interval_value})
    SystemSetting.set(_MAINTENANCE_JOB_CATEGORY_KEY, json.dumps(clean, ensure_ascii=True))
    return clean


def _get_vehicle_family_oil_change_limits():
    """Load per-family oil change config from settings.

    Format:
      { "Family": { "limit_km": 5000, "near_percent": 90 } }
    Backward compatibility:
      { "Family": 5000 }  # old format
    """
    raw = (SystemSetting.get(_VEHICLE_FAMILY_OIL_LIMITS_KEY, '') or '').strip()
    default_near = _get_vehicle_family_oil_near_percent()

    def _norm_near(v):
        try:
            n = int(float(v))
        except Exception:
            n = default_near
        if n < 50:
            return 50
        if n > 99:
            return 99
        return n

    out = {}
    if raw:
        try:
            parsed = json.loads(raw)
            if isinstance(parsed, dict):
                for k, v in parsed.items():
                    fam = (str(k) or '').strip()
                    if not fam:
                        continue
                    km = None
                    near = default_near
                    if isinstance(v, dict):
                        try:
                            km = int(float(v.get('limit_km')))
                        except Exception:
                            km = None
                        near = _norm_near(v.get('near_percent'))
                    else:
                        try:
                            km = int(float(v))
                        except Exception:
                            km = None
                    if km and km > 0:
                        out[fam] = {
                            'limit_km': km,
                            'near_percent': near,
                        }
        except Exception:
            out = {}
    return out


def _save_vehicle_family_oil_change_limits(limits):
    clean = {}
    default_near = _get_vehicle_family_oil_near_percent()

    def _norm_near(v):
        try:
            n = int(float(v))
        except Exception:
            n = default_near
        if n < 50:
            return 50
        if n > 99:
            return 99
        return n

    for fam, payload in (limits or {}).items():
        fam_name = (fam or '').strip()
        if not fam_name:
            continue
        km_val = None
        near_val = default_near
        if isinstance(payload, dict):
            try:
                km_val = int(float(payload.get('limit_km')))
            except Exception:
                km_val = None
            near_val = _norm_near(payload.get('near_percent'))
        else:
            try:
                km_val = int(float(payload))
            except Exception:
                km_val = None
        if km_val and km_val > 0:
            clean[fam_name] = {
                'limit_km': km_val,
                'near_percent': near_val,
            }
    SystemSetting.set(_VEHICLE_FAMILY_OIL_LIMITS_KEY, json.dumps(clean, ensure_ascii=True))
    return clean


def _get_vehicle_family_oil_near_percent():
    raw = (SystemSetting.get(_VEHICLE_FAMILY_OIL_NEAR_PERCENT_KEY, '90') or '90').strip()
    try:
        val = int(raw)
    except Exception:
        val = 90
    if val < 50:
        return 50
    if val > 99:
        return 99
    return val


def _save_vehicle_family_oil_near_percent(percent):
    try:
        val = int(percent)
    except Exception:
        val = 90
    if val < 50:
        val = 50
    if val > 99:
        val = 99
    SystemSetting.set(_VEHICLE_FAMILY_OIL_NEAR_PERCENT_KEY, str(val))
    return val


def _fetch_text_url(url, timeout=12):
    import urllib.request
    req = urllib.request.Request(url, headers={
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    })
    with urllib.request.urlopen(req, timeout=timeout) as r:
        raw = r.read()
    return raw.decode('utf-8', errors='ignore')


def _to_float_price(s):
    try:
        return float(str(s).replace(',', '').strip())
    except Exception:
        return None


def _extract_price_with_patterns(text, patterns):
    for pat in patterns:
        m = re.search(pat, text, flags=re.I | re.S)
        if m:
            val = _to_float_price(m.group(1))
            if val is not None:
                return val
    return None


def _scan_pso_rates():
    url = 'https://psopk.com/en/fuels/fuel-prices'
    html = _fetch_text_url(url)
    petrol = _extract_price_with_patterns(html, [
        r'PREMIER\s*EURO\s*5[^0-9]{0,100}([0-9][0-9,]*\.?[0-9]*)\s*/?\s*Ltr',
        r'Premier[^0-9]{0,120}([0-9][0-9,]*\.?[0-9]*)\s*/?\s*Ltr',
    ])
    diesel = _extract_price_with_patterns(html, [
        r'HI-?CETANE\s*DIESEL\s*EURO\s*5[^0-9]{0,100}([0-9][0-9,]*\.?[0-9]*)\s*/?\s*Ltr',
        r'Hi-?Cetane\s*Diesel[^0-9]{0,120}([0-9][0-9,]*\.?[0-9]*)\s*/?\s*Ltr',
    ])
    ok = petrol is not None and diesel is not None
    return {
        'source': 'PSO',
        'ok': ok,
        'petrol': petrol,
        'diesel': diesel,
        'updated_at': '',
        'error': '' if ok else 'Unable to parse Premier/Hi-Cetane Diesel from PSO page.',
    }


def _read_fuel_market_scan():
    raw = (SystemSetting.get(_FUEL_MARKET_SCAN_KEY, '') or '').strip()
    if not raw:
        return {}
    try:
        data = json.loads(raw)
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def _save_fuel_market_scan(data):
    SystemSetting.set(_FUEL_MARKET_SCAN_KEY, json.dumps(data or {}, ensure_ascii=True))


def _fuel_expense_last_saved_rec(workspace_employee_id, last_id=None):
    """Latest saved fuel expense for the add-form Last Saved strip."""
    if last_id:
        rec = db.session.get(FuelExpense, last_id)
        if rec and (not workspace_employee_id or not rec.employee_id or rec.employee_id == workspace_employee_id):
            return rec
    q = FuelExpense.query
    if workspace_employee_id:
        q = q.filter(FuelExpense.employee_id == workspace_employee_id)
    return q.order_by(FuelExpense.id.desc()).first()


def _fuel_expense_last_entry_payload(rec):
    if not rec:
        return None
    pump_name = ''
    if rec.workspace_pump:
        pump_name = rec.workspace_pump.name or ''
    elif rec.fuel_pump:
        pump_name = rec.fuel_pump.name or ''
    elif rec.workspace_pump_id:
        wp = db.session.get(WorkspaceParty, rec.workspace_pump_id)
        if wp:
            pump_name = wp.name or ''
    fuel_type = (rec.fuel_type or '').strip()
    if rec.vehicle and not fuel_type and rec.vehicle.fuel_type:
        fuel_type = (rec.vehicle.fuel_type or '').strip()
    km_out_task = float(rec.km_out_task) if rec.km_out_task is not None else None
    km_in_task = float(rec.km_in_task) if rec.km_in_task is not None else None
    if km_out_task is None and km_in_task is None and rec.fueling_date and rec.vehicle_id:
        km_out_task, km_in_task = _fuel_expense_task_readings(rec.vehicle_id, rec.fueling_date)
        km_out_task = float(km_out_task) if km_out_task is not None else None
        km_in_task = float(km_in_task) if km_in_task is not None else None
    km_val = float(rec.km) if rec.km is not None else None
    if km_val is None and rec.previous_reading is not None and rec.current_reading is not None:
        km_val = float(rec.current_reading) - float(rec.previous_reading)
    return {
        'id': rec.id,
        'fueling_date': rec.fueling_date.strftime('%d-%m-%Y') if rec.fueling_date else '',
        'card_swipe_date': rec.card_swipe_date.strftime('%d-%m-%Y') if rec.card_swipe_date else '',
        'district': rec.district.name if rec.district else '',
        'project': rec.project.name if rec.project else '',
        'vehicle_no': rec.vehicle.vehicle_no if rec.vehicle else '',
        'payment_type': (rec.payment_type or '').strip(),
        'slip_no': (rec.slip_no or '').strip(),
        'fuel_type': fuel_type or 'Petrol',
        'pump_name': pump_name,
        'amount': float(rec.amount) if rec.amount is not None else None,
        'liters': float(rec.liters) if rec.liters is not None else None,
        'mpg': float(rec.mpg) if rec.mpg is not None else None,
        'fuel_price': float(rec.fuel_price) if rec.fuel_price is not None else None,
        'km': km_val,
        'current_reading': float(rec.current_reading) if rec.current_reading is not None else None,
        'previous_reading': float(rec.previous_reading) if rec.previous_reading is not None else None,
        'km_out_task': km_out_task,
        'km_in_task': km_in_task,
        'meter_reading_matched': (rec.meter_reading_matched or '').strip(),
        'upload_status': rec.upload_status or '',
        'media_url': url_for('fuel_expense_media', pk=rec.id),
        'has_media': rec.attachments.count() > 0,
        'edit_url': url_for('fuel_expense_edit', pk=rec.id),
    }


def _fuel_expense_add_form_ctx(workspace_employee_id, last_id=None):
    from fuel_expense_settings import fuel_expense_settings_payload
    return {
        'location_cascade': _fuel_expense_location_cascade_dict(),
        'last_rec': _fuel_expense_last_saved_rec(workspace_employee_id, last_id),
        'from_save': last_id is not None,
        'fuel_market_scan': _read_fuel_market_scan() or None,
        'fuel_expense_settings': fuel_expense_settings_payload(),
    }


def _fuel_expense_month_mpg(vehicle_id, fueling_date, workspace_employee_id=None, exclude_id=None):
    """Month-to-date MPG for a vehicle (MPG report formula)."""
    if not vehicle_id or not fueling_date:
        return None
    month_start = fueling_date.replace(day=1)
    month_end = fueling_date
    fuel_q = FuelExpense.query.filter(
        FuelExpense.vehicle_id == vehicle_id,
        FuelExpense.fueling_date >= month_start,
        FuelExpense.fueling_date <= month_end,
    )
    if workspace_employee_id:
        fuel_q = fuel_q.filter(FuelExpense.employee_id == workspace_employee_id)
    if exclude_id:
        fuel_q = fuel_q.filter(FuelExpense.id != exclude_id)
    rows = fuel_q.order_by(
        FuelExpense.fueling_date.asc(),
        db.case((FuelExpense.current_reading.is_(None), 1), else_=0).asc(),
        FuelExpense.current_reading.asc(),
        FuelExpense.id.asc(),
    ).all()
    if not rows:
        return None
    same_start_rows = [r for r in rows if r.fueling_date == month_start]
    same_end_rows = [r for r in rows if r.fueling_date == month_end]
    first_row = same_start_rows[0] if same_start_rows else rows[0]
    last_row = same_end_rows[-1] if same_end_rows else rows[-1]
    previous_reading = float(first_row.previous_reading) if first_row.previous_reading is not None else None
    current_reading = float(last_row.current_reading) if last_row.current_reading is not None else None
    km = (current_reading - previous_reading) if (previous_reading is not None and current_reading is not None) else None
    total_ltr = sum(float(r.liters or 0) for r in rows)
    mpg = round(km / total_ltr, 2) if (km is not None and total_ltr > 0) else None
    return {
        'mpg': mpg,
        'from_date': month_start.strftime('%d-%m-%Y'),
        'to_date': month_end.strftime('%d-%m-%Y'),
        'km': round(km, 2) if km is not None else None,
        'total_liters': round(total_ltr, 2),
        'entries': len(rows),
    }


def _scan_fuel_market_rates(force=False):
    """Scan PSO for today's rate. Stores per-date rates in a 'rates' dict."""
    today_s = pk_date().strftime('%Y-%m-%d')
    current = _read_fuel_market_scan()
    rates = current.get('rates') or {}
    if not force and today_s in rates and rates[today_s].get('ok'):
        current['scan_date'] = today_s
        return current

    try:
        pso = _scan_pso_rates()
    except Exception as exc:
        pso = {
            'source': 'PSO',
            'ok': False,
            'petrol': None,
            'diesel': None,
            'updated_at': '',
            'error': str(exc)[:220],
        }

    scanned_at = pk_now().strftime('%d-%m-%Y %I:%M:%S %p')
    rates[today_s] = {
        'ok': pso.get('ok', False),
        'petrol': pso.get('petrol'),
        'diesel': pso.get('diesel'),
        'scanned_at': scanned_at,
        'error': pso.get('error', ''),
    }

    scan_payload = {
        'scan_date': today_s,
        'scanned_at': scanned_at,
        'status': 'ok' if pso.get('ok') else 'error',
        'sources': {'pso': pso},
        'rates': rates,
    }
    _save_fuel_market_scan(scan_payload)
    return scan_payload





def _attendance_local_time():
    return pk_time()

def _attendance_local_date():
    return pk_date()

def _attendance_local_now():
    return pk_now()


def _get_checked_in_vehicle_ids(attendance_date=None):
    """Vehicle IDs with any open session (check-out pending), including overnight from prior dates."""
    q = db.session.query(Driver.vehicle_id).join(
        DriverAttendance, Driver.id == DriverAttendance.driver_id
    ).filter(
        DriverAttendance.check_in.isnot(None),
        DriverAttendance.check_out.is_(None),
        Driver.vehicle_id.isnot(None),
    )
    if attendance_date is not None:
        q = q.filter(DriverAttendance.attendance_date <= attendance_date)
    rows = q.distinct().all()
    return {r[0] for r in rows}


def _other_driver_open_session_on_vehicle(driver_id, vehicle_id):
    """True if another driver on this vehicle has check-in without check-out (any date)."""
    if not vehicle_id:
        return False
    return (
        db.session.query(DriverAttendance.id)
        .join(Driver, Driver.id == DriverAttendance.driver_id)
        .filter(
            Driver.vehicle_id == vehicle_id,
            Driver.id != driver_id,
            DriverAttendance.check_in.isnot(None),
            DriverAttendance.check_out.is_(None),
        )
        .first()
        is not None
    )


def _vehicle_oldest_pending_checkout(vehicle_id):
    """First open session on vehicle (check-in hai, check-out nahi) — kisi bhi driver / date."""
    if not vehicle_id:
        return None, None
    row = (
        db.session.query(DriverAttendance, Driver)
        .join(Driver, Driver.id == DriverAttendance.driver_id)
        .filter(
            Driver.vehicle_id == vehicle_id,
            DriverAttendance.check_in.isnot(None),
            DriverAttendance.check_out.is_(None),
        )
        .order_by(DriverAttendance.attendance_date.asc(), DriverAttendance.attendance_segment.asc())
        .first()
    )
    if not row:
        return None, None
    return row[0], row[1]


def _vehicle_pending_checkout_block_message(driver_id, vehicle):
    """Block new check-in until every open session on this vehicle is checked out."""
    if not vehicle or not getattr(vehicle, 'id', None):
        return None
    rec, d = _vehicle_oldest_pending_checkout(vehicle.id)
    if not rec:
        return None
    vno = (getattr(vehicle, 'vehicle_no', None) or '').strip() or 'Vehicle'
    dt = rec.attendance_date.strftime('%d-%m-%Y') if rec.attendance_date else ''
    ci = format_time_ampm(rec.check_in) if rec.check_in else ''
    if d and d.id == driver_id:
        return (
            f'{vno}: aap ka {dt} ka check-in ({ci}) abhi check-out pending hai — '
            'pehle us session ka check-out complete karein, phir naya check-in.'
        )
    dname = (d.name if d else '') or 'Driver'
    return (
        f'{vno}: {dname} ka {dt} ka check-in ({ci}) check-out pending hai — '
        'pehle us session ka check-out complete karein, phir is gari par kisi driver ka naya check-in.'
    )


def _attendance_checkin_stamp(rec):
    dt = rec.attendance_date.strftime('%d-%m-%Y') if rec and rec.attendance_date else ''
    ci = format_time_ampm(rec.check_in) if rec and rec.check_in else ''
    return dt, ci


def _vehicle_label(vehicle):
    return (getattr(vehicle, 'vehicle_no', None) or '').strip() or 'Vehicle'


def _vehicle_capacity_value(vehicle):
    if vehicle is None:
        return 1
    try:
        c = getattr(vehicle, 'driver_capacity', None)
        if c is None:
            return 1
        return max(1, int(c))
    except (TypeError, ValueError):
        return 1


def _time_to_minutes_day(t):
    """Minute-of-day 0–1439 from datetime.time."""
    if t is None:
        return None
    return t.hour * 60 + t.minute


def _minutes_span_contains(m, start_m, end_m):
    """Inclusive minute span on the clock; overnight wrap when start_m > end_m."""
    if start_m is None or end_m is None:
        return False
    if start_m <= end_m:
        return start_m <= m <= end_m
    return m >= start_m or m <= end_m


def _checkin_window_slot_from_time(tw, check_in_time):
    """Morning vs night check-in window from Time Overrides + actual check-in time (not assigned shift)."""
    if not tw or not check_in_time:
        return None
    m = _time_to_minutes_day(check_in_time)
    if m is None:
        return None
    ms = _time_to_minutes_day(tw.get('morning_start'))
    me = _time_to_minutes_day(tw.get('morning_end'))
    ns = _time_to_minutes_day(tw.get('night_start'))
    ne = _time_to_minutes_day(tw.get('night_end'))
    has_m = ms is not None and me is not None
    has_n = ns is not None and ne is not None
    in_m = _minutes_span_contains(m, ms, me) if has_m else False
    in_n = _minutes_span_contains(m, ns, ne) if has_n else False
    if in_m and not in_n:
        return 'morning'
    if in_n and not in_m:
        return 'night'
    if in_m and in_n:
        return 'morning'
    if has_m and not has_n:
        return 'morning' if in_m else None
    if has_n and not has_m:
        return 'night' if in_n else None
    return 'morning' if (6 * 60) <= m < (13 * 60) else 'night'


def _checkin_window_active_now(tw, slot, now_time):
    """Is the given slot's check-in window active at now_time?"""
    if not tw or not now_time or slot not in ('morning', 'night'):
        return False
    if slot == 'morning':
        start_t, end_t = tw.get('morning_start'), tw.get('morning_end')
    else:
        start_t, end_t = tw.get('night_start'), tw.get('night_end')
    if start_t is None or end_t is None:
        return False
    return _attendance_time_in_window(now_time, start_t, end_t)


def _alternate_checkin_window_active(tw, session_check_in_time, now_time):
    """
    True when the other check-in window (relative to session check-in) is active now.
    Uses Time Overrides only — not driver assigned shift.
    """
    slot = _checkin_window_slot_from_time(tw, session_check_in_time)
    if slot == 'morning':
        return _checkin_window_active_now(tw, 'night', now_time)
    if slot == 'night':
        return _checkin_window_active_now(tw, 'morning', now_time)
    return False


def _duty_shift_fallback_from_minutes(m):
    """If check-in falls outside configured windows: coarse Morning vs Evening band."""
    return 'Morning duty' if (6 * 60) <= m < (13 * 60) else 'Evening duty'


def _duty_shift_label_from_check_in(driver, vehicle, check_in_time):
    """Morning duty vs Evening duty from actual check-in + hierarchical Morning/Night windows."""
    m = _time_to_minutes_day(check_in_time)
    if m is None:
        return None
    win = _get_effective_time_window(driver=driver, vehicle_id=vehicle.id if vehicle else None)
    ms = _time_to_minutes_day(win.get('morning_start'))
    me = _time_to_minutes_day(win.get('morning_end'))
    ns = _time_to_minutes_day(win.get('night_start'))
    ne = _time_to_minutes_day(win.get('night_end'))
    has_m = ms is not None and me is not None
    has_n = ns is not None and ne is not None
    in_m = _minutes_span_contains(m, ms, me) if has_m else False
    in_n = _minutes_span_contains(m, ns, ne) if has_n else False
    if in_m and not in_n:
        return 'Morning duty'
    if in_n and not in_m:
        return 'Evening duty'
    if in_m and in_n:
        return 'Morning duty'
    if has_m and not has_n:
        return 'Morning duty' if in_m else 'Evening duty'
    if has_n and not has_m:
        return 'Evening duty' if in_n else 'Morning duty'
    return _duty_shift_fallback_from_minutes(m)


def _duty_shift_label(driver, vehicle, attendance_segment, rec=None):
    """Duty column: check-in time vs Morning/Night windows first; else segment slot or assigned shift."""
    if rec is not None and getattr(rec, 'check_in', None) is not None:
        return _duty_shift_label_from_check_in(driver, vehicle, rec.check_in)
    seg = int(attendance_segment or 1)
    cap = _vehicle_capacity_value(vehicle)
    if cap <= 1:
        return (driver.shift or '-') if driver else '-'
    slot_names = ('Morning duty', 'Evening duty')
    i = seg - 1
    if 0 <= i < len(slot_names):
        return slot_names[i]
    return 'Duty #%s' % seg


def _duty_shift_passes_filter(duty_label, filt):
    """filt: '', 'morning', or 'evening' — matches Duty shift column (Morning/Evening duty, assigned Morning/Night)."""
    if not filt or filt == 'all':
        return True
    s = (duty_label or '').strip().lower()
    is_morning = ('morning duty' in s) or (s == 'morning')
    is_evening = ('evening duty' in s) or (s == 'night') or ('evening' in s)
    if filt == 'morning':
        return bool(is_morning)
    if filt == 'evening':
        return bool(is_evening)
    return True


def _filter_attendance_rows_by_duty_shift(flat_rows, filt):
    if not flat_rows or not filt:
        return flat_rows
    return [r for r in flat_rows if _duty_shift_passes_filter(r.get('duty_shift'), filt)]


def _parse_duty_shift_filter_request():
    """Request arg duty_shift: '', 'morning', or 'evening'."""
    filt = (request.args.get('duty_shift') or '').strip().lower()
    if filt not in ('', 'morning', 'evening'):
        return ''
    return filt


def _parse_attendance_status_filter_request():
    """status_filter query: comma-separated Present,Leave,... (multi-select)."""
    allowed = {v for v, _ in ATTENDANCE_LIST_STATUS_FILTER_OPTIONS}
    raw = (request.args.get('status_filter') or '').strip()
    if not raw:
        parts = request.args.getlist('status')
    else:
        parts = [p.strip() for p in raw.split(',') if p.strip()]
    seen = set()
    out = []
    for p in parts:
        if p in allowed and p not in seen:
            seen.add(p)
            out.append(p)
    return out


def _filter_attendance_rows_by_status(flat_rows, status_filters):
    if not flat_rows or not status_filters:
        return flat_rows
    allowed = {s.strip() for s in status_filters}
    return [r for r in flat_rows if r.get('rec') and (r['rec'].status or '') in allowed]


def _driver_attendance_record_allowed_for_user(rec, uc):
    """Scope check for a DriverAttendance row (same idea as Attendance List)."""
    if not rec or not rec.driver:
        return False
    if uc.get('is_master_or_admin'):
        return True
    d = rec.driver
    veh = d.vehicle
    allowed_projects = uc.get('allowed_projects') or set()
    allowed_districts = uc.get('allowed_districts') or set()
    allowed_vehicles = uc.get('allowed_vehicles') or set()
    allowed_shifts = uc.get('allowed_shifts') or set()
    if allowed_projects and d.project_id not in allowed_projects:
        return False
    if allowed_districts:
        vd = veh.district_id if veh else None
        if d.district_id not in allowed_districts and vd not in allowed_districts:
            return False
    if allowed_vehicles and (not veh or veh.id not in allowed_vehicles):
        return False
    if allowed_shifts and d.shift not in allowed_shifts:
        return False
    return True


def _attendance_list_manual_edit_allowed():
    """Role permission for Edit check-in / Edit check-out links from Attendance List."""
    return user_can_access(session.get('permissions') or [], 'driver_attendance_list_manual_edit')


def _attendance_list_manual_delete_allowed():
    """Role permission for Delete check-in / Delete check-out from Attendance List."""
    return user_can_access(session.get('permissions') or [], 'driver_attendance_list_manual_delete')


_ATTENDANCE_GPS_CAM_REMARK = 'Ye GPS + Camera se attendance lagi hai.'


def _attendance_mark_record_clearable(rec):
    """Whether Leave/Late/Half-Day/Off/Absent mark can be removed from the mark form."""
    if not rec:
        return False
    if rec.remarks and _ATTENDANCE_GPS_CAM_REMARK in (rec.remarks or ''):
        return False
    if rec.check_in_latitude is not None or rec.check_in_longitude is not None:
        return False
    if rec.check_in_photo_path:
        return False
    if rec.check_out_latitude is not None or rec.check_out_longitude is not None:
        return False
    if rec.check_out_photo_path:
        return False
    return True


def _attendance_is_empty_gps_shell(rec):
    """
    Row with no check-in/out times or GPS media left, but GPS/Camera history in remarks.
    Shown after Attendance List delete check-in; should not stay as Present with dashes.
    """
    if not rec:
        return False
    if rec.check_in or rec.check_out:
        return False
    if (rec.check_in_photo_path or '').strip() or (rec.check_out_photo_path or '').strip():
        return False
    if rec.check_in_latitude is not None or rec.check_in_longitude is not None:
        return False
    if rec.check_out_latitude is not None or rec.check_out_longitude is not None:
        return False
    if rec.parking_station_id:
        return False
    rem = (rec.remarks or '')
    if 'GPS' in rem or 'GPS+Cam' in rem or 'Camera' in rem or 'List delete' in rem:
        return True
    return False


def _attendance_present_session_complete(rec):
    """Present session counts in reports only when both check-in and check-out exist."""
    if not rec or _attendance_is_empty_gps_shell(rec):
        return False
    if (rec.status or '').strip() != 'Present':
        return False
    return rec.check_in is not None and rec.check_out is not None


def _attendance_record_counts_in_report(rec):
    """Whether an attendance row contributes to monthly/daily report totals."""
    if not rec or _attendance_is_empty_gps_shell(rec):
        return False
    status = (rec.status or '').strip()
    if status == 'Present':
        return _attendance_present_session_complete(rec)
    return status in ('Absent', 'Leave', 'Late', 'Half-Day', 'Off')


def _attendance_status_abbr(status):
    """Short label for day-wise attendance grid cells."""
    s = (status or '').strip()
    return {
        'Present': 'P',
        'Absent': 'A',
        'Leave': 'L',
        'Late': 'Lt',
        'Half-Day': 'H',
        'Off': 'O',
    }.get(s, (s[:2] if s else ''))


def _vehicle_for_driver_on_date(driver_id, on_date):
    """Vehicle assignment as of a date from vehicle transfer history (ignores shift-only moves)."""
    if not driver_id or not on_date:
        return None
    t = (
        DriverTransfer.query.filter(
            DriverTransfer.driver_id == driver_id,
            DriverTransfer.is_shift_only.is_(False),
            DriverTransfer.transfer_date <= on_date,
        )
        .order_by(DriverTransfer.transfer_date.desc(), DriverTransfer.id.desc())
        .first()
    )
    if t and t.new_vehicle_id:
        return db.session.get(Vehicle, t.new_vehicle_id)
    first = (
        DriverTransfer.query.filter(
            DriverTransfer.driver_id == driver_id,
            DriverTransfer.is_shift_only.is_(False),
        )
        .order_by(DriverTransfer.transfer_date.asc(), DriverTransfer.id.asc())
        .first()
    )
    if first and first.old_vehicle_id:
        return db.session.get(Vehicle, first.old_vehicle_id)
    d = db.session.get(Driver, driver_id)
    return d.vehicle if d else None


def _vehicle_for_attendance_record(rec, driver=None):
    """Vehicle for a check-in — parking station first, then transfer history."""
    driver = driver or getattr(rec, 'driver', None)
    if rec and rec.parking_station_id:
        q = Vehicle.query.filter(Vehicle.parking_station_id == rec.parking_station_id)
        if rec.project_id:
            veh = q.filter(Vehicle.project_id == rec.project_id).first()
            if veh:
                return veh
        candidates = q.all()
        if len(candidates) == 1:
            return candidates[0]
    if driver:
        veh = _vehicle_for_driver_on_date(driver.id, rec.attendance_date)
        if veh:
            return veh
        return driver.vehicle
    return None


def _attendance_daily_slot_key(rec, driver=None, vehicle=None):
    """Morning (M) vs Evening (E) column — from check-in duty window, not assigned driver shift."""
    if not rec:
        return 'M'
    drv = driver or rec.driver
    veh = vehicle or (drv.vehicle if drv else None)

    if rec.check_in and drv:
        label = _duty_shift_label_from_check_in(drv, veh, rec.check_in) or ''
        low = label.lower()
        if 'evening' in low or 'night' in low:
            return 'E'
        if 'morning' in low:
            return 'M'

    seg = int(getattr(rec, 'attendance_segment', None) or 1)
    return 'E' if seg >= 2 else 'M'


def _format_attendance_time_display(t):
    """12-hour time for day-wise grid tooltips."""
    if not t:
        return None
    s = t.strftime('%I:%M %p')
    return s[1:] if s.startswith('0') else s


def _attendance_daily_cell_tooltip(rec):
    """Hover text for day-wise attendance grid cells."""
    if not rec:
        return ''
    status = (rec.status or '').strip()
    parts = []
    ci = _format_attendance_time_display(rec.check_in)
    co = _format_attendance_time_display(rec.check_out)
    if status and status != 'Present':
        parts.append(status)
    if ci:
        parts.append(f'Check-in: {ci}')
    if co:
        parts.append(f'Check-out: {co}')
    if not parts:
        return status or ''
    return ', '.join(parts)


def _count_month_present_days(grid):
    """Total Present (P) cells in the driver row — Morning + Evening, any shift."""
    total = 0
    for _day_num, slots in (grid or {}).items():
        if not isinstance(slots, dict):
            continue
        for slot_key in ('M', 'E'):
            cell = slots.get(slot_key) or {}
            abbr = cell.get('v', '') if isinstance(cell, dict) else (cell or '')
            if abbr == 'P':
                total += 1
    return total


DAILY_ATTENDANCE_EX_MARKER = 'Ex'
DAILY_ATTENDANCE_INACTIVE_MARKER = 'NA'
DAILY_ATTENDANCE_TRANSFER_MARKER = 'Tf'


def _daily_attendance_grid_cell_value(cell):
    """Display value from a day-wise grid cell (dict or legacy string)."""
    if isinstance(cell, dict):
        return (cell.get('v') or '').strip()
    return (cell or '').strip()


def _daily_attendance_grid_cell_skip_totals(cell):
    """Synthetic cells (Ex / NA / Tf) must not count in footer or status totals."""
    return isinstance(cell, dict) and cell.get('kind') in ('left', 'inactive', 'transfer')


def _daily_attendance_pre_segment_tip(driver, segment, tra_cache, seg_start):
    """Why days before segment_start are NA on this row."""
    rejoin_rec = tra_cache.rejoin_in_month.get(driver.id)
    if rejoin_rec and rejoin_rec.change_date and rejoin_rec.change_date == seg_start:
        return f'Before rejoin — Rejoined {rejoin_rec.change_date.strftime("%d-%b-%Y")}'
    transfer_in = segment.get('transfer_in')
    if transfer_in and transfer_in.transfer_date and seg_start:
        new_vehicle_start = transfer_in.transfer_date + timedelta(days=1)
        if seg_start >= new_vehicle_start:
            old_v = transfer_in.old_vehicle.vehicle_no if transfer_in.old_vehicle else '?'
            new_v = transfer_in.new_vehicle.vehicle_no if transfer_in.new_vehicle else '?'
            return (
                f'Before transfer — Moved from {old_v} to {new_v} on '
                f'{transfer_in.transfer_date.strftime("%d-%b-%Y")}'
            )
    if driver.assign_date and driver.assign_date == seg_start:
        return f'Before assignment — Joined {driver.assign_date.strftime("%d-%b-%Y")}'
    return f'Not on this vehicle before {seg_start.strftime("%d-%b-%Y")}'


def _daily_attendance_post_segment_transfer_tip(segment, seg_end):
    """Transferred away after segment_end on this vehicle row."""
    transfer_out = segment.get('transfer_out')
    if not transfer_out:
        return None
    old_v = transfer_out.old_vehicle.vehicle_no if transfer_out.old_vehicle else '?'
    new_v = transfer_out.new_vehicle.vehicle_no if transfer_out.new_vehicle else '?'
    return (
        f'Transferred — From {old_v} to {new_v} on '
        f'{transfer_out.transfer_date.strftime("%d-%b-%Y")} (duty ended {seg_end.strftime("%d-%b-%Y")})'
    )


def _daily_attendance_put_grid_marker(grid, day_num, slot, value, tip, kind):
    if day_num not in grid:
        grid[day_num] = {}
    existing = grid[day_num].get(slot)
    if _daily_attendance_grid_cell_value(existing):
        return
    grid[day_num][slot] = {'v': value, 'tip': tip, 'kind': kind}


def _daily_attendance_fill_segment_boundary_cells(
    grid, year, month, seg_start, seg_end, segment, driver, left_date, tra_cache,
):
    """Mark NA (pre-duty), Tf (transferred out), and Ex (job left) on empty M/E slots."""
    from calendar import monthrange

    _, ndays = monthrange(year, month)
    pre_tip = _daily_attendance_pre_segment_tip(driver, segment, tra_cache, seg_start)
    transfer_tip = _daily_attendance_post_segment_transfer_tip(segment, seg_end)
    ex_tip = (
        f'Ex-employee — Job Left {left_date.strftime("%d-%b-%Y")}'
        if left_date else None
    )

    for day_num in range(1, ndays + 1):
        att_d = date(year, month, day_num)
        for slot in ('M', 'E'):
            if att_d < seg_start:
                _daily_attendance_put_grid_marker(
                    grid, day_num, slot,
                    DAILY_ATTENDANCE_INACTIVE_MARKER, pre_tip, 'inactive',
                )
            elif left_date and att_d > left_date and att_d <= seg_end and ex_tip:
                _daily_attendance_put_grid_marker(
                    grid, day_num, slot,
                    DAILY_ATTENDANCE_EX_MARKER, ex_tip, 'left',
                )
            elif att_d > seg_end and transfer_tip:
                _daily_attendance_put_grid_marker(
                    grid, day_num, slot,
                    DAILY_ATTENDANCE_TRANSFER_MARKER, transfer_tip, 'transfer',
                )


def _daily_attendance_row_lifecycle_badges(driver, segment, tra_cache, start_d, end_d, seg_start):
    """Tags under driver name: Assigned / Rejoined / Transferred (same rules as TRA segments)."""
    badges = []
    rejoin_rec = tra_cache.rejoin_in_month.get(driver.id)
    transfer_in = segment.get('transfer_in')
    transfer_out = segment.get('transfer_out')

    if driver.assign_date and start_d <= driver.assign_date <= end_d and driver.assign_date == seg_start:
        badges.append({
            'type': 'assign',
            'text': f'Assigned {driver.assign_date.strftime("%d-%b-%Y")}',
            'title': f'Driver assigned on {driver.assign_date.strftime("%d-%b-%Y")}',
        })
    if rejoin_rec and rejoin_rec.change_date and rejoin_rec.change_date == seg_start:
        badges.append({
            'type': 'rejoin',
            'text': f'Rejoined {rejoin_rec.change_date.strftime("%d-%b-%Y")}',
            'title': f'Driver rejoined on {rejoin_rec.change_date.strftime("%d-%b-%Y")}',
        })
    if transfer_in:
        old_v = transfer_in.old_vehicle.vehicle_no if transfer_in.old_vehicle else '?'
        new_v = transfer_in.new_vehicle.vehicle_no if transfer_in.new_vehicle else '?'
        t_d = transfer_in.transfer_date
        badges.append({
            'type': 'transfer',
            'text': f'Transferred {t_d.strftime("%d-%b-%Y")} · {old_v} → {new_v}',
            'title': f'Transferred from {old_v} to {new_v} on {t_d.strftime("%d-%b-%Y")}',
        })
    elif transfer_out:
        old_v = transfer_out.old_vehicle.vehicle_no if transfer_out.old_vehicle else '?'
        new_v = transfer_out.new_vehicle.vehicle_no if transfer_out.new_vehicle else '?'
        t_d = transfer_out.transfer_date
        badges.append({
            'type': 'transfer',
            'text': f'Transferred {t_d.strftime("%d-%b-%Y")} · {old_v} → {new_v}',
            'title': (
                f'Transferred from {old_v} to {new_v} on {t_d.strftime("%d-%b-%Y")} '
                f'(last day on {old_v}: {(t_d - timedelta(days=1)).strftime("%d-%b-%Y")})'
            ),
        })
    return badges


def _driver_attendance_mark_redirect_url():
    """Rebuild mark form URL after clear/save."""
    date_str = (request.form.get('attendance_date') or request.args.get('date') or '').strip()
    params = {}
    if date_str:
        params['date'] = date_str
    for k in ('project_id', 'district_id', 'vehicle_id', 'shift', 'search', 'driver_id'):
        v = request.form.get(k) or request.args.get(k)
        if v is not None and str(v).strip() not in ('', '0'):
            params[k] = v
    return url_for('driver_attendance_mark', **_preserve_nav_from(params))


def _nav_back_ctx(default_url, default_label=None, show_without_nav_from=True):
    from nav_back import nav_back_context
    return nav_back_context(default_url, default_label, show_without_nav_from=show_without_nav_from)


def _master_nav_back():
    return _nav_back_ctx(url_for('module_hub', hub_slug='master-data'))


def _assignments_nav_back():
    return _nav_back_ctx(url_for('module_hub', hub_slug='assignments'))


def _transfers_nav_back():
    return _nav_back_ctx(url_for('module_hub', hub_slug='transfers'))


def _workforce_nav_back():
    return _nav_back_ctx(url_for('module_hub', hub_slug='workforce'))


def _finance_nav_back():
    return _nav_back_ctx(url_for('module_hub', hub_slug='finance'))


def _payroll_nav_back():
    return _nav_back_ctx(url_for('module_hub', hub_slug='payroll'))


def _books_nav_back():
    return _nav_back_ctx(url_for('module_hub', hub_slug='books'))


def _notifications_nav_back():
    return _nav_back_ctx(url_for('module_hub', hub_slug='notifications'))


def _administration_nav_back():
    return _nav_back_ctx(url_for('module_hub', hub_slug='administration'))


def _preserve_nav_from(params=None):
    from nav_back import preserve_nav_from
    return preserve_nav_from(params)


def _redirect_with_nav(endpoint, **params):
    """Redirect and keep nav_from so Back stays on hub / Report Centre after refresh."""
    return redirect(url_for(endpoint, **_preserve_nav_from(params)))


def _resolve_nav_back_url(default_url):
    from nav_back import back_url_for_request
    return back_url_for_request(default_url=default_url)


def _attendance_list_manual_checkout_allowed():
    """Role permission for Check-out button (missing checkout) from Attendance List."""
    return user_can_access(session.get('permissions') or [], 'driver_attendance_list_manual_checkout')


def _delete_stored_attendance_photo(path):
    """Remove R2 object by URL or local file under UPLOAD_FOLDER (paths like /uploads/attendance/...)."""
    if not path:
        return
    fp = (path or '').strip()
    if fp.startswith('http://') or fp.startswith('https://'):
        try:
            from r2_storage import delete_file_by_url
            delete_file_by_url(fp)
        except Exception:
            pass
        return
    rel = fp.lstrip('/')
    if rel.startswith('uploads/'):
        rel = rel[len('uploads/'):]
    full_path = os.path.join(app.config['UPLOAD_FOLDER'], rel.replace('/', os.sep))
    if os.path.isfile(full_path):
        try:
            os.remove(full_path)
        except OSError:
            pass


def _driver_attendance_list_redirect_params_from_form():
    """Rebuild Attendance List query args from POST hidden fields (filter bar state)."""
    keys = (
        'project_id', 'district_id', 'vehicle_id', 'shift', 'search', 'from_date', 'to_date',
        'driver_id', 'duty_shift', 'status_filter', 'list_filter', 'page', 'per_page',
    )
    params = {}
    for k in keys:
        v = request.form.get(k)
        if v is None or str(v).strip() == '':
            continue
        params[k] = str(v).strip()
    return params


def _build_attendance_media_gallery_items(flat_rows, gallery_shift, photo_kind):
    """Build media_items for maintenance_expense_media-style gallery."""
    gs = (gallery_shift or 'both').strip().lower()
    pk = (photo_kind or 'both').strip().lower()
    if gs not in ('morning', 'evening', 'both'):
        gs = 'both'
    if pk not in ('checkin', 'checkout', 'both'):
        pk = 'both'
    items = []
    for row in flat_rows or []:
        duty = row.get('duty_shift') or ''
        if gs != 'both' and not _duty_shift_passes_filter(duty, gs):
            continue
        rec = row.get('rec')
        d = row.get('driver')
        if not rec or not d:
            continue
        date_s = rec.attendance_date.strftime('%d-%m-%Y') if rec.attendance_date else ''
        if pk in ('checkin', 'both') and (rec.check_in_photo_path or '').strip():
            p = rec.check_in_photo_path.strip()
            url = media_url_filter(p)
            if url:
                items.append({
                    'url': url,
                    'type': 'image',
                    'name': f"{d.name or 'Driver'} ({d.driver_id or '-'}) — {date_s} — Check-in",
                    'stored_path': p,
                    'created_at': '',
                    'created_at_iso': '',
                    'size_bytes': None,
                    'size_label': '',
                    'download_url': url_for('driver_attendance_media_item_download', rec_id=rec.id, kind='checkin'),
                    'is_local_file': bool(_maintenance_attachment_local_full_path(p)),
                })
        if pk in ('checkout', 'both') and (rec.check_out_photo_path or '').strip():
            p = rec.check_out_photo_path.strip()
            url = media_url_filter(p)
            if url:
                items.append({
                    'url': url,
                    'type': 'image',
                    'name': f"{d.name or 'Driver'} ({d.driver_id or '-'}) — {date_s} — Check-out",
                    'stored_path': p,
                    'created_at': '',
                    'created_at_iso': '',
                    'size_bytes': None,
                    'size_label': '',
                    'download_url': url_for('driver_attendance_media_item_download', rec_id=rec.id, kind='checkout'),
                    'is_local_file': bool(_maintenance_attachment_local_full_path(p)),
                })
    return items


def _driver_attendance_flat_rows(
    project_id=None,
    district_id=None,
    vehicle_id=None,
    shift=None,
    search=None,
    driver_id=None,
    user_context=None,
    *,
    from_date=None,
    to_date=None,
    single_date=None,
):
    """One dict per DriverAttendance row: driver, rec, duty_shift. Same filters as Attendance List."""
    from auth_utils import get_user_context as _guc

    if user_context is None:
        uid = session.get('user_id')
        user_context = _guc(uid) if uid else {}
    allowed_projects = user_context.get('allowed_projects', set())
    allowed_districts = user_context.get('allowed_districts', set())
    allowed_vehicles = user_context.get('allowed_vehicles', set())
    allowed_shifts = user_context.get('allowed_shifts', set())
    is_master_or_admin = user_context.get('is_master_or_admin', False)

    q = DriverAttendance.query.options(
        db.joinedload(DriverAttendance.driver).joinedload(Driver.vehicle).joinedload(Vehicle.parking_station),
        db.joinedload(DriverAttendance.driver).joinedload(Driver.project),
        db.joinedload(DriverAttendance.driver).joinedload(Driver.district),
        db.joinedload(DriverAttendance.parking_station),
    ).join(Driver, DriverAttendance.driver_id == Driver.id).outerjoin(Vehicle, Driver.vehicle_id == Vehicle.id)

    if from_date and to_date:
        q = q.filter(DriverAttendance.attendance_date.between(from_date, to_date))
    elif from_date:
        q = q.filter(DriverAttendance.attendance_date >= from_date)
    elif to_date:
        q = q.filter(DriverAttendance.attendance_date <= to_date)
    elif single_date:
        q = q.filter(DriverAttendance.attendance_date == single_date)

    if driver_id:
        q = q.filter(DriverAttendance.driver_id == driver_id)

    if project_id:
        q = q.filter(
            db.or_(
                DriverAttendance.project_id == project_id,
                db.and_(DriverAttendance.project_id.is_(None), Driver.project_id == project_id),
            )
        )
    elif not is_master_or_admin and allowed_projects:
        q = q.filter(Driver.project_id.in_(list(allowed_projects)))

    if district_id:
        q = q.filter(db.or_(Driver.district_id == district_id, Vehicle.district_id == district_id))
    elif not is_master_or_admin and allowed_districts:
        q = q.filter(
            db.or_(
                Driver.district_id.in_(list(allowed_districts)),
                Vehicle.district_id.in_(list(allowed_districts)),
            )
        )

    if vehicle_id:
        q = q.filter(Driver.vehicle_id == vehicle_id)

    if shift:
        q = q.filter(Driver.shift == shift)

    if not is_master_or_admin and allowed_vehicles:
        q = q.filter(Driver.vehicle_id.in_(list(allowed_vehicles)))

    if not is_master_or_admin and allowed_shifts:
        q = q.filter(Driver.shift.in_(list(allowed_shifts)))

    if search:
        q = (
            q.outerjoin(Project, Driver.project_id == Project.id)
            .outerjoin(District, Vehicle.district_id == District.id)
            .outerjoin(ParkingStation, Vehicle.parking_station_id == ParkingStation.id)
        )
        flt = _multi_word_filter(
            search,
            Driver.name,
            Driver.driver_id,
            Driver.shift,
            Vehicle.vehicle_no,
            Vehicle.vehicle_type,
            Project.name,
            District.name,
            ParkingStation.name,
        )
        if flt is not None:
            q = q.filter(flt)

    records = q.order_by(DriverAttendance.attendance_date, Driver.name, DriverAttendance.attendance_segment).all()
    out = []
    for rec in records:
        if _attendance_is_empty_gps_shell(rec):
            continue
        d = rec.driver
        if not d:
            continue
        veh = d.vehicle
        duty = _duty_shift_label(d, veh, getattr(rec, 'attendance_segment', None), rec)
        out.append({'driver': d, 'rec': rec, 'duty_shift': duty})
    return out


def _driver_has_open_segment(driver_id, attendance_date):
    return db.session.query(DriverAttendance.id).filter(
        DriverAttendance.driver_id == driver_id,
        DriverAttendance.attendance_date == attendance_date,
        DriverAttendance.check_in.isnot(None),
        DriverAttendance.check_out.is_(None),
    ).first() is not None


def _count_driver_segments_with_checkin(driver_id, attendance_date):
    return (
        db.session.query(func.count(DriverAttendance.id))
        .filter(
            DriverAttendance.driver_id == driver_id,
            DriverAttendance.attendance_date == attendance_date,
            DriverAttendance.check_in.isnot(None),
        )
        .scalar()
        or 0
    )


def _next_attendance_segment(driver_id, attendance_date):
    mx = db.session.query(func.coalesce(func.max(DriverAttendance.attendance_segment), 0)).filter(
        DriverAttendance.driver_id == driver_id,
        DriverAttendance.attendance_date == attendance_date,
    ).scalar()
    return int(mx or 0) + 1


def _gps_marked_attendance_row(rec):
    if not rec:
        return False
    rem = (rec.remarks or '') or ''
    has_photo = bool((rec.check_in_photo_path or '').strip())
    has_remarks = 'GPS' in rem or 'Ye GPS' in rem or 'Camera' in rem or 'GPS+Cam' in rem
    return bool(has_photo or has_remarks)


def _open_gps_driver_attendance_session(driver_id, today):
    rows = (
        DriverAttendance.query.filter(
            DriverAttendance.driver_id == driver_id,
            DriverAttendance.attendance_date == today,
            DriverAttendance.check_in.isnot(None),
            DriverAttendance.check_out.is_(None),
        )
        .order_by(DriverAttendance.attendance_segment.desc())
        .all()
    )
    for rec in rows:
        if _gps_marked_attendance_row(rec):
            return rec
    return None


def _open_gps_driver_attendance_for_checkout(driver_id, today):
    """Today's open GPS-marked session, or overnight open session from yesterday."""
    rec = _open_gps_driver_attendance_session(driver_id, today)
    if rec:
        return rec
    yesterday = today - timedelta(days=1)
    rows = (
        DriverAttendance.query.filter(
            DriverAttendance.driver_id == driver_id,
            DriverAttendance.attendance_date == yesterday,
            DriverAttendance.check_in.isnot(None),
            DriverAttendance.check_out.is_(None),
        )
        .order_by(DriverAttendance.attendance_segment.desc())
        .all()
    )
    for r in rows:
        if _gps_marked_attendance_row(r):
            return r
    return None


def _open_driver_attendance_for_manual_checkout(driver_id, attendance_date):
    return (
        DriverAttendance.query.filter(
            DriverAttendance.driver_id == driver_id,
            DriverAttendance.attendance_date == attendance_date,
            DriverAttendance.check_in.isnot(None),
            DriverAttendance.check_out.is_(None),
        )
        .order_by(DriverAttendance.attendance_segment.asc())
        .first()
    )


def _pending_blocked_by_other_driver_on_vehicle(driver_id, vehicle_id, view_date, cap):
    if not vehicle_id:
        return False
    if _other_driver_open_session_on_vehicle(driver_id, vehicle_id):
        return True
    if cap <= 1:
        return (
            db.session.query(DriverAttendance.id)
            .join(Driver, Driver.id == DriverAttendance.driver_id)
            .filter(
                Driver.vehicle_id == vehicle_id,
                Driver.id != driver_id,
                DriverAttendance.attendance_date == view_date,
                DriverAttendance.check_in.isnot(None),
            )
            .first()
            is not None
        )
    return False


def _driver_marked_duty_off_no_checkin(driver_id, view_date):
    """True when attendance for this date is Off with no check-in (e.g. Bulk Status Off)."""
    return (
        db.session.query(DriverAttendance.id)
        .filter(
            DriverAttendance.driver_id == driver_id,
            DriverAttendance.attendance_date == view_date,
            DriverAttendance.status == 'Off',
            DriverAttendance.check_in.is_(None),
        )
        .first()
        is not None
    )


def _driver_excluded_from_missing_checkin_list(driver, view_date):
    if _driver_marked_duty_off_no_checkin(driver.id, view_date):
        return True
    cap = _vehicle_capacity_value(getattr(driver, 'vehicle', None))
    if _driver_has_open_segment(driver.id, view_date):
        return True
    if _count_driver_segments_with_checkin(driver.id, view_date) >= cap:
        return True
    if _pending_blocked_by_other_driver_on_vehicle(driver.id, driver.vehicle_id, view_date, cap):
        return True
    return False


def _manual_checkin_blocked_by_vehicle_rules(driver_id, vehicle, view_date):
    if not vehicle or not getattr(vehicle, 'id', None):
        return None
    vid = vehicle.id
    cap = _vehicle_capacity_value(vehicle)
    pending_msg = _vehicle_pending_checkout_block_message(driver_id, vehicle)
    if pending_msg:
        return pending_msg
    if cap <= 1:
        other = (
            db.session.query(DriverAttendance.id)
            .join(Driver, Driver.id == DriverAttendance.driver_id)
            .filter(
                Driver.vehicle_id == vid,
                Driver.id != driver_id,
                DriverAttendance.attendance_date == view_date,
                DriverAttendance.check_in.isnot(None),
            )
            .first()
        )
        if other:
            return (
                'Is vehicle (capacity 1) par aaj kisi aur driver ka check-in record hai — '
                'dusra manual check-in allow nahi.'
            )
    return None


@app.before_request
def require_login():
    """Redirect to login if user not logged in. Check permission for endpoint. Session timeout."""
    from routes_misc import _login_next_path
    def _api_error(payload, status):
        if request.path.startswith('/api/'):
            return jsonify(payload), status
        return None

    endpoint = request.endpoint or ''
    if endpoint.startswith('static'):
        return
    if endpoint in (
        'login', 'pwa_manifest', 'service_worker', 'biometric_login', 'app_logout', 'mobile_init', 'session_ping', 'app_check_update',
        'health_check', 'report_driver_profile_public',
    ):
        return
    if endpoint == 'set_new_password' and session.get('must_set_password_user_id'):
        return
    _uid = session.get('user_id')
    if not _uid:
        api_resp = _api_error({'ok': False, 'error': 'Session expired. Please login again.'}, 401)
        if api_resp:
            return api_resp
        return redirect(url_for('login', next=_login_next_path()))
    if not session.get('user'):
        try:
            _u = db.session.get(User, _uid)
            if _u and _u.is_active:
                session['user'] = _u.full_name or _u.username
            else:
                session.clear()
                api_resp = _api_error({'ok': False, 'error': 'Session expired. Please login again.'}, 401)
                if api_resp:
                    return api_resp
                return redirect(url_for('login', next=_login_next_path()))
        except Exception:
            app.logger.warning('require_login: failed to refresh session user for uid=%s', _uid, exc_info=True)
    if _uid and not session.get('permissions') and not session.get('is_master'):
        try:
            _perm_u = db.session.get(User, _uid)
            if _perm_u and _perm_u.is_active:
                _role = (_perm_u.role.name if _perm_u.role else '').strip()
                if _role == 'Master':
                    session['permissions'] = [p.code for p in Permission.query.all()]
                    session['is_master'] = True
                else:
                    _perms = _perm_u.permission_codes()
                    try:
                        from permissions_config import expand_login_permissions
                        _perms = expand_login_permissions(_perms)
                    except Exception:
                        app.logger.warning('require_login: expand_login_permissions failed', exc_info=True)
                    session['permissions'] = _perms
        except Exception:
            app.logger.warning('require_login: failed to refresh permissions for uid=%s', _uid, exc_info=True)
    if getattr(session, 'permanent', False):
        _uid = session.get('user_id')
        if _uid:
            _cached_ts = session.get('_user_active_ts')
            import time as _tm
            _now_ts = _tm.time()
            if not _cached_ts or (_now_ts - _cached_ts) > 300:
                try:
                    _u = db.session.get(User, _uid)
                    if not _u or not _u.is_active:
                        session.clear()
                        api_resp = _api_error({'ok': False, 'error': 'Your account has been deactivated. Please contact administrator.'}, 403)
                        if api_resp:
                            return api_resp
                        flash('Your account has been deactivated. Please contact administrator.', 'danger')
                        return redirect(url_for('login'))
                    session['_user_active_ts'] = _now_ts
                except Exception:
                    app.logger.warning('require_login: active-check failed for uid=%s', _uid, exc_info=True)
    # Session timeout (unless "remember me" made session permanent)
    timeout_mins = app.config.get('SESSION_TIMEOUT_MINUTES', 60)
    last = session.get('last_activity')
    now = pk_now()
    session['last_activity'] = now
    if last and not getattr(session, 'permanent', False):
        try:
            last_cmp = last
            now_cmp = now
            if getattr(last_cmp, 'tzinfo', None) is not None:
                last_cmp = last_cmp.replace(tzinfo=None)
            if getattr(now_cmp, 'tzinfo', None) is not None:
                now_cmp = now_cmp.replace(tzinfo=None)
            if (now_cmp - last_cmp).total_seconds() > timeout_mins * 60:
                session.clear()
                api_resp = _api_error({'ok': False, 'error': 'Session expired. Please login again.'}, 401)
                if api_resp:
                    return api_resp
                flash('Session expired. Please login again.', 'info')
                return redirect(url_for('login', next=_login_next_path()))
        except (TypeError, ValueError, AttributeError):
            pass
    # Role-based access: check if user has permission for this endpoint
    # Master ke liye role ki value nahi: sab routes allow, koi permission check nahi
    if session.get('is_master'):
        session['last_activity'] = now
        return
    required = get_required_permission(endpoint)
    # For master-data forms, switch Add → Edit permission when editing existing record (id present)
    if request.view_args and request.view_args.get('id'):
        if endpoint == 'company_form':
            required = 'companies_edit'
        elif endpoint == 'project_form':
            required = 'projects_edit'
        elif endpoint == 'district_form':
            required = 'districts_edit'
        elif endpoint == 'vehicle_form':
            required = 'vehicles_edit'
        elif endpoint == 'parking_form':
            required = 'parking_edit'
        elif endpoint == 'driver_form':
            required = 'drivers_edit'
        elif endpoint == 'employee_form':
            required = 'employees_edit'
        elif endpoint == 'driver_post_form':
            required = 'driver_post_edit'
        elif endpoint == 'product_form':
            required = 'product_edit'
    # Report Centre → Company Profile uses /companies/?mode=report (same route as master list).
    _explicitly_allowed = False
    if endpoint == 'companies' and request.args.get('mode') == 'report':
        perms_cp = session.get('permissions') or []
        if not user_can_access(perms_cp, 'report_company_profile') and not user_can_access(perms_cp, 'companies_list'):
            session['show_no_access'] = True
            api_resp = _api_error({'ok': False, 'error': 'You do not have permission for this action.'}, 403)
            if api_resp:
                return api_resp
            return redirect(url_for('login'))
        required = None
        _explicitly_allowed = True
    # S-01: Default-deny — unknown endpoints without permission mapping are blocked
    if required is None and not _explicitly_allowed and not is_endpoint_allowed_for_any_authed(endpoint):
        session['show_no_access'] = True
        api_resp = _api_error({'ok': False, 'error': 'You do not have permission for this action.'}, 403)
        if api_resp:
            return api_resp
        return redirect(url_for('login'))
    if required:
        perms = session.get('permissions') or []

        # ── Smart Dashboard Route Guard ───────────────────────────────────────
        # Having 'dashboard' (full) OR any individual card/feature permission is
        # sufficient to reach the dashboard page. The template then shows only
        # the elements the user's specific permissions allow.
        if required == 'dashboard':
            perms_set = set(perms)
            has_dashboard_access = (
                'dashboard' in perms_set
                or any(p.startswith('dashboard_card_') for p in perms_set)
                or 'view_fleet_map' in perms_set
                or 'global_search' in perms_set
            )
            if has_dashboard_access:
                return  # allow
            session['show_no_access'] = True
            api_resp = _api_error({'ok': False, 'error': 'You do not have permission for this action.'}, 403)
            if api_resp:
                return api_resp
            return redirect(url_for('login'))

        # Explicit: assignment (full) grants all assignment sub-pages (Vehicle to Parking, etc.)
        if not user_can_access(perms, required):
            if required and required.startswith('assign_') and ('assignment' in perms):
                pass  # allow
            else:
                session['show_no_access'] = True  # show once on login page, not per-request flash
                api_resp = _api_error({'ok': False, 'error': 'You do not have permission for this action.'}, 403)
                if api_resp:
                    return api_resp
                return redirect(url_for('login'))


def _form_control_tab_allowed(tab_key):
    """Per-tab Setting permission (Master = all; legacy form_control = all tabs)."""
    from permissions_config import user_has_form_control_tab
    return user_has_form_control_tab(
        session.get('permissions') or [],
        tab_key,
        is_master=bool(session.get('is_master')),
    )


def _any_form_control_tab_allowed():
    from permissions_config import user_has_any_form_control_tab
    return user_has_any_form_control_tab(
        session.get('permissions') or [],
        is_master=bool(session.get('is_master')),
    )


def _resolve_form_control_settings_tab(requested):
    from permissions_config import FORM_CONTROL_TAB_KEYS
    allowed = [k for k in FORM_CONTROL_TAB_KEYS if _form_control_tab_allowed(k)]
    if not allowed:
        return None
    req = (requested or 'attendance').strip()
    if req in allowed:
        return req
    return allowed[0]


def _form_control_tab_guard(tab_key, message=None):
    if _form_control_tab_allowed(tab_key):
        return None
    flash(message or 'You do not have permission for this settings tab.', 'danger')
    first_tab = _resolve_form_control_settings_tab(None)
    return redirect(url_for('form_control', settings_tab=first_tab or 'attendance'))


@app.before_request
def enforce_data_freeze():
    endpoint = request.endpoint or ''
    cfg = get_freeze_config()
    if not is_freeze_protected_request(request, endpoint, cfg):
        return
    if not cfg.get('is_effective'):
        return
    effective_date = extract_effective_date(request)
    blocked, rule_text = evaluate_freeze(cfg, effective_date)
    if not blocked:
        return

    msg = (
        f"Data freeze is active ({rule_text}). "
        f"Selected/derived date {effective_date.strftime('%d-%m-%Y')} cannot be modified."
    )
    if cfg.get('reason'):
        msg += f" Reason: {cfg['reason']}"

    if request.path.startswith('/api/'):
        return jsonify({'ok': False, 'message': msg, 'rule': rule_text}), 423

    flash(msg, 'danger')
    return redirect(request.referrer or url_for('form_control', settings_tab='freeze'))


def _endpoint_description(endpoint, method):
    """Human-readable short description for activity log."""
    if not endpoint:
        return None
    ep = (endpoint or '').replace('_', ' ').title()
    if method == 'GET':
        return f'Viewed {ep}'
    if method == 'POST':
        return f'Submitted {ep}'
    return f'{method} {ep}'


@app.after_request
def log_activity(response):
    """Log authenticated user activity for activity report (skip static, login, logout, report itself)."""
    try:
        endpoint = request.endpoint or ''
        if endpoint.startswith('static') or endpoint in ('login', 'logout'):
            return response
        if endpoint == 'activity_log_report':
            return response
        user_id = session.get('user_id')
        if not user_id:
            return response
        login_log_id = session.get('login_log_id')
        path = (request.path or '')[:500]
        desc = _endpoint_description(endpoint, request.method)
        act = ActivityLog(
            user_id=user_id,
            login_log_id=login_log_id,
            endpoint=endpoint[:120] if endpoint else None,
            method=request.method[:10] if request.method else None,
            path=path or None,
            description=desc[:500] if desc else None,
        )
        db.session.add(act)
        db.session.commit()
    except Exception:
        try:
            db.session.rollback()
        except Exception:
            pass
    return response


# ────────────────────────────────────────────────
# API: CNIC / License duplicate check (for driver form)
# ────────────────────────────────────────────────
def _cnic_digits(cnic):
    if not cnic:
        return ''
    return re.sub(r'[-\s]', '', str(cnic).strip())


DEFAULT_FIRST_PASSWORD = '123'


# ────────────────────────────────────────────────
# Serve uploaded files (vehicles/drivers documents and photos)
# ────────────────────────────────────────────────

@app.template_filter('media_url')
def media_url_filter(path):
    """Convert a stored file path (local relative or R2 full URL) to a usable URL."""
    if not path:
        return ''
    if path.startswith('http://') or path.startswith('https://'):
        return path
    return url_for('uploaded_file', filename=path)


def _expense_attachment_r2_ready():
    try:
        import r2_storage as _r2
        return bool(
            _r2.R2_PUBLIC_URL
            and _r2.R2_ACCESS_KEY_ID
            and _r2.R2_SECRET_ACCESS_KEY
            and _r2.R2_ENDPOINT_URL
            and _r2.R2_BUCKET_NAME
        )
    except Exception:
        return False


def _save_expense_attachment_path(file_storage, file_type, original_fn, r2_folder, upload_root, rel_prefix):
    """Store image/video on R2 when configured, else under upload_root/rel_prefix. Returns DB file_path value."""
    if _expense_attachment_r2_ready():
        try:
            from r2_storage import upload_image_file, upload_binary_file, upload_pdf_file
            file_storage.seek(0)
            if file_type == 'image':
                url = upload_image_file(file_storage, folder=r2_folder)
            elif file_type == 'pdf':
                url = upload_pdf_file(file_storage, folder=r2_folder)
            else:
                url = upload_binary_file(file_storage, folder=r2_folder, original_filename=original_fn)
            if url:
                return url
        except Exception as e:
            app.logger.warning('R2 expense media upload failed (%s): %s', r2_folder, e)
    file_storage.seek(0)
    fn = secure_filename(original_fn or '') or 'file'
    base, ext = os.path.splitext(fn)
    if not base:
        base = 'file'
    unique = f"{base}_{pk_now().strftime('%Y%m%d%H%M%S')}{ext}"
    subdir = os.path.join(upload_root, rel_prefix.replace('/', os.sep))
    os.makedirs(subdir, exist_ok=True)
    path = os.path.join(subdir, unique)
    file_storage.save(path)
    return '/'.join((rel_prefix.strip('/'), unique))


def _expense_attachment_max_bytes():
    try:
        mb = int(os.environ.get('EXPENSE_ATTACHMENT_MAX_MB', '300'))
    except ValueError:
        mb = 300
    return max(1, mb) * 1024 * 1024


def _filestorage_byte_size(file_storage):
    """Best-effort uploaded size; None if stream is not measurable."""
    try:
        file_storage.seek(0)
        stream = file_storage.stream
        stream.seek(0, 2)
        n = int(stream.tell())
        stream.seek(0)
        file_storage.seek(0)
        return n
    except Exception:
        try:
            file_storage.seek(0)
        except Exception:
            pass
        return None


def _add_expense_attachments_from_request(files, *, r2_folder, attachment_model, fk_kwargs):
    """Append attachment rows; caller commits. Returns list of skipped filenames + reasons."""
    allowed_image_ct = {'image/jpeg', 'image/png', 'image/gif', 'image/webp'}
    allowed_video_ct = {'video/mp4', 'video/webm', 'video/quicktime'}
    upload_root = app.config['UPLOAD_FOLDER']
    max_bytes = _expense_attachment_max_bytes()
    failed = []
    for f in files or []:
        if not f or not f.filename:
            continue
        fn = secure_filename(f.filename)
        if not fn:
            continue
        sz = _filestorage_byte_size(f)
        if sz is not None and sz > max_bytes:
            failed.append(f'{fn} (max {max_bytes // (1024 * 1024)} MB)')
            continue
        ext_lo = os.path.splitext(fn)[1].lower()
        content_type = (f.content_type or '').lower()
        if content_type in allowed_image_ct or ext_lo in {'.jpg', '.jpeg', '.png', '.gif', '.webp'}:
            ft = 'image'
        elif content_type in allowed_video_ct or ext_lo in {'.mp4', '.webm', '.mov'}:
            ft = 'video'
        else:
            continue
        fk_id = next(iter(fk_kwargs.values()))
        rel_prefix = f"{r2_folder}/{fk_id}"
        try:
            stored = _save_expense_attachment_path(f, ft, fn, r2_folder, upload_root, rel_prefix)
            row = attachment_model(file_path=stored, file_type=ft, original_name=fn, **fk_kwargs)
            db.session.add(row)
        except Exception as ex:
            app.logger.warning('Expense attachment skipped (%s): %s', fn, ex)
            failed.append(fn)
    return failed


def _add_expense_attachments_direct_client(payload, *, attachment_model, fk_kwargs, folder_prefix):
    """
    Persist attachment rows after the browser uploaded objects via presigned PUT to R2.
    Validates that public_url matches this app's R2_PUBLIC_URL + key and key prefix.
    """
    failed = []
    if not isinstance(payload, list):
        return ['Invalid attachment payload']
    try:
        import r2_storage as _r2
        base = _r2.R2_PUBLIC_URL.rstrip('/')
    except Exception:
        return ['R2 not configured']
    prefix = (folder_prefix or '').strip().rstrip('/') + '/'
    for item in payload:
        if not isinstance(item, dict):
            failed.append('invalid row')
            continue
        key = (item.get('key') or '').strip()
        public_url = (item.get('public_url') or '').strip()
        ft = (item.get('file_type') or '').strip().lower()
        orig = ((item.get('original_name') or '')[:255]).strip() or None
        if ft not in ('image', 'video'):
            failed.append(orig or key or 'bad file_type')
            continue
        if not key.startswith(prefix):
            failed.append(orig or 'bad key prefix')
            continue
        if f'{base}/{key}' != public_url:
            failed.append(orig or key or 'url mismatch')
            continue
        try:
            row = attachment_model(
                file_path=public_url,
                file_type=ft,
                original_name=orig,
                **fk_kwargs,
            )
            db.session.add(row)
        except Exception as ex:
            app.logger.warning('Direct attachment row failed (%s): %s', key, ex)
            failed.append(orig or key)
    return failed


_expense_upload_lock = threading.Lock()
_expense_upload_active = set()
_expense_delete_lock = threading.Lock()
_expense_delete_active = set()


def _expense_upload_queue_root(kind):
    root = os.path.join(app.config['UPLOAD_FOLDER'], 'expense_upload_queue', str(kind).strip().lower())
    os.makedirs(root, exist_ok=True)
    return root


def _prepare_expense_upload_manifest(files, kind, expense_id):
    """Persist request files quickly and return manifest for background upload."""
    allowed_image_ct = {'image/jpeg', 'image/png', 'image/gif', 'image/webp'}
    allowed_video_ct = {'video/mp4', 'video/webm', 'video/quicktime'}
    max_bytes = _expense_attachment_max_bytes()
    kind = str(kind or '').strip().lower() or 'maintenance'
    batch_id = f"{kind}_{expense_id}_{pk_now().strftime('%Y%m%d%H%M%S%f')}"
    batch_dir = os.path.join(_expense_upload_queue_root(kind), batch_id)
    os.makedirs(batch_dir, exist_ok=True)
    manifest = []
    skipped = []
    for f in files or []:
        if not f or not getattr(f, 'filename', None):
            continue
        fn = secure_filename(f.filename)
        if not fn:
            continue
        sz = _filestorage_byte_size(f)
        if sz is not None and sz > max_bytes:
            skipped.append(f'{fn} (max {max_bytes // (1024 * 1024)} MB)')
            continue
        ext_lo = os.path.splitext(fn)[1].lower()
        content_type = (f.content_type or '').lower()
        if content_type in allowed_image_ct or ext_lo in {'.jpg', '.jpeg', '.png', '.gif', '.webp'}:
            ftype = 'image'
        elif content_type in allowed_video_ct or ext_lo in {'.mp4', '.webm', '.mov'}:
            ftype = 'video'
        else:
            skipped.append(f'{fn} (unsupported)')
            continue
        unique = f"{uuid.uuid4().hex}_{fn}"
        temp_path = os.path.join(batch_dir, unique)
        f.seek(0)
        f.save(temp_path)
        manifest.append({
            'temp_path': temp_path,
            'original_name': fn,
            'file_type': ftype,
            'content_type': (f.content_type or mimetypes.guess_type(fn)[0] or '').strip(),
        })
    return manifest, skipped


def _expense_upload_kind_config(kind):
    kind = str(kind or '').strip().lower()
    mapping = {
        'maintenance': {
            'record_model': MaintenanceExpense,
            'attachment_model': MaintenanceExpenseAttachment,
            'fk_field': 'maintenance_expense_id',
            'folder': 'maintenance_expense',
            'log_prefix': 'Maintenance',
        },
        'fuel': {
            'record_model': FuelExpense,
            'attachment_model': FuelExpenseAttachment,
            'fk_field': 'fuel_expense_id',
            'folder': 'fuel_expense',
            'log_prefix': 'Fuel',
        },
        'oil': {
            'record_model': OilExpense,
            'attachment_model': OilExpenseAttachment,
            'fk_field': 'oil_expense_id',
            'folder': 'oil_expense',
            'log_prefix': 'Oil',
        },
        'work_order': {
            'record_model': MaintenanceWorkOrder,
            'attachment_model': MaintenanceWorkOrderAttachment,
            'fk_field': 'work_order_id',
            'folder': 'maintenance_work_order',
            'log_prefix': 'WorkOrder',
        },
    }
    return mapping.get(kind)


def _start_expense_upload_worker(kind, expense_id: int):
    kind = str(kind or '').strip().lower()
    key = (kind, int(expense_id))
    with _expense_upload_lock:
        if key in _expense_upload_active:
            return False
        _expense_upload_active.add(key)

    def _runner():
        try:
            _process_expense_upload_job(kind, expense_id)
        finally:
            with _expense_upload_lock:
                _expense_upload_active.discard(key)

    t = threading.Thread(target=_runner, daemon=True, name=f'{kind}-upload-{expense_id}')
    t.start()
    return True


def _process_expense_upload_job(kind, expense_id: int):
    kind = str(kind or '').strip().lower()
    cfg = _expense_upload_kind_config(kind)
    if not cfg:
        return
    with app.app_context():
        # Render PostgreSQL drops idle connections; dispose stale pool before
        # any DB work in this background thread so psycopg2 gets a fresh socket.
        try:
            db.engine.dispose()
        except Exception:
            pass

        def _safe_commit():
            """Commit and release the scoped session so next query gets a fresh connection."""
            try:
                db.session.commit()
            except Exception:
                db.session.rollback()
                raise
            finally:
                db.session.remove()

        rec = db.session.get(cfg['record_model'], expense_id)
        if not rec:
            return
        try:
            manifest = json.loads(rec.upload_manifest_json or '[]')
        except Exception:
            manifest = []
        if not manifest:
            rec.upload_status = 'success'
            rec.upload_failed = 0
            rec.upload_error = None
            rec.upload_finished_at = pk_now()
            _safe_commit()
            return
        rec.upload_status = 'processing'
        rec.upload_started_at = pk_now()
        rec.upload_finished_at = None
        rec.upload_error = None
        _safe_commit()

        # Re-fetch after session was released
        rec = db.session.get(cfg['record_model'], expense_id)
        if not rec:
            return

        remaining = []
        errors = []
        done = int(rec.upload_done or 0)
        for item in manifest:
            temp_path = (item.get('temp_path') or '').strip()
            original_name = (item.get('original_name') or '').strip()
            ftype = (item.get('file_type') or '').strip().lower()
            if not temp_path or not os.path.isfile(temp_path):
                errors.append(f'{original_name or "file"}: temp file missing')
                remaining.append(item)
                continue
            try:
                from werkzeug.datastructures import FileStorage
                with open(temp_path, 'rb') as fp:
                    fs = FileStorage(stream=fp, filename=original_name, content_type=(item.get('content_type') or None))
                    stored = _save_expense_attachment_path(
                        fs,
                        ftype if ftype in ('image', 'video') else 'image',
                        original_name,
                        cfg['folder'],
                        app.config['UPLOAD_FOLDER'],
                        f"{cfg['folder']}/{rec.id}",
                    )
                db.session.add(cfg['attachment_model'](
                    **{
                        cfg['fk_field']: rec.id,
                        'file_path': stored,
                        'file_type': (ftype if ftype in ('image', 'video') else None),
                        'original_name': (original_name[:255] if original_name else None),
                    }
                ))
                done += 1
                try:
                    os.remove(temp_path)
                except OSError:
                    pass
            except Exception as ex:
                app.logger.warning('%s queued upload failed (%s): %s', cfg['log_prefix'], original_name or temp_path, ex)
                errors.append(f'{original_name or "file"}: {ex}')
                remaining.append(item)
            rec.upload_done = done
            rec.upload_failed = len(remaining)
            rec.upload_manifest_json = json.dumps(remaining)
            rec.upload_error = '\n'.join(errors[-10:]) if errors else None
            try:
                _safe_commit()
                # Re-fetch rec after session release so next iteration has live object
                rec = db.session.get(cfg['record_model'], expense_id)
                if not rec:
                    return
            except Exception as _commit_ex:
                app.logger.warning('%s commit failed mid-loop: %s', cfg['log_prefix'], _commit_ex)
                break

        rec = db.session.get(cfg['record_model'], expense_id)
        if not rec:
            return
        rec.upload_failed = len(remaining)
        rec.upload_manifest_json = json.dumps(remaining)
        rec.upload_error = '\n'.join(errors[-10:]) if errors else None
        rec.upload_finished_at = pk_now()
        if remaining and done > 0:
            rec.upload_status = 'partial'
        elif remaining:
            rec.upload_status = 'error'
        else:
            rec.upload_status = 'success'
            rec.upload_error = None
        _safe_commit()


def _prepare_maintenance_upload_manifest(files, expense_id):
    return _prepare_expense_upload_manifest(files, 'maintenance', expense_id)


def _append_expense_upload_manifest(rec, kind, files, *, start_worker=True):
    """Queue uploaded files on disk for background R2 upload; append to existing manifest."""
    manifest, skipped = _prepare_expense_upload_manifest(files, kind, rec.id)
    if not manifest:
        return 0, skipped
    try:
        existing = json.loads(rec.upload_manifest_json or '[]')
    except Exception:
        existing = []
    combined = existing + manifest
    rec.upload_total = int(rec.upload_total or 0) + len(manifest)
    rec.upload_manifest_json = json.dumps(combined)
    rec.upload_status = 'processing'
    rec.upload_error = None
    if not rec.upload_started_at:
        rec.upload_started_at = pk_now()
    rec.upload_finished_at = None
    db.session.commit()
    if start_worker:
        _start_expense_upload_worker(kind, rec.id)
    return len(manifest), skipped


def _start_maintenance_upload_worker(expense_id: int):
    return _start_expense_upload_worker('maintenance', expense_id)


def _prepare_work_order_upload_manifest(files, work_order_id):
    return _prepare_expense_upload_manifest(files, 'work_order', work_order_id)


def _start_work_order_upload_worker(work_order_id: int):
    return _start_expense_upload_worker('work_order', work_order_id)


def _prepare_fuel_upload_manifest(files, expense_id):
    return _prepare_expense_upload_manifest(files, 'fuel', expense_id)


def _start_fuel_upload_worker(expense_id: int):
    return _start_expense_upload_worker('fuel', expense_id)


def _prepare_oil_upload_manifest(files, expense_id):
    return _prepare_expense_upload_manifest(files, 'oil', expense_id)


def _start_oil_upload_worker(expense_id: int):
    return _start_expense_upload_worker('oil', expense_id)


def _delete_stored_expense_attachment(file_path):
    if not file_path:
        return
    fp = (file_path or '').strip()
    if fp.startswith('http://') or fp.startswith('https://'):
        try:
            from r2_storage import delete_file_by_url
            delete_file_by_url(fp)
        except Exception:
            pass
        return
    full_path = os.path.join(app.config['UPLOAD_FOLDER'], fp)
    if os.path.isfile(full_path):
        try:
            os.remove(full_path)
        except OSError:
            pass


def _expense_cleanup_list_link(kind):
    kind = str(kind or '').strip().lower()
    mapping = {
        'fuel': '/expenses/fuel',
        'oil': '/oil-expenses',
        'maintenance': '/maintenance-expenses',
    }
    return mapping.get(kind) or '/workspace'


def _expense_cleanup_permission(kind):
    kind = str(kind or '').strip().lower()
    mapping = {
        'fuel': 'fuel_expense',
        'oil': 'oil_expense',
        'maintenance': 'maintenance_expense',
    }
    return mapping.get(kind) or 'workspace_dashboard'


def _latest_expense_cleanup_status(kind, employee_id):
    if not employee_id:
        return None
    kind = str(kind or '').strip().lower()
    row = (
        ExpenseDeleteCleanupJob.query
        .filter_by(employee_id=employee_id, expense_kind=kind)
        .order_by(ExpenseDeleteCleanupJob.updated_at.desc(), ExpenseDeleteCleanupJob.id.desc())
        .first()
    )
    if not row:
        return None
    status = (row.status or '').strip().lower() or 'processing'
    badge_map = {
        'success': ('success', 'Cleanup Success'),
        'partial': ('warning', 'Cleanup Partial'),
        'error': ('danger', 'Cleanup Failed'),
        'processing': ('warning', 'Cleanup Running'),
    }
    badge_class, label = badge_map.get(status, ('secondary', 'Cleanup Status'))
    retry_url = None
    if status in ('partial', 'error') and (row.failed_files or 0) > 0:
        retry_url = url_for('expense_delete_cleanup_retry', job_id=row.id)
    return {
        'label': label,
        'badge_class': badge_class,
        'deleted_files': int(row.deleted_files or 0),
        'total_files': int(row.total_files or 0),
        'failed_files': int(row.failed_files or 0),
        'retry_url': retry_url,
    }


def _start_expense_delete_cleanup_worker(kind, expense_id: int, file_paths, employee_id=None, initiated_by_user_id=None, cleanup_job_id=None):
    kind = str(kind or '').strip().lower()
    expense_id = int(expense_id)
    if not cleanup_job_id and not employee_id:
        return False

    with app.app_context():
        job = None
        if cleanup_job_id:
            job = ExpenseDeleteCleanupJob.query.filter_by(id=int(cleanup_job_id)).first()
        if not job:
            cleanup_paths = [str(p).strip() for p in (file_paths or []) if str(p).strip()]
            if not cleanup_paths:
                return False
            job = ExpenseDeleteCleanupJob(
                employee_id=int(employee_id),
                expense_kind=kind,
                expense_id=expense_id,
                status='processing',
                total_files=len(cleanup_paths),
                deleted_files=0,
                failed_files=0,
                pending_paths_json=json.dumps(cleanup_paths),
                initiated_by_user_id=initiated_by_user_id,
                retry_count=0,
            )
            db.session.add(job)
            db.session.commit()
        else:
            try:
                existing_failed = json.loads(job.pending_paths_json or '[]')
            except Exception:
                existing_failed = []
            if not existing_failed:
                return False
            job.status = 'processing'
            job.last_error = None
            job.finished_at = None
            job.retry_count = int(job.retry_count or 0) + 1
            db.session.commit()

        key = ('job', int(job.id))
        with _expense_delete_lock:
            if key in _expense_delete_active:
                return False
            _expense_delete_active.add(key)

    def _runner():
        try:
            with app.app_context():
                current = ExpenseDeleteCleanupJob.query.filter_by(id=int(job.id)).first()
                if not current:
                    return
                try:
                    cleanup_paths = json.loads(current.pending_paths_json or '[]')
                except Exception:
                    cleanup_paths = []
                cleanup_paths = [str(p).strip() for p in cleanup_paths if str(p).strip()]
                total = len(cleanup_paths)
                failed_paths = []
                for p in cleanup_paths:
                    try:
                        _delete_stored_expense_attachment(p)
                    except Exception:
                        failed_paths.append(p)
                failed = len(failed_paths)
                success = max(0, total - failed)
                current.total_files = total
                current.deleted_files = success
                current.failed_files = failed
                current.pending_paths_json = json.dumps(failed_paths)
                current.finished_at = pk_now()
                if failed == 0:
                    current.status = 'success'
                    current.last_error = None
                elif success > 0:
                    current.status = 'partial'
                    current.last_error = f'{failed} file(s) failed to delete.'
                else:
                    current.status = 'error'
                    current.last_error = 'All file deletions failed.'
                db.session.commit()
        except Exception as exc:
            with app.app_context():
                current = ExpenseDeleteCleanupJob.query.filter_by(id=int(job.id)).first()
                if current:
                    current.status = 'error'
                    current.last_error = str(exc)
                    current.finished_at = pk_now()
                    db.session.commit()
            app.logger.exception('%s delete cleanup failed for expense #%s', kind, expense_id)
        finally:
            with _expense_delete_lock:
                _expense_delete_active.discard(('job', int(job.id)))

    t = threading.Thread(target=_runner, daemon=True, name=f'{kind}-delete-{expense_id}-{job.id}')
    t.start()
    return True


# ────────────────────────────────────────────────
# Dashboard / Home
# ────────────────────────────────────────────────










def _load_pdf_writer_reader():
    """Import PDF merger classes with fallback names."""
    try:
        from pypdf import PdfReader, PdfWriter
        return PdfReader, PdfWriter
    except Exception:
        from PyPDF2 import PdfReader, PdfWriter
        return PdfReader, PdfWriter


def _html_to_pdf_bytes(html_content, landscape=True):
    """Render HTML to PDF bytes via headless Chromium (Playwright)."""
    from playwright.sync_api import sync_playwright

    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=True)
        try:
            page = browser.new_page()
            page.set_content(html_content or '', wait_until='load')
            page.emulate_media(media='print')
            pdf_bytes = page.pdf(
                format='A4',
                landscape=bool(landscape),
                print_background=True,
                margin={'top': '8mm', 'right': '8mm', 'bottom': '8mm', 'left': '8mm'},
            )
            if not pdf_bytes or len(pdf_bytes) < 900:
                raise ValueError('Generated PDF is empty')
            return pdf_bytes
        finally:
            browser.close()







def _require_master_admin():
    if not session.get('is_master'):
        flash('Only master admin can access Personal Tools.', 'danger')
        return False
    return True


def _validate_csrf_exempt_origin():
    """Validate Origin/Referer on CSRF-exempt endpoints to prevent cross-site attacks.
    Allows same-origin and Capacitor WebView (https://localhost) requests.
    Returns True if safe, False if cross-origin."""
    origin = (request.headers.get('Origin') or '').rstrip('/')
    referer = (request.headers.get('Referer') or '').rstrip('/')
    expected = request.host_url.rstrip('/')
    allowed_prefixes = [expected, 'https://localhost', 'http://localhost']
    if origin:
        return any(origin.startswith(p) for p in allowed_prefixes)
    if referer:
        return any(referer.startswith(p) for p in allowed_prefixes)
    return True



















# ────────────────────────────────────────────────
# Notifications (user-created broadcast; all users see and mark read)
# ────────────────────────────────────────────────


# ────────────────────────────────────────────────
# Reminders (personal; each user sees only their own)
# ────────────────────────────────────────────────
def _parse_time(s):
    """Parse HH:MM, H:MM, HH:MM:SS, or 12h AM/PM to time object."""
    if not s or not str(s).strip():
        return None
    s = str(s).strip().upper()
    ampm = None
    if s.endswith(' AM'):
        ampm = 'AM'
        s = s[:-3].strip()
    elif s.endswith(' PM'):
        ampm = 'PM'
        s = s[:-3].strip()
    elif s.endswith('AM'):
        ampm = 'AM'
        s = s[:-2].strip()
    elif s.endswith('PM'):
        ampm = 'PM'
        s = s[:-2].strip()
    parts = s.replace('.', ':').split(':')
    if len(parts) >= 2:
        try:
            h = int(parts[0])
            m = int(''.join(ch for ch in parts[1] if ch.isdigit()) or '0')
            if ampm == 'PM' and h < 12:
                h += 12
            elif ampm == 'AM' and h == 12:
                h = 0
            if 0 <= h <= 23 and 0 <= m <= 59:
                return time(h, m, 0)
        except ValueError:
            pass
    return None


def _time_input_value(raw):
    """Normalize stored/query time to HH:MM for HTML time inputs."""
    parsed = _parse_time(raw)
    if parsed:
        return parsed.strftime('%H:%M')
    return (raw or '').strip()







# ────────────────────────────────────────────────
# Account (profile, change password, biometric, session)
# ────────────────────────────────────────────────



def _biometric_token_valid(user, token):
    """Validate biometric token — v1 only (v0 legacy sunset)."""
    from routes_misc import _biometric_hmac_token
    import hmac as _hmac
    if not user or not token:
        return False
    return _hmac.compare_digest(token, _biometric_hmac_token(user))












# ────────────────────────────────────────────────
# Companies
# ────────────────────────────────────────────────





# ────────────────────────────────────────────────
# Projects
# ────────────────────────────────────────────────







# ────────────────────────────────────────────────
# Vehicles
# ────────────────────────────────────────────────


# ────────────────────────────────────────────────
# Accounts - Actual routes are now in routes_finance.py and registered in app.py
# ────────────────────────────────────────────────
# Placeholder routes removed - finance routes now active


# ────────────────────────────────────────────────
# Backup System (Download, Email, Save to path)
# ────────────────────────────────────────────────





















# ────────────────────────────────────────────────
# Drivers
# ────────────────────────────────────────────────

def _driver_update_field_data(driver):
    def _val(v):
        if v is None:
            return ''
        if hasattr(v, 'strftime'):
            return v.strftime('%d-%m-%Y')
        return str(v).strip()

    def _safe(text):
        s = (text or '').strip()
        return s if s else '—'

    district = driver.driver_district or (driver.district.name if getattr(driver, 'district', None) else '') or '—'
    phone = format_phone(driver.phone1) or '—'
    cnic = format_cnic(driver.cnic_no) or '—'
    address = (driver.address or '').replace('\r\n', ' ').replace('\n', ' ').strip() or '—'

    return {
        'district': _safe(district),
        'name': _safe(driver.name),
        'father_name': _safe(driver.father_name),
        'cnic': _safe(cnic),
        'dob': _safe(_val(driver.dob)),
        'license_type': _safe(driver.license_type),
        'license_no': _safe(driver.license_no),
        'issue_district': _safe(driver.issue_district),
        'license_issue_date': _safe(_val(driver.license_issue_date)),
        'license_expiry_date': _safe(_val(driver.license_expiry_date)),
        'address': _safe(address),
        'phone': _safe(phone),
        'bank_name': _safe(driver.bank_name),
        'account_no': _safe(driver.account_no),
        'account_title': _safe(driver.account_title),
        'remarks': _safe(driver.remarks),
    }


def _format_driver_update_whatsapp_body(fields):
    def _bullet(label, value):
        return f'• *{label}:* {value}'

    sep = '━━━━━━━━━━━━━━━━'
    f = fields
    detail_lines = [
        '',
        '👤 *DRIVER DETAIL*',
        sep,
        _bullet('District', f['district']),
        _bullet('Name', f['name']),
        _bullet('Father Name', f['father_name']),
        _bullet('CNIC', f['cnic']),
        _bullet('Date of Birth', f['dob']),
        '',
        '📋 *LICENSE DETAIL*',
        sep,
        _bullet('Type', f['license_type']),
        _bullet('Number', f['license_no']),
        _bullet('Issue District', f['issue_district']),
        _bullet('Issue Date', f['license_issue_date']),
        _bullet('Expiry Date', f['license_expiry_date']),
        '',
        '🏠 *ADDRESS*',
        f['address'],
        '',
        '📞 *CONTACT*',
        sep,
        f'📱 *Mobile:* {f["phone"]}',
        f'💬 *WhatsApp:* {f["phone"]}',
        '',
        '🏦 *BANK DETAIL*',
        sep,
        _bullet('Bank Name', f['bank_name']),
        _bullet('Account No', f['account_no']),
        _bullet('Account Title', f['account_title']),
        '',
        f'📝 *Remarks:* {f["remarks"]}',
        '',
        '*Please update this Driver.*',
    ]
    return '\n'.join(detail_lines)


def _build_driver_update_whatsapp_parts(driver):
    vehicle_no = (driver.vehicle.vehicle_no if getattr(driver, 'vehicle', None) else '') or '-'
    today = pk_date().strftime('%d-%m-%Y')
    fields = _driver_update_field_data(driver)
    detail_text = _format_driver_update_whatsapp_body(fields)
    return {
        'vehicle_no': vehicle_no,
        'report_date': today,
        'fields': fields,
        'detail_text': detail_text,
    }


def _build_driver_update_whatsapp_header(vehicle_no=None, report_date=None):
    vehicle = (vehicle_no or '-').strip() or '-'
    date = (report_date or '').strip() or '—'
    sep = '━━━━━━━━━━━━━━━━'
    return '\n'.join([
        '*DRIVER UPDATE*',
        sep,
        f'🚙 *Vehicle:* {vehicle}',
        f'📅 *Date:* {date}',
        '',
    ])


def _build_driver_update_whatsapp_text(driver, vehicle_no=None, report_date=None):
    parts = _build_driver_update_whatsapp_parts(driver)
    v = (vehicle_no if vehicle_no is not None else parts['vehicle_no']) or '-'
    d = (report_date if report_date is not None else parts['report_date']) or ''
    header = _build_driver_update_whatsapp_header(v, d)
    return header + parts['detail_text']







# Employees (non-driver staff)
# ────────────────────────────────────────────────






# ────────────────────────────────────────────────
# Employee Lifecycle: Assign / Deassign / Left / Rejoin / History
# ────────────────────────────────────────────────

def _log_employee_assignment(emp_id, action, district_id=None, project_id=None,
                              effective_date=None, reason=None, remarks=None, user_id=None):
    ea = EmployeeAssignment(
        employee_id=emp_id, action=action,
        district_id=district_id, project_id=project_id,
        effective_date=effective_date or pk_date(),
        reason=reason, remarks=remarks,
        created_by_user_id=user_id or session.get('user_id'),
    )
    db.session.add(ea)
    return ea
















# ────────────────────────────────────────────────
# Login (User model + permissions in session)
# ────────────────────────────────────────────────


def _persist_client_diagnostic(
    event_type,
    *,
    page_path=None,
    message=None,
    duration_ms=None,
    status_code=None,
    device_id=None,
    device_model=None,
    os_version=None,
    network_type=None,
):
    """Store one diagnostic row; never raises to caller."""
    user_id = session.get('user_id')
    if not user_id:
        return
    try:
        ua = (request.headers.get('User-Agent') or '')[:500] or None
        row = ClientDiagnosticLog(
            user_id=user_id,
            login_log_id=session.get('login_log_id'),
            device_id=(device_id or '')[:80] or None,
            user_agent=ua,
            event_type=(event_type or 'unknown')[:40],
            page_path=(page_path or request.path or '')[:500] or None,
            message=(message or '')[:2000] or None,
            duration_ms=int(duration_ms) if duration_ms is not None else None,
            status_code=int(status_code) if status_code is not None else None,
            device_model=(device_model or '')[:120] or None,
            os_version=(os_version or '')[:80] or None,
            network_type=(network_type or '')[:40] or None,
        )
        db.session.add(row)
        db.session.commit()
    except Exception:
        db.session.rollback()



# ────────────────────────────────────────────────
# User & Role Management (requires users_manage permission)
# ────────────────────────────────────────────────











def _get_user_scope():
    """
    Login ke waqt session me store ki hui user scope (projects/districts/vehicles/shifts) ko read karein.
    Master/Admin ke liye None return hota hai taa ke queries unko na filter karein.
    """
    from routes_auth import _user_has_full_scope
    if _user_has_full_scope():
        return None, None, None, None
    return (
        session.get('allowed_projects') or [],
        session.get('allowed_districts') or [],
        session.get('allowed_vehicles') or [],
        session.get('allowed_shifts') or [],
    )




def _sync_user_active_by_cnic(cnic, is_active):
    """User (username=CNIC) ka is_active set karo – Employee Inactive/Active ya Driver Left/Rejoin pe."""
    if not cnic or not (cnic or '').strip():
        return
    username = (cnic or '').strip()
    u = User.query.filter(func.lower(User.username) == username.lower()).first()
    if u and u.is_active != is_active:
        u.is_active = is_active
        try:
            db.session.commit()
        except Exception:
            db.session.rollback()


def _sync_user_full_name_by_cnic(cnic, full_name):
    """User (username=CNIC) ka full_name employee/driver record ke naam se sync karo — sirf display name."""
    if not cnic or not (cnic or '').strip():
        return
    username = (cnic or '').strip()
    if len(username) < 5:
        return
    new_name = (full_name or '').strip() or None
    u = User.query.filter(func.lower(User.username) == username.lower()).first()
    if not u:
        return
    if (u.full_name or '').strip() == (new_name or '').strip():
        return
    u.full_name = new_name
    try:
        db.session.commit()
    except Exception:
        db.session.rollback()



def _create_user_for_employee_or_driver(username_cnic, full_name, employee_post_id=None, role_id=None):
    """Auto-create login user when Employee or Driver is added. Username = CNIC, password = 123, must change on first login."""
    from routes_auth import _role_name
    if not username_cnic or not (username_cnic or '').strip():
        return None
    username = (username_cnic or '').strip()
    if len(username) < 5:  # CNIC at least 5 chars
        return None
    existing = User.query.filter(func.lower(User.username) == username.lower()).first()
    if existing:
        # User already bana hua hai → sirf sync/update karein (post/role/name) taa ke Roles & Permissions view me sahi count aaye
        updated = False
        synced_name = (full_name or '').strip() or username
        if synced_name and (existing.full_name or '').strip() != synced_name:
            existing.full_name = synced_name
            updated = True
        if employee_post_id and existing.employee_post_id != employee_post_id:
            existing.employee_post_id = employee_post_id
            updated = True
        if role_id and existing.role_id != role_id:
            # Safety: yahan Master/Admin auto-assign na karein (wo sirf manual user_form se ho)
            role = db.session.get(Role, role_id)
            if role and _role_name(role) not in ('Master', 'Admin'):
                existing.role_id = role_id
                updated = True
        if updated:
            try:
                db.session.commit()
            except Exception:
                db.session.rollback()
        return existing
    try:
        user = User(
            username=username,
            password_hash=generate_password_hash(DEFAULT_FIRST_PASSWORD),
            full_name=(full_name or '').strip() or username,
            employee_post_id=employee_post_id,
            role_id=role_id,
            is_active=True,
            force_password_change=True,
        )
        db.session.add(user)
        db.session.commit()
        return user
    except Exception:
        db.session.rollback()
        return None








# ────────────────────────────────────────────────
# Form Control (User & Access) — Attendance time windows, etc.
# ────────────────────────────────────────────────










# ────────────────────────────────────────────────
# Parking Stations
# ────────────────────────────────────────────────







# ────────────────────────────────────────────────
# Districts
# ────────────────────────────────────────────────






# ────────────────────────────────────────────────
# Assignment: Project → Company
# ────────────────────────────────────────────────
def _assign_project_to_company_data(search=None, sort_by='assign_date', sort_order='desc'):
    """Query assigned projects (optionally filtered by search). Returns list of Project."""
    q = Project.query.filter(Project.company_id.isnot(None))
    if search:
        flt = _multi_word_filter(search, Project.name, Company.name)
        if flt is not None:
            q = q.outerjoin(Company, Project.company_id == Company.id).filter(flt)
    
    # Apply sorting - database level for performance
    if sort_by == 'project':
        order_col = Project.name
    elif sort_by == 'company':
        q = q.join(Company, Project.company_id == Company.id)
        order_col = Company.name
    elif sort_by == 'assign_date':
        order_col = Project.assign_date
    else:
        order_col = Project.assign_date
    
    if sort_order == 'asc':
        q = q.order_by(order_col.asc())
    else:
        q = q.order_by(order_col.desc())
    
    return q.all()








# ────────────────────────────────────────────────
# Assignment: Project → District
# ────────────────────────────────────────────────

def _assign_project_to_district_data(search=None, project_id=None, district_id=None):
    """Shared query for list, print, export. Optional filter by project_id and/or district_id."""
    query = db.session.query(
        Project, District, project_district.c.assign_date, project_district.c.remarks
    ).join(project_district, Project.id == project_district.c.project_id).join(
        District, District.id == project_district.c.district_id
    )
    if project_id:
        query = query.filter(Project.id == project_id)
    if district_id:
        query = query.filter(District.id == district_id)
    if search:
        flt = _multi_word_filter(search,
            Project.name, District.name,
            project_district.c.remarks,
            cast(project_district.c.assign_date, SAString))
        if flt is not None:
            query = query.filter(flt)
    results = query.order_by(Project.name).all()
    assigned_structured = []
    for proj, dist, a_date, a_remarks in results:
        assigned_structured.append(({'assign_date': a_date, 'remarks': a_remarks}, proj, dist))
    return assigned_structured








# ────────────────────────────────────────────────
# Assignment: Vehicle → District
# ────────────────────────────────────────────────
def _assign_vehicle_to_district_data(search=None, project_id=None, district_id=None, sort_by='assign_date', sort_order='desc'):
    """Query assigned vehicles (optionally filtered by search, project_id, district_id). Returns list of Vehicle."""
    query = Vehicle.query.filter(Vehicle.district_id.isnot(None))
    if project_id:
        query = query.filter(Vehicle.project_id == project_id)
    if district_id:
        query = query.filter(Vehicle.district_id == district_id)
    if search:
        query = query.outerjoin(Project,  Vehicle.project_id  == Project.id) \
                     .outerjoin(District, Vehicle.district_id == District.id)
        flt = _multi_word_filter(search, Vehicle.vehicle_no, Vehicle.model, Project.name, District.name)
        if flt is not None:
            query = query.filter(flt)
    
    # Apply sorting - database level for performance
    if sort_by == 'vehicle':
        order_col = Vehicle.vehicle_no
    elif sort_by == 'project':
        query = query.join(Project, Vehicle.project_id == Project.id)
        order_col = Project.name
    elif sort_by == 'district':
        query = query.join(District, Vehicle.district_id == District.id)
        order_col = District.name
    elif sort_by == 'assign_date':
        order_col = Vehicle.assign_to_district_date
    else:
        order_col = Vehicle.assign_to_district_date
    
    if sort_order == 'asc':
        query = query.order_by(order_col.asc())
    else:
        query = query.order_by(order_col.desc())
    
    return query.all()













def _assign_vehicle_to_parking_data(search=None, project_id=None, district_id=None, from_date=None, to_date=None, sort_by='assign_date', sort_order='desc'):
    """Query vehicles with parking assigned (optionally filtered by search, project_id, district_id, date range). Returns list of Vehicle."""
    query = Vehicle.query.filter(Vehicle.parking_station_id.isnot(None))
    if project_id:
        query = query.filter(Vehicle.project_id == project_id)
    if district_id:
        query = query.filter(Vehicle.district_id == district_id)
    if from_date:
        query = query.filter(Vehicle.parking_assign_date >= from_date)
    if to_date:
        query = query.filter(Vehicle.parking_assign_date <= to_date)

    # Track which joins have already been applied to avoid duplicates
    _joined_project  = False
    _joined_district = False
    _joined_parking  = False

    if search:
        # Use outerjoin so vehicles with NULL project_id / district_id are NOT excluded
        query = query.outerjoin(Project,        Vehicle.project_id         == Project.id) \
                     .outerjoin(District,       Vehicle.district_id        == District.id) \
                     .outerjoin(ParkingStation, Vehicle.parking_station_id == ParkingStation.id)
        _joined_project  = True
        _joined_district = True
        _joined_parking  = True
        flt = _multi_word_filter(search,
            Vehicle.vehicle_no, Vehicle.model,
            Project.name, District.name, ParkingStation.name)
        if flt is not None:
            query = query.filter(flt)

    # Apply sorting — reuse existing joins if already present
    if sort_by == 'vehicle':
        order_col = Vehicle.vehicle_no
    elif sort_by == 'project':
        if not _joined_project:
            query = query.outerjoin(Project, Vehicle.project_id == Project.id)
        order_col = Project.name
    elif sort_by == 'district':
        if not _joined_district:
            query = query.outerjoin(District, Vehicle.district_id == District.id)
        order_col = District.name
    elif sort_by == 'parking':
        if not _joined_parking:
            query = query.outerjoin(ParkingStation, Vehicle.parking_station_id == ParkingStation.id)
        order_col = ParkingStation.name
    elif sort_by == 'assign_date':
        order_col = Vehicle.parking_assign_date
    else:
        order_col = Vehicle.parking_assign_date

    query = query.order_by(order_col.asc() if sort_order == 'asc' else order_col.desc())
    return query.all()






def _assign_driver_to_vehicle_data(search=None, project_id=None, district_id=None, sort_by='driver', sort_order='asc'):
    """Query (Driver, Vehicle) pairs for assigned drivers. Optional filter by project_id, district_id."""
    query = db.session.query(Driver, Vehicle).join(Vehicle, Driver.vehicle_id == Vehicle.id)
    if project_id:
        query = query.filter(Driver.project_id == project_id)
    if district_id:
        query = query.filter(Vehicle.district_id == district_id)

    _joined_project  = False
    _joined_district = False

    if search:
        query = query.outerjoin(Project,  Driver.project_id   == Project.id) \
                     .outerjoin(District, Vehicle.district_id == District.id)
        _joined_project  = True
        _joined_district = True
        flt = _multi_word_filter(search,
            Driver.name, Driver.driver_id, Driver.cnic_no, Driver.shift,
            Vehicle.vehicle_no, Vehicle.model,
            Project.name, District.name)
        if flt is not None:
            query = query.filter(flt)

    # Apply sorting — reuse joins already applied
    if sort_by == 'driver':
        order_col = Driver.name
    elif sort_by == 'vehicle':
        order_col = Vehicle.vehicle_no
    elif sort_by == 'project':
        if not _joined_project:
            query = query.outerjoin(Project, Driver.project_id == Project.id)
        order_col = Project.name
    elif sort_by == 'district':
        if not _joined_district:
            query = query.outerjoin(District, Vehicle.district_id == District.id)
        order_col = District.name
    else:
        order_col = Driver.name

    query = query.order_by(order_col.asc() if sort_order == 'asc' else order_col.desc())
    return query.all()




# Transfer Section - Project
# ==========================================
# VEHICLE TRANSFER ROUTES
# ==========================================

# ==========================================
# DRIVER TRANSFER ROUTES & APIs
# ==========================================

# ── Active Driver Summary Report ─────────────────────────────────────────
def _active_drivers_data(project_id=0, district_id=0, vehicle_id=0, shift='',
                         from_date_val=None, to_date_val=None,
                         allowed_vehicles=None, allowed_projects=None,
                         allowed_districts=None, is_master_or_admin=True):
    """Shared query for active (vehicle-assigned) drivers with optional scope/filters."""
    rejoin_sub = db.session.query(
        DriverStatusChange.driver_id,
        func.max(DriverStatusChange.change_date).label('rejoin_date')
    ).filter(
        DriverStatusChange.action_type == 'rejoin'
    ).group_by(DriverStatusChange.driver_id).subquery()

    query = db.session.query(
        Driver, Vehicle, Project, District, rejoin_sub.c.rejoin_date
    ).join(
        Vehicle, Driver.vehicle_id == Vehicle.id
    ).outerjoin(
        Project, Driver.project_id == Project.id
    ).outerjoin(
        District, Vehicle.district_id == District.id
    ).outerjoin(
        rejoin_sub, Driver.id == rejoin_sub.c.driver_id
    ).filter(
        Driver.vehicle_id.isnot(None),
        Driver.status != 'Left'
    )

    # User scope enforcement (non-master/admin)
    if not is_master_or_admin:
        if allowed_projects:
            query = query.filter(
                or_(Driver.project_id.in_(list(allowed_projects)),
                    Vehicle.project_id.in_(list(allowed_projects)))
            )
        if allowed_districts:
            query = query.filter(
                or_(Driver.district_id.in_(list(allowed_districts)),
                    Vehicle.district_id.in_(list(allowed_districts)))
            )
        if allowed_vehicles:
            query = query.filter(Driver.vehicle_id.in_(list(allowed_vehicles)))

    # Filter params (district is on Vehicle, use OR for both driver and vehicle columns)
    if project_id:
        query = query.filter(
            or_(Driver.project_id == project_id, Vehicle.project_id == project_id)
        )
    if district_id:
        query = query.filter(
            or_(Driver.district_id == district_id, Vehicle.district_id == district_id)
        )
    if vehicle_id:
        query = query.filter(Driver.vehicle_id == vehicle_id)
    if shift:
        query = query.filter(Driver.shift == shift)
    if from_date_val:
        query = query.filter(Driver.assign_date >= from_date_val)
    if to_date_val:
        query = query.filter(Driver.assign_date <= to_date_val)

    return query.order_by(*vehicle_order_by(Project.name, District.name), Driver.name).all()


def _parse_active_driver_filters():
    """Parse common URL args for active-driver-report routes."""
    project_id   = request.args.get('project_id', type=int) or 0
    district_id  = request.args.get('district_id', type=int) or 0
    vehicle_id   = request.args.get('vehicle_id', type=int) or 0
    shift        = (request.args.get('shift') or '').strip()
    from_date_str = (request.args.get('from_date') or '').strip()
    to_date_str   = (request.args.get('to_date') or '').strip()
    from_date_val = None
    to_date_val   = None
    try:
        if from_date_str:
            from_date_val = datetime.strptime(from_date_str, '%d-%m-%Y').date()
    except ValueError:
        from_date_str = ''
    try:
        if to_date_str:
            to_date_val = datetime.strptime(to_date_str, '%d-%m-%Y').date()
    except ValueError:
        to_date_str = ''
    return project_id, district_id, vehicle_id, shift, from_date_str, to_date_str, from_date_val, to_date_val


def _parse_salary_slip_filters():
    district_id = request.args.get('district_id', type=int) or 0
    project_id = request.args.get('project_id', type=int) or 0
    vehicle_id = request.args.get('vehicle_id', type=int) or 0
    driver_id = request.args.get('driver_id', type=int) or 0
    return district_id, project_id, vehicle_id, driver_id


def _driver_accessible_for_salary_slip(driver, user_context):
    """Align with _active_drivers_data / active driver list (scoped project, district, vehicle)."""
    if not driver or driver.status != 'Active' or not driver.vehicle_id:
        return False
    is_master_or_admin = user_context.get('is_master_or_admin', False)
    if is_master_or_admin:
        return True
    allowed_projects = user_context.get('allowed_projects', set())
    allowed_districts = user_context.get('allowed_districts', set())
    allowed_vehicles = user_context.get('allowed_vehicles', set())
    v = driver.vehicle
    if allowed_projects:
        dp, vp = driver.project_id, (v.project_id if v else None)
        if not ((dp in allowed_projects) or (vp is not None and vp in allowed_projects)):
            return False
    if allowed_districts:
        dd, vd = driver.district_id, (v.district_id if v else None)
        if not ((dd in allowed_districts) or (vd is not None and vd in allowed_districts)):
            return False
    if allowed_vehicles and driver.vehicle_id not in allowed_vehicles:
        return False
    return True


def _driver_to_salary_slip_payload(driver):
    """Build JSON-serializable dict of auto-filled (driver) fields for the salary slip UI."""
    v = driver.vehicle
    dist = None
    if driver.district_id and driver.district:
        dist = driver.district
    elif v and v.district_id and v.district:
        dist = v.district
    proj = driver.project
    if not proj and v and v.project_id:
        proj = v.project
    company_name = ''
    if proj and proj.company:
        company_name = proj.company.name or ''
    dstr = lambda d: d.strftime('%d-%m-%Y') if d else ''
    return {
        'id': driver.id,
        'driver_code': (driver.driver_id or '').strip(),
        'name': driver.name or '',
        'father_name': driver.father_name or '',
        'cnic': format_cnic(driver.cnic_no) if driver.cnic_no else '',
        'dob': dstr(driver.dob),
        'address': (driver.address or '').strip(),
        'phone1': (driver.phone1 or '').strip(),
        'post': (driver.post or 'Driver').strip() or 'Driver',
        'shift': (driver.shift or '').strip(),
        'assign_date': dstr(driver.assign_date),
        'license_no': (driver.license_no or '').strip(),
        'bank_name': (driver.bank_name or '').strip(),
        'account_no': (driver.account_no or '').strip(),
        'account_title': (driver.account_title or '').strip(),
        'district': dist.name if dist else '',
        'project': proj.name if proj else '',
        'company': company_name,
        'vehicle_no': (v.vehicle_no if v else '') or '',
        'vehicle_model': (v.model if v else '') or '',
        'vehicle_type': (v.vehicle_type if v else '') or '',
    }




def _coerce_salary_slip_chain(district_id, project_id, vehicle_id, selected_driver_id, user_context):
    """Drop invalid / inconsistent filter combinations (cascade: district → project → vehicle → driver)."""
    if not district_id:
        return 0, 0, 0, 0
    if project_id:
        if not db.session.query(project_district).filter_by(district_id=district_id, project_id=project_id).first():
            project_id = 0
    if not project_id:
        return district_id, 0, 0, 0
    if vehicle_id:
        v = db.session.get(Vehicle, vehicle_id)
        if (not v
                or v.district_id != district_id
                or v.project_id != project_id):
            vehicle_id = 0
    if not vehicle_id:
        return district_id, project_id, 0, 0
    if selected_driver_id:
        dr = db.session.get(Driver, selected_driver_id)
        if (not dr
                or dr.vehicle_id != vehicle_id
                or (dr.status or '') != 'Active'
                or not _driver_accessible_for_salary_slip(dr, user_context)):
            selected_driver_id = 0
    return district_id, project_id, vehicle_id, selected_driver_id


def _build_salary_slip_driver_choices(vehicle_id, user_context):
    if not vehicle_id:
        return []
    q = (
        Driver.query.filter(Driver.vehicle_id == vehicle_id, Driver.status == 'Active')
        .order_by(Driver.name)
    )
    out = []
    for d in q.all():
        if _driver_accessible_for_salary_slip(d, user_context):
            out.append((d.id, f"{d.name} ({d.driver_id})"))
    return out






# ── Oil Change Alert Report ──────────────────────────────────────────────
def _apply_workspace_employee_scope_for_expense(query, model_cls, workspace_employee_id):
    if workspace_employee_id:
        query = query.filter(
            db.or_(
                model_cls.employee_id == workspace_employee_id,
                model_cls.employee_id.is_(None),
            )
        )
    return query


def _vehicle_latest_meter_reading(vehicle_id, workspace_employee_id=None):
    task_row = VehicleDailyTask.query.filter(
        VehicleDailyTask.vehicle_id == vehicle_id,
        VehicleDailyTask.close_reading.isnot(None),
    ).order_by(VehicleDailyTask.task_date.desc(), VehicleDailyTask.id.desc()).first()
    if task_row and task_row.close_reading is not None:
        return float(task_row.close_reading), task_row.task_date, 'Task Close Reading'
    return None, None, None


def _vehicle_last_oil_change_base(vehicle_id, workspace_employee_id=None):
    oil_q = OilExpense.query.filter(
        OilExpense.vehicle_id == vehicle_id,
        OilExpense.current_reading.isnot(None),
    )
    oil_q = _apply_workspace_employee_scope_for_expense(oil_q, OilExpense, workspace_employee_id)
    oil_row = oil_q.order_by(OilExpense.expense_date.desc(), OilExpense.id.desc()).first()
    if oil_row and oil_row.current_reading is not None:
        return float(oil_row.current_reading), oil_row.expense_date, 'Last Oil Change'
    if workspace_employee_id:
        fallback = _fallback_vehicle_previous_reading(workspace_employee_id, vehicle_id, 'oil')
        if fallback is not None:
            return float(fallback), None, 'Reading Setup'
    setup = WorkspaceVehicleReadingSetup.query.filter(
        WorkspaceVehicleReadingSetup.vehicle_id == vehicle_id,
        WorkspaceVehicleReadingSetup.oil_previous_reading.isnot(None),
    ).order_by(WorkspaceVehicleReadingSetup.updated_at.desc()).first()
    if setup and setup.oil_previous_reading is not None:
        return float(setup.oil_previous_reading), setup.setup_date, 'Reading Setup'
    return None, None, None


def _vehicle_last_oil_change_date(vehicle_id, workspace_employee_id=None):
    oil_q = OilExpense.query.filter(
        OilExpense.vehicle_id == vehicle_id,
        OilExpense.expense_date.isnot(None),
    )
    oil_q = _apply_workspace_employee_scope_for_expense(oil_q, OilExpense, workspace_employee_id)
    oil_row = oil_q.order_by(OilExpense.expense_date.desc(), OilExpense.id.desc()).first()
    if oil_row and oil_row.expense_date:
        return oil_row.expense_date
    return None


def _oil_alert_tehsil_for_vehicle(vehicle):
    """Tehsil from vehicle parking station (Oil Change Alert Report)."""
    ps = getattr(vehicle, 'parking_station', None) if vehicle else None
    if ps and (getattr(ps, 'tehsil', None) or '').strip():
        return (ps.tehsil or '').strip()
    return '-'


def _oil_change_alert_rows(project_id=0, district_id=0, vehicle_family='',
                           from_date=None, to_date=None,
                           statuses=None,
                           custom_km=None,
                           custom_km_mode='',
                           workspace_employee_id=None,
                           allowed_projects=None, allowed_districts=None, allowed_vehicles=None,
                           is_master_or_admin=True):
    limits = _get_vehicle_family_oil_change_limits()

    query = db.session.query(Vehicle, Project, District).options(
        joinedload(Vehicle.parking_station),
    ).outerjoin(
        Project, Vehicle.project_id == Project.id
    ).outerjoin(
        District, Vehicle.district_id == District.id
    )
    if not is_master_or_admin:
        if allowed_projects:
            query = query.filter(Vehicle.project_id.in_(list(allowed_projects)))
        if allowed_districts:
            query = query.filter(Vehicle.district_id.in_(list(allowed_districts)))
        if allowed_vehicles:
            query = query.filter(Vehicle.id.in_(list(allowed_vehicles)))
    if project_id:
        query = query.filter(Vehicle.project_id == project_id)
    if district_id:
        query = query.filter(Vehicle.district_id == district_id)
    if vehicle_family:
        query = query.filter(Vehicle.vehicle_family == vehicle_family)

    statuses_set = set(statuses or [])
    rows = []
    for vehicle, project, district in query.order_by(*vehicle_order_by(Project.name, District.name)).all():
        family_name = (vehicle.vehicle_family or '').strip()
        family_cfg = limits.get(family_name) or {}
        km_limit = family_cfg.get('limit_km')
        near_percent = family_cfg.get('near_percent')
        if not family_name or not km_limit or not near_percent:
            continue

        base_reading, base_date, base_source = _vehicle_last_oil_change_base(vehicle.id, workspace_employee_id)
        if base_reading is None:
            continue
        last_oil_change_date = _vehicle_last_oil_change_date(vehicle.id, workspace_employee_id)
        if from_date and (not last_oil_change_date or last_oil_change_date < from_date):
            continue
        if to_date and (not last_oil_change_date or last_oil_change_date > to_date):
            continue

        current_reading, current_date, current_source = _vehicle_latest_meter_reading(vehicle.id, workspace_employee_id)
        if current_reading is None:
            current_reading = base_reading
            current_date = base_date
            current_source = base_source

        kms_after_oil = round(max(0.0, float(current_reading) - float(base_reading)), 2)
        limit_f = float(km_limit)
        near_ratio = float(near_percent) / 100.0
        near_at = round(limit_f * near_ratio, 2)
        remaining = round(limit_f - kms_after_oil, 2)

        if kms_after_oil >= limit_f:
            status_code = 'crossed'
        elif kms_after_oil >= near_at:
            status_code = 'near'
        else:
            status_code = 'safe'

        # Multi-status filter: if any statuses selected, row must match one.
        if statuses_set and status_code not in statuses_set:
            continue

        custom_state = None
        custom_diff = None
        if custom_km is not None:
            custom_diff = round(kms_after_oil - float(custom_km), 2)
            if custom_diff > 0:
                custom_state = 'ahead'
            elif custom_diff < 0:
                custom_state = 'behind'
            else:
                custom_state = 'equal'
            if custom_km_mode == 'above' and custom_state != 'ahead':
                continue
            if custom_km_mode == 'below' and custom_state != 'behind':
                continue

        rows.append({
            'vehicle': vehicle,
            'project': project,
            'district': district,
            'tehsil': _oil_alert_tehsil_for_vehicle(vehicle),
            'vehicle_family': family_name,
            'limit_km': int(limit_f),
            'near_percent': int(near_percent),
            'base_reading': round(float(base_reading), 2),
            'base_date': base_date,
            'last_oil_change_date': last_oil_change_date,
            'base_source': base_source,
            'current_reading': round(float(current_reading), 2),
            'current_date': current_date,
            'current_source': current_source,
            'kms_after_oil': kms_after_oil,
            'remaining_km': remaining,
            'status': status_code,
            'custom_state': custom_state,
            'custom_diff': custom_diff,
        })

    rows.sort(key=lambda x: (0 if x['status'] == 'crossed' else 1 if x['status'] == 'near' else 2, x['remaining_km']))
    return rows





# ── Driver Seat Available Report ───────────────────────────────────────────
def _parse_activity_datetime(raw):
    if not raw:
        return None
    s = str(raw).strip()
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%d-%m-%Y %H:%M:%S', '%d-%m-%Y %H:%M'):
        try:
            return datetime.strptime(s, fmt)
        except Exception:
            continue
    return None


def _speed_monitoring_rows(from_date=None, to_date=None, project_id=0, district_id=0, vehicle_id=0,
                           check_type='', speed_limit=None,
                           allowed_projects=None, allowed_districts=None, allowed_vehicles=None,
                           is_master_or_admin=True):
    query = db.session.query(
        VehicleActivityRecord, Vehicle, Project, District
    ).outerjoin(
        Vehicle, Vehicle.vehicle_no == VehicleActivityRecord.vehicle_no
    ).outerjoin(
        Project, Vehicle.project_id == Project.id
    ).outerjoin(
        District, Vehicle.district_id == District.id
    )

    if from_date:
        query = query.filter(VehicleActivityRecord.task_date >= from_date)
    if to_date:
        query = query.filter(VehicleActivityRecord.task_date <= to_date)

    if not is_master_or_admin:
        if allowed_projects:
            query = query.filter(Vehicle.project_id.in_(list(allowed_projects)))
        if allowed_districts:
            query = query.filter(Vehicle.district_id.in_(list(allowed_districts)))
        if allowed_vehicles:
            query = query.filter(Vehicle.id.in_(list(allowed_vehicles)))

    if project_id:
        query = query.filter(Vehicle.project_id == project_id)
    if district_id:
        query = query.filter(Vehicle.district_id == district_id)
    if vehicle_id:
        query = query.filter(Vehicle.id == vehicle_id)

    out = []
    for rec, vehicle, project, district in query.order_by(VehicleActivityRecord.task_date.desc(), VehicleActivityRecord.id.desc()).all():
        speed_val = float(rec.speed or 0)
        if speed_limit is not None:
            if check_type == 'above' and not (speed_val > speed_limit):
                continue
            if check_type == 'below' and not (speed_val < speed_limit):
                continue

        dt = _parse_activity_datetime(rec.record_date_time)
        if from_date and dt and dt.date() < from_date:
            continue
        if to_date and dt and dt.date() > to_date:
            continue

        check_result = None
        if speed_limit is not None:
            if speed_val > speed_limit:
                check_result = f'Above (+{speed_val - speed_limit:.2f})'
            elif speed_val < speed_limit:
                check_result = f'Below (-{speed_limit - speed_val:.2f})'
            else:
                check_result = 'Equal (0.00)'

        location_text = (rec.location or '').strip()
        if '||' in location_text:
            parts = location_text.split('||', 1)
            if len(parts) == 2 and parts[1].strip():
                location_text = parts[1].strip()

        out.append({
            'rec': rec,
            'vehicle': vehicle,
            'project': project,
            'district': district,
            'record_dt': dt,
            'speed': speed_val,
            'location_text': location_text,
            'check_result': check_result,
        })

    out.sort(key=lambda r: r['record_dt'] or datetime.min, reverse=True)
    return out




def _speed_monitoring_report_preview_context():
    from auth_utils import get_user_context
    user_id = session.get('user_id')
    user_context = get_user_context(user_id) if user_id else {}
    allowed_projects = user_context.get('allowed_projects', set())
    allowed_districts = user_context.get('allowed_districts', set())
    allowed_vehicles = user_context.get('allowed_vehicles', set())
    is_master_or_admin = user_context.get('is_master_or_admin', False)

    from_date = parse_date(request.args.get('from_date')) or pk_date()
    to_date = parse_date(request.args.get('to_date')) or pk_date()
    if from_date and to_date and from_date > to_date:
        from_date, to_date = to_date, from_date

    project_id = request.args.get('project_id', type=int) or 0
    district_id = request.args.get('district_id', type=int) or 0
    vehicle_id = request.args.get('vehicle_id', type=int) or 0
    check_type = (request.args.get('check_type') or '').strip().lower()
    if check_type not in ('', 'above', 'below'):
        check_type = ''

    speed_limit_raw = (request.args.get('speed_limit') or '').strip()
    speed_limit = None
    if speed_limit_raw:
        try:
            speed_limit = float(speed_limit_raw)
            if speed_limit < 0:
                speed_limit = None
        except Exception:
            speed_limit = None

    rows = _speed_monitoring_rows(
        from_date=from_date,
        to_date=to_date,
        project_id=project_id,
        district_id=district_id,
        vehicle_id=vehicle_id,
        check_type=check_type,
        speed_limit=speed_limit,
        allowed_projects=allowed_projects,
        allowed_districts=allowed_districts,
        allowed_vehicles=allowed_vehicles,
        is_master_or_admin=is_master_or_admin,
    )
    table_search = (request.args.get('table_search') or '').strip().lower()
    if table_search:
        def _matches_search(r):
            blob = ' '.join([
                r['district'].name if r.get('district') else '',
                r['project'].name if r.get('project') else '',
                r['vehicle'].vehicle_no if r.get('vehicle') else (r['rec'].vehicle_no or ''),
                r['record_dt'].strftime('%d-%m-%Y %I:%M %p') if r.get('record_dt') else (r['rec'].record_date_time or ''),
                f"{r.get('speed', 0):.2f}",
                r['rec'].reason or '',
                r.get('location_text') or '',
                r.get('check_result') or '',
            ]).lower()
            return table_search in blob
        rows = [r for r in rows if _matches_search(r)]

    return {
        'rows': rows,
        'total': len(rows),
        'from_date': from_date,
        'to_date': to_date,
        'speed_limit': speed_limit_raw,
        'check_type': check_type,
        'now': datetime.now,
    }




# ── Mileage Report ───────────────────────────────────────────
def _resolve_task_readings(task_row, vehicle_id, prev_close_cache=None):
    close_r = float(task_row.close_reading or 0)
    start_r = float(task_row.start_reading) if task_row.start_reading is not None else None

    # If start reading is missing/zero, fallback to previous day's close for the same vehicle.
    if start_r is None or start_r == 0:
        cache_key = (vehicle_id, task_row.task_date) if vehicle_id and task_row.task_date else None
        prev_close = None
        if prev_close_cache is not None and cache_key in prev_close_cache:
            prev_close = prev_close_cache.get(cache_key)
        else:
            prev = VehicleDailyTask.query.filter(
                VehicleDailyTask.vehicle_id == vehicle_id,
                VehicleDailyTask.task_date < task_row.task_date,
                VehicleDailyTask.close_reading.isnot(None),
            ).order_by(
                VehicleDailyTask.task_date.desc(),
                VehicleDailyTask.id.desc()
            ).first()
            prev_close = float(prev.close_reading) if prev and prev.close_reading is not None else None
            if prev_close_cache is not None and cache_key is not None:
                prev_close_cache[cache_key] = prev_close
        if prev_close is not None:
            start_r = prev_close

    if start_r is None:
        start_r = 0.0
    return round(start_r, 2), round(close_r, 2)


def _mileage_report_rows(from_date=None, to_date=None, project_id=0, district_id=0, vehicle_id=0,
                         check_type='', km_limit=None,
                         allowed_projects=None, allowed_districts=None, allowed_vehicles=None,
                         is_master_or_admin=True):
    query = db.session.query(
        VehicleDailyTask, Vehicle, Project, District
    ).join(
        Vehicle, Vehicle.id == VehicleDailyTask.vehicle_id
    ).outerjoin(
        Project, Vehicle.project_id == Project.id
    ).outerjoin(
        District, Vehicle.district_id == District.id
    )

    if from_date:
        query = query.filter(VehicleDailyTask.task_date >= from_date)
    if to_date:
        query = query.filter(VehicleDailyTask.task_date <= to_date)

    if not is_master_or_admin:
        if allowed_projects:
            query = query.filter(Vehicle.project_id.in_(list(allowed_projects)))
        if allowed_districts:
            query = query.filter(Vehicle.district_id.in_(list(allowed_districts)))
        if allowed_vehicles:
            query = query.filter(Vehicle.id.in_(list(allowed_vehicles)))

    if project_id:
        query = query.filter(Vehicle.project_id == project_id)
    if district_id:
        query = query.filter(Vehicle.district_id == district_id)
    if vehicle_id:
        query = query.filter(Vehicle.id == vehicle_id)

    out = []
    prev_close_cache = {}
    for rec, vehicle, project, district in query.order_by(VehicleDailyTask.task_date.desc(), VehicleDailyTask.id.desc()).all():
        start_r, close_r = _resolve_task_readings(rec, vehicle.id if vehicle else None, prev_close_cache)
        total_km = round(close_r - start_r, 2)

        if km_limit is not None:
            if check_type == 'above' and not (total_km > km_limit):
                continue
            if check_type == 'below' and not (total_km < km_limit):
                continue

        check_result = None
        if km_limit is not None:
            if total_km > km_limit:
                check_result = f'Above (+{total_km - km_limit:.2f})'
            elif total_km < km_limit:
                check_result = f'Below (-{km_limit - total_km:.2f})'
            else:
                check_result = 'Equal (0.00)'

        out.append({
            'rec': rec,
            'vehicle': vehicle,
            'project': project,
            'district': district,
            'start_reading': start_r,
            'close_reading': close_r,
            'total_km': total_km,
            'tasks_count': int(rec.tasks_count or 0),
            'check_result': check_result,
        })
    return out




def _mileage_report_preview_context():
    from auth_utils import get_user_context
    user_id = session.get('user_id')
    user_context = get_user_context(user_id) if user_id else {}
    allowed_projects = user_context.get('allowed_projects', set())
    allowed_districts = user_context.get('allowed_districts', set())
    allowed_vehicles = user_context.get('allowed_vehicles', set())
    is_master_or_admin = user_context.get('is_master_or_admin', False)

    from_date = parse_date(request.args.get('from_date')) or pk_date()
    to_date = parse_date(request.args.get('to_date')) or pk_date()
    if from_date and to_date and from_date > to_date:
        from_date, to_date = to_date, from_date
    project_id = request.args.get('project_id', type=int) or 0
    district_id = request.args.get('district_id', type=int) or 0
    vehicle_id = request.args.get('vehicle_id', type=int) or 0
    check_type = (request.args.get('check_type') or '').strip().lower()
    if check_type not in ('', 'above', 'below'):
        check_type = ''
    km_limit_raw = (request.args.get('km_limit') or '').strip()
    km_limit = None
    if km_limit_raw:
        try:
            km_limit = float(km_limit_raw)
            if km_limit < 0:
                km_limit = None
        except Exception:
            km_limit = None

    rows = _mileage_report_rows(
        from_date=from_date, to_date=to_date, project_id=project_id, district_id=district_id,
        vehicle_id=vehicle_id, check_type=check_type, km_limit=km_limit,
        allowed_projects=allowed_projects, allowed_districts=allowed_districts, allowed_vehicles=allowed_vehicles,
        is_master_or_admin=is_master_or_admin,
    )
    table_search = (request.args.get('table_search') or '').strip().lower()
    if table_search:
        def _m(r):
            v = r.get('vehicle')
            tehsil_txt = ''
            vtype_txt = ''
            if v:
                ps = getattr(v, 'parking_station', None)
                if ps:
                    tehsil_txt = ps.tehsil or ''
                vtype_txt = v.vehicle_type or ''
            blob = ' '.join([
                r['rec'].task_date.strftime('%d-%m-%Y') if r['rec'].task_date else '',
                r['district'].name if r.get('district') else '',
                tehsil_txt,
                r['project'].name if r.get('project') else '',
                vtype_txt,
                v.vehicle_no if v else '',
                format_reading(r['start_reading']),
                format_reading(r['close_reading']),
                f"{r['total_km']:.2f}",
                str(r['tasks_count']),
                r.get('check_result') or '',
            ]).lower()
            return table_search in blob
        rows = [r for r in rows if _m(r)]

    return {
        'rows': rows,
        'total': len(rows),
        'from_date': from_date,
        'to_date': to_date,
        'km_limit': km_limit_raw,
        'check_type': check_type,
        'now': datetime.now,
    }




# ── Unauthorized Movement Report ──────────────────────────────────────────
def _unauthorized_movement_rows(from_date=None, to_date=None, project_id=0, district_id=0, vehicle_id=0,
                                check_type='', without_task_move_limit=None,
                                allowed_projects=None, allowed_districts=None, allowed_vehicles=None,
                                is_master_or_admin=True):
    def _haversine_meters(lat1, lon1, lat2, lon2):
        try:
            lat1 = float(lat1)
            lon1 = float(lon1)
            lat2 = float(lat2)
            lon2 = float(lon2)
        except (TypeError, ValueError):
            return None
        r = 6371000.0
        p1 = math.radians(lat1)
        p2 = math.radians(lat2)
        dphi = math.radians(lat2 - lat1)
        dlambda = math.radians(lon2 - lon1)
        a = math.sin(dphi / 2.0) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dlambda / 2.0) ** 2
        return 2.0 * r * math.asin(min(1.0, math.sqrt(a)))

    allowed_projects = set(allowed_projects or [])
    allowed_districts = set(allowed_districts or [])
    allowed_vehicles = set(allowed_vehicles or [])

    vehicle_q = db.session.query(Vehicle, Project, District).outerjoin(
        Project, Vehicle.project_id == Project.id
    ).outerjoin(
        District, Vehicle.district_id == District.id
    )

    if not is_master_or_admin:
        if allowed_projects:
            vehicle_q = vehicle_q.filter(Vehicle.project_id.in_(list(allowed_projects)))
        if allowed_districts:
            vehicle_q = vehicle_q.filter(Vehicle.district_id.in_(list(allowed_districts)))
        if allowed_vehicles:
            vehicle_q = vehicle_q.filter(Vehicle.id.in_(list(allowed_vehicles)))

    if project_id:
        vehicle_q = vehicle_q.filter(Vehicle.project_id == project_id)
    if district_id:
        vehicle_q = vehicle_q.filter(Vehicle.district_id == district_id)
    if vehicle_id:
        vehicle_q = vehicle_q.filter(Vehicle.id == vehicle_id)

    vehicles = vehicle_q.order_by(*vehicle_order_by()).all()
    if not vehicles:
        return []

    by_vehicle_id = {}
    by_vehicle_no = {}
    for v, p, d in vehicles:
        key_no = (v.vehicle_no or '').strip().upper()
        by_vehicle_id[v.id] = {
            'vehicle': v,
            'project': p,
            'district': d,
            'km_driven': 0.0,
            'tracker_km': 0.0,
            'task_running_km': 0.0,
            'task_count': 0,
            'vehicle_no_key': key_no,
            'last_task_running_km': 0.0,
            'return_to_parking_km': 0.0,
        }
        if key_no:
            by_vehicle_no[key_no] = by_vehicle_id[v.id]

    vehicle_ids = list(by_vehicle_id.keys())
    vehicle_nos = [k for k in by_vehicle_no.keys() if k]

    if from_date and to_date and vehicle_ids:
        task_rows = VehicleDailyTask.query.filter(
            VehicleDailyTask.task_date >= from_date,
            VehicleDailyTask.task_date <= to_date,
            VehicleDailyTask.vehicle_id.in_(vehicle_ids),
        ).order_by(VehicleDailyTask.task_date.desc(), VehicleDailyTask.id.desc()).all()

        prev_close_cache = {}
        for rec in task_rows:
            agg = by_vehicle_id.get(rec.vehicle_id)
            if not agg:
                continue
            start_r, close_r = _resolve_task_readings(rec, rec.vehicle_id, prev_close_cache)
            agg['km_driven'] += float(close_r - start_r)

        tracker_rows = VehicleMileageRecord.query.filter(
            VehicleMileageRecord.task_date >= from_date,
            VehicleMileageRecord.task_date <= to_date,
            VehicleMileageRecord.reg_no.in_(vehicle_nos),
        ).all()
        for rec in tracker_rows:
            key_no = (rec.reg_no or '').strip().upper()
            agg = by_vehicle_no.get(key_no)
            if not agg:
                continue
            agg['tracker_km'] += float(rec.effective_km() or 0)

        emg_rows = EmergencyTaskRecord.query.filter(
            EmergencyTaskRecord.task_date >= from_date,
            EmergencyTaskRecord.task_date <= to_date,
            EmergencyTaskRecord.amb_reg_no.in_(vehicle_nos),
            EmergencyTaskRecord.category.in_(['Green', 'Yellow']),
            EmergencyTaskRecord.completed_date_time.isnot(None),
            EmergencyTaskRecord.excel_created_date.isnot(None),
        ).all()
        task_windows = {}
        task_meta = {}
        for emg in emg_rows:
            key_no = (emg.amb_reg_no or '').strip().upper()
            if key_no not in by_vehicle_no:
                continue
            assign_dt = _parse_emg_datetime(emg.excel_created_date)
            close_dt = _parse_emg_datetime(emg.completed_date_time)
            if not assign_dt or not close_dt or close_dt < assign_dt:
                continue
            idx = len(task_windows.setdefault(key_no, []))
            task_windows[key_no].append((assign_dt, close_dt))
            task_meta.setdefault(key_no, []).append({
                'idx': idx,
                'assign_dt': assign_dt,
                'close_dt': close_dt,
                'running_km': 0.0,
            })
            by_vehicle_no[key_no]['task_count'] += 1

        if task_windows:
            activity_rows = VehicleActivityRecord.query.filter(
                VehicleActivityRecord.task_date >= from_date,
                VehicleActivityRecord.task_date <= to_date,
                VehicleActivityRecord.vehicle_no.in_(list(task_windows.keys())),
            ).all()

            activity_by_vehicle = {}
            for act in activity_rows:
                key_no = (act.vehicle_no or '').strip().upper()
                windows = task_windows.get(key_no) or []
                if not windows:
                    continue
                act_dt = _parse_activity_datetime(act.record_date_time)
                if not act_dt:
                    continue
                dist_val = float(act.distance or 0)
                activity_by_vehicle.setdefault(key_no, []).append((act_dt, dist_val, act))
                for assign_dt, close_dt in windows:
                    if assign_dt <= act_dt <= close_dt:
                        by_vehicle_no[key_no]['task_running_km'] += dist_val
                        for tm in (task_meta.get(key_no) or []):
                            if tm['assign_dt'] == assign_dt and tm['close_dt'] == close_dt:
                                tm['running_km'] += dist_val
                                break
                        break

            # Return-to-parking calculation (same rules as timeline detail):
            # for EACH task window:
            # 1) start after task close
            # 2) stop at min(next task start, close+120m)
            # 3) stop on geofence entry <=500m
            # 4) cap by that task's running KM
            for key_no, activities in activity_by_vehicle.items():
                agg = by_vehicle_no.get(key_no)
                if not agg:
                    continue
                v = agg.get('vehicle')
                parking = getattr(v, 'parking_station', None) if v else None
                base_lat = getattr(parking, 'latitude', None) if parking else None
                base_lon = getattr(parking, 'longitude', None) if parking else None
                if base_lat is None or base_lon is None:
                    continue

                metas = sorted((task_meta.get(key_no) or []), key=lambda x: x['assign_dt'])
                if not metas:
                    continue

                activities_sorted = sorted(activities, key=lambda x: x[0])
                total_return_km = 0.0
                for i, tm in enumerate(metas):
                    close_dt = tm['close_dt']
                    next_assign_dt = metas[i + 1]['assign_dt'] if i + 1 < len(metas) else None
                    cutoff_dt = close_dt + timedelta(minutes=120)
                    window_end = cutoff_dt if not next_assign_dt else min(cutoff_dt, next_assign_dt)
                    cap_km = max(float(tm.get('running_km') or 0.0), 0.0)
                    if cap_km <= 0 or window_end <= close_dt:
                        continue

                    ret_km = 0.0
                    for act_dt, dist_val, act in activities_sorted:
                        if act_dt <= close_dt:
                            continue
                        if act_dt >= window_end:
                            break
                        # Keep return window clean: points inside any task belong to task running.
                        in_task = any(tt['assign_dt'] <= act_dt <= tt['close_dt'] for tt in metas)
                        if in_task:
                            continue

                        d_m = _haversine_meters(
                            getattr(act, 'latitude', None),
                            getattr(act, 'longitude', None),
                            base_lat,
                            base_lon,
                        )
                        if d_m is not None and d_m <= 500.0:
                            break
                        if dist_val > 0:
                            ret_km += dist_val
                        if ret_km >= cap_km:
                            ret_km = cap_km
                            break
                    total_return_km += max(0.0, ret_km)

                agg['return_to_parking_km'] = max(0.0, total_return_km)

    rows = []
    for agg in by_vehicle_id.values():
        km_driven = round(float(agg['km_driven']), 2)
        tracker_km = round(float(agg['tracker_km']), 2)
        task_running_km = round(float(agg['task_running_km']), 2)
        return_to_parking_km = round(float(agg.get('return_to_parking_km') or 0.0), 2)
        without_task_move = round(tracker_km - (task_running_km + return_to_parking_km), 2)

        if without_task_move_limit is not None:
            if check_type == 'above' and not (without_task_move > without_task_move_limit):
                continue
            if check_type == 'below' and not (without_task_move < without_task_move_limit):
                continue

        # Practical view: only keep vehicles with any movement footprint.
        if km_driven == 0 and tracker_km == 0 and task_running_km == 0 and without_task_move == 0:
            continue

        rows.append({
            'vehicle': agg['vehicle'],
            'project': agg['project'],
            'district': agg['district'],
            'km_driven': km_driven,
            'tracker_km': tracker_km,
            'task_running_km': task_running_km,
            'task_count': int(agg['task_count'] or 0),
            'return_to_parking_km': return_to_parking_km,
            'without_task_move': without_task_move,
        })

    rows.sort(key=lambda r: (
        (r['district'].name if r.get('district') else '').lower(),
        (r['project'].name if r.get('project') else '').lower(),
        (r['vehicle'].vehicle_no if r.get('vehicle') else '').lower(),
    ))
    return rows


def _filter_unauthorized_movement_rows(rows, table_search):
    s = (table_search or '').strip().lower()
    if not s:
        return rows
    out = []
    for row in rows:
        blob = ' '.join([
            row['district'].name if row.get('district') else '',
            row['project'].name if row.get('project') else '',
            row['vehicle'].vehicle_no if row.get('vehicle') else '',
            f"{row['km_driven']:.2f}",
            f"{row['tracker_km']:.2f}",
            f"{row['task_running_km']:.2f}",
            str(row.get('task_count') or 0),
            f"{row['return_to_parking_km']:.2f}",
            f"{row['without_task_move']:.2f}",
        ]).lower()
        if s in blob:
            out.append(row)
    return out


def _build_unauthorized_movement_timeline(vehicle, from_date, to_date):
    if not vehicle:
        return {'segments': [], 'totals': {}, 'tasks': []}

    vehicle_no = (vehicle.vehicle_no or '').strip().upper()
    if not vehicle_no:
        return {'segments': [], 'totals': {}, 'tasks': []}

    def _haversine_meters(lat1, lon1, lat2, lon2):
        try:
            lat1 = float(lat1)
            lon1 = float(lon1)
            lat2 = float(lat2)
            lon2 = float(lon2)
        except (TypeError, ValueError):
            return None
        r = 6371000.0
        p1 = math.radians(lat1)
        p2 = math.radians(lat2)
        dphi = math.radians(lat2 - lat1)
        dlambda = math.radians(lon2 - lon1)
        a = math.sin(dphi / 2.0) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dlambda / 2.0) ** 2
        return 2.0 * r * math.asin(min(1.0, math.sqrt(a)))

    emg_rows = EmergencyTaskRecord.query.filter(
        EmergencyTaskRecord.task_date >= from_date,
        EmergencyTaskRecord.task_date <= to_date,
        EmergencyTaskRecord.amb_reg_no == vehicle_no,
        EmergencyTaskRecord.category.in_(['Green', 'Yellow']),
        EmergencyTaskRecord.completed_date_time.isnot(None),
        EmergencyTaskRecord.excel_created_date.isnot(None),
    ).all()

    tasks = []
    for emg in emg_rows:
        assign_dt = _parse_emg_datetime(emg.excel_created_date)
        close_dt = _parse_emg_datetime(emg.completed_date_time)
        if not assign_dt or not close_dt or close_dt < assign_dt:
            continue
        tasks.append({
            'task_id': emg.task_id_ext or '',
            'assign_dt': assign_dt,
            'close_dt': close_dt,
            'running_km': 0.0,
        })
    tasks.sort(key=lambda x: x['assign_dt'])

    activity_rows = VehicleActivityRecord.query.filter(
        VehicleActivityRecord.task_date >= from_date,
        VehicleActivityRecord.task_date <= to_date,
        VehicleActivityRecord.vehicle_no == vehicle_no,
    ).all()
    acts = []
    for act in activity_rows:
        adt = _parse_activity_datetime(act.record_date_time)
        if not adt:
            continue
        try:
            dist_val = float(act.distance or 0)
        except (TypeError, ValueError):
            dist_val = 0.0
        acts.append((adt, dist_val, act))
    acts.sort(key=lambda x: x[0])

    # Running KM per task (for KM cap on return window)
    for adt, dist_val, _act in acts:
        for t in tasks:
            if t['assign_dt'] <= adt <= t['close_dt']:
                t['running_km'] += dist_val
                break

    parking = getattr(vehicle, 'parking_station', None)
    base_lat = getattr(parking, 'latitude', None) if parking else None
    base_lon = getattr(parking, 'longitude', None) if parking else None

    # Build return windows for each task, with stop reason.
    return_windows = []
    for i, t in enumerate(tasks):
        close_dt = t['close_dt']
        next_assign_dt = tasks[i + 1]['assign_dt'] if i + 1 < len(tasks) else None
        cutoff_dt = close_dt + timedelta(minutes=120)
        window_end = cutoff_dt if not next_assign_dt else min(cutoff_dt, next_assign_dt)
        cap_km = max(float(t.get('running_km') or 0.0), 0.0)
        if cap_km <= 0 or window_end <= close_dt:
            continue

        ret_km = 0.0
        stop_dt = window_end
        stop_reason = 'cutoff_120m' if window_end == cutoff_dt else 'next_task_started'
        for adt, dist_val, act in acts:
            if adt <= close_dt:
                continue
            if adt >= window_end:
                break
            # Keep return window clean: movement inside any task belongs to task running.
            in_task = any(tt['assign_dt'] <= adt <= tt['close_dt'] for tt in tasks)
            if in_task:
                continue
            d_m = _haversine_meters(getattr(act, 'latitude', None), getattr(act, 'longitude', None), base_lat, base_lon) \
                if (base_lat is not None and base_lon is not None) else None
            if d_m is not None and d_m <= 500.0:
                stop_dt = adt
                stop_reason = 'geofence_500m'
                break
            if dist_val > 0:
                ret_km += dist_val
            if ret_km >= cap_km:
                ret_km = cap_km
                stop_dt = adt
                stop_reason = 'km_cap_reached'
                break
        if stop_dt > close_dt:
            return_windows.append({
                'start': close_dt,
                'end': stop_dt,
                'task_id': t.get('task_id') or '',
                'km': round(ret_km, 2),
                'reason': stop_reason,
            })

    # Classify each activity point.
    timeline_points = []
    totals = {'task_running_km': 0.0, 'return_to_parking_km': 0.0, 'without_task_km': 0.0}
    for adt, dist_val, _act in acts:
        label = 'without_task'
        task_id = ''
        for t in tasks:
            if t['assign_dt'] <= adt <= t['close_dt']:
                label = 'task_running'
                task_id = t.get('task_id') or ''
                break
        if label != 'task_running':
            for w in return_windows:
                if w['start'] < adt < w['end']:
                    label = 'return_to_parking'
                    task_id = w.get('task_id') or ''
                    break

        if label == 'task_running':
            totals['task_running_km'] += dist_val
        elif label == 'return_to_parking':
            totals['return_to_parking_km'] += dist_val
        else:
            totals['without_task_km'] += dist_val

        timeline_points.append({'dt': adt, 'km': dist_val, 'label': label, 'task_id': task_id})

    # Merge into readable time segments.
    segments = []
    current = None
    for p in timeline_points:
        if current and current['label'] == p['label'] and current.get('task_id', '') == p.get('task_id', ''):
            current['end'] = p['dt']
            current['km'] += p['km']
        else:
            if current:
                segments.append(current)
            current = {
                'label': p['label'],
                'task_id': p.get('task_id') or '',
                'start': p['dt'],
                'end': p['dt'],
                'km': p['km'],
            }
    if current:
        segments.append(current)

    fmt = '%d-%m-%Y %I:%M:%S %p'
    return {
        'segments': [{
            'label': s['label'],
            'task_id': s['task_id'],
            'start': s['start'].strftime(fmt) if s.get('start') else '-',
            'end': s['end'].strftime(fmt) if s.get('end') else '-',
            'km': round(float(s.get('km') or 0.0), 2),
        } for s in segments],
        'totals': {
            'task_running_km': round(totals['task_running_km'], 2),
            'return_to_parking_km': round(totals['return_to_parking_km'], 2),
            'without_task_km': round(totals['without_task_km'], 2),
        },
        'tasks': [{
            'task_id': t.get('task_id') or '-',
            'assign': t['assign_dt'].strftime(fmt),
            'close': t['close_dt'].strftime(fmt),
            'running_km': round(float(t.get('running_km') or 0.0), 2),
        } for t in tasks],
        'return_windows': [{
            'task_id': w.get('task_id') or '-',
            'start': w['start'].strftime(fmt),
            'end': w['end'].strftime(fmt),
            'km': round(float(w.get('km') or 0.0), 2),
            'reason': w.get('reason') or '',
        } for w in return_windows],
    }


def _umr_timeline_reason_label(reason):
    return {
        'next_task_started': 'Next task started',
        'geofence_500m': 'Geofence <= 500m',
        'km_cap_reached': 'KM cap reached',
        'cutoff_120m': '120 min cutoff',
    }.get(reason or '', reason or '-')


def _umr_segment_type_label(label):
    return {
        'task_running': 'Task Running',
        'return_to_parking': 'Return to Parking',
        'without_task': 'Without Task',
    }.get(label or '', label or '-')


def _split_umr_datetime(ts):
    if not ts or ts == '-':
        return '-', '-'
    m = re.match(r'^(\d{2}-\d{2}-\d{4})\s+(.+)$', str(ts).strip())
    if m:
        return m.group(1), m.group(2)
    return str(ts), '-'


def _build_umr_timeline_excel(payload, *, vehicle_no, project_name, district_name, from_date, to_date):
    output = io.BytesIO()
    wb = xlsxwriter.Workbook(output, {'in_memory': True})
    safe_sheet = re.sub(r'[^\w\- ]', '_', (vehicle_no or 'Timeline'))[:28] or 'Timeline'
    ws = wb.add_worksheet(safe_sheet)
    ws.hide_gridlines(2)
    ws.set_landscape()
    ws.set_margins(left=0.5, right=0.5, top=0.5, bottom=0.5)
    ws.set_default_row(18)

    font = 'Segoe UI'
    cols = 6
    exported_at = pk_now().strftime('%d/%m/%Y, %I:%M:%S %p')
    period_txt = f'{from_date.strftime("%d-%m-%Y")} — {to_date.strftime("%d-%m-%Y")}'
    totals = payload.get('totals') or {}
    tasks = payload.get('tasks') or []
    return_windows = payload.get('return_windows') or []
    segments = payload.get('segments') or []
    task_km = float(totals.get('task_running_km') or 0.0)
    return_km = float(totals.get('return_to_parking_km') or 0.0)
    without_km = float(totals.get('without_task_km') or 0.0)
    total_km = round(task_km + return_km + without_km, 2)
    seg_total = round(sum(float(s.get('km') or 0.0) for s in segments), 2)

    def pct(part):
        return round(float(part or 0) * 100.0 / total_km, 1) if total_km > 0 else 0.0

    def fmt(**kw):
        base = {'font_name': font, 'valign': 'vcenter'}
        base.update(kw)
        return wb.add_format(base)

    header_fmt = fmt(bold=True, font_size=15, font_color='#FFFFFF', bg_color='#D97706', align='left', indent=1)
    header_period_fmt = fmt(font_size=10, font_color='#FFFFFF', bg_color='#D97706', align='right')
    header_sub_fmt = fmt(font_size=10, font_color='#FFEDD5', bg_color='#D97706', align='left', indent=1)

    bd_title_fmt = fmt(bold=True, font_size=10, font_color='#64748B', bg_color='#FFFFFF', align='left', indent=1)
    bd_total_fmt = fmt(bold=True, font_size=10, font_color='#0F172A', bg_color='#F1F5F9', align='right',
                       top=1, bottom=1, left=1, right=1, top_color='#E2E8F0', bottom_color='#E2E8F0',
                       left_color='#E2E8F0', right_color='#E2E8F0')

    def stat_lbl_f(accent):
        return fmt(font_size=9, font_color='#64748B', bg_color='#FAFAFA', align='left', bold=True, indent=1,
                   left=3, left_color=accent, right=1, right_color='#E2E8F0', top=1, top_color='#E2E8F0')

    def stat_val_f(accent, warn=False):
        return fmt(bold=True, font_size=15, font_color='#DC2626' if warn else '#0F172A', bg_color='#FAFAFA',
                   align='left', indent=1, left=3, left_color=accent, right=1, right_color='#E2E8F0')

    def stat_pct_f(accent):
        return fmt(font_size=9, font_color='#64748B', bg_color='#FAFAFA', align='left', indent=1,
                   left=3, left_color=accent, right=1, right_color='#E2E8F0', bottom=1, bottom_color='#E2E8F0')

    sec_fmt = fmt(bold=True, font_size=11, font_color='#0F172A', bg_color='#FFFFFF', align='left', indent=1,
                  top=1, bottom=1, top_color='#E2E8F0', bottom_color='#E2E8F0')
    th_fmt = fmt(bold=True, font_size=9, font_color='#475569', bg_color='#F8FAFC', align='left',
                 bottom=1, bottom_color='#DBE3EE')
    th_c_fmt = fmt(bold=True, font_size=9, font_color='#475569', bg_color='#F8FAFC', align='center',
                   bottom=1, bottom_color='#DBE3EE')
    th_r_fmt = fmt(bold=True, font_size=9, font_color='#475569', bg_color='#F8FAFC', align='right',
                   bottom=1, bottom_color='#DBE3EE')
    td_fmt = fmt(font_size=10, font_color='#0F172A', align='left', bottom=1, bottom_color='#DBE3EE', indent=1)
    td_c_fmt = fmt(font_size=10, font_color='#0F172A', align='center', bottom=1, bottom_color='#DBE3EE')
    td_km_fmt = fmt(bold=True, font_size=10, font_color='#0F172A', align='right', bottom=1, bottom_color='#DBE3EE', num_format='0.00')
    td_warn_fmt = fmt(font_size=10, font_color='#0F172A', align='left', bottom=1, bottom_color='#DBE3EE',
                      bg_color='#FFFBEB', indent=1)
    td_warn_c_fmt = fmt(font_size=10, font_color='#0F172A', align='center', bottom=1, bottom_color='#DBE3EE', bg_color='#FFFBEB')
    td_warn_km_fmt = fmt(bold=True, font_size=10, font_color='#92400E', align='right', bottom=1, bottom_color='#DBE3EE',
                         bg_color='#FFFBEB', num_format='0.00')
    badge_task = fmt(bold=True, font_size=9, font_color='#1D4ED8', bg_color='#EFF6FF', align='center',
                     bottom=1, bottom_color='#DBE3EE')
    badge_return = fmt(bold=True, font_size=9, font_color='#15803D', bg_color='#F0FDF4', align='center',
                       bottom=1, bottom_color='#DBE3EE')
    badge_without = fmt(bold=True, font_size=9, font_color='#B91C1C', bg_color='#FEF2F2', align='center',
                          bottom=1, bottom_color='#DBE3EE')
    foot_fmt = fmt(bold=True, font_size=10, font_color='#475569', bg_color='#F8FAFC', align='right', bottom=1, bottom_color='#DBE3EE')
    foot_km_fmt = fmt(bold=True, font_size=10, font_color='#92400E', bg_color='#F8FAFC', align='right',
                      bottom=1, bottom_color='#DBE3EE', num_format='0.00')
    stamp_fmt = fmt(font_size=9, font_color='#64748B', italic=True, align='right')
    empty_fmt = fmt(font_size=10, font_color='#94A3B8', italic=True, align='center', bottom=1, bottom_color='#DBE3EE')

    def write_simple_table(title, headers, widths, rows_data, row_fn, foot_fn=None):
        nonlocal row
        ws.set_row(row, 24)
        ws.merge_range(row, 0, row, cols - 1, title, sec_fmt)
        row += 1
        ws.set_row(row, 22)
        for i, h in enumerate(headers):
            f = th_r_fmt if h.endswith('KM') or h == 'KM' else (th_c_fmt if h in ('Task ID', 'Sr') else th_fmt)
            ws.write(row, i, h, f)
        row += 1
        if not rows_data:
            ws.merge_range(row, 0, row, len(headers) - 1, 'No data', empty_fmt)
            row += 1
        else:
            for item in rows_data:
                ws.set_row(row, 20)
                row_fn(row, item)
                row += 1
        if foot_fn:
            ws.set_row(row, 22)
            foot_fn(row)
            row += 1
        ws.set_row(row, 10)
        row += 1
        for i, w in enumerate(widths):
            ws.set_column(i, i, w)

    row = 0
    ws.set_row(row, 32)
    ws.merge_range(row, 0, row, 3, 'Movement Timeline Detail', header_fmt)
    ws.merge_range(row, 4, row, cols - 1, period_txt, header_period_fmt)
    row += 1
    meta_bits = [f'Vehicle: {vehicle_no or "-"}']
    if project_name and project_name != '-':
        meta_bits.append(f'Project: {project_name}')
    if district_name and district_name != '-':
        meta_bits.append(f'District: {district_name}')
    ws.set_row(row, 20)
    ws.merge_range(row, 0, row, cols - 1, '   ·   '.join(meta_bits), header_sub_fmt)
    row += 1
    ws.set_row(row, 10)
    row += 1

    ws.set_row(row, 24)
    ws.write(row, 0, 'MOVEMENT BREAKDOWN', bd_title_fmt)
    ws.merge_range(row, 4, row, cols - 1, f'Total {total_km:.2f} KM', bd_total_fmt)
    row += 1
    stat_row = row
    cards = [
        (0, 1, '#2563EB', 'Task Running', task_km, pct(task_km), False),
        (2, 3, '#16A34A', 'Return to Parking', return_km, pct(return_km), False),
        (4, 5, '#DC2626', 'Without Task', without_km, pct(without_km), True),
    ]
    for c0, c1, accent, label, val, pc, warn in cards:
        ws.merge_range(stat_row, c0, stat_row, c1, label.upper(), stat_lbl_f(accent))
        ws.merge_range(stat_row + 1, c0, stat_row + 1, c1, f'{val:.2f} KM', stat_val_f(accent, warn))
        ws.merge_range(stat_row + 2, c0, stat_row + 2, c1, f'{pc:.1f}% of movement', stat_pct_f(accent))
    for rr in range(3):
        ws.set_row(stat_row + rr, 18 if rr != 1 else 24)
    row = stat_row + 3
    ws.set_row(row, 12)
    row += 1

    write_simple_table(
        'Task Windows',
        ['Task ID', 'Assign', 'Close', 'Task Running KM'],
        [16, 26, 26, 14],
        tasks,
        lambda r, t: (
            ws.write(r, 0, t.get('task_id') or '-', td_c_fmt),
            ws.write(r, 1, t.get('assign') or '-', td_fmt),
            ws.write(r, 2, t.get('close') or '-', td_fmt),
            ws.write(r, 3, float(t.get('running_km') or 0.0), td_km_fmt),
        ),
    )

    write_simple_table(
        'Return Windows',
        ['Task ID', 'Start', 'End', 'KM', 'Stop Reason'],
        [16, 24, 24, 10, 22],
        return_windows,
        lambda r, w: (
            ws.write(r, 0, w.get('task_id') or '-', td_c_fmt),
            ws.write(r, 1, w.get('start') or '-', td_fmt),
            ws.write(r, 2, w.get('end') or '-', td_fmt),
            ws.write(r, 3, float(w.get('km') or 0.0), td_km_fmt),
            ws.write(r, 4, _umr_timeline_reason_label(w.get('reason')), td_fmt),
        ),
    )

    def seg_row_fn(r, s):
        label = s.get('label') or 'without_task'
        is_without = label == 'without_task'
        bf = badge_task if label == 'task_running' else (badge_return if label == 'return_to_parking' else badge_without)
        bl = 'Task Running' if label == 'task_running' else ('Return to Parking' if label == 'return_to_parking' else 'Without Task')
        if is_without:
            ws.write(r, 0, bl, badge_without)
            ws.write(r, 1, s.get('task_id') or '-', td_warn_c_fmt)
            ws.write(r, 2, s.get('start') or '-', td_warn_fmt)
            ws.write(r, 3, s.get('end') or '-', td_warn_fmt)
            ws.write(r, 4, float(s.get('km') or 0.0), td_warn_km_fmt)
        else:
            ws.write(r, 0, bl, bf)
            ws.write(r, 1, s.get('task_id') or '-', td_c_fmt)
            ws.write(r, 2, s.get('start') or '-', td_fmt)
            ws.write(r, 3, s.get('end') or '-', td_fmt)
            ws.write(r, 4, float(s.get('km') or 0.0), td_km_fmt)

    def seg_foot_fn(r):
        ws.merge_range(r, 0, r, 3, 'Visible segment total', foot_fmt)
        ws.write(r, 4, seg_total, foot_km_fmt)

    write_simple_table(
        'Activity Timeline',
        ['Type', 'Task ID', 'From Time', 'To Time', 'KM'],
        [18, 16, 26, 26, 12],
        segments,
        seg_row_fn,
        seg_foot_fn,
    )

    ws.merge_range(row, 0, row, cols - 1, f'Exported {exported_at}', stamp_fmt)
    ws.freeze_panes(3, 0)
    try:
        ws.autofit()
    except Exception:
        pass
    wb.close()
    output.seek(0)
    return output





def _unauthorized_movement_preview_context():
    from auth_utils import get_user_context
    user_id = session.get('user_id')
    user_context = get_user_context(user_id) if user_id else {}
    allowed_projects = user_context.get('allowed_projects', set())
    allowed_districts = user_context.get('allowed_districts', set())
    allowed_vehicles = user_context.get('allowed_vehicles', set())
    is_master_or_admin = user_context.get('is_master_or_admin', False)

    from_date = parse_date(request.args.get('from_date')) or pk_date()
    to_date = parse_date(request.args.get('to_date')) or pk_date()
    if from_date and to_date and from_date > to_date:
        from_date, to_date = to_date, from_date
    project_id = request.args.get('project_id', type=int) or 0
    district_id = request.args.get('district_id', type=int) or 0
    vehicle_id = request.args.get('vehicle_id', type=int) or 0
    check_type = (request.args.get('check_type') or '').strip().lower()
    if check_type not in ('', 'above', 'below'):
        check_type = ''
    without_task_move_limit_raw = (request.args.get('without_task_move_limit') or '').strip()
    without_task_move_limit = None
    if without_task_move_limit_raw:
        try:
            without_task_move_limit = float(without_task_move_limit_raw)
            if without_task_move_limit < 0:
                without_task_move_limit = None
        except Exception:
            without_task_move_limit = None

    rows = _unauthorized_movement_rows(
        from_date=from_date,
        to_date=to_date,
        project_id=project_id,
        district_id=district_id,
        vehicle_id=vehicle_id,
        check_type=check_type,
        without_task_move_limit=without_task_move_limit,
        allowed_projects=allowed_projects,
        allowed_districts=allowed_districts,
        allowed_vehicles=allowed_vehicles,
        is_master_or_admin=is_master_or_admin,
    )
    rows = _filter_unauthorized_movement_rows(rows, request.args.get('table_search'))

    return {
        'rows': rows,
        'total': len(rows),
        'from_date': from_date,
        'to_date': to_date,
        'now': datetime.now,
    }




# ── Driver Response Time Report (task assign → first vehicle activity) ───
def _format_task_delay_display(delay_minutes):
    if delay_minutes is None:
        return '-'
    try:
        m = int(round(float(delay_minutes)))
    except (TypeError, ValueError):
        return '-'
    if m < 0:
        m = 0
    h, r = divmod(m, 60)
    if h:
        return f'{h}h {r:02d}m'
    return f'{m}m'


def _build_vehicle_activity_index(from_date, to_date, vehicle_nos):
    """vehicle_no (upper) -> list of (parsed_dt, VehicleActivityRecord), sorted by dt"""
    if not vehicle_nos:
        return {}
    activity_rows = VehicleActivityRecord.query.filter(
        VehicleActivityRecord.task_date >= from_date,
        VehicleActivityRecord.task_date <= to_date,
        func.upper(func.trim(VehicleActivityRecord.vehicle_no)).in_(
            [v.strip().upper() for v in vehicle_nos if v]
        ),
    ).all()
    d = {}
    for act in activity_rows:
        kn = (act.vehicle_no or '').strip().upper()
        adt = _parse_activity_datetime(act.record_date_time)
        if not adt:
            continue
        d.setdefault(kn, []).append((adt, act))
    for kn, lst in d.items():
        lst.sort(key=lambda x: x[0])
    return d


def _first_activity_after_task_assign(sorted_acts, assign_dt, close_dt):
    """
    Pehli activity jahan assign/close ke darmiyan ho, assign ke baat ya usi time,
    aur jis line ka distance > 0 ho (movement) — yahi 'Vehicle Start' time.
    """
    for adt, rec in sorted_acts:
        if adt < assign_dt:
            continue
        if adt > close_dt:
            break
        try:
            d_km = float(rec.distance or 0)
        except (TypeError, ValueError):
            d_km = 0.0
        if d_km <= 0:
            continue
        return adt, rec
    return None, None


def _task_start_delay_report_filters_from_request(args):
    """Parse shared query args for Driver Response Time Report."""
    from_date = parse_date(args.get('from_date')) or pk_date()
    to_date = parse_date(args.get('to_date')) or pk_date()
    if from_date and to_date and from_date > to_date:
        from_date, to_date = to_date, from_date
    project_id = args.get('project_id', type=int) or 0
    district_id = args.get('district_id', type=int) or 0
    vehicle_id = args.get('vehicle_id', type=int) or 0
    check_type = (args.get('check_type') or '').strip().lower()
    if check_type not in ('', 'above', 'below'):
        check_type = ''
    delay_limit_raw = (args.get('delay_limit') or '').strip()
    delay_limit = None
    if delay_limit_raw:
        try:
            delay_limit = float(delay_limit_raw)
            if delay_limit < 0:
                delay_limit = None
                delay_limit_raw = ''
        except Exception:
            delay_limit = None
            delay_limit_raw = ''
    filter_group = (args.get('filter_group') or '').strip().lower()
    if filter_group not in ('', 'time', 'delay'):
        filter_group = ''
    start_time_raw = (args.get('time_from') or args.get('start_time') or '').strip()
    end_time_raw = (args.get('time_to') or args.get('end_time') or '').strip()
    start_time_limit = _parse_time(start_time_raw)
    end_time_limit = _parse_time(end_time_raw)
    has_time = bool(start_time_raw) or bool(end_time_raw)
    has_delay = bool(check_type) or bool(delay_limit_raw)

    status = (args.get('status') or '').strip().lower()
    if status not in ('', '-', 'late_only', 'early_only'):
        status = '-'
    if status == '':
        status = '-'

    if filter_group == 'time' or (has_time and filter_group != 'delay'):
        time_mode = True
        delay_mode = False
        apply_delay = None
        check_type = ''
        delay_limit_raw = ''
        # late_only needs a start reference; early_only needs an end reference
        if status == 'late_only' and start_time_limit is None:
            status = '-'
        if status == 'early_only' and end_time_limit is None:
            status = '-'
    elif filter_group == 'delay' or has_delay:
        delay_mode = True
        time_mode = False
        apply_delay = delay_limit if (check_type in ('above', 'below') and delay_limit is not None) else None
        start_time_raw = ''
        end_time_raw = ''
        start_time_limit = None
        end_time_limit = None
        status = '-'
    else:
        delay_mode = False
        time_mode = False
        apply_delay = None
        status = '-'
    return {
        'from_date': from_date,
        'to_date': to_date,
        'project_id': project_id,
        'district_id': district_id,
        'vehicle_id': vehicle_id,
        'check_type': check_type,
        'delay_limit_raw': delay_limit_raw,
        'delay_limit': apply_delay,
        'start_time_raw': start_time_raw,
        'start_time_limit': start_time_limit,
        'end_time_raw': end_time_raw,
        'end_time_limit': end_time_limit,
        'delay_mode': delay_mode,
        'time_mode': time_mode,
        'filter_group': filter_group,
        'status': status,
    }


def _vehicle_start_in_time_window(v_start_clock, start_time_limit, end_time_limit):
    """True when vehicle start clock time falls in the optional start/end window."""
    if start_time_limit is None and end_time_limit is None:
        return True
    if start_time_limit is not None and end_time_limit is not None:
        if start_time_limit <= end_time_limit:
            return start_time_limit <= v_start_clock <= end_time_limit
        return v_start_clock >= start_time_limit or v_start_clock <= end_time_limit
    if start_time_limit is not None:
        return v_start_clock >= start_time_limit
    return v_start_clock <= end_time_limit


def _task_start_delay_rows(from_date, to_date, project_id=0, district_id=0, vehicle_id=0,
                           check_type='', delay_limit=None, start_time_limit=None, end_time_limit=None,
                           allowed_projects=None, allowed_districts=None, allowed_vehicles=None,
                           is_master_or_admin=True, status='-'):
    allowed_projects = set(allowed_projects or [])
    allowed_districts = set(allowed_districts or [])
    allowed_vehicles = set(allowed_vehicles or [])

    def _norm_vno(vno):
        return (vno or '').strip().upper()

    all_emg = EmergencyTaskRecord.query.filter(
        EmergencyTaskRecord.task_date >= from_date,
        EmergencyTaskRecord.task_date <= to_date,
        EmergencyTaskRecord.category.in_(['Green', 'Yellow']),
        EmergencyTaskRecord.completed_date_time.isnot(None),
        EmergencyTaskRecord.excel_created_date.isnot(None),
    ).order_by(EmergencyTaskRecord.task_date.desc(), EmergencyTaskRecord.id.desc()).all()
    if not all_emg:
        return []

    vnos = list({_norm_vno(e.amb_reg_no) for e in all_emg if e.amb_reg_no})
    db_vehicles = Vehicle.query.options(
        joinedload(Vehicle.parking_station),
        joinedload(Vehicle.project),
        joinedload(Vehicle.district),
    ).filter(Vehicle.vehicle_no.in_(vnos)).all() if vnos else []
    vehicle_by_no = {_norm_vno(v.vehicle_no): v for v in db_vehicles}

    if vehicle_id:
        v = Vehicle.query.options(joinedload(Vehicle.parking_station)).get(vehicle_id)
        target_no = (v.vehicle_no if v else None) and _norm_vno(v.vehicle_no)
    else:
        target_no = None

    filtered = []
    for emg in all_emg:
        vno = _norm_vno(emg.amb_reg_no)
        if not vno:
            continue
        if target_no and vno != target_no:
            continue
        v = vehicle_by_no.get(vno)
        if not v:
            continue
        if not is_master_or_admin:
            if allowed_vehicles and v.id not in allowed_vehicles:
                continue
            if allowed_districts and v.district_id not in allowed_districts:
                continue
            if allowed_projects and v.project_id not in allowed_projects:
                continue
        if project_id and v.project_id != project_id:
            continue
        if district_id and v.district_id != district_id:
            continue
        filtered.append((emg, v))

    if not filtered:
        return []

    need_nos = list({(emg.amb_reg_no or '').strip().upper() for emg, v in filtered})
    act_index = _build_vehicle_activity_index(from_date, to_date, need_nos)

    out = []
    for emg, v in filtered:
        kn = (emg.amb_reg_no or '').strip().upper()
        assign_dt = _parse_emg_datetime(emg.excel_created_date)
        close_dt = _parse_emg_datetime(emg.completed_date_time)
        if not assign_dt or not close_dt or close_dt < assign_dt:
            continue
        sorted_acts = act_index.get(kn) or []
        v_start_dt, _v_act = _first_activity_after_task_assign(sorted_acts, assign_dt, close_dt)
        if v_start_dt is None:
            delay_minutes = None
        else:
            if status == 'late_only' and start_time_limit is not None:
                expected_dt = datetime.combine(v_start_dt.date(), start_time_limit)
                delay_minutes = (v_start_dt - expected_dt).total_seconds() / 60.0
            elif status == 'early_only' and end_time_limit is not None:
                expected_dt = datetime.combine(v_start_dt.date(), end_time_limit)
                delay_minutes = (expected_dt - v_start_dt).total_seconds() / 60.0
            else:
                delay_minutes = (v_start_dt - assign_dt).total_seconds() / 60.0
            if delay_minutes < 0:
                delay_minutes = 0.0

        if delay_limit is not None:
            if check_type == 'above' and (delay_minutes is None or not (delay_minutes > delay_limit)):
                continue
            if check_type == 'below' and (delay_minutes is None or not (delay_minutes < delay_limit)):
                continue

        if start_time_limit is not None or end_time_limit is not None:
            if v_start_dt is None:
                continue
            if not _vehicle_start_in_time_window(v_start_dt.time(), start_time_limit, end_time_limit):
                continue

        if delay_minutes is None:
            delay_display = '-'
            delay_kind = ''
        elif status == 'late_only':
            formatted = _format_task_delay_display(delay_minutes)
            delay_display = formatted + ' late' if formatted != '0m' else '0m'
            delay_kind = 'late'
        elif status == 'early_only':
            formatted = _format_task_delay_display(delay_minutes)
            delay_display = formatted + ' early' if formatted != '0m' else '0m'
            delay_kind = 'early'
        else:
            delay_display = _format_task_delay_display(delay_minutes)
            delay_kind = 'normal' if delay_minutes == 0 else 'late'

        p = v.project
        d = v.district
        out.append({
            'emg': emg,
            'vehicle': v,
            'project': p,
            'district': d,
            'task_id': (emg.task_id_ext or '').strip() or '-',
            'category': (emg.category or '').strip() or '-',
            'assign_dt': assign_dt,
            'close_dt': close_dt,
            'vehicle_start_dt': v_start_dt,
            'delay_minutes': None if delay_minutes is None else round(float(delay_minutes), 2),
            'delay_display': delay_display,
            'status': status,
            'delay_kind': delay_kind,
        })
    return out


def _filter_task_start_delay_rows(rows, table_search):
    s = (table_search or '').strip().lower()
    if not s:
        return rows
    out = []
    for row in rows:
        emg = row['emg']
        blob = ' '.join([
            row['district'].name if row.get('district') else '',
            row['project'].name if row.get('project') else '',
            row['vehicle'].parking_station.name if row.get('vehicle') and row['vehicle'].parking_station else '',
            row['vehicle'].vehicle_no if row.get('vehicle') else '',
            str(row.get('task_id') or ''),
            str(row.get('category') or ''),
            emg.task_date.strftime('%d-%m-%Y') if emg.task_date else '',
            row['assign_dt'].strftime('%d-%m-%Y %I:%M %p') if row.get('assign_dt') else '',
            row['close_dt'].strftime('%d-%m-%Y %I:%M %p') if row.get('close_dt') else '',
            row['vehicle_start_dt'].strftime('%d-%m-%Y %I:%M %p') if row.get('vehicle_start_dt') else '',
            str(row.get('delay_display') or ''),
            f"{row['delay_minutes']:.2f}" if row.get('delay_minutes') is not None else '',
        ]).lower()
        if s in blob:
            out.append(row)
    return out




def _task_start_delay_report_preview_context():
    from auth_utils import get_user_context
    user_id = session.get('user_id')
    user_context = get_user_context(user_id) if user_id else {}
    allowed_projects = user_context.get('allowed_projects', set())
    allowed_districts = user_context.get('allowed_districts', set())
    allowed_vehicles = user_context.get('allowed_vehicles', set())
    is_master_or_admin = user_context.get('is_master_or_admin', False)

    filters = _task_start_delay_report_filters_from_request(request.args)
    rows = _task_start_delay_rows(
        from_date=filters['from_date'], to_date=filters['to_date'],
        project_id=filters['project_id'], district_id=filters['district_id'], vehicle_id=filters['vehicle_id'],
        check_type=filters['check_type'], delay_limit=filters['delay_limit'],
        start_time_limit=filters['start_time_limit'], end_time_limit=filters['end_time_limit'],
        allowed_projects=allowed_projects, allowed_districts=allowed_districts, allowed_vehicles=allowed_vehicles,
        is_master_or_admin=is_master_or_admin, status=filters['status'],
    )
    rows = _filter_task_start_delay_rows(rows, request.args.get('table_search'))
    return {
        'rows': rows,
        'total': len(rows),
        'from_date': filters['from_date'],
        'to_date': filters['to_date'],
        'check_type': filters['check_type'],
        'delay_limit': filters['delay_limit_raw'],
        'start_time': _time_input_value(filters['start_time_raw']),
        'end_time': _time_input_value(filters['end_time_raw']),
        'delay_mode': filters['delay_mode'],
        'time_mode': filters['time_mode'],
        'status': filters['status'],
        'filter_group': 'time' if filters['time_mode'] else ('delay' if filters['delay_mode'] else ''),
        'now': datetime.now,
    }




def _driver_response_vehicle_activity_rows(vehicle_no, from_date, to_date, start_time=None, end_time=None):
    """All vehicle activity records (start/stop/ignition + km) in range, time-sorted.
    Optional start_time / end_time are datetime.time objects that further restrict the
    record time of day (same semantics as the report's Start/End time window)."""
    acts = VehicleActivityRecord.query.filter(
        func.upper(func.trim(VehicleActivityRecord.vehicle_no)) == (vehicle_no or '').strip().upper(),
        VehicleActivityRecord.task_date >= from_date,
        VehicleActivityRecord.task_date <= to_date,
    ).all()
    parsed = []
    for a in acts:
        adt = _parse_activity_datetime(a.record_date_time)
        if adt:
            parsed.append((adt, a))
    parsed.sort(key=lambda x: x[0])

    def _in_time_window(t):
        if start_time is None and end_time is None:
            return True
        if start_time is not None and end_time is not None:
            if start_time <= end_time:
                return start_time <= t <= end_time
            return t >= start_time or t <= end_time
        if start_time is not None:
            return t >= start_time
        return t <= end_time

    rows = []
    total_km = 0.0
    moving_km = 0.0
    for adt, a in parsed:
        if not _in_time_window(adt.time()):
            continue
        try:
            km = float(a.distance or 0)
        except (TypeError, ValueError):
            km = 0.0
        try:
            spd = float(a.speed or 0)
        except (TypeError, ValueError):
            spd = 0.0
        total_km += km
        if km > 0:
            moving_km += km
        rows.append({
            'record_dt': adt.strftime('%d-%m-%Y %I:%M %p'),
            'reason': (a.reason or '-').strip() or '-',
            'speed': round(spd, 2),
            'distance': round(km, 2),
            'travel_time': (a.travel_time or '-') or '-',
            'stop_time': (a.stop_time or '-') or '-',
            'location': (a.location or '-') or '-',
        })
    return rows, round(total_km, 2), round(moving_km, 2)


def _driver_response_vehicle_activity_request():
    """Parse + scope-check shared args for vehicle activity detail (API + export)."""
    from auth_utils import get_user_context
    user_id = session.get('user_id')
    user_context = get_user_context(user_id) if user_id else {}
    allowed_vehicles = user_context.get('allowed_vehicles', set())
    allowed_districts = user_context.get('allowed_districts', set())
    allowed_projects = user_context.get('allowed_projects', set())
    is_master_or_admin = user_context.get('is_master_or_admin', False)

    vehicle_no = (request.args.get('vehicle_no') or '').strip()
    from_date = parse_date(request.args.get('from_date'))
    to_date = parse_date(request.args.get('to_date'))
    if not vehicle_no or not from_date or not to_date:
        return None, 'Missing vehicle or date range.'
    if from_date > to_date:
        from_date, to_date = to_date, from_date

    start_time_raw = (request.args.get('time_from') or '').strip()
    end_time_raw = (request.args.get('time_to') or '').strip()
    start_time = _parse_time(start_time_raw)
    end_time = _parse_time(end_time_raw)

    if not is_master_or_admin:
        v = Vehicle.query.filter(func.upper(Vehicle.vehicle_no) == vehicle_no.upper()).first()
        if not v:
            return None, 'Vehicle not found.'
        if allowed_vehicles and v.id not in allowed_vehicles:
            return None, 'Not allowed.'
        if allowed_districts and v.district_id not in allowed_districts:
            return None, 'Not allowed.'
        if allowed_projects and v.project_id not in allowed_projects:
            return None, 'Not allowed.'

    rows, total_km, moving_km = _driver_response_vehicle_activity_rows(
        vehicle_no, from_date, to_date, start_time=start_time, end_time=end_time)
    return {
        'vehicle_no': vehicle_no,
        'from_date': from_date,
        'to_date': to_date,
        'time_from': start_time_raw,
        'time_to': end_time_raw,
        'rows': rows,
        'total_km': total_km,
        'moving_km': moving_km,
    }, None




def _format_task_duration_hhmm(duration_minutes):
    """Total minutes between task assign and close → 'HH:MM' (hours can exceed 24)."""
    if duration_minutes is None:
        return '-'
    try:
        m = int(round(float(duration_minutes)))
    except (TypeError, ValueError):
        return '-'
    if m < 0:
        m = 0
    h, r = divmod(m, 60)
    return f'{h:02d}:{r:02d}'


def _parse_duration_limit_hhmm(s):
    """
    Filter value: 'H:MM' or 'HH:MM' (e.g. 1:30 → 90 minutes), or plain number as minutes.
    Returns float minutes, or None if empty/invalid.
    """
    s = (s or '').strip()
    if not s:
        return None
    m = re.match(r'^(\d{1,4})\s*:\s*(\d{1,2})\s*$', s)
    if m:
        h = int(m.group(1))
        mi = int(m.group(2))
        if mi > 59 or h < 0:
            return None
        return float(h * 60 + mi)
    try:
        v = float(s)
        if v < 0:
            return None
        return v
    except (TypeError, ValueError):
        return None


def _task_turnaround_rows(from_date, to_date, project_id=0, district_id=0, vehicle_id=0,
                          check_type='', time_limit=None,
                          allowed_projects=None, allowed_districts=None, allowed_vehicles=None,
                          is_master_or_admin=True):
    """
    Task assign (excel create) se task close tak ka duration — vehicle activity not used.
    """
    allowed_projects = set(allowed_projects or [])
    allowed_districts = set(allowed_districts or [])
    allowed_vehicles = set(allowed_vehicles or [])

    def _norm_vno(vno):
        return (vno or '').strip().upper()

    all_emg = EmergencyTaskRecord.query.filter(
        EmergencyTaskRecord.task_date >= from_date,
        EmergencyTaskRecord.task_date <= to_date,
        EmergencyTaskRecord.category.in_(['Green', 'Yellow']),
        EmergencyTaskRecord.completed_date_time.isnot(None),
        EmergencyTaskRecord.excel_created_date.isnot(None),
    ).order_by(EmergencyTaskRecord.task_date.desc(), EmergencyTaskRecord.id.desc()).all()
    if not all_emg:
        return []

    vnos = list({_norm_vno(e.amb_reg_no) for e in all_emg if e.amb_reg_no})
    db_vehicles = Vehicle.query.filter(Vehicle.vehicle_no.in_(vnos)).all() if vnos else []
    vehicle_by_no = {_norm_vno(v.vehicle_no): v for v in db_vehicles}

    if vehicle_id:
        v = db.session.get(Vehicle, vehicle_id)
        target_no = (v.vehicle_no if v else None) and _norm_vno(v.vehicle_no)
    else:
        target_no = None

    filtered = []
    for emg in all_emg:
        vno = _norm_vno(emg.amb_reg_no)
        if not vno:
            continue
        if target_no and vno != target_no:
            continue
        v = vehicle_by_no.get(vno)
        if not v:
            continue
        if not is_master_or_admin:
            if allowed_vehicles and v.id not in allowed_vehicles:
                continue
            if allowed_districts and v.district_id not in allowed_districts:
                continue
            if allowed_projects and v.project_id not in allowed_projects:
                continue
        if project_id and v.project_id != project_id:
            continue
        if district_id and v.district_id != district_id:
            continue
        filtered.append((emg, v))

    if not filtered:
        return []

    out = []
    for emg, v in filtered:
        assign_dt = _parse_emg_datetime(emg.excel_created_date)
        close_dt = _parse_emg_datetime(emg.completed_date_time)
        if not assign_dt or not close_dt or close_dt < assign_dt:
            continue
        duration_minutes = (close_dt - assign_dt).total_seconds() / 60.0
        if duration_minutes < 0:
            duration_minutes = 0.0

        if time_limit is not None:
            if check_type == 'above' and not (duration_minutes > time_limit):
                continue
            if check_type == 'below' and not (duration_minutes < time_limit):
                continue

        p = v.project
        d = v.district
        out.append({
            'emg': emg,
            'vehicle': v,
            'project': p,
            'district': d,
            'task_id': (emg.task_id_ext or '').strip() or '-',
            'category': (emg.category or '').strip() or '-',
            'assign_dt': assign_dt,
            'close_dt': close_dt,
            'duration_minutes': round(float(duration_minutes), 2),
            'duration_display': _format_task_duration_hhmm(duration_minutes),
        })
    return out


def _filter_task_turnaround_rows(rows, table_search):
    s = (table_search or '').strip().lower()
    if not s:
        return rows
    out = []
    for row in rows:
        emg = row['emg']
        blob = ' '.join([
            row['district'].name if row.get('district') else '',
            row['project'].name if row.get('project') else '',
            row['vehicle'].vehicle_no if row.get('vehicle') else '',
            str(row.get('task_id') or ''),
            str(row.get('category') or ''),
            emg.task_date.strftime('%d-%m-%Y') if emg.task_date else '',
            row['assign_dt'].strftime('%d-%m-%Y %I:%M %p') if row.get('assign_dt') else '',
            row['close_dt'].strftime('%d-%m-%Y %I:%M %p') if row.get('close_dt') else '',
            str(row.get('duration_display') or ''),
            f"{row['duration_minutes']:.2f}" if row.get('duration_minutes') is not None else '',
        ]).lower()
        if s in blob:
            out.append(row)
    return out




def _task_turnaround_report_preview_context():
    from auth_utils import get_user_context
    user_id = session.get('user_id')
    user_context = get_user_context(user_id) if user_id else {}
    allowed_projects = user_context.get('allowed_projects', set())
    allowed_districts = user_context.get('allowed_districts', set())
    allowed_vehicles = user_context.get('allowed_vehicles', set())
    is_master_or_admin = user_context.get('is_master_or_admin', False)

    from_date = parse_date(request.args.get('from_date')) or pk_date()
    to_date = parse_date(request.args.get('to_date')) or pk_date()
    if from_date and to_date and from_date > to_date:
        from_date, to_date = to_date, from_date
    project_id = request.args.get('project_id', type=int) or 0
    district_id = request.args.get('district_id', type=int) or 0
    vehicle_id = request.args.get('vehicle_id', type=int) or 0
    check_type = (request.args.get('check_type') or '').strip().lower()
    if check_type not in ('', 'above', 'below'):
        check_type = ''
    duration_limit_raw = (request.args.get('duration_limit') or '').strip()
    time_limit = _parse_duration_limit_hhmm(duration_limit_raw)
    apply_limit = time_limit if (check_type in ('above', 'below') and time_limit is not None) else None

    rows = _task_turnaround_rows(
        from_date=from_date, to_date=to_date, project_id=project_id, district_id=district_id, vehicle_id=vehicle_id,
        check_type=check_type, time_limit=apply_limit,
        allowed_projects=allowed_projects, allowed_districts=allowed_districts, allowed_vehicles=allowed_vehicles,
        is_master_or_admin=is_master_or_admin,
    )
    rows = _filter_task_turnaround_rows(rows, request.args.get('table_search'))
    return {
        'rows': rows,
        'total': len(rows),
        'from_date': from_date,
        'to_date': to_date,
        'check_type': check_type,
        'duration_limit': duration_limit_raw,
        'now': datetime.now,
    }




# ── Tracker Difference Report ───────────────────────────────────────────
def _tracker_difference_rows(from_date=None, to_date=None, project_id=0, district_id=0, vehicle_id=0,
                             check_type='', diff_pct_limit=None,
                             allowed_projects=None, allowed_districts=None, allowed_vehicles=None,
                             is_master_or_admin=True):
    query = db.session.query(
        VehicleDailyTask, Vehicle, Project, District
    ).join(
        Vehicle, Vehicle.id == VehicleDailyTask.vehicle_id
    ).outerjoin(
        Project, Vehicle.project_id == Project.id
    ).outerjoin(
        District, Vehicle.district_id == District.id
    )

    if from_date:
        query = query.filter(VehicleDailyTask.task_date >= from_date)
    if to_date:
        query = query.filter(VehicleDailyTask.task_date <= to_date)

    if not is_master_or_admin:
        if allowed_projects:
            query = query.filter(Vehicle.project_id.in_(list(allowed_projects)))
        if allowed_districts:
            query = query.filter(Vehicle.district_id.in_(list(allowed_districts)))
        if allowed_vehicles:
            query = query.filter(Vehicle.id.in_(list(allowed_vehicles)))

    if project_id:
        query = query.filter(Vehicle.project_id == project_id)
    if district_id:
        query = query.filter(Vehicle.district_id == district_id)
    if vehicle_id:
        query = query.filter(Vehicle.id == vehicle_id)

    out = []
    prev_close_cache = {}
    for rec, vehicle, project, district in query.order_by(VehicleDailyTask.task_date.desc(), VehicleDailyTask.id.desc()).all():
        start_r, close_r = _resolve_task_readings(rec, vehicle.id if vehicle else None, prev_close_cache)
        total_km = round(close_r - start_r, 2)

        tracker_rec = VehicleMileageRecord.query.filter_by(
            task_date=rec.task_date, reg_no=(vehicle.vehicle_no if vehicle else None)
        ).first()
        tracker_mileage = float(tracker_rec.effective_km()) if tracker_rec else 0.0

        # Match Vehicles Daily Task Report logic:
        # Diff Km's = Driven (Total KMs) - Tracker Mileage
        # % Diff = Diff Km's / Driven (Total KMs) * 100
        diff_km = round(total_km - tracker_mileage, 2)
        diff_pct = round((diff_km / total_km) * 100.0, 2) if total_km not in (0, 0.0) else None

        if diff_pct_limit is not None:
            if diff_pct is None:
                continue
            if check_type == 'above' and not (diff_pct > diff_pct_limit):
                continue
            if check_type == 'below' and not (diff_pct < diff_pct_limit):
                continue

        check_result = None
        if diff_pct_limit is not None and diff_pct is not None:
            if diff_pct > diff_pct_limit:
                check_result = f'Above (+{diff_pct - diff_pct_limit:.2f}%)'
            elif diff_pct < diff_pct_limit:
                check_result = f'Below (-{diff_pct_limit - diff_pct:.2f}%)'
            else:
                check_result = 'Equal (0.00%)'

        out.append({
            'rec': rec,
            'vehicle': vehicle,
            'project': project,
            'district': district,
            'start_reading': start_r,
            'close_reading': close_r,
            'total_km': total_km,
            'tracker_mileage': tracker_mileage,
            'diff_km': diff_km,
            'diff_pct': diff_pct,
            'check_result': check_result,
        })
    return out




def _tracker_difference_report_preview_context():
    from auth_utils import get_user_context
    user_id = session.get('user_id')
    user_context = get_user_context(user_id) if user_id else {}
    allowed_projects = user_context.get('allowed_projects', set())
    allowed_districts = user_context.get('allowed_districts', set())
    allowed_vehicles = user_context.get('allowed_vehicles', set())
    is_master_or_admin = user_context.get('is_master_or_admin', False)

    from_date = parse_date(request.args.get('from_date')) or pk_date()
    to_date = parse_date(request.args.get('to_date')) or pk_date()
    if from_date and to_date and from_date > to_date:
        from_date, to_date = to_date, from_date
    project_id = request.args.get('project_id', type=int) or 0
    district_id = request.args.get('district_id', type=int) or 0
    vehicle_id = request.args.get('vehicle_id', type=int) or 0
    check_type = (request.args.get('check_type') or '').strip().lower()
    if check_type not in ('', 'above', 'below'):
        check_type = ''
    diff_pct_limit_raw = (request.args.get('diff_pct_limit') or '').strip()
    diff_pct_limit = None
    if diff_pct_limit_raw:
        try:
            diff_pct_limit = float(diff_pct_limit_raw)
            if diff_pct_limit < 0:
                diff_pct_limit = None
        except Exception:
            diff_pct_limit = None

    rows = _tracker_difference_rows(
        from_date=from_date, to_date=to_date, project_id=project_id, district_id=district_id,
        vehicle_id=vehicle_id, check_type=check_type, diff_pct_limit=diff_pct_limit,
        allowed_projects=allowed_projects, allowed_districts=allowed_districts, allowed_vehicles=allowed_vehicles,
        is_master_or_admin=is_master_or_admin,
    )
    table_search = (request.args.get('table_search') or '').strip().lower()
    if table_search:
        def _m(r):
            blob = ' '.join([
                r['rec'].task_date.strftime('%d-%m-%Y') if r['rec'].task_date else '',
                r['district'].name if r.get('district') else '',
                r['project'].name if r.get('project') else '',
                r['vehicle'].vehicle_no if r.get('vehicle') else '',
                format_reading(r['start_reading']),
                format_reading(r['close_reading']),
                f"{r['total_km']:.2f}",
                f"{r['tracker_mileage']:.2f}",
                f"{r['diff_km']:.2f}",
                f"{r['diff_pct']:.2f}%" if r['diff_pct'] is not None else '',
                r.get('check_result') or '',
            ]).lower()
            return table_search in blob
        rows = [r for r in rows if _m(r)]

    return {
        'rows': rows,
        'total': len(rows),
        'from_date': from_date,
        'to_date': to_date,
        'diff_pct_limit': diff_pct_limit_raw,
        'check_type': check_type,
        'now': datetime.now,
    }




# ── Driver Seat Available Report ───────────────────────────────────────────

def _seat_available_data(project_id=0, district_id=0, vehicle_type='',
                         allowed_vehicles=None, allowed_projects=None,
                         allowed_districts=None, is_master_or_admin=True):
    """
    Vehicles where assigned_driver_count < driver_capacity.
    Returns list of (Vehicle, Project, District, capacity, assigned, vacant).
    """
    assigned_sub = db.session.query(
        Driver.vehicle_id,
        func.count(Driver.id).label('assigned_count')
    ).filter(
        Driver.vehicle_id.isnot(None),
        Driver.status != 'Left'
    ).group_by(Driver.vehicle_id).subquery()

    query = db.session.query(
        Vehicle, Project, District,
        Vehicle.driver_capacity,
        func.coalesce(assigned_sub.c.assigned_count, 0).label('assigned_count')
    ).outerjoin(
        Project, Vehicle.project_id == Project.id
    ).outerjoin(
        District, Vehicle.district_id == District.id
    ).outerjoin(
        assigned_sub, Vehicle.id == assigned_sub.c.vehicle_id
    ).filter(
        func.coalesce(Vehicle.driver_capacity, 1) > func.coalesce(assigned_sub.c.assigned_count, 0)
    )

    if not is_master_or_admin:
        if allowed_projects:
            query = query.filter(Vehicle.project_id.in_(list(allowed_projects)))
        if allowed_districts:
            query = query.filter(Vehicle.district_id.in_(list(allowed_districts)))
        if allowed_vehicles:
            query = query.filter(Vehicle.id.in_(list(allowed_vehicles)))

    if project_id:
        query = query.filter(Vehicle.project_id == project_id)
    if district_id:
        query = query.filter(Vehicle.district_id == district_id)
    if vehicle_type:
        query = query.filter(Vehicle.vehicle_type == vehicle_type)

    rows = query.order_by(*vehicle_order_by(Project.name, District.name)).all()

    results = []
    for vehicle, project, district, capacity, assigned in rows:
        cap = capacity if capacity and capacity > 0 else 1
        asgn = assigned or 0
        vacant = cap - asgn
        if vacant > 0:
            results.append((vehicle, project, district, cap, asgn, vacant))
    return results





# ── Driver Missing Documents Report ───────────────────────────────────────

# ── Missing Documents Report – Print View ─────────────────────────────────

# API: Ek makhsoos gaari (Vehicle) par assign drivers nikalna

# --- 1. REJOIN MAIN ROUTE ---
# 2. EDIT ROUTE
# 3. DELETE ROUTE
# routes.py mein ye function add karein
def _attendance_media_gallery_flat_and_items_from_request():
    """Shared list filters + gallery_shift/gallery_photo for gallery, zip, and share JSON."""
    from auth_utils import get_user_context

    uid = session.get('user_id')
    uc = get_user_context(uid) if uid else {}
    view_date = _attendance_local_date()
    from_date_str = (request.args.get('from_date') or '').strip()
    to_date_str = (request.args.get('to_date') or '').strip()
    from_date = parse_date(from_date_str) if from_date_str else None
    to_date = parse_date(to_date_str) if to_date_str else None
    if request.args.get('date') and not from_date and not to_date:
        view_date = parse_date(request.args.get('date')) or view_date
        from_date = to_date = view_date

    def _iq(name):
        v = request.args.get(name)
        if v is None or v == '':
            return None
        try:
            n = int(v)
            return n if n != 0 else None
        except (TypeError, ValueError):
            return None

    project_id = _iq('project_id')
    district_id = _iq('district_id')
    vehicle_id = _iq('vehicle_id')
    driver_id = _iq('driver_id')
    shift = (request.args.get('shift') or '').strip()
    search = (request.args.get('search') or '').strip()
    duty_shift_filter = (request.args.get('duty_shift') or '').strip().lower()
    if duty_shift_filter not in ('', 'morning', 'evening'):
        duty_shift_filter = ''

    gallery_shift = (request.args.get('gallery_shift') or 'both').strip().lower()
    gallery_photo = (request.args.get('gallery_photo') or 'both').strip().lower()
    if gallery_shift not in ('morning', 'evening', 'both'):
        gallery_shift = 'both'
    if gallery_photo not in ('checkin', 'checkout', 'both'):
        gallery_photo = 'both'

    flat = _driver_attendance_flat_rows(
        project_id,
        district_id,
        vehicle_id,
        shift,
        search,
        driver_id,
        uc,
        from_date=from_date,
        to_date=to_date,
        single_date=view_date if not from_date and not to_date else None,
    )
    if duty_shift_filter:
        flat = _filter_attendance_rows_by_duty_shift(flat, duty_shift_filter)

    media_items = _build_attendance_media_gallery_items(flat, gallery_shift, gallery_photo)
    return media_items, gallery_shift, gallery_photo






def _attendance_check_in_remarks(rec):
    """Check-in remarks: GPS+Camera auto text, else manual form reason."""
    if not rec:
        return ''
    if rec.check_in_photo_path or (rec.check_in_latitude is not None and rec.check_in_longitude is not None):
        return 'Check-in via GPS & Camera'
    r = rec.remarks or ''
    if 'Manual check-in:' in r:
        part = r.split('Manual check-in:')[1].split(' | ')[0].strip()
        return part or 'Manual check-in'
    return ''


def _attendance_check_out_remarks(rec):
    """Check-out remarks: GPS+Camera auto text, else manual form reason."""
    if not rec:
        return ''
    if rec.check_out_photo_path or (rec.check_out_latitude is not None and rec.check_out_longitude is not None):
        return 'Check-out via GPS & Camera'
    r = rec.remarks or ''
    if 'Auto check-out:' in r:
        part = r.split('Auto check-out:', 1)[1].split(' | ')[0].strip()
        return ('Auto check-out: ' + part).strip() if part else GPS_AUTO_CHECKOUT_REMARK
    if 'Manual check-out' in r:
        part = r.split('Manual check-out', 1)[1].split(' | ')[0].strip()
        if part.startswith(': '):
            part = part[2:].strip()
        return part or 'Manual check-out'
    return ''








def _missing_checkout_records(view_date, project_id, district_id, vehicle_id, shift, driver_id, search, user_context):
    """Drivers with check-in but no check-out for view_date, scoped like Missing Check-outs filters."""
    allowed_projects = user_context.get('allowed_projects', set())
    allowed_districts = user_context.get('allowed_districts', set())
    is_master_or_admin = user_context.get('is_master_or_admin', False)
    need_vehicle_join = bool(district_id or search or (not is_master_or_admin and allowed_districts))
    query = DriverAttendance.query.filter(
        DriverAttendance.attendance_date == view_date,
        DriverAttendance.check_in.isnot(None),
        DriverAttendance.check_out.is_(None),
    ).join(Driver, DriverAttendance.driver_id == Driver.id).options(db.joinedload(DriverAttendance.driver))
    if need_vehicle_join:
        query = query.outerjoin(Vehicle, Driver.vehicle_id == Vehicle.id)
    if project_id:
        query = query.filter(Driver.project_id == project_id)
    elif not is_master_or_admin and allowed_projects:
        query = query.filter(Driver.project_id.in_(list(allowed_projects)))
    if district_id:
        query = query.filter(db.or_(Driver.district_id == district_id, Vehicle.district_id == district_id))
    elif not is_master_or_admin and allowed_districts:
        query = query.filter(
            db.or_(
                Driver.district_id.in_(list(allowed_districts)),
                Vehicle.district_id.in_(list(allowed_districts)),
            )
        )
    if vehicle_id:
        query = query.filter(Driver.vehicle_id == vehicle_id)
    if shift:
        query = query.filter(Driver.shift == shift)
    if driver_id:
        query = query.filter(Driver.id == driver_id)
    if search:
        flt = _multi_word_filter(search, Driver.name, Driver.driver_id, Vehicle.vehicle_no)
        if flt is not None:
            query = query.filter(flt)
    return query.order_by(Driver.name).all()



def _manual_attendance_driver_id():
    """Resolve driver_id when query string repeats keys (e.g. filter driver_id=0 + row driver_id). Last positive id wins."""
    merged = []
    merged.extend(request.args.getlist('driver_id'))
    merged.extend(request.form.getlist('driver_id'))
    for raw in reversed(merged):
        try:
            n = int(raw)
            if n > 0:
                return n
        except (TypeError, ValueError):
            continue
    return None






# ────────────────────────────────────────────────
# Geofenced attendance: driver at parking station + selfie
# ────────────────────────────────────────────────




def _ov_to_dict(ov, source_label=None):
    """Convert an AttendanceTimeOverride row to a time-window dict."""
    d = {
        'morning_start': ov.morning_start, 'morning_end': ov.morning_end,
        'night_start': ov.night_start, 'night_end': ov.night_end,
        'source': source_label or getattr(ov, 'scope_label', ''),
    }
    for f in ('morning_checkout_start', 'morning_checkout_end', 'night_checkout_start', 'night_checkout_end'):
        d[f] = getattr(ov, f, None)
    return d


def _attendance_time_in_window(t, start_t, end_t):
    if start_t is None or end_t is None:
        return True
    if end_t < start_t:
        return t >= start_t or t <= end_t
    return start_t <= t <= end_t


def _attendance_allow_morning_driver_night_gps_checkin():
    glob = AttendanceTimeOverride.query.filter_by(scope='global').first()
    return bool(glob and glob.allow_morning_driver_night_gps_checkin)


def _attendance_allow_night_driver_morning_gps_checkin():
    glob = AttendanceTimeOverride.query.filter_by(scope='global').first()
    return bool(glob and glob.allow_night_driver_morning_gps_checkin)


def _attendance_auto_gps_checkout_on_window_end_enabled():
    glob = AttendanceTimeOverride.query.filter_by(scope='global').first()
    return bool(glob and glob.auto_gps_checkout_on_window_end)


def _attendance_capacity_one_checkin_mode():
    """Global policy for capacity-1 vehicles: both / morning_only / night_only."""
    glob = AttendanceTimeOverride.query.filter_by(scope='global').first()
    raw = ((glob.capacity_one_checkin_mode if glob else None) or 'both').strip().lower()
    if raw in ('morning', 'morning_only', 'morning-only'):
        return 'morning_only'
    if raw in ('night', 'evening', 'night_only', 'night-only', 'evening_only', 'evening-only'):
        return 'night_only'
    return 'both'


def _checkin_in_morning_window(tw, check_in_time):
    return _attendance_time_in_window(
        check_in_time, tw.get('morning_start'), tw.get('morning_end'),
    )


def _checkin_in_night_window(tw, check_in_time):
    return _attendance_time_in_window(
        check_in_time, tw.get('night_start'), tw.get('night_end'),
    )


def _gps_checkin_was_cross_shift(driver, check_in_time, tw):
    """True when GPS check-in used the opposite window (toggle path), not assigned shift window."""
    if not driver or not check_in_time:
        return False
    shift_l = (driver.shift or '').strip().lower()
    in_m = _checkin_in_morning_window(tw, check_in_time)
    in_n = _checkin_in_night_window(tw, check_in_time)
    if shift_l == 'morning':
        return bool(
            _attendance_allow_morning_driver_night_gps_checkin()
            and in_n
            and not in_m
        )
    if shift_l == 'night':
        return bool(
            _attendance_allow_night_driver_morning_gps_checkin()
            and in_m
            and not in_n
        )
    return False


def _gps_checkout_window_bounds(driver, tw, check_in_time):
    """GPS check-out window from check-in context: cross-shift uses opposite Check-OUT override."""
    shift_l = (driver.shift or '').strip().lower() if driver else ''
    cross = _gps_checkin_was_cross_shift(driver, check_in_time, tw)
    co_s = co_e = None
    if shift_l == 'morning':
        if cross:
            co_s = tw.get('night_checkout_start')
            co_e = tw.get('night_checkout_end')
            if not co_s and not co_e:
                co_s = tw.get('night_start')
                co_e = tw.get('night_end')
        else:
            co_s = tw.get('morning_checkout_start')
            co_e = tw.get('morning_checkout_end')
            if not co_s and not co_e:
                co_s = tw.get('night_start')
                co_e = tw.get('night_end')
    elif shift_l == 'night':
        if cross:
            co_s = tw.get('morning_checkout_start')
            co_e = tw.get('morning_checkout_end')
            if not co_s and not co_e:
                co_s = tw.get('night_start')
                co_e = tw.get('night_end')
        else:
            co_s = tw.get('night_checkout_start')
            co_e = tw.get('night_checkout_end')
            if not co_s and not co_e:
                co_s = tw.get('morning_start')
                co_e = tw.get('morning_end')
    return co_s, co_e, cross


def _checkout_window_end_datetime(attendance_date, check_in_time, start_t, end_t):
    """Datetime when this session's checkout window ends (must be after check-in)."""
    if not attendance_date or not end_t:
        return None
    if start_t and end_t and end_t < start_t:
        end_date = attendance_date + timedelta(days=1)
    else:
        end_date = attendance_date
    end_dt = datetime.combine(end_date, end_t)
    if check_in_time:
        check_in_dt = datetime.combine(attendance_date, check_in_time)
        if end_dt <= check_in_dt:
            end_dt = datetime.combine(end_date + timedelta(days=1), end_t)
    return end_dt


def _checkout_window_end_passed(attendance_date, check_in_time, start_t, end_t, now_dt):
    end_dt = _checkout_window_end_datetime(attendance_date, check_in_time, start_t, end_t)
    if not end_dt or not now_dt:
        return False
    return now_dt >= end_dt


GPS_AUTO_CHECKOUT_REMARK = 'Auto check-out: Checkout window ended (system setting)'


def _gps_checkout_window_ok(driver, now_time, tw, check_in_time):
    co_s, co_e, cross = _gps_checkout_window_bounds(driver, tw, check_in_time)
    if not _attendance_time_in_window(now_time, co_s, co_e):
        shift_l = (driver.shift or '').strip().lower() if driver else ''
        if cross and shift_l == 'morning':
            return False, 'Is session ki check-in night window mein thi — night check-out window ke dauran check-out karein.'
        if cross and shift_l == 'night':
            return False, 'Is session ki check-in morning window mein thi — morning check-out window ke dauran check-out karein.'
        if shift_l == 'morning':
            return False, 'Morning check-out time-window allowed nahi.'
        if shift_l == 'night':
            return False, 'Night check-out time-window allowed nahi.'
        return False, 'Check-out time-window allowed nahi.'
    return True, None


def _gps_checkin_shift_window_ok(shift, now_time, tw, driver=None, vehicle=None):
    """GPS check-in allowed for assigned shift and configured windows."""
    v = vehicle or (driver.vehicle if driver is not None else None)
    cap = _vehicle_capacity_value(v)
    if cap == 1:
        cap_mode = _attendance_capacity_one_checkin_mode()
        if cap_mode == 'morning_only':
            if _attendance_time_in_window(now_time, tw.get('morning_start'), tw.get('morning_end')):
                return True, None
            return False, 'Capacity-1 vehicle: attendance sirf morning window mein allowed hai (settings).'
        if cap_mode == 'night_only':
            if _attendance_time_in_window(now_time, tw.get('night_start'), tw.get('night_end')):
                return True, None
            return False, 'Capacity-1 vehicle: attendance sirf evening/night window mein allowed hai (settings).'

    shift_l = (shift or '').strip().lower()
    if shift_l == 'morning':
        if _attendance_time_in_window(now_time, tw.get('morning_start'), tw.get('morning_end')):
            return True, None
        if _attendance_allow_morning_driver_night_gps_checkin() and _attendance_time_in_window(
            now_time, tw.get('night_start'), tw.get('night_end')
        ):
            return True, None
        if _attendance_allow_morning_driver_night_gps_checkin():
            return False, (
                'Morning shift driver: abhi na morning na night check-in window mein. '
                'Settings → Attendance → GPS Check-in settings aur Night window check karein.'
            )
        return False, 'Morning shift ki attendance sirf morning time window mein lag sakti hai.'
    if shift_l == 'night':
        if _attendance_time_in_window(now_time, tw.get('night_start'), tw.get('night_end')):
            return True, None
        if _attendance_allow_night_driver_morning_gps_checkin() and _attendance_time_in_window(
            now_time, tw.get('morning_start'), tw.get('morning_end')
        ):
            return True, None
        if _attendance_allow_night_driver_morning_gps_checkin():
            return False, (
                'Night shift driver: abhi na night na morning check-in window mein. '
                'Settings → Attendance → GPS Check-in settings aur Morning window check karein.'
            )
        return False, 'Night shift ki attendance sirf night time window mein lag sakti hai.'
    return True, None


def _get_effective_time_window(driver=None, vehicle_id=None, project_id=None):
    """Hierarchical time lookup: Vehicle > District > Project > Global.
    Accepts driver object OR explicit vehicle_id/project_id for early lookup
    before a driver is selected."""
    _vehicle = None
    _project_id = project_id
    _district_id = None

    if driver:
        _vehicle = driver.vehicle if driver.vehicle_id else None
        if not _project_id:
            _project_id = driver.project_id
    if vehicle_id and not _vehicle:
        _vehicle = db.session.get(Vehicle, vehicle_id)
    if _vehicle:
        if not _project_id:
            _project_id = _vehicle.project_id
        _district_id = _vehicle.district_id

    if _vehicle:
        ov = AttendanceTimeOverride.query.filter_by(scope='vehicle', vehicle_id=_vehicle.id).first()
        if ov:
            return _ov_to_dict(ov)
    if _district_id and _project_id:
        ov = AttendanceTimeOverride.query.filter_by(scope='district', project_id=_project_id, district_id=_district_id).first()
        if ov:
            return _ov_to_dict(ov)
    if _project_id:
        ov = AttendanceTimeOverride.query.filter_by(scope='project', project_id=_project_id).first()
        if ov:
            return _ov_to_dict(ov)

    glob = AttendanceTimeOverride.query.filter_by(scope='global').first()
    if glob:
        return _ov_to_dict(glob, 'Global Default')
    ctrl = AttendanceTimeControl.query.first()
    if ctrl:
        return {
            'morning_start': ctrl.morning_start, 'morning_end': ctrl.morning_end,
            'night_start': ctrl.night_start, 'night_end': ctrl.night_end,
            'morning_checkout_start': None, 'morning_checkout_end': None,
            'night_checkout_start': None, 'night_checkout_end': None,
            'source': 'Global Default (Legacy)',
        }
    return {
        'morning_start': None, 'morning_end': None,
        'night_start': None, 'night_end': None,
        'morning_checkout_start': None, 'morning_checkout_end': None,
        'night_checkout_start': None, 'night_checkout_end': None,
        'source': 'None',
    }



def _gps_checkin_submit_status(driver_id, vehicle_id=None, project_id=None):
    """Whether GPS check-in can be submitted now (prevents duplicate selfie + misleading local pending)."""
    today = _attendance_local_date()
    now_time = _attendance_local_time()
    driver = Driver.query.options(joinedload(Driver.vehicle)).get(driver_id) if driver_id else None
    if driver_id and not driver:
        return {'ok': False, 'can_submit': False, 'state': 'blocked', 'message': 'Invalid driver.'}
    vehicle = None
    if vehicle_id:
        vehicle = db.session.get(Vehicle, vehicle_id)
    if not vehicle and driver:
        vehicle = driver.vehicle
    if not driver_id and vehicle_id:
        pending_msg = _vehicle_pending_checkout_block_message(0, vehicle)
        if pending_msg:
            return {
                'ok': True,
                'can_submit': False,
                'state': 'blocked',
                'message': pending_msg,
                'vehicle_blocked': True,
            }
        return {'ok': True, 'can_submit': True, 'state': 'allowed', 'message': '', 'vehicle_ok': True}
    if _driver_marked_duty_off_no_checkin(driver_id, today):
        return {
            'ok': True,
            'can_submit': False,
            'state': 'blocked',
            'message': 'Aaj ki date par is driver ki duty Off mark hai — GPS/Camera se attendance nahi lag sakti.',
        }

    tw = _get_effective_time_window(driver=driver, vehicle_id=vehicle_id, project_id=project_id)
    vno = _vehicle_label(vehicle)
    cap = _vehicle_capacity_value(vehicle)

    open_rec_self = _open_gps_driver_attendance_session(driver_id, today)
    if open_rec_self and _gps_marked_attendance_row(open_rec_self):
        ci_t = open_rec_self.check_in.strftime('%H:%M') if open_rec_self.check_in else None
        dt_s, ci_s = _attendance_checkin_stamp(open_rec_self)
        if not _alternate_checkin_window_active(tw, open_rec_self.check_in, now_time):
            return {
                'ok': True,
                'can_submit': False,
                'state': 'complete',
                'message': (
                    'Aap ka check-in complete ho gaya hai. Dubara Mark Attendance ki zaroorat nahi — '
                    'check-out ke liye Check-out page use karein.'
                ),
                'check_in_time': ci_t,
                'has_open_session': True,
                'awaiting_checkout': True,
            }
        return {
            'ok': True,
            'can_submit': False,
            'state': 'checkout_pending',
            'message': (
                f'{vno}: aap ka {dt_s} ka check-in ({ci_s}) abhi check-out pending hai — '
                'pehle us session ka check-out complete karein, phir naya check-in ho ga.'
            ),
            'check_in_time': ci_t,
            'has_open_session': True,
        }

    pending_rec, pending_driver = (None, None)
    if vehicle and getattr(vehicle, 'id', None):
        pending_rec, pending_driver = _vehicle_oldest_pending_checkout(vehicle.id)

    if pending_rec and pending_driver and pending_driver.id != driver_id:
        pd_name = (pending_driver.name or '').strip() or 'Driver'
        dt_s, ci_s = _attendance_checkin_stamp(pending_rec)
        other_shift_on = _alternate_checkin_window_active(tw, pending_rec.check_in, now_time)
        if other_shift_on:
            return {
                'ok': True,
                'can_submit': False,
                'state': 'blocked',
                'message': (
                    f'{vno}: {pd_name} ka check-out abhi pending hai ({dt_s} check-in {ci_s}) — '
                    'un se check-out karwaen, phir aap ka check-in ho ga.'
                ),
                'vehicle_blocked': True,
                'pending_driver_name': pd_name,
            }
        return {
            'ok': True,
            'can_submit': False,
            'state': 'blocked',
            'message': (
                f'Is gari ke dusre driver {pd_name} ne {dt_s} ko {ci_s} par attendance laga di hai. '
                'Jab tak wo check-out nahi karte, aap ki attendance nahi lag sakti.'
            ),
            'vehicle_blocked': True,
            'pending_driver_name': pd_name,
        }

    blocked_msg = _manual_checkin_blocked_by_vehicle_rules(driver_id, vehicle, today)
    if blocked_msg:
        return {'ok': True, 'can_submit': False, 'state': 'blocked', 'message': blocked_msg}

    cap = _vehicle_capacity_value(vehicle)
    if _count_driver_segments_with_checkin(driver_id, today) >= cap:
        last = (
            DriverAttendance.query.filter(
                DriverAttendance.driver_id == driver_id,
                DriverAttendance.attendance_date == today,
                DriverAttendance.check_in.isnot(None),
            )
            .order_by(DriverAttendance.attendance_segment.desc(), DriverAttendance.id.desc())
            .first()
        )
        ci_t = last.check_in.strftime('%H:%M') if last and last.check_in else None
        return {
            'ok': True,
            'can_submit': False,
            'state': 'complete',
            'message': (
                'Aaj ka check-in pehle ho chuka hai'
                + ((' (' + format_time_ampm(last.check_in) + ')') if last and last.check_in else '')
                + '. Dubara selfie ya check-in ki zaroorat nahi.'
            ),
            'check_in_time': ci_t,
            'segments_used': int(_count_driver_segments_with_checkin(driver_id, today)),
            'capacity': cap,
        }
    shift = (driver.shift or '').strip().lower() if driver else ''
    ci_ok, ci_msg = _gps_checkin_shift_window_ok(shift, now_time, tw, driver=driver, vehicle=vehicle)
    if not ci_ok:
        return {'ok': True, 'can_submit': False, 'state': 'blocked', 'message': ci_msg}
    return {'ok': True, 'can_submit': True, 'state': 'allowed', 'message': ''}


def _gps_checkout_submit_status(driver_id, vehicle_id=None, project_id=None):
    """Whether GPS check-out can be submitted now."""
    today = _attendance_local_date()
    driver = db.session.get(Driver, driver_id)
    if not driver:
        return {'ok': False, 'can_submit': False, 'state': 'blocked', 'message': 'Invalid driver.'}
    open_rec = _open_gps_driver_attendance_for_checkout(driver_id, today)
    if open_rec:
        ci_t = open_rec.check_in.strftime('%H:%M') if open_rec.check_in else None
        return {
            'ok': True,
            'can_submit': True,
            'state': 'checkout_pending',
            'message': '',
            'check_in_time': ci_t,
            'has_open_session': True,
        }
    done = (
        DriverAttendance.query.filter(
            DriverAttendance.driver_id == driver_id,
            DriverAttendance.attendance_date == today,
            DriverAttendance.check_in.isnot(None),
            DriverAttendance.check_out.isnot(None),
        )
        .order_by(DriverAttendance.attendance_segment.desc(), DriverAttendance.id.desc())
        .first()
    )
    if done and _gps_marked_attendance_row(done):
        co_t = done.check_out.strftime('%H:%M') if done.check_out else None
        return {
            'ok': True,
            'can_submit': False,
            'state': 'complete',
            'message': (
                'Aaj ka check-out pehle ho chuka hai'
                + ((' (' + co_t + ')') if co_t else '')
                + '. Dubara selfie ki zaroorat nahi.'
            ),
            'check_out_time': co_t,
            'check_in_time': done.check_in.strftime('%H:%M') if done.check_in else None,
        }
    return {
        'ok': True,
        'can_submit': False,
        'state': 'no_checkin',
        'message': 'Pehle Mark Attendance se check-in karein, phir check-out karein.',
    }




def _attendance_media_payload(rec, kind):
    path_val = ''
    time_val = None
    if kind == 'checkin':
        path_val = (rec.check_in_photo_path or '').strip()
        time_val = rec.check_in
    else:
        path_val = (rec.check_out_photo_path or '').strip()
        time_val = rec.check_out
    media_url = media_url_filter(path_val) if path_val else None
    return {
        'has_media': bool(path_val),
        'media_path': path_val or None,
        'media_url': media_url,
        'uploaded': bool(path_val),
        'time': time_val.strftime('%H:%M') if time_val else None,
    }



def _decode_attendance_photo_b64(photo_b64):
    """Decode a data-URL or raw base64 string into image bytes. Returns None if invalid."""
    import base64

    if not (photo_b64 or '').strip():
        return None
    s = photo_b64.strip()
    m = re.match(r'data:image/[^;]+;base64,(.+)', s, re.DOTALL)
    raw_b64 = m.group(1) if m else s
    try:
        return base64.b64decode(raw_b64)
    except Exception:
        return None


def _upload_attendance_image_bytes_with_fallback(data, folder='attendance'):
    """Upload via R2 (WebP); if R2 misconfigured or fails, save JPEG bytes under uploads/."""
    if not data:
        return None
    try:
        return upload_image_bytes(data, folder=folder)
    except Exception as exc:
        app.logger.warning('Attendance photo R2 upload failed (%s), using disk fallback', exc)
    try:
        root = os.path.dirname(os.path.abspath(__file__))
        upload_dir = os.path.join(root, 'uploads', folder)
        os.makedirs(upload_dir, exist_ok=True)
        fname = uuid.uuid4().hex + '.jpg'
        fpath = os.path.join(upload_dir, fname)
        with open(fpath, 'wb') as out:
            out.write(data)
        return '/uploads/%s/%s' % (folder, fname)
    except Exception as exc2:
        app.logger.exception('Attendance photo local save failed: %s', exc2)
        return None


def _upload_attendance_photo_from_form_or_b64(photo_file, photo_b64, *, required=True):
    """Resolve a werkzeug FileStorage and/or base64 payload to a stored path URL."""
    data = None
    if photo_file and getattr(photo_file, 'filename', None):
        photo_file.stream.seek(0)
        data = photo_file.read()
    elif photo_b64:
        data = _decode_attendance_photo_b64(photo_b64)
    photo_path = _upload_attendance_image_bytes_with_fallback(data) if data else None
    if not photo_path:
        if required:
            raise ValueError('Image required.')
        return None
    return photo_path






def _build_driver_daily_attendance_report_payload(
    month,
    year,
    project_id,
    district_id,
    vehicle_id=None,
    driver_id_filter=0,
    search='',
    scope_projects=None,
    scope_districts=None,
    scope_vehicles=None,
    scope_shifts=None,
):
    """Build day-wise attendance grid data for page render or Excel export."""
    from calendar import monthrange

    scope_projects = scope_projects or []
    scope_districts = scope_districts or []
    scope_vehicles = scope_vehicles or []
    scope_shifts = scope_shifts or []
    status_columns = ['Present', 'Absent', 'Leave', 'Late', 'Half-Day', 'Off']

    if not month or not year or not project_id or not district_id:
        return None

    _, ndays = monthrange(year, month)
    start_d = date(year, month, 1)
    end_d = date(year, month, ndays)
    day_headers = [date(year, month, d) for d in range(1, ndays + 1)]
    report_title = f'Day Wise Attendance — {start_d.strftime("%d-%b-%Y")} to {end_d.strftime("%d-%b-%Y")}'
    empty_totals = {s: 0 for s in status_columns}

    # Same driver pool + assignment rules as TRA Attendance Sheet
    eligible_ids = _tra_report_driver_ids_for_month(
        start_d,
        end_d,
        project_id,
        district_id,
        vehicle_id,
        None,
        scope_projects,
        scope_districts,
        scope_vehicles,
        scope_shifts,
    )
    if driver_id_filter:
        eligible_ids = {driver_id_filter} if driver_id_filter in eligible_ids else set()
    if not eligible_ids:
        return {
            'report': [],
            'day_headers': day_headers,
            'report_title': report_title,
            'ndays': ndays,
            'grand_totals': empty_totals,
            'status_columns': status_columns,
        }

    drivers = (
        Driver.query.filter(Driver.id.in_(eligible_ids))
        .order_by(Driver.name)
        .all()
    )
    tra_cache = _TraMonthCache(start_d, end_d, eligible_ids)
    tra_cache.load()
    tra_cache.prewarm(drivers)

    att_by_driver = {did: [] for did in eligible_ids}
    att_rows = (
        DriverAttendance.query.options(
            joinedload(DriverAttendance.driver).joinedload(Driver.vehicle),
        )
        .filter(
            DriverAttendance.driver_id.in_(eligible_ids),
            DriverAttendance.attendance_date >= start_d,
            DriverAttendance.attendance_date <= end_d,
        )
        .order_by(DriverAttendance.attendance_date, DriverAttendance.attendance_segment)
        .all()
    )
    for r in att_rows:
        if not _attendance_record_counts_in_report(r):
            continue
        att_by_driver[r.driver_id].append(r)

    report = []
    grand_totals = {s: 0 for s in status_columns}
    for d in drivers:
        if not tra_cache.driver_on_duty_in_month(d):
            continue

        records = att_by_driver.get(d.id, [])
        for segment in tra_cache.segments(d):
            eff = segment['eff']
            seg_start = segment['segment_start']
            seg_end = segment['segment_end']
            eff_vehicle_id = eff.get('vehicle_id')

            if search and not _tra_search_matches(search, d, eff):
                continue
            if not _tra_driver_matches_scope(
                eff,
                project_id,
                district_id,
                vehicle_id,
                None,
                scope_projects,
                scope_districts,
                scope_vehicles,
                scope_shifts,
            ):
                continue

            vehicle = eff.get('vehicle')
            display_district = eff.get('district')
            display_project = eff.get('project')
            grid = {}
            status_totals = {s: 0 for s in status_columns}

            for r in records:
                att_d = r.attendance_date
                if att_d < seg_start or att_d > seg_end:
                    continue
                if not tra_cache.driver_duty_on_date(d.id, att_d):
                    continue
                if tra_cache.vehicle_id_on_date(d.id, att_d) != eff_vehicle_id:
                    continue

                day_num = att_d.day
                slot = _attendance_daily_slot_key(r, driver=d, vehicle=vehicle)
                status = (r.status or '').strip()
                abbr = _attendance_status_abbr(status)
                if day_num not in grid:
                    grid[day_num] = {}
                grid[day_num][slot] = {'v': abbr, 'tip': _attendance_daily_cell_tooltip(r)}
                if status in status_totals:
                    status_totals[status] += 1
                    grand_totals[status] += 1

            left_rec = tra_cache.left_rec(d.id)
            left_date = None
            if left_rec and left_rec.change_date and start_d <= left_rec.change_date <= end_d:
                left_date = left_rec.change_date
            _daily_attendance_fill_segment_boundary_cells(
                grid, year, month, seg_start, seg_end, segment, d, left_date, tra_cache,
            )
            lifecycle_badges = _daily_attendance_row_lifecycle_badges(
                d, segment, tra_cache, start_d, end_d, seg_start,
            )

            report.append({
                'district_name': (
                    (display_district.name if display_district else None)
                    or (vehicle.district.name if vehicle and vehicle.district else None)
                    or '-'
                ),
                'project_name': (
                    (display_project.name if display_project else None)
                    or (vehicle.project.name if vehicle and vehicle.project else None)
                    or '-'
                ),
                'vehicle_no': (vehicle.vehicle_no if vehicle else '-') or '-',
                'shift': (eff.get('shift') or d.shift or '-') or '-',
                'driver_name': (d.name or '-') or '-',
                'date_of_leaving': left_date,
                'lifecycle_badges': lifecycle_badges,
                'month_present_days': _count_month_present_days(grid),
                'days_in_month': ndays,
                'grid': grid,
                'status_totals': status_totals,
            })

    report.sort(key=lambda row: (row.get('driver_name') or '', row.get('vehicle_no') or ''))

    return {
        'report': report,
        'day_headers': day_headers,
        'report_title': report_title,
        'ndays': ndays,
        'grand_totals': grand_totals,
        'status_columns': status_columns,
    }


def _daily_attendance_slot_day_totals(report, ndays):
    """Per-day M/E counts across all drivers (for Excel total row)."""
    totals = {(day, slot): 0 for day in range(1, ndays + 1) for slot in ('M', 'E')}
    for row in report:
        grid = row.get('grid') or {}
        for day in range(1, ndays + 1):
            slots = grid.get(day) or {}
            for slot in ('M', 'E'):
                cell = slots.get(slot) or {}
                if _daily_attendance_grid_cell_skip_totals(cell):
                    continue
                if _daily_attendance_grid_cell_value(cell):
                    totals[(day, slot)] += 1
    return totals


def _generate_driver_daily_attendance_excel(payload):
    """Export formatted Day Wise Attendance Report (.xlsx)."""
    report = payload['report']
    day_headers = payload['day_headers']
    report_title = payload['report_title']
    ndays = payload['ndays']
    grand_totals = payload['grand_totals']
    status_columns = payload['status_columns']

    info_cols = 6
    day_slot_cols = ndays * 2
    sum_cols = len(status_columns)
    last_col = info_cols + day_slot_cols + sum_cols - 1
    hdr_row_top = 2
    hdr_row_sub = 3
    data_start = 4

    output = BytesIO()
    wb = xlsxwriter.Workbook(output, {'in_memory': True})
    ws = wb.add_worksheet('Day Wise Attendance')

    dark = '#1B4332'
    dark2 = '#2D6A4F'
    white = '#FFFFFF'
    border_white = '#FFFFFF'
    border_grey = '#CBD5E1'
    zebra = '#F8FAFC'
    total_bg = '#D1FAE5'

    title_fmt = wb.add_format({
        'bold': True, 'font_size': 16, 'font_color': white,
        'bg_color': dark, 'align': 'center', 'valign': 'vcenter',
    })
    sub_fmt = wb.add_format({
        'bold': True, 'font_size': 12, 'font_color': white,
        'bg_color': dark2, 'align': 'center', 'valign': 'vcenter',
    })
    hdr_fmt = wb.add_format({
        'bold': True, 'font_color': white, 'bg_color': dark,
        'align': 'center', 'valign': 'vcenter', 'border': 1, 'border_color': border_white,
    })
    cell_fmt = wb.add_format({
        'border': 1, 'border_color': border_grey, 'align': 'center', 'valign': 'vcenter',
    })
    cell_left_fmt = wb.add_format({
        'border': 1, 'border_color': border_grey, 'align': 'left', 'valign': 'vcenter',
    })
    zebra_fmt = wb.add_format({
        'border': 1, 'border_color': border_grey, 'align': 'center', 'valign': 'vcenter',
        'bg_color': zebra,
    })
    zebra_left_fmt = wb.add_format({
        'border': 1, 'border_color': border_grey, 'align': 'left', 'valign': 'vcenter',
        'bg_color': zebra,
    })
    ex_cell_fmt = wb.add_format({
        'border': 1, 'border_color': border_grey, 'align': 'center', 'valign': 'vcenter',
        'font_color': '#64748B', 'bg_color': '#E2E8F0', 'bold': True,
    })
    ex_cell_zebra_fmt = wb.add_format({
        'border': 1, 'border_color': border_grey, 'align': 'center', 'valign': 'vcenter',
        'font_color': '#64748B', 'bg_color': '#CBD5E1', 'bold': True,
    })
    inactive_cell_fmt = wb.add_format({
        'border': 1, 'border_color': border_grey, 'align': 'center', 'valign': 'vcenter',
        'font_color': '#94A3B8', 'bg_color': '#F8FAFC', 'italic': True,
    })
    inactive_cell_zebra_fmt = wb.add_format({
        'border': 1, 'border_color': border_grey, 'align': 'center', 'valign': 'vcenter',
        'font_color': '#94A3B8', 'bg_color': '#F1F5F9', 'italic': True,
    })
    transfer_cell_fmt = wb.add_format({
        'border': 1, 'border_color': border_grey, 'align': 'center', 'valign': 'vcenter',
        'font_color': '#1D4ED8', 'bg_color': '#DBEAFE', 'bold': True,
    })
    transfer_cell_zebra_fmt = wb.add_format({
        'border': 1, 'border_color': border_grey, 'align': 'center', 'valign': 'vcenter',
        'font_color': '#1D4ED8', 'bg_color': '#BFDBFE', 'bold': True,
    })
    total_fmt = wb.add_format({
        'bold': True, 'bg_color': total_bg, 'border': 1, 'border_color': border_grey,
        'align': 'center', 'valign': 'vcenter',
    })
    total_left_fmt = wb.add_format({
        'bold': True, 'bg_color': total_bg, 'border': 1, 'border_color': border_grey,
        'align': 'left', 'valign': 'vcenter',
    })

    ws.set_row(0, 22)
    ws.set_row(1, 18)
    ws.merge_range(0, 0, 0, last_col, 'Day Wise Attendance Report', title_fmt)
    ws.merge_range(1, 0, 1, last_col, report_title, sub_fmt)

    info_labels = ['Sr', 'District', 'Project', 'Vehicle', 'Shift', 'Driver']
    for c, label in enumerate(info_labels):
        ws.merge_range(hdr_row_top, c, hdr_row_sub, c, label, hdr_fmt)

    for i, _d in enumerate(day_headers):
        c0 = info_cols + i * 2
        ws.merge_range(hdr_row_top, c0, hdr_row_top, c0 + 1, f'{_d.day:02d}', hdr_fmt)
        ws.write(hdr_row_sub, c0, 'M', hdr_fmt)
        ws.write(hdr_row_sub, c0 + 1, 'E', hdr_fmt)

    for j, st in enumerate(status_columns):
        c = info_cols + day_slot_cols + j
        ws.merge_range(hdr_row_top, c, hdr_row_sub, c, st, hdr_fmt)

    slot_totals = _daily_attendance_slot_day_totals(report, ndays)
    row_idx = data_start
    for i, row in enumerate(report, 1):
        use_zebra = i % 2 == 0
        cf = zebra_fmt if use_zebra else cell_fmt
        lf = zebra_left_fmt if use_zebra else cell_left_fmt
        driver_label = (
            f"{row['driver_name']} ({row['month_present_days']}/{row['days_in_month']})"
        )
        ws.write(row_idx, 0, i, cf)
        ws.write(row_idx, 1, row['district_name'], lf)
        ws.write(row_idx, 2, row['project_name'], lf)
        ws.write(row_idx, 3, row['vehicle_no'], lf)
        ws.write(row_idx, 4, row['shift'], lf)
        ws.write(row_idx, 5, driver_label, lf)
        grid = row.get('grid') or {}
        for day in range(1, ndays + 1):
            slots = grid.get(day) or {}
            for si, slot in enumerate(('M', 'E')):
                cell = slots.get(slot) or {}
                v = _daily_attendance_grid_cell_value(cell)
                if isinstance(cell, dict) and cell.get('kind') == 'left':
                    slot_cf = ex_cell_zebra_fmt if use_zebra else ex_cell_fmt
                elif isinstance(cell, dict) and cell.get('kind') == 'inactive':
                    slot_cf = inactive_cell_zebra_fmt if use_zebra else inactive_cell_fmt
                elif isinstance(cell, dict) and cell.get('kind') == 'transfer':
                    slot_cf = transfer_cell_zebra_fmt if use_zebra else transfer_cell_fmt
                else:
                    slot_cf = cf
                ws.write(row_idx, info_cols + (day - 1) * 2 + si, v or '', slot_cf)
        for j, st in enumerate(status_columns):
            ws.write(row_idx, info_cols + day_slot_cols + j, row['status_totals'].get(st, 0), cf)
        row_idx += 1

    ws.write(row_idx, 0, 'Total', total_left_fmt)
    for c in range(1, info_cols):
        ws.write(row_idx, c, '', total_fmt)
    for day in range(1, ndays + 1):
        for si, slot in enumerate(('M', 'E')):
            ws.write(row_idx, info_cols + (day - 1) * 2 + si, slot_totals.get((day, slot), 0), total_fmt)
    for j, st in enumerate(status_columns):
        ws.write(row_idx, info_cols + day_slot_cols + j, grand_totals.get(st, 0), total_fmt)

    data_end = row_idx
    att_first = info_cols
    att_last = info_cols + day_slot_cols - 1
    if data_end >= data_start and att_last >= att_first:
        green_fmt = wb.add_format({'font_color': '#166534', 'border': 1, 'border_color': border_grey})
        red_fmt = wb.add_format({'font_color': '#991b1b', 'border': 1, 'border_color': border_grey})
        ws.conditional_format(data_start, att_first, data_end, att_last, {
            'type': 'text', 'criteria': 'containing', 'value': 'P', 'format': green_fmt,
        })
        ws.conditional_format(data_start, att_first, data_end, att_last, {
            'type': 'text', 'criteria': 'containing', 'value': 'A', 'format': red_fmt,
        })

    ws.freeze_panes(data_start, info_cols)
    ws.autofilter(hdr_row_sub, 0, data_end, last_col)

    ws.set_column(0, 0, 5)
    ws.set_column(1, 4, 14)
    ws.set_column(5, 5, 32)
    if day_slot_cols:
        ws.set_column(info_cols, info_cols + day_slot_cols - 1, 3.5)
    if sum_cols:
        ws.set_column(info_cols + day_slot_cols, last_col, 9)

    wb.close()
    output.seek(0)
    fname = f"Day_Wise_Attendance_{pk_now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        download_name=fname,
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )




def _tra_pack_rejoin_assignment(rejoin_rec):
    """Assignment captured on Driver Rejoin form."""
    return {
        'vehicle': rejoin_rec.new_vehicle,
        'vehicle_id': rejoin_rec.new_vehicle_id,
        'project': rejoin_rec.new_project,
        'district': rejoin_rec.new_district,
        'shift': rejoin_rec.new_shift,
    }


def _tra_status_events_up_to(driver_id, on_date, status_events=None):
    """Left/rejoin timeline up to and including on_date (ascending)."""
    if status_events is not None:
        return [ev for ev in status_events if ev.change_date and ev.change_date <= on_date]
    return (
        DriverStatusChange.query.filter(
            DriverStatusChange.driver_id == driver_id,
            DriverStatusChange.action_type.in_(['left', 'rejoin']),
            DriverStatusChange.change_date <= on_date,
        )
        .order_by(DriverStatusChange.change_date.asc(), DriverStatusChange.id.asc())
        .all()
    )


def _tra_applicable_rejoin_on_date(driver_id, on_date, status_events=None):
    """Latest rejoin effective on on_date (must follow a prior left)."""
    events = _tra_status_events_up_to(driver_id, on_date, status_events)
    last_left = None
    applicable = None
    for ev in events:
        if ev.action_type == 'left':
            last_left = ev
            applicable = None
        elif (
            ev.action_type == 'rejoin'
            and last_left
            and ev.change_date
            and last_left.change_date
            and ev.change_date > last_left.change_date
        ):
            applicable = ev
    if applicable and applicable.change_date and applicable.change_date <= on_date:
        return applicable
    return None


def _tra_driver_duty_active_on_date(driver, on_date, status_events=None):
    """On duty on a calendar day — respects assign_date and left/rejoin cycles."""
    if driver.assign_date and on_date < driver.assign_date:
        return False
    events = _tra_status_events_up_to(driver.id, on_date, status_events)
    last_left = None
    last_rejoin = None
    for ev in events:
        if ev.action_type == 'left' and ev.change_date and ev.change_date <= on_date:
            last_left = ev.change_date
        elif ev.action_type == 'rejoin' and ev.change_date and ev.change_date <= on_date:
            last_rejoin = ev.change_date
    if last_left is None:
        return True
    if last_rejoin and last_rejoin >= last_left:
        return True
    if (
        driver.vehicle_id
        and driver.assign_date
        and driver.assign_date > last_left
        and driver.assign_date <= on_date
    ):
        return True
    return False


def _tra_effective_assignment(driver, left_rec=None, rejoin_rec=None):
    """Project/vehicle/district/shift for TRA row when Job Left cleared current assignment."""
    if driver and driver.vehicle_id and driver.vehicle:
        return {
            'vehicle': driver.vehicle,
            'vehicle_id': driver.vehicle_id,
            'project': driver.project,
            'district': driver.district,
            'shift': driver.shift,
        }
    if rejoin_rec and rejoin_rec.new_vehicle_id:
        return _tra_pack_rejoin_assignment(rejoin_rec)
    if left_rec:
        veh = left_rec.left_vehicle
        return {
            'vehicle': veh,
            'vehicle_id': left_rec.left_vehicle_id,
            'project': left_rec.left_project,
            'district': left_rec.left_district,
            'shift': left_rec.left_shift,
        }
    return {
        'vehicle': None,
        'vehicle_id': None,
        'project': None,
        'district': None,
        'shift': None,
    }


def _tra_pack_transfer_side(transfer, side):
    if side == 'old':
        return {
            'vehicle': transfer.old_vehicle,
            'vehicle_id': transfer.old_vehicle_id,
            'project': transfer.old_project,
            'district': transfer.old_district,
            'shift': transfer.old_shift,
        }
    return {
        'vehicle': transfer.new_vehicle,
        'vehicle_id': transfer.new_vehicle_id,
        'project': transfer.new_project,
        'district': transfer.new_district,
        'shift': transfer.new_shift,
    }


def _tra_non_shift_transfer_filter():
    return db.or_(DriverTransfer.is_shift_only == False, DriverTransfer.is_shift_only.is_(None))


def _tra_assignment_for_report_month(driver, left_rec, start_d, end_d):
    """Vehicle/project/district/shift as they were during the report month (not current assignment)."""
    rejoin_during = DriverStatusChange.query.filter(
        DriverStatusChange.driver_id == driver.id,
        DriverStatusChange.action_type == 'rejoin',
        DriverStatusChange.change_date >= start_d,
        DriverStatusChange.change_date <= end_d,
    ).order_by(DriverStatusChange.change_date.desc(), DriverStatusChange.id.desc()).first()
    if not driver.vehicle_id and rejoin_during and rejoin_during.new_vehicle_id:
        return _tra_pack_rejoin_assignment(rejoin_during)
    if not driver.vehicle_id and left_rec:
        if left_rec.change_date and left_rec.change_date > end_d:
            return _tra_effective_assignment(driver, left_rec)
        if left_rec.change_date and left_rec.change_date < start_d:
            return _tra_effective_assignment(driver, left_rec)
        if left_rec.change_date and start_d <= left_rec.change_date <= end_d:
            return _tra_effective_assignment(driver, left_rec)

    sf = _tra_non_shift_transfer_filter()

    future_t = DriverTransfer.query.filter(
        DriverTransfer.driver_id == driver.id,
        sf,
        DriverTransfer.transfer_date > end_d,
    ).order_by(DriverTransfer.transfer_date.asc(), DriverTransfer.id.asc()).first()
    if future_t:
        return _tra_pack_transfer_side(future_t, 'old')

    during_t = DriverTransfer.query.filter(
        DriverTransfer.driver_id == driver.id,
        sf,
        DriverTransfer.transfer_date >= start_d,
        DriverTransfer.transfer_date <= end_d,
    ).order_by(DriverTransfer.transfer_date.desc(), DriverTransfer.id.desc()).first()
    if during_t:
        return _tra_pack_transfer_side(during_t, 'new')

    before_t = DriverTransfer.query.filter(
        DriverTransfer.driver_id == driver.id,
        sf,
        DriverTransfer.transfer_date < start_d,
    ).order_by(DriverTransfer.transfer_date.desc(), DriverTransfer.id.desc()).first()
    if before_t:
        return _tra_pack_transfer_side(before_t, 'new')

    return _tra_effective_assignment(driver, left_rec)


def _tra_assignment_as_of(driver, left_rec, as_of_date):
    """Vehicle assignment on a specific calendar date."""
    if not driver.vehicle_id and left_rec:
        if left_rec.change_date and left_rec.change_date > as_of_date:
            return _tra_effective_assignment(driver, left_rec)
        if left_rec.change_date and left_rec.change_date <= as_of_date:
            return _tra_effective_assignment(driver, left_rec)

    sf = _tra_non_shift_transfer_filter()
    future_t = DriverTransfer.query.filter(
        DriverTransfer.driver_id == driver.id,
        sf,
        DriverTransfer.transfer_date > as_of_date,
    ).order_by(DriverTransfer.transfer_date.asc(), DriverTransfer.id.asc()).first()
    if future_t:
        return _tra_pack_transfer_side(future_t, 'old')

    past_t = DriverTransfer.query.filter(
        DriverTransfer.driver_id == driver.id,
        sf,
        DriverTransfer.transfer_date <= as_of_date,
    ).order_by(DriverTransfer.transfer_date.desc(), DriverTransfer.id.desc()).first()
    if past_t:
        if past_t.transfer_date == as_of_date:
            return _tra_pack_transfer_side(past_t, 'old')
        return _tra_pack_transfer_side(past_t, 'new')

    return _tra_effective_assignment(driver, left_rec)


def _tra_driver_duty_on_date(driver, left_rec, on_date):
    return _tra_driver_duty_active_on_date(driver, on_date)


def _tra_vehicle_id_from_transfer_list(transfers, on_date, driver=None, left_rec=None):
    """Vehicle on a calendar date from non-shift transfer history (transfer day = last day on old vehicle)."""
    future = [t for t in transfers if t.transfer_date > on_date]
    if future:
        return future[0].old_vehicle_id
    past = [t for t in transfers if t.transfer_date <= on_date]
    if past:
        t = past[-1]
        if t.transfer_date == on_date:
            return t.old_vehicle_id
        return t.new_vehicle_id
    if driver and driver.vehicle_id:
        return driver.vehicle_id
    if left_rec:
        return left_rec.left_vehicle_id
    return None


def _tra_build_in_month_transfer_segments(
    transfers,
    start_d,
    end_d,
    driver,
    left_date,
    rejoin_rec,
):
    """
    Build vehicle segments for in-month transfers.
    Transfer date D is the last day on the old vehicle; new vehicle starts D+1.
    """
    segments = []
    cursor = start_d
    for t in transfers:
        t_date = t.transfer_date
        old_eff = _tra_pack_transfer_side(t, 'old')
        if t_date >= cursor and old_eff.get('vehicle_id'):
            seg_s, seg_e = _tra_clamp_segment_bounds(
                cursor, t_date, start_d, end_d, driver.assign_date, left_date,
            )
            if rejoin_rec and rejoin_rec.change_date and seg_s and rejoin_rec.change_date > seg_s:
                seg_s = max(seg_s, rejoin_rec.change_date)
            if seg_s is not None and seg_s <= seg_e:
                segments.append({
                    'eff': old_eff,
                    'segment_start': seg_s,
                    'segment_end': seg_e,
                    'transfer_in': None,
                    'transfer_out': t,
                })
        cursor = t_date + timedelta(days=1)

    if transfers and cursor <= end_d:
        last_t = transfers[-1]
        new_eff = _tra_pack_transfer_side(last_t, 'new')
        if new_eff.get('vehicle_id'):
            seg_s, seg_e = _tra_clamp_segment_bounds(
                cursor, end_d, start_d, end_d, driver.assign_date, left_date,
            )
            if rejoin_rec and rejoin_rec.change_date and seg_s and rejoin_rec.change_date > seg_s:
                seg_s = max(seg_s, rejoin_rec.change_date)
            if seg_s is not None and seg_s <= seg_e:
                segments.append({
                    'eff': new_eff,
                    'segment_start': seg_s,
                    'segment_end': seg_e,
                    'transfer_in': last_t,
                    'transfer_out': None,
                })
    return segments


def _tra_clamp_segment_bounds(seg_start, seg_end, month_start, month_end, assign_date, left_date):
    seg_s = max(seg_start, month_start)
    seg_e = min(seg_end, month_end)
    if assign_date and assign_date > seg_s:
        if assign_date > seg_e:
            return None, None
        seg_s = assign_date
    if left_date and left_date < seg_e:
        seg_e = left_date
    if seg_s > seg_e:
        return None, None
    return seg_s, seg_e


def _tra_driver_vehicle_segments(driver, left_rec, start_d, end_d):
    """One segment per vehicle stint when driver transfers mid-month."""
    sf = _tra_non_shift_transfer_filter()
    left_date = None
    if left_rec and left_rec.change_date and start_d <= left_rec.change_date <= end_d:
        left_date = left_rec.change_date

    rejoin_rec = DriverStatusChange.query.filter(
        DriverStatusChange.driver_id == driver.id,
        DriverStatusChange.action_type == 'rejoin',
        DriverStatusChange.change_date >= start_d,
        DriverStatusChange.change_date <= end_d,
    ).order_by(DriverStatusChange.change_date.desc()).first()

    transfers = DriverTransfer.query.filter(
        DriverTransfer.driver_id == driver.id,
        sf,
        DriverTransfer.transfer_date >= start_d,
        DriverTransfer.transfer_date <= end_d,
    ).order_by(DriverTransfer.transfer_date.asc(), DriverTransfer.id.asc()).all()

    segments = []

    if not transfers:
        eff = _tra_assignment_for_report_month(driver, left_rec, start_d, end_d)
        if not eff.get('vehicle_id'):
            return []
        seg_s, seg_e = _tra_clamp_segment_bounds(
            start_d, end_d, start_d, end_d, driver.assign_date, left_date
        )
        if rejoin_rec and rejoin_rec.change_date and seg_s and rejoin_rec.change_date > seg_s:
            seg_s = max(seg_s, rejoin_rec.change_date)
        if seg_s is None or seg_s > seg_e:
            return []
        segments.append({
            'eff': eff,
            'segment_start': seg_s,
            'segment_end': seg_e,
            'transfer_in': None,
            'transfer_out': None,
        })
        return segments

    return _tra_build_in_month_transfer_segments(
        transfers, start_d, end_d, driver, left_date, rejoin_rec,
    )


def _tra_count_drivers_on_vehicle_day(vehicle_id, on_date, month_start, month_end):
    """How many drivers were assigned to this vehicle on a calendar date."""
    candidate_ids = _tra_vehicle_driver_ids_in_month(vehicle_id, month_start, month_end)
    count = 0
    for pid in candidate_ids:
        partner = db.session.get(Driver, pid)
        if not partner:
            continue
        p_left = _tra_status_rec(pid, 'left')
        if not _tra_driver_duty_on_date(partner, p_left, on_date):
            continue
        eff = _tra_assignment_as_of(partner, p_left, on_date)
        if eff.get('vehicle_id') == vehicle_id:
            count += 1
    return count


def _tra_segment_solo_days(vehicle_id, seg_start, seg_end, month_start, month_end):
    """Days in segment with only one driver on vehicle (eligible for double duty)."""
    solo = 0
    d = seg_start
    while d <= seg_end:
        if _tra_count_drivers_on_vehicle_day(vehicle_id, d, month_start, month_end) == 1:
            solo += 1
        d += timedelta(days=1)
    return solo


def _tra_search_matches(search, driver, eff):
    tokens = [t for t in (search or '').split() if t]
    if not tokens:
        return True
    veh_no = eff.get('vehicle').vehicle_no if eff.get('vehicle') else ''
    haystacks = [driver.name or '', driver.driver_id or '', veh_no]
    return all(
        any(tok.lower() in (h or '').lower() for h in haystacks)
        for tok in tokens
    )


def _tra_driver_matches_scope(
    eff,
    project_id,
    district_id,
    vehicle_id,
    shift,
    scope_projects,
    scope_districts,
    scope_vehicles,
    scope_shifts,
):
    """Whether effective assignment matches TRA report filters."""
    veh = eff.get('vehicle')
    veh_id = eff.get('vehicle_id')
    proj_id = eff.get('project').id if eff.get('project') else None
    dist_id = eff.get('district').id if eff.get('district') else (
        veh.district_id if veh else None
    )
    eff_shift = eff.get('shift')

    if scope_projects and proj_id not in scope_projects and not (
        veh and veh.project_id in scope_projects
    ):
        return False
    if scope_districts and dist_id not in scope_districts:
        return False
    if scope_vehicles and veh_id not in scope_vehicles:
        return False
    if scope_shifts and eff_shift not in scope_shifts:
        return False
    if project_id and proj_id != project_id and not (veh and veh.project_id == project_id):
        return False
    if district_id and dist_id != district_id:
        return False
    if vehicle_id and veh_id != vehicle_id:
        return False
    if shift and eff_shift != shift:
        return False
    return True


def _tra_driver_on_duty_in_month(driver_id, start_d, end_d, assign_date):
    """Whether driver belongs on TRA sheet for this calendar month."""
    if assign_date and assign_date > end_d:
        has_att = DriverAttendance.query.filter(
            DriverAttendance.driver_id == driver_id,
            DriverAttendance.attendance_date >= start_d,
            DriverAttendance.attendance_date <= end_d,
        ).first()
        if has_att:
            return True
        sf = _tra_non_shift_transfer_filter()
        transferred_in_month = DriverTransfer.query.filter(
            DriverTransfer.driver_id == driver_id,
            sf,
            DriverTransfer.transfer_date >= start_d,
            DriverTransfer.transfer_date <= end_d,
        ).first()
        if transferred_in_month:
            return True
        rejoined_in_month = DriverStatusChange.query.filter(
            DriverStatusChange.driver_id == driver_id,
            DriverStatusChange.action_type == 'rejoin',
            DriverStatusChange.change_date >= start_d,
            DriverStatusChange.change_date <= end_d,
        ).first()
        if rejoined_in_month:
            return True
        left_in_month = DriverStatusChange.query.filter(
            DriverStatusChange.driver_id == driver_id,
            DriverStatusChange.action_type == 'left',
            DriverStatusChange.change_date >= start_d,
            DriverStatusChange.change_date <= end_d,
        ).first()
        return left_in_month is not None

    left_before_month = DriverStatusChange.query.filter(
        DriverStatusChange.driver_id == driver_id,
        DriverStatusChange.action_type == 'left',
        DriverStatusChange.change_date < start_d,
    ).order_by(DriverStatusChange.change_date.desc()).first()

    if left_before_month:
        driver = db.session.get(Driver, driver_id)
        if (
            driver
            and driver.vehicle_id
            and assign_date
            and assign_date > left_before_month.change_date
            and assign_date <= end_d
        ):
            return True
        rejoined = DriverStatusChange.query.filter(
            DriverStatusChange.driver_id == driver_id,
            DriverStatusChange.action_type == 'rejoin',
            DriverStatusChange.change_date > left_before_month.change_date,
            DriverStatusChange.change_date <= end_d,
        ).first()
        if not rejoined:
            return False

    return True


def _tra_report_driver_ids_for_month(
    start_d,
    end_d,
    project_id,
    district_id,
    vehicle_id,
    shift,
    scope_projects,
    scope_districts,
    scope_vehicles,
    scope_shifts,
):
    """Drivers on TRA sheet: currently assigned or who worked during the selected month."""
    ids = set()

    assigned_q = Driver.query.filter(
        Driver.vehicle_id.isnot(None),
    ).outerjoin(Vehicle, Driver.vehicle_id == Vehicle.id)
    if scope_projects:
        assigned_q = assigned_q.filter(
            db.or_(Driver.project_id.in_(scope_projects), Vehicle.project_id.in_(scope_projects))
        )
    if scope_districts:
        assigned_q = assigned_q.filter(
            db.or_(Driver.district_id.in_(scope_districts), Vehicle.district_id.in_(scope_districts))
        )
    if scope_vehicles:
        assigned_q = assigned_q.filter(Driver.vehicle_id.in_(scope_vehicles))
    if scope_shifts:
        assigned_q = assigned_q.filter(Driver.shift.in_(scope_shifts))
    if project_id:
        assigned_q = assigned_q.filter(
            db.or_(Driver.project_id == project_id, Vehicle.project_id == project_id)
        )
    if district_id:
        assigned_q = assigned_q.filter(
            db.or_(Driver.district_id == district_id, Vehicle.district_id == district_id)
        )
    if vehicle_id:
        assigned_q = assigned_q.filter(Driver.vehicle_id == vehicle_id)
    if shift:
        assigned_q = assigned_q.filter(Driver.shift == shift)
    for row in assigned_q.with_entities(Driver.id).distinct():
        ids.add(row[0])

    att_q = DriverAttendance.query.filter(
        DriverAttendance.attendance_date >= start_d,
        DriverAttendance.attendance_date <= end_d,
    )
    if project_id:
        att_q = att_q.filter(DriverAttendance.project_id == project_id)
    for row in att_q.with_entities(DriverAttendance.driver_id).distinct():
        ids.add(row[0])

    worked_left_q = DriverStatusChange.query.join(Driver).filter(
        DriverStatusChange.action_type == 'left',
        db.or_(
            db.and_(
                DriverStatusChange.change_date >= start_d,
                DriverStatusChange.change_date <= end_d,
            ),
            db.and_(
                DriverStatusChange.change_date > end_d,
                db.or_(Driver.assign_date.is_(None), Driver.assign_date <= end_d),
            ),
        ),
    )
    if scope_projects:
        worked_left_q = worked_left_q.filter(
            DriverStatusChange.left_project_id.in_(scope_projects)
        )
    if scope_districts:
        worked_left_q = worked_left_q.filter(
            DriverStatusChange.left_district_id.in_(scope_districts)
        )
    if scope_vehicles:
        worked_left_q = worked_left_q.filter(
            DriverStatusChange.left_vehicle_id.in_(scope_vehicles)
        )
    if scope_shifts:
        worked_left_q = worked_left_q.filter(
            DriverStatusChange.left_shift.in_(scope_shifts)
        )
    if project_id:
        worked_left_q = worked_left_q.filter(
            DriverStatusChange.left_project_id == project_id
        )
    if district_id:
        worked_left_q = worked_left_q.filter(
            DriverStatusChange.left_district_id == district_id
        )
    if vehicle_id:
        worked_left_q = worked_left_q.filter(
            DriverStatusChange.left_vehicle_id == vehicle_id
        )
    if shift:
        worked_left_q = worked_left_q.filter(
            DriverStatusChange.left_shift == shift
        )
    for row in worked_left_q.with_entities(DriverStatusChange.driver_id).distinct():
        ids.add(row[0])

    rejoin_q = DriverStatusChange.query.join(Driver).filter(
        DriverStatusChange.action_type == 'rejoin',
        DriverStatusChange.change_date >= start_d,
        DriverStatusChange.change_date <= end_d,
    )
    if scope_projects:
        rejoin_q = rejoin_q.filter(
            DriverStatusChange.new_project_id.in_(scope_projects)
        )
    if scope_districts:
        rejoin_q = rejoin_q.filter(
            DriverStatusChange.new_district_id.in_(scope_districts)
        )
    if scope_vehicles:
        rejoin_q = rejoin_q.filter(
            DriverStatusChange.new_vehicle_id.in_(scope_vehicles)
        )
    if scope_shifts:
        rejoin_q = rejoin_q.filter(
            DriverStatusChange.new_shift.in_(scope_shifts)
        )
    if project_id:
        rejoin_q = rejoin_q.filter(
            DriverStatusChange.new_project_id == project_id
        )
    if district_id:
        rejoin_q = rejoin_q.filter(
            DriverStatusChange.new_district_id == district_id
        )
    if vehicle_id:
        rejoin_q = rejoin_q.filter(
            DriverStatusChange.new_vehicle_id == vehicle_id
        )
    if shift:
        rejoin_q = rejoin_q.filter(
            DriverStatusChange.new_shift == shift
        )
    for row in rejoin_q.with_entities(DriverStatusChange.driver_id).distinct():
        ids.add(row[0])

    vehicle_ids_to_scan = []
    if vehicle_id:
        vehicle_ids_to_scan = [vehicle_id]
    elif scope_vehicles:
        vehicle_ids_to_scan = list(scope_vehicles)
    for vid in vehicle_ids_to_scan:
        ids.update(_tra_vehicle_driver_ids_in_month(vid, start_d, end_d))

    return ids


class _TraMonthCache:
    """Preload month data once to avoid N+1 queries in TRA report."""

    def __init__(self, start_d, end_d, driver_ids):
        self.start_d = start_d
        self.end_d = end_d
        self.driver_ids = set(driver_ids or [])
        self.drivers = {}
        self.left_by_driver = {}
        self.rejoin_in_month = {}
        self.status_events_by_driver = {}
        self.transfers_by_driver = {}
        self._att_in_month = set()
        self._vehicle_on_date = {}
        self._segments = {}
        self._vehicle_candidates = {}
        self._vehicle_day_counts = {}
        self._maint_by_vehicle = {}
        self._loaded = False

    def load(self):
        if self._loaded:
            return
        if not self.driver_ids:
            self._loaded = True
            return

        self.drivers = {
            d.id: d
            for d in Driver.query.filter(Driver.id.in_(self.driver_ids)).all()
        }
        ids = list(self.driver_ids)

        self.status_events_by_driver = {did: [] for did in ids}
        for sc in DriverStatusChange.query.filter(
            DriverStatusChange.driver_id.in_(ids),
            DriverStatusChange.action_type.in_(['left', 'rejoin']),
        ).order_by(DriverStatusChange.change_date.asc(), DriverStatusChange.id.asc()):
            self.status_events_by_driver.setdefault(sc.driver_id, []).append(sc)
        for sc in reversed(
            DriverStatusChange.query.filter(
                DriverStatusChange.driver_id.in_(ids),
                DriverStatusChange.action_type.in_(['left', 'rejoin']),
            ).order_by(DriverStatusChange.change_date.desc(), DriverStatusChange.id.desc()).all()
        ):
            if sc.action_type == 'left' and sc.driver_id not in self.left_by_driver:
                self.left_by_driver[sc.driver_id] = sc
            elif (
                sc.action_type == 'rejoin'
                and sc.change_date
                and self.start_d <= sc.change_date <= self.end_d
                and sc.driver_id not in self.rejoin_in_month
            ):
                self.rejoin_in_month[sc.driver_id] = sc

        self.transfers_by_driver = {did: [] for did in ids}
        for t in DriverTransfer.query.filter(
            DriverTransfer.driver_id.in_(ids),
            _tra_non_shift_transfer_filter(),
        ).order_by(DriverTransfer.transfer_date.asc(), DriverTransfer.id.asc()):
            self.transfers_by_driver.setdefault(t.driver_id, []).append(t)

        self._att_in_month = {
            row[0]
            for row in DriverAttendance.query.filter(
                DriverAttendance.driver_id.in_(ids),
                DriverAttendance.attendance_date >= self.start_d,
                DriverAttendance.attendance_date <= self.end_d,
            ).with_entities(DriverAttendance.driver_id).distinct()
        }
        self._loaded = True

    def _add_drivers(self, new_ids):
        new_ids = {i for i in new_ids if i and i not in self.drivers}
        if not new_ids:
            return
        self.driver_ids.update(new_ids)
        for d in Driver.query.filter(Driver.id.in_(new_ids)).all():
            self.drivers[d.id] = d
        for sc in DriverStatusChange.query.filter(
            DriverStatusChange.driver_id.in_(new_ids),
            DriverStatusChange.action_type.in_(['left', 'rejoin']),
        ).order_by(DriverStatusChange.change_date.asc(), DriverStatusChange.id.asc()):
            self.status_events_by_driver.setdefault(sc.driver_id, []).append(sc)
        for sc in reversed(
            DriverStatusChange.query.filter(
                DriverStatusChange.driver_id.in_(new_ids),
                DriverStatusChange.action_type.in_(['left', 'rejoin']),
            ).order_by(DriverStatusChange.change_date.desc(), DriverStatusChange.id.desc()).all()
        ):
            if sc.action_type == 'left' and sc.driver_id not in self.left_by_driver:
                self.left_by_driver[sc.driver_id] = sc
            elif (
                sc.action_type == 'rejoin'
                and sc.change_date
                and self.start_d <= sc.change_date <= self.end_d
                and sc.driver_id not in self.rejoin_in_month
            ):
                self.rejoin_in_month[sc.driver_id] = sc
        for t in DriverTransfer.query.filter(
            DriverTransfer.driver_id.in_(new_ids),
            _tra_non_shift_transfer_filter(),
        ).order_by(DriverTransfer.transfer_date.asc(), DriverTransfer.id.asc()):
            self.transfers_by_driver.setdefault(t.driver_id, []).append(t)
        for row in DriverAttendance.query.filter(
            DriverAttendance.driver_id.in_(new_ids),
            DriverAttendance.attendance_date >= self.start_d,
            DriverAttendance.attendance_date <= self.end_d,
        ).with_entities(DriverAttendance.driver_id).distinct():
            self._att_in_month.add(row[0])
        self._segments = {}
        self._vehicle_candidates = {}
        self._vehicle_day_counts = {}

    def left_rec(self, driver_id):
        return self.left_by_driver.get(driver_id)

    def driver_on_duty_in_month(self, driver):
        assign_date = driver.assign_date
        if assign_date and assign_date > self.end_d:
            if driver.id in self._att_in_month:
                return True
            if any(
                t.transfer_date
                and self.start_d <= t.transfer_date <= self.end_d
                for t in self.transfers_by_driver.get(driver.id, [])
            ):
                return True
            rejoin_rec = self.rejoin_in_month.get(driver.id)
            if rejoin_rec and rejoin_rec.change_date:
                return True
            return any(
                sc.change_date and self.start_d <= sc.change_date <= self.end_d
                for sc in [self.left_by_driver.get(driver.id)]
                if sc
            )
        left_rec = self.left_by_driver.get(driver.id)
        if left_rec and left_rec.change_date:
            if (
                driver.vehicle_id
                and assign_date
                and assign_date > left_rec.change_date
                and assign_date <= self.end_d
            ):
                return True
            if left_rec.change_date < self.start_d:
                rejoined = any(
                    sc.change_date
                    and sc.change_date > left_rec.change_date
                    and sc.change_date <= self.end_d
                    for sc in [self.rejoin_in_month.get(driver.id)]
                    if sc
                )
                if not rejoined:
                    return False
        return True

    def vehicle_id_on_date(self, driver_id, on_date):
        key = (driver_id, on_date)
        if key in self._vehicle_on_date:
            return self._vehicle_on_date[key]
        driver = self.drivers.get(driver_id)
        if not driver:
            self._vehicle_on_date[key] = None
            return None
        status_events = self.status_events_by_driver.get(driver_id, [])
        rejoin = _tra_applicable_rejoin_on_date(driver_id, on_date, status_events)
        if rejoin and rejoin.new_vehicle_id and rejoin.change_date and on_date >= rejoin.change_date:
            self._vehicle_on_date[key] = rejoin.new_vehicle_id
            return rejoin.new_vehicle_id
        left = self.left_rec(driver_id)
        if not driver.vehicle_id and left and left.change_date and left.change_date > on_date:
            vid = left.left_vehicle_id
            self._vehicle_on_date[key] = vid
            return vid

        transfers = self.transfers_by_driver.get(driver_id, [])
        vid = _tra_vehicle_id_from_transfer_list(transfers, on_date, driver, left)
        self._vehicle_on_date[key] = vid
        return vid

    def driver_duty_on_date(self, driver_id, on_date):
        driver = self.drivers.get(driver_id)
        if not driver:
            return False
        return _tra_driver_duty_active_on_date(
            driver, on_date, self.status_events_by_driver.get(driver_id),
        )

    def segments(self, driver):
        if driver.id in self._segments:
            return self._segments[driver.id]
        left_rec = self.left_rec(driver.id)
        rejoin_rec = self.rejoin_in_month.get(driver.id)
        left_date = None
        if left_rec and left_rec.change_date and self.start_d <= left_rec.change_date <= self.end_d:
            left_date = left_rec.change_date

        transfers = [
            t for t in self.transfers_by_driver.get(driver.id, [])
            if self.start_d <= t.transfer_date <= self.end_d
        ]
        segments = []

        if not transfers:
            eff = _tra_assignment_for_report_month(driver, left_rec, self.start_d, self.end_d)
            if not eff.get('vehicle_id'):
                self._segments[driver.id] = []
                return []
            seg_s, seg_e = _tra_clamp_segment_bounds(
                self.start_d, self.end_d, self.start_d, self.end_d,
                driver.assign_date, left_date,
            )
            if rejoin_rec and rejoin_rec.change_date and seg_s and rejoin_rec.change_date > seg_s:
                seg_s = max(seg_s, rejoin_rec.change_date)
            if seg_s is None or seg_s > seg_e:
                self._segments[driver.id] = []
                return []
            segments.append({
                'eff': eff,
                'segment_start': seg_s,
                'segment_end': seg_e,
                'transfer_in': None,
                'transfer_out': None,
            })
            self._segments[driver.id] = segments
            return segments

        segments = _tra_build_in_month_transfer_segments(
            transfers, self.start_d, self.end_d, driver, left_date, rejoin_rec,
        )
        self._segments[driver.id] = segments
        return segments

    def _expand_for_vehicles(self, vehicle_ids):
        if not vehicle_ids:
            return
        sf = _tra_non_shift_transfer_filter()
        new_ids = set()
        for vid in vehicle_ids:
            for row in Driver.query.filter(Driver.vehicle_id == vid).with_entities(Driver.id):
                new_ids.add(row[0])
            for row in DriverTransfer.query.filter(
                sf, DriverTransfer.new_vehicle_id == vid, DriverTransfer.transfer_date <= self.end_d,
            ).with_entities(DriverTransfer.driver_id):
                new_ids.add(row[0])
            for row in DriverTransfer.query.filter(
                sf,
                DriverTransfer.old_vehicle_id == vid,
                db.or_(
                    DriverTransfer.transfer_date >= self.start_d,
                    DriverTransfer.transfer_date > self.end_d,
                ),
            ).with_entities(DriverTransfer.driver_id):
                new_ids.add(row[0])
            for row in DriverStatusChange.query.filter(
                DriverStatusChange.action_type == 'left',
                DriverStatusChange.left_vehicle_id == vid,
                DriverStatusChange.change_date >= self.start_d,
            ).with_entities(DriverStatusChange.driver_id):
                new_ids.add(row[0])
        self._add_drivers(new_ids)

    def _vehicle_candidate_ids(self, vehicle_id):
        if vehicle_id in self._vehicle_candidates:
            return self._vehicle_candidates[vehicle_id]
        ids = set()
        for did, driver in self.drivers.items():
            if not self.driver_on_duty_in_month(driver):
                continue
            for seg in self.segments(driver):
                if seg['eff'].get('vehicle_id') == vehicle_id:
                    ids.add(did)
                    break
        self._vehicle_candidates[vehicle_id] = ids
        return ids

    def ensure_vehicle_day_counts(self, vehicle_ids):
        vehicle_ids = [vid for vid in (vehicle_ids or []) if vid]
        missing = [
            vid for vid in vehicle_ids
            if not any(k[0] == vid for k in self._vehicle_day_counts)
        ]
        if not missing:
            return
        self._expand_for_vehicles(set(missing))
        for vid in missing:
            self._vehicle_candidates.pop(vid, None)
            candidates = self._vehicle_candidate_ids(vid)
            d = self.start_d
            while d <= self.end_d:
                count = 0
                for pid in candidates:
                    if not self.driver_duty_on_date(pid, d):
                        continue
                    if self.vehicle_id_on_date(pid, d) == vid:
                        count += 1
                self._vehicle_day_counts[(vid, d)] = count
                d += timedelta(days=1)

    def prewarm(self, drivers):
        """Build segment + vehicle-day caches before the report row loop."""
        vehicle_ids = set()
        for d in drivers:
            for seg in self.segments(d):
                vid = seg['eff'].get('vehicle_id')
                if vid:
                    vehicle_ids.add(vid)
        if vehicle_ids:
            self.ensure_vehicle_day_counts(list(vehicle_ids))

    def segment_solo_days(self, vehicle_id, seg_start, seg_end):
        self.ensure_vehicle_day_counts([vehicle_id])
        solo = 0
        d = seg_start
        while d <= seg_end:
            if self._vehicle_day_counts.get((vehicle_id, d), 0) == 1:
                solo += 1
            d += timedelta(days=1)
        return solo

    def ensure_maint_data(self, vehicle_ids):
        missing = [vid for vid in vehicle_ids if vid and vid not in self._maint_by_vehicle]
        if not missing:
            return
        for row in MaintenanceExpense.query.filter(
            MaintenanceExpense.vehicle_id.in_(missing),
            MaintenanceExpense.expense_date >= self.start_d,
            MaintenanceExpense.expense_date <= self.end_d,
        ).with_entities(MaintenanceExpense.vehicle_id, MaintenanceExpense.expense_date):
            self._maint_by_vehicle.setdefault(row[0], set()).add(row[1])

    def maint_count(self, vehicle_id, seg_start, seg_end):
        if not vehicle_id:
            return 0
        self.ensure_maint_data([vehicle_id])
        dates = self._maint_by_vehicle.get(vehicle_id, set())
        return sum(1 for d in dates if seg_start <= d <= seg_end)


def _tra_status_rec(driver_id, action_type):
    return DriverStatusChange.query.filter(
        DriverStatusChange.driver_id == driver_id,
        DriverStatusChange.action_type == action_type,
    ).order_by(DriverStatusChange.change_date.desc()).first()


def _tra_transfer_rec(driver_id, vehicle_id, side, start_d, end_d):
    if not vehicle_id:
        return None
    q = DriverTransfer.query.filter(
        DriverTransfer.driver_id == driver_id,
        _tra_non_shift_transfer_filter(),
        DriverTransfer.transfer_date >= start_d,
        DriverTransfer.transfer_date <= end_d,
    )
    if side == 'in':
        q = q.filter(DriverTransfer.new_vehicle_id == vehicle_id)
    else:
        q = q.filter(DriverTransfer.old_vehicle_id == vehicle_id)
    return q.order_by(DriverTransfer.transfer_date.desc()).first()


def _tra_duty_window(
    start_d,
    end_d,
    assign_date,
    rejoin_date,
    transfer_in_date,
    left_date,
    transfer_out_date,
):
    drv_start = start_d
    drv_end = end_d
    if assign_date and assign_date > start_d:
        drv_start = assign_date
    if rejoin_date and start_d <= rejoin_date <= end_d and rejoin_date > drv_start:
        drv_start = rejoin_date
    if transfer_in_date and start_d <= transfer_in_date <= end_d and transfer_in_date > drv_start:
        drv_start = transfer_in_date
    if left_date and start_d <= left_date <= end_d and left_date < drv_end:
        drv_end = left_date
    if transfer_out_date and start_d <= transfer_out_date <= end_d:
        t_last = transfer_out_date - timedelta(days=1)
        if t_last >= start_d and t_last < drv_end:
            drv_end = t_last
    if drv_start > drv_end:
        return drv_start, drv_end, 0
    return drv_start, drv_end, (drv_end - drv_start).days + 1


def _tra_vehicle_driver_ids_in_month(vehicle_id, start_d, end_d):
    candidates = set()
    sf = _tra_non_shift_transfer_filter()
    for row in DriverTransfer.query.filter(sf, DriverTransfer.new_vehicle_id == vehicle_id).with_entities(
        DriverTransfer.driver_id, DriverTransfer.transfer_date
    ):
        if row[1] <= end_d:
            candidates.add(row[0])
    for row in DriverTransfer.query.filter(
        sf,
        DriverTransfer.old_vehicle_id == vehicle_id,
        DriverTransfer.transfer_date >= start_d,
        DriverTransfer.transfer_date <= end_d,
    ).with_entities(DriverTransfer.driver_id):
        candidates.add(row[0])
    for row in DriverTransfer.query.filter(
        sf,
        DriverTransfer.old_vehicle_id == vehicle_id,
        DriverTransfer.transfer_date > end_d,
    ).with_entities(DriverTransfer.driver_id):
        candidates.add(row[0])
    for row in Driver.query.filter(Driver.vehicle_id == vehicle_id).with_entities(Driver.id):
        candidates.add(row[0])
    for row in DriverStatusChange.query.filter(
        DriverStatusChange.action_type == 'left',
        DriverStatusChange.left_vehicle_id == vehicle_id,
        DriverStatusChange.change_date >= start_d,
    ).with_entities(DriverStatusChange.driver_id):
        candidates.add(row[0])

    ids = set()
    for pid in candidates:
        partner = db.session.get(Driver, pid)
        if not partner:
            continue
        p_left = _tra_status_rec(pid, 'left')
        if not _tra_driver_on_duty_in_month(pid, start_d, end_d, partner.assign_date):
            continue
        on_vehicle = False
        for seg in _tra_driver_vehicle_segments(partner, p_left, start_d, end_d):
            if seg['eff'].get('vehicle_id') == vehicle_id:
                on_vehicle = True
                break
        if on_vehicle:
            ids.add(pid)
    return ids


def _tra_build_segment_remarks(
    driver,
    eff,
    segment,
    active_days,
    working_days,
    solo_days,
    ndays,
    start_d,
    end_d,
    rejoin_in_month,
    left_rec,
    left_date,
):
    parts = []
    veh_no = eff.get('vehicle').vehicle_no if eff.get('vehicle') else '-'
    seg_s = segment['segment_start']
    seg_e = segment['segment_end']
    transfer_in_rec = segment.get('transfer_in')
    transfer_out_rec = segment.get('transfer_out')

    if transfer_in_rec:
        old_v = transfer_in_rec.old_vehicle.vehicle_no if transfer_in_rec.old_vehicle else '?'
        new_v = transfer_in_rec.new_vehicle.vehicle_no if transfer_in_rec.new_vehicle else '?'
        t_date = transfer_in_rec.transfer_date.strftime('%d-%m-%Y')
        parts.append(
            f'Transferred from {old_v} to {new_v} on {t_date}. '
            f'Duty on {new_v}: {active_days} day(s) ({seg_s.strftime("%d-%m-%Y")} to {seg_e.strftime("%d-%m-%Y")}).'
        )
        if transfer_in_rec.remarks:
            parts.append(transfer_in_rec.remarks.strip())
    elif transfer_out_rec:
        old_v = transfer_out_rec.old_vehicle.vehicle_no if transfer_out_rec.old_vehicle else '?'
        new_v = transfer_out_rec.new_vehicle.vehicle_no if transfer_out_rec.new_vehicle else '?'
        t_date = transfer_out_rec.transfer_date.strftime('%d-%m-%Y')
        parts.append(
            f'Transferred from {old_v} to {new_v} on {t_date}. '
            f'Worked {active_days} day(s) on {old_v} ({seg_s.strftime("%d-%m-%Y")} to {seg_e.strftime("%d-%m-%Y")}).'
        )
        if transfer_out_rec.remarks:
            parts.append(transfer_out_rec.remarks.strip())

    if rejoin_in_month and rejoin_in_month.change_date and not transfer_in_rec:
        if seg_s <= rejoin_in_month.change_date <= seg_e:
            r_date = rejoin_in_month.change_date.strftime('%d-%m-%Y')
            parts.append(f'Rejoined on {r_date}. Working days after rejoin on {veh_no}: {active_days}.')

    if left_date and seg_s <= left_date <= seg_e and not transfer_out_rec:
        l_date = left_date.strftime('%d-%m-%Y')
        msg = f'Left on {l_date}. Worked {active_days} day(s) on {veh_no} before leaving.'
        if left_rec and left_rec.reason:
            msg += f' Reason: {left_rec.reason}.'
        parts.append(msg)
    elif (
        driver.assign_date
        and start_d <= driver.assign_date <= end_d
        and seg_s <= driver.assign_date <= seg_e
        and not transfer_in_rec
        and not rejoin_in_month
        and active_days > 0
    ):
        parts.append(
            f'Joined duty on {driver.assign_date.strftime("%d-%m-%Y")}. '
            f'Working days on {veh_no}: {active_days}.'
        )
    elif active_days < ndays and not transfer_in_rec and not transfer_out_rec and not left_date:
        parts.append(
            f'Duty on {veh_no} from {seg_s.strftime("%d-%m-%Y")} to {seg_e.strftime("%d-%m-%Y")} '
            f'({active_days} day(s)).'
        )

    if solo_days > 0:
        if solo_days >= active_days:
            parts.append(
                f'No second driver on {veh_no} for this period; double duty for '
                f'{solo_days} day(s) ({working_days} total working days).'
            )
        else:
            paired_days = active_days - solo_days
            parts.append(
                f'Second driver present for {paired_days} day(s). '
                f'No second driver for remaining {solo_days} day(s); double duty applied '
                f'({working_days} total working days).'
            )

    return ' '.join(p for p in parts if p)


def _tra_compute_segment_metrics(driver, segment, start_d, end_d, ndays, cache=None):
    eff = segment['eff']
    vehicle = eff.get('vehicle')
    vehicle_id = eff.get('vehicle_id')
    seg_s = segment['segment_start']
    seg_e = segment['segment_end']
    active_days = (seg_e - seg_s).days + 1

    if cache:
        left_rec = cache.left_rec(driver.id)
        rejoin_in_month = cache.rejoin_in_month.get(driver.id)
    else:
        left_rec = _tra_status_rec(driver.id, 'left')
        rejoin_rec = _tra_status_rec(driver.id, 'rejoin')
        rejoin_in_month = (
            rejoin_rec
            if rejoin_rec and rejoin_rec.change_date and start_d <= rejoin_rec.change_date <= end_d
            else None
        )
    left_date = None
    if left_rec and left_rec.change_date and start_d <= left_rec.change_date <= end_d:
        left_date = left_rec.change_date

    working_days = active_days
    solo_days = 0
    capacity = vehicle.driver_capacity if vehicle else 1

    if capacity >= 2 and vehicle_id and active_days > 0:
        if cache:
            solo_days = cache.segment_solo_days(vehicle_id, seg_s, seg_e)
        else:
            solo_days = _tra_segment_solo_days(vehicle_id, seg_s, seg_e, start_d, end_d)
        working_days = active_days + solo_days

    remarks = _tra_build_segment_remarks(
        driver,
        eff,
        segment,
        active_days,
        working_days,
        solo_days,
        ndays,
        start_d,
        end_d,
        rejoin_in_month,
        left_rec,
        left_date,
    )

    transfer_in = segment.get('transfer_in')
    transfer_out = segment.get('transfer_out')
    return {
        'working_days': working_days,
        'active_days': active_days,
        'remarks': remarks,
        'transfer_date': (
            transfer_in.transfer_date if transfer_in
            else (transfer_out.transfer_date if transfer_out else None)
        ),
        'date_of_leaving': left_date if left_date and seg_s <= left_date <= seg_e else None,
        'date_of_rejoining': (
            rejoin_in_month.change_date
            if rejoin_in_month and seg_s <= rejoin_in_month.change_date <= seg_e
            else None
        ),
    }


# ────────────────────────────────────────────────
# TRA Attendance Sheet (Monthly with Transfer/Rejoin)
# ────────────────────────────────────────────────

# ────────────────────────────────────────────────
# Leave Approval Workflow
# ────────────────────────────────────────────────



# ─── Task Report ─────────────────────────────────


def _fuel_expense_location_cascade_dict():
    """
    One-shot data for Add/Edit Fuel form: all projects by district (same as get_projects_by_district)
    plus all vehicles the user is allowed to see (same scoping as get_vehicles_by_project_district).
    """
    pbd = {}
    q_pd = (
        db.session.query(Project.id, Project.name, project_district.c.district_id)
        .join(project_district, Project.id == project_district.c.project_id)
        .order_by(project_district.c.district_id, Project.name)
    )
    for pid, name, did in q_pd.all():
        key = str(did)
        if key not in pbd:
            pbd[key] = []
        pbd[key].append({"id": int(pid), "name": name})

    scope_projects, scope_districts, scope_vehicles, _ = _get_user_scope()
    vq = Vehicle.query
    if scope_projects:
        vq = vq.filter(Vehicle.project_id.in_(list(scope_projects)))
    if scope_districts:
        vq = vq.filter(Vehicle.district_id.in_(list(scope_districts)))
    if scope_vehicles:
        vq = vq.filter(Vehicle.id.in_(list(scope_vehicles)))
    vehicles = []
    for v in vq.order_by(*vehicle_order_by()).all():
        vehicles.append(
            {
                "id": v.id,
                "vehicle_no": v.vehicle_no or "",
                "project_id": v.project_id,
                "district_id": v.district_id,
                "fuel_type": (v.fuel_type or "Petrol"),
                "fuel_tank_capacity": float(v.fuel_tank_capacity or 0),
                "vehicle_family": (v.vehicle_family or ""),
            }
        )
    return {"projects_by_district": pbd, "vehicles": vehicles}



def _vehicle_query_task_report_scope(is_master_or_admin, allowed_projects, allowed_districts, allowed_vehicles):
    """Restrict Vehicle rows to user's district/project/vehicle assignments (same AND logic as fleet reports)."""
    q = Vehicle.query
    if is_master_or_admin:
        return q
    ap = set(allowed_projects or [])
    ad = set(allowed_districts or [])
    av = set(allowed_vehicles or [])
    if not ap and not ad and not av:
        return q.filter(Vehicle.id.in_([-1]))
    if ap:
        q = q.filter(Vehicle.project_id.in_(list(ap)))
    if ad:
        q = q.filter(Vehicle.district_id.in_(list(ad)))
    if av:
        q = q.filter(Vehicle.id.in_(list(av)))
    return q




def _vehicle_period_detail_rows(from_date, to_date, project_id, district_id, vehicle_id,
                                is_master_or_admin, allowed_projects, allowed_districts, allowed_vehicles):
    """One summary row per vehicle for the date range (logbook-style) with daily-task column metrics."""
    if not project_id:
        return []
    vq = Vehicle.query.filter(Vehicle.project_id == project_id)
    if district_id:
        vq = vq.filter(Vehicle.district_id == district_id)
    if vehicle_id:
        vq = vq.filter(Vehicle.id == vehicle_id)
    if not is_master_or_admin:
        ap, ad, av = allowed_projects, allowed_districts, allowed_vehicles
        if not ap and not ad and not av:
            return []
        if av:
            vq = vq.filter(Vehicle.id.in_(list(av)))
        if ap:
            vq = vq.filter(Vehicle.project_id.in_(list(ap)))
        if ad:
            vq = vq.filter(Vehicle.district_id.in_(list(ad)))
    rows = []
    for v in vq.order_by(*vehicle_order_by()).all():
        tasks = VehicleDailyTask.query.filter(
            VehicleDailyTask.vehicle_id == v.id,
            VehicleDailyTask.task_date >= from_date,
            VehicleDailyTask.task_date <= to_date,
        ).order_by(VehicleDailyTask.task_date.asc()).all()
        if not tasks:
            continue
        first_d = tasks[0].task_date
        prev = VehicleDailyTask.query.filter(
            VehicleDailyTask.vehicle_id == v.id,
            VehicleDailyTask.task_date < first_d,
        ).order_by(VehicleDailyTask.task_date.desc()).first()
        if prev and prev.close_reading is not None:
            start_reading = float(prev.close_reading)
        elif tasks[0].start_reading is not None:
            start_reading = float(tasks[0].start_reading)
        else:
            start_reading = 0
        close_reading = float(tasks[-1].close_reading)
        kms_driven = close_reading - start_reading
        if kms_driven < 0:
            kms_driven = 0
        tasks_count = sum(int(t.tasks_count or 0) for t in tasks)
        emg_tasks = 0
        tracker_km = 0.0
        odometer_photo_path = ''
        for t in tasks:
            task_d = t.task_date
            emg_tasks += EmergencyTaskRecord.query.filter(
                EmergencyTaskRecord.task_date == task_d,
                EmergencyTaskRecord.amb_reg_no == v.vehicle_no,
                EmergencyTaskRecord.category.in_(['Green', 'Yellow']),
            ).count()
            _mil_rec = VehicleMileageRecord.query.filter_by(task_date=task_d, reg_no=v.vehicle_no).first()
            tracker_km += _mil_rec.effective_km() if _mil_rec else 0
            ph = (getattr(t, 'odometer_photo_path', None) or '').strip()
            if ph:
                odometer_photo_path = ph
        kms_diff = kms_driven - tracker_km
        pct_diff = round((kms_diff / kms_driven) * 100, 1) if kms_driven else None
        rows.append({
            'vehicle': v,
            'start_reading': start_reading,
            'close_reading': close_reading,
            'kms_driven': round(kms_driven, 2),
            'tasks_count': tasks_count,
            'emg_tasks': emg_tasks,
            'tracker_km': round(tracker_km, 2),
            'kms_diff': round(kms_diff, 2),
            'pct_diff': pct_diff,
            'odometer_photo_path': odometer_photo_path,
        })
    return rows


def _task_report_vehicle_period_detail_impl(redirect_endpoint, template_name, export_pdf=False):
    from auth_utils import get_user_context

    user_id = session.get('user_id')
    user_context = get_user_context(user_id) if user_id else {}
    allowed_projects = user_context.get('allowed_projects', set())
    allowed_districts = user_context.get('allowed_districts', set())
    allowed_vehicles = user_context.get('allowed_vehicles', set())
    is_master_or_admin = user_context.get('is_master_or_admin', False)

    form = TaskReportFilterForm()
    district_q = District.query
    project_q = Project.query
    if not is_master_or_admin:
        ap, ad, av = allowed_projects, allowed_districts, allowed_vehicles
        if not ap and not ad and not av:
            district_q = district_q.filter(District.id.in_([-1]))
            project_q = project_q.filter(Project.id.in_([-1]))
        else:
            if ad:
                district_q = district_q.filter(District.id.in_(list(ad)))
            elif ap:
                district_q = (
                    district_q.join(project_district, project_district.c.district_id == District.id)
                    .filter(project_district.c.project_id.in_(list(ap)))
                    .distinct()
                )
            elif av:
                d_ids = [
                    r[0] for r in db.session.query(Vehicle.district_id)
                    .filter(Vehicle.id.in_(list(av)), Vehicle.district_id.isnot(None))
                    .distinct().all()
                ]
                district_q = district_q.filter(District.id.in_(d_ids or [-1]))
            if ap:
                project_q = project_q.filter(Project.id.in_(list(ap)))
            if ad:
                project_q = (
                    project_q.join(project_district, project_district.c.project_id == Project.id)
                    .filter(project_district.c.district_id.in_(list(ad)))
                    .distinct()
                )
            if not ap and not ad and av:
                p_ids = [
                    r[0] for r in db.session.query(Vehicle.project_id)
                    .filter(Vehicle.id.in_(list(av)), Vehicle.project_id.isnot(None))
                    .distinct().all()
                ]
                project_q = project_q.filter(Project.id.in_(p_ids or [-1]))
    form.district_id.choices = [(0, '-- All Districts --')] + [(d.id, d.name) for d in district_q.order_by(District.name).all()]
    form.project_id.choices = [(0, '-- All Projects --')] + [(p.id, p.name) for p in project_q.order_by(Project.name).all()]

    today = pk_date()
    from_date = today
    to_date = today
    lad = set(allowed_districts) if allowed_districts else set()
    lap = set(allowed_projects) if allowed_projects else set()
    lav = set(allowed_vehicles) if allowed_vehicles else set()

    def coerce_task_report_scope(did, pid, vid=0):
        vd = {c[0] for c in form.district_id.choices}
        vp = {c[0] for c in form.project_id.choices}
        lk = {'lock_district': False, 'lock_project': False, 'lock_vehicle': False}
        if did and did not in vd:
            did = 0
        if pid and pid not in vp:
            pid = 0
        if not is_master_or_admin:
            if len(lad) == 1:
                only_d = next(iter(lad))
                if only_d in vd:
                    did = only_d
                    lk['lock_district'] = True
            if len(lad) == 1 and len(lap) == 1 and len(lav) == 1:
                only_p = next(iter(lap))
                if only_p in vp:
                    pid = only_p
                    lk['lock_project'] = True
        vehicle_q = _vehicle_query_task_report_scope(
            is_master_or_admin, allowed_projects, allowed_districts, allowed_vehicles
        )
        if did:
            vehicle_q = vehicle_q.filter(Vehicle.district_id == did)
        if pid:
            vehicle_q = vehicle_q.filter(Vehicle.project_id == pid)
        form.vehicle_id.choices = [(0, '-- All Vehicles --')] + [
            (v.id, v.vehicle_no) for v in vehicle_q.order_by(*vehicle_order_by()).all()
        ]
        vv = {c[0] for c in form.vehicle_id.choices}
        if vid and vid not in vv:
            vid = 0
        if not is_master_or_admin and len(lav) == 1:
            only_v = next(iter(lav))
            if only_v in vv:
                vid = only_v
                lk['lock_vehicle'] = True
        return did, pid, vid, lk

    if request.method == 'POST':
        from_date = parse_date(request.form.get('from_date')) or today
        to_date = parse_date(request.form.get('to_date')) or today
        district_id = request.form.get('district_id', type=int) or 0
        project_id = request.form.get('project_id', type=int) or 0
        vehicle_id = request.form.get('vehicle_id', type=int) or 0
        if from_date and to_date and from_date > to_date:
            from_date, to_date = to_date, from_date
        district_id, project_id, vehicle_id, _ = coerce_task_report_scope(district_id, project_id, vehicle_id)
        return redirect(url_for(
            redirect_endpoint,
            from_date=from_date.strftime('%d-%m-%Y') if from_date else '',
            to_date=to_date.strftime('%d-%m-%Y') if to_date else '',
            district_id=district_id,
            project_id=project_id,
            vehicle_id=vehicle_id,
        ))

    district_id = request.args.get('district_id', type=int) or 0
    project_id = request.args.get('project_id', type=int) or 0
    vehicle_id = request.args.get('vehicle_id', type=int) or 0
    from_str = request.args.get('from_date', '')
    to_str = request.args.get('to_date', '')
    if from_str:
        from_date = parse_date(from_str) or from_date
    if to_str:
        to_date = parse_date(to_str) or to_date
    if from_date and to_date and from_date > to_date:
        from_date, to_date = to_date, from_date
    district_id, project_id, vehicle_id, task_report_filter_lock = coerce_task_report_scope(
        district_id, project_id, vehicle_id
    )
    if district_id:
        projects = Project.query.join(project_district).filter(
            project_district.c.district_id == district_id
        ).order_by(Project.name).all()
        scoped = [(0, '-- All Projects --')] + [(p.id, p.name) for p in projects]
        if not is_master_or_admin and allowed_projects:
            scoped = [(0, '-- All Projects --')] + [
                (p.id, p.name) for p in projects if p.id in allowed_projects
            ]
        form.project_id.choices = scoped
    form.from_date.data = from_date
    form.to_date.data = to_date
    form.district_id.data = district_id
    form.project_id.data = project_id
    form.vehicle_id.data = vehicle_id

    rows = []
    if project_id:
        rows = _vehicle_period_detail_rows(
            from_date, to_date, project_id, district_id, vehicle_id,
            is_master_or_admin, allowed_projects, allowed_districts, allowed_vehicles,
        )

    search = (request.args.get('search') or '').strip()
    if search:
        tokens = [t.lower() for t in search.split() if t]

        def _match(r):
            v = r['vehicle']
            blob = ' '.join([
                v.vehicle_no,
                v.district.name if v.district else '',
                v.parking_station.tehsil if v.parking_station else '',
                v.parking_station.name if v.parking_station else '',
                v.vehicle_type or '',
                str(r['kms_driven']), str(r['tasks_count']), str(r['emg_tasks']),
                str(r.get('odometer_photo_path') or ''),
            ]).lower()
            return all(tok in blob for tok in tokens)

        rows = [r for r in rows if _match(r)]

    total_kms = sum(r['kms_driven'] for r in rows)
    total_tracker = sum(r['tracker_km'] for r in rows)
    total_diff = round(total_kms - total_tracker, 2)
    total_pct = round((total_diff / total_kms * 100), 1) if total_kms else None
    total_tasks = sum(r['tasks_count'] for r in rows)
    total_emg = sum(r['emg_tasks'] for r in rows)
    total_task_diff = total_tasks - total_emg

    if export_pdf:
        if not project_id:
            return jsonify({'error': 'Pehle Project select karein.'}), 400
        if not rows:
            return jsonify({'error': 'Export ke liye koi row nahi mili.'}), 400
        try:
            html = render_template(
                'task_report_vehicle_period_detail_print.html',
                rows=rows,
                from_date=from_date,
                to_date=to_date,
                total_kms=total_kms,
                total_tracker=total_tracker,
                total_diff=total_diff,
                total_pct=total_pct,
                total_tasks=total_tasks,
                total_emg=total_emg,
                total_task_diff=total_task_diff,
            )
            pdf_bytes = _html_to_pdf_bytes(html, landscape=True)
        except Exception as exc:
            app.logger.exception('Vehicle period task detail PDF export failed')
            return jsonify({'error': f'PDF generation failed: {exc}'}), 500
        fname = (
            f'Daily_Task_Period_Detail_{from_date.strftime("%d-%m-%Y")}_to_'
            f'{to_date.strftime("%d-%m-%Y")}.pdf'
        )
        return Response(
            pdf_bytes,
            mimetype='application/pdf',
            headers={'Content-Disposition': f'attachment; filename="{fname}"'},
        )

    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    pagination = SimplePagination(rows, page, per_page)
    rows = pagination.items
    return render_template(
        template_name,
        form=form,
        rows=rows,
        from_date=from_date,
        to_date=to_date,
        total_kms=total_kms,
        total_tracker=total_tracker,
        total_diff=total_diff,
        total_pct=total_pct,
        total_tasks=total_tasks,
        total_emg=total_emg,
        total_task_diff=total_task_diff,
        pagination=pagination,
        per_page=per_page,
        search=search,
        task_report_filter_lock=task_report_filter_lock,
        **_nav_back_ctx(url_for('module_hub', hub_slug='task-logbook'), show_without_nav_from=True),
    )




def _logbook_vehicle_aggregate(vehicle_id, from_date, to_date):
    """For one vehicle, return start_reading, close_reading, total_kms, total_task for the date range."""
    tasks = VehicleDailyTask.query.filter(
        VehicleDailyTask.vehicle_id == vehicle_id,
        VehicleDailyTask.task_date >= from_date,
        VehicleDailyTask.task_date <= to_date
    ).order_by(VehicleDailyTask.task_date.asc()).all()
    if not tasks:
        return None
    first = tasks[0]
    last = tasks[-1]
    prev = VehicleDailyTask.query.filter(
        VehicleDailyTask.vehicle_id == vehicle_id,
        VehicleDailyTask.task_date < from_date
    ).order_by(VehicleDailyTask.task_date.desc()).first()
    if prev and prev.close_reading is not None:
        start_reading = float(prev.close_reading)
    elif first.start_reading is not None:
        start_reading = float(first.start_reading)
    else:
        start_reading = 0
    close_reading = float(last.close_reading)
    total_kms = close_reading - start_reading
    if total_kms < 0:
        total_kms = 0
    total_task = sum(t.tasks_count for t in tasks)
    return {
        'start_reading': start_reading,
        'close_reading': close_reading,
        'total_kms': round(total_kms, 2),
        'total_task': total_task,
    }





def _show_task_batch_totals(user_context, row_list):
    """Footer batch totals: show for employees with >1 assigned vehicle, or admin with >1 row; hide for drivers."""
    if not row_list:
        return False
    if user_context.get('is_driver'):
        return False
    if user_context.get('is_master_or_admin'):
        return len(row_list) > 1
    if user_context.get('is_employee'):
        return len(user_context.get('allowed_vehicles') or ()) > 1
    return len(row_list) > 1


def _task_entry_resolve_start_reading(v, task_date, form_dict):
    """Same start-reading resolution as _build_vehicle_rows (for save validation)."""
    form_dict = form_dict or {}
    prev = VehicleDailyTask.query.filter(
        VehicleDailyTask.vehicle_id == v.id,
        VehicleDailyTask.task_date < task_date,
    ).order_by(VehicleDailyTask.task_date.desc()).first()
    has_prev = prev is not None and prev.close_reading is not None
    start_reading = float(prev.close_reading) if has_prev else 0
    existing = VehicleDailyTask.query.filter_by(vehicle_id=v.id, task_date=task_date).first()
    if existing and existing.start_reading is not None and not has_prev:
        start_reading = float(existing.start_reading)
    key_start = 'vehicle_%s_start_reading' % v.id
    if not has_prev and key_start in form_dict and form_dict[key_start] not in (None, ''):
        try:
            start_reading = float(form_dict[key_start])
        except (TypeError, ValueError):
            pass
    return start_reading


def _parse_hhmm_time_optional(s):
    if not s or not str(s).strip():
        return None
    try:
        return datetime.strptime(str(s).strip()[:5], '%H:%M').time()
    except ValueError:
        return None


def _default_task_entry_date_for_project(project_id):
    """24h projects: before configured time, default task date = yesterday."""
    today = pk_date()
    if not project_id:
        return today
    project = db.session.get(Project, project_id)
    if not project or not project.task_entry_yesterday_default_until:
        return today
    if _attendance_local_time() < project.task_entry_yesterday_default_until:
        return today - timedelta(days=1)
    return today


def _task_entry_date_save_ok(project, task_date):
    """Block saving today's task on 24h-style projects during grace window (common mistake)."""
    if not project or not project.task_entry_yesterday_default_until or not task_date:
        return True, None
    today = pk_date()
    if task_date != today:
        return True, None
    if _attendance_local_time() >= project.task_entry_yesterday_default_until:
        return True, None
    y = today - timedelta(days=1)
    return False, (
        f'"{project.name}": subah {project.task_entry_yesterday_default_until.strftime("%H:%M")} se pehle '
        f'aaj ({today.strftime("%d-%m-%Y")}) ki date par save nahi — kal ({y.strftime("%d-%m-%Y")}) select karein.'
    )


def _task_entry_date_hint_for_project(project, view_date):
    if not project or not project.task_entry_yesterday_default_until:
        return None
    today = pk_date()
    y = today - timedelta(days=1)
    until_s = project.task_entry_yesterday_default_until.strftime('%H:%M')
    if _attendance_local_time() < project.task_entry_yesterday_default_until:
        if view_date == y:
            return (
                f'"{project.name}": abhi {until_s} se pehle — task report zyada tar '
                f'kal ({y.strftime("%d-%m-%Y")}) ki duty ke liye hoti hai.'
            )
        return (
            f'"{project.name}": aaj subah {until_s} se pehle kal ki date ({y.strftime("%d-%m-%Y")}) '
            f'use karein — raat ki duty ki report calendar date badalne ke baad bhi kal ki hoti hai.'
        )
    return None





def _task_entry_record_in_user_scope(rec, is_master_or_admin, allowed_projects, allowed_districts, allowed_vehicles):
    """True if this VehicleDailyTask's vehicle passes the same scope filters as task report entry."""
    if is_master_or_admin:
        return True
    v = rec.vehicle
    if not v:
        return False
    ap = set(allowed_projects or [])
    ad = set(allowed_districts or [])
    av = set(allowed_vehicles or [])
    if not ap and not ad and not av:
        return False
    if ap and v.project_id not in ap:
        return False
    if ad and v.district_id not in ad:
        return False
    if av and v.id not in av:
        return False
    return True



def _build_vehicle_rows(vehicles, task_date, form=None):
    form = form or {}
    rows = []
    for v in vehicles:
        prev = VehicleDailyTask.query.filter(
            VehicleDailyTask.vehicle_id == v.id,
            VehicleDailyTask.task_date < task_date
        ).order_by(VehicleDailyTask.task_date.desc()).first()
        has_prev = prev is not None and prev.close_reading is not None
        start_reading = float(prev.close_reading) if has_prev else 0
        emg_tasks = EmergencyTaskRecord.query.filter(
            EmergencyTaskRecord.task_date == task_date,
            EmergencyTaskRecord.amb_reg_no == v.vehicle_no,
            EmergencyTaskRecord.category.in_(['Green', 'Yellow']),
        ).count()
        _mil_rec = VehicleMileageRecord.query.filter_by(task_date=task_date, reg_no=v.vehicle_no).first()
        tracker_km = _mil_rec.effective_km() if _mil_rec else 0
        existing = VehicleDailyTask.query.filter_by(vehicle_id=v.id, task_date=task_date).first()
        if existing and existing.start_reading is not None and not has_prev:
            start_reading = float(existing.start_reading)
        existing_close = float(existing.close_reading) if existing and existing.close_reading is not None else None
        existing_tasks = existing.tasks_count if existing else None
        if form:
            key_close = 'vehicle_%s_close_reading' % v.id
            key_tasks = 'vehicle_%s_tasks_count' % v.id
            key_start = 'vehicle_%s_start_reading' % v.id
            if key_close in form and form[key_close] not in (None, ''):
                try:
                    existing_close = float(form[key_close])
                except (TypeError, ValueError):
                    pass
            if key_tasks in form and form[key_tasks] not in (None, ''):
                try:
                    existing_tasks = int(float(form[key_tasks]))
                except (TypeError, ValueError):
                    pass
            if not has_prev and key_start in form and form[key_start] not in (None, ''):
                try:
                    start_reading = float(form[key_start])
                except (TypeError, ValueError):
                    pass
        rows.append({
            'vehicle': v,
            'start_reading': start_reading,
            'start_editable': not has_prev,
            'emg_tasks': emg_tasks,
            'tracker_km': round(tracker_km, 2),
            'close_reading': existing_close,
            'tasks_count': existing_tasks,
            'saved': existing,
            'odometer_photo_path': ((existing.odometer_photo_path or '').strip()) if existing else '',
        })
    return rows


def _task_row_is_pending(entry_row):
    """Pending = Close Reading missing and/or Task's not entered (empty field, not 0)."""
    if entry_row.get('close_reading') is None:
        return True
    if entry_row.get('tasks_count') is None:
        return True
    return False


def _filter_pending_task_rows(rows):
    return [r for r in rows if _task_row_is_pending(r)]


def _task_report_entry_scope_context(user_context):
    """Districts/projects and scope flags shared by New Task Entry + Pending report."""
    allowed_projects = user_context.get('allowed_projects', set())
    allowed_districts = user_context.get('allowed_districts', set())
    allowed_vehicles = user_context.get('allowed_vehicles', set())
    is_master_or_admin = user_context.get('is_master_or_admin', False)

    district_q = District.query
    project_q_all = Project.query
    if not is_master_or_admin:
        ap, ad, av = allowed_projects, allowed_districts, allowed_vehicles
        if not ap and not ad and not av:
            district_q = district_q.filter(District.id.in_([-1]))
            project_q_all = project_q_all.filter(Project.id.in_([-1]))
        else:
            if ad:
                district_q = district_q.filter(District.id.in_(list(ad)))
            elif ap:
                district_q = (
                    district_q.join(project_district, project_district.c.district_id == District.id)
                    .filter(project_district.c.project_id.in_(list(ap)))
                    .distinct()
                )
            elif av:
                d_ids = [
                    r[0] for r in db.session.query(Vehicle.district_id)
                    .filter(Vehicle.id.in_(list(av)), Vehicle.district_id.isnot(None))
                    .distinct().all()
                ]
                district_q = district_q.filter(District.id.in_(d_ids or [-1]))
            if ap:
                project_q_all = project_q_all.filter(Project.id.in_(list(ap)))
            if ad:
                project_q_all = (
                    project_q_all.join(project_district, project_district.c.project_id == Project.id)
                    .filter(project_district.c.district_id.in_(list(ad)))
                    .distinct()
                )
            if not ap and not ad and av:
                p_ids = [
                    r[0] for r in db.session.query(Vehicle.project_id)
                    .filter(Vehicle.id.in_(list(av)), Vehicle.project_id.isnot(None))
                    .distinct().all()
                ]
                project_q_all = project_q_all.filter(Project.id.in_(p_ids or [-1]))

    districts = district_q.order_by(District.name).all()
    valid_district_ids = {d.id for d in districts}
    scoped_project_ids = None
    if not is_master_or_admin:
        scoped_project_ids = {p.id for p in project_q_all.order_by(Project.name).all()}
    return {
        'allowed_projects': allowed_projects,
        'allowed_districts': allowed_districts,
        'allowed_vehicles': allowed_vehicles,
        'is_master_or_admin': is_master_or_admin,
        'districts': districts,
        'valid_district_ids': valid_district_ids,
        'scoped_project_ids': scoped_project_ids,
    }









# Excel header -> model field mapping for EmergencyTaskRecord
_EMG_HEADER_MAP = {
    'taskid': 'task_id_ext',
    'requestfrom': 'request_from',
    'phone': 'phone',
    'cli': 'cli',
    'name': 'name',
    'husband': 'husband',
    'address': 'address',
    'location': 'location',
    'housecolor': 'house_color',
    'doorcolor': 'door_color',
    'nearestlandmark': 'nearest_landmark',
    'edd': 'edd',
    'clinicaldetails': 'clinical_details',
    'districtname': 'district_name',
    'tehsilname': 'tehsil_name',
    'ucname': 'uc_name',
    'ambregno': 'amb_reg_no',
    'status': 'status',
    'receivedby': 'received_by',
    'category': 'category',
    'subcategory': 'sub_category',
    'facilityname': 'facility_name',
    'facilitycode': 'facility_code',
    'facilitytype': 'facility_type',
    'changefacilitycomments': 'change_facility_comments',
    'createddate': 'excel_created_date',
    'completeddatetime': 'completed_date_time',
    'firsttranfercreateddate': 'first_transfer_created_date',
    'firsttranferclinicaldetails': 'first_transfer_clinical_details',
    'firsttranferfacilityname': 'first_transfer_facility_name',
    'firsttranferfacilitytype': 'first_transfer_facility_type',
    'firsttranferdoctordetail': 'first_transfer_doctor_detail',
    'secondtranfercreateddate': 'second_transfer_created_date',
    'secondtranferclinicaldetails': 'second_transfer_clinical_details',
    'secondtranferfacilityname': 'second_transfer_facility_name',
    'secondtranferfacilitytype': 'second_transfer_facility_type',
    'secondtranferdoctordetail': 'second_transfer_doctor_detail',
    'createdby': 'created_by',
    'createddate1': 'created_date1',
    'createdtime': 'created_time',
    'pregnancymonth': 'pregnancy_month',
    'closingremarks': 'closing_remarks',
    'pregnancymonthclosing': 'pregnancy_month_closing',
    'cliclosing': 'cli_closing',
    'taskclosedby': 'task_closed_by',
    'patientcnic': 'patient_cnic',
    'patientadmissionno': 'patient_admission_no',
    'requestfor': 'request_for',
    'closed_by': 'closed_by',
    'callername': 'caller_name',
    'taskstartlat': 'task_start_lat',
    'taskstartlon': 'task_start_lon',
    'taskendlat': 'task_end_lat',
    'taskendlon': 'task_end_lon',
    'rascow': 'ras_cow',
    'distanceinkm': 'distance_in_km',
    'nearresthealthfacility': 'nearrest_health_facility',
}


import re as _re

_VEHICLE_SUFFIX_RE = _re.compile(
    r'[\s\-]+(COW|USG\+P|USG|RAS|MNHC|EMS|NHP)\s*$',
    _re.IGNORECASE,
)
_ACTIVITY_TITLE_RE = _re.compile(r'activity\s*report\s*\(([^)]+)\)', _re.IGNORECASE)


_STRING_LIMIT_CACHE = {}


def _get_table_string_limits(table_name, model_cls=None):
    """Return {column_name: max_length} for VARCHAR-like columns."""
    if table_name in _STRING_LIMIT_CACHE:
        return _STRING_LIMIT_CACHE[table_name]

    limits = {}
    try:
        for col in inspect(db.engine).get_columns(table_name):
            col_name = col.get('name')
            col_type = col.get('type')
            max_len = getattr(col_type, 'length', None)
            if col_name and isinstance(max_len, int) and max_len > 0:
                limits[col_name] = max_len
    except Exception:
        # Fallback to ORM metadata if DB inspection is unavailable.
        if model_cls is not None:
            for col in model_cls.__table__.columns:
                max_len = getattr(col.type, 'length', None)
                if isinstance(max_len, int) and max_len > 0:
                    limits[col.name] = max_len

    _STRING_LIMIT_CACHE[table_name] = limits
    return limits


def _validate_string_lengths(table_name, model_cls, values, row_no, report_label):
    """Fail fast with clear row/column details before DB flush."""
    limits = _get_table_string_limits(table_name, model_cls=model_cls)
    if not limits:
        return

    violations = []
    for field, raw_val in values.items():
        if raw_val is None:
            continue
        max_len = limits.get(field)
        if not max_len:
            continue
        text_val = str(raw_val)
        actual_len = len(text_val)
        if actual_len > max_len:
            violations.append((field, max_len, actual_len, text_val[:140]))

    if not violations:
        return

    log_bits = '; '.join(
        f"{field}={actual}/{max_len} sample='{sample}'"
        for field, max_len, actual, sample in violations[:5]
    )
    app.logger.warning(
        "%s validation failed at row %s: %s",
        report_label,
        row_no,
        log_bits,
    )

    user_bits = ', '.join(
        f"{field} ({actual}/{max_len})"
        for field, max_len, actual, _ in violations[:3]
    )
    raise ValueError(
        f"{report_label} row {row_no}: value too long for {user_bits}. "
        "Please correct this row in Excel and upload again."
    )


def _build_upload_error_message(label, exc):
    raw = (str(exc) or exc.__class__.__name__).strip()
    low = raw.lower()
    if isinstance(exc, ValueError):
        return raw
    if isinstance(exc, (DataError, IntegrityError)) or 'value too long for type character varying' in low:
        return (
            f"{label} upload failed: one or more Excel values exceed database limits. "
            "Please verify row values and retry."
        )
    return f"{label} upload failed ({exc.__class__.__name__}). Please verify file format and retry."


def _normalize_vehicle_no(raw):
    """Strip known project suffixes from vehicle numbers in uploaded reports.
    e.g. 'LEG-17-2191 COW' -> 'LEG-17-2191', 'GBD-24-395-COW' -> 'GBD-24-395'
    """
    if not raw:
        return raw
    s = str(raw).strip()
    return _VEHICLE_SUFFIX_RE.sub('', s).strip()


def _read_rows_auto(file_obj):
    """Read rows from uploaded file; auto-detect XLSX vs TSV/CSV."""
    import io, openpyxl
    raw = file_obj.read()
    file_obj.seek(0)
    if raw[:4] == b'PK\x03\x04':
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
        rows = [list(r) for r in ws.iter_rows(min_row=1, values_only=True)]
        wb.close()
        return rows
    text = raw.decode('utf-8', errors='replace')
    lines = [l for l in text.splitlines() if l.strip()]
    if not lines:
        return []
    sep = '\t' if '\t' in lines[0] else ','
    return [line.split(sep) for line in lines]


def _parse_emergency_excel(f, task_date):
    """Parse EmergencyTaskReport (XLSX or TSV) and store all columns."""
    rows = _read_rows_auto(f)
    if not rows:
        return 0
    headers_raw = [str(c).strip() if c else '' for c in rows[0]]
    norm = lambda s: ''.join(ch for ch in s.lower() if ch.isalnum() or ch == '_')
    col_map = {}
    for idx, h in enumerate(headers_raw):
        key = norm(h)
        if key in _EMG_HEADER_MAP:
            col_map[idx] = _EMG_HEADER_MAP[key]
    if not col_map:
        raise ValueError("No recognised column headers found in EmergencyTaskReport file.")

    EmergencyTaskRecord.query.filter_by(task_date=task_date).delete()

    count = 0
    today = pk_date()
    for row_no, row in enumerate(rows[1:], start=2):
        vals = {}
        for idx, field in col_map.items():
            raw_val = row[idx] if idx < len(row) else None
            vals[field] = str(raw_val).strip() if raw_val is not None and str(raw_val).strip() else None
        if not vals.get('amb_reg_no') and not vals.get('task_id_ext'):
            continue
        if vals.get('amb_reg_no'):
            vals['amb_reg_no'] = _normalize_vehicle_no(vals['amb_reg_no'])
        _validate_string_lengths('emergency_task_record', EmergencyTaskRecord, vals, row_no, 'EmergencyTaskReport')
        rec = EmergencyTaskRecord(task_date=task_date, upload_date=today, **vals)
        db.session.add(rec)
        count += 1
    return count


def _parse_mileage_excel(f, task_date):
    """Parse Vehicle Mileage Report (XLSX with headers at row 10)."""
    import io, openpyxl
    raw = f.read()
    f.seek(0)
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    all_rows = [list(r) for r in ws.iter_rows(min_row=1, values_only=True)]
    wb.close()
    if not all_rows:
        return 0

    header_row_idx = None
    for ri, row in enumerate(all_rows):
        for cell in row:
            if cell and 'regno' in str(cell).lower().replace(' ', '').replace('_', ''):
                header_row_idx = ri
                break
        if header_row_idx is not None:
            break
    if header_row_idx is None:
        header_row_idx = 9

    headers = all_rows[header_row_idx] if header_row_idx < len(all_rows) else []
    norm_headers = [str(h).strip().lower().replace(' ', '').replace('_', '') if h else '' for h in headers]

    reg_col = mil_col = ptop_col = None
    dt_cols = []
    for i, nh in enumerate(norm_headers):
        if 'regno' in nh:
            reg_col = i
        elif nh in ('mileage', 'runningkms', 'runningkm') or 'mileage' in nh:
            mil_col = i
        elif 'ptop' in nh:
            ptop_col = i
        elif 'date' in nh or 'time' in nh:
            dt_cols.append(i)

    if reg_col is None:
        reg_col = 1
    if mil_col is None:
        mil_col = 6
    if ptop_col is None:
        ptop_col = 7

    VehicleMileageRecord.query.filter_by(task_date=task_date).delete()

    count = 0
    today = pk_date()
    for excel_row_no, row in enumerate(all_rows[header_row_idx + 1:], start=header_row_idx + 2):
        if not row or (reg_col < len(row) and row[reg_col] is None):
            continue
        v_no = _normalize_vehicle_no(str(row[reg_col]).strip()) if reg_col < len(row) and row[reg_col] else ''
        if not v_no:
            continue

        def _safe_float(val):
            if val is None:
                return 0
            try:
                return float(val)
            except (TypeError, ValueError):
                return 0

        def _safe_str(val):
            return str(val).strip() if val is not None and str(val).strip() else None

        dt_c = _safe_str(row[dt_cols[0]]) if len(dt_cols) > 0 and dt_cols[0] < len(row) else None
        dt_d = _safe_str(row[dt_cols[1]]) if len(dt_cols) > 1 and dt_cols[1] < len(row) else None
        dt_e = _safe_str(row[dt_cols[2]]) if len(dt_cols) > 2 and dt_cols[2] < len(row) else None
        dt_f = _safe_str(row[dt_cols[3]]) if len(dt_cols) > 3 and dt_cols[3] < len(row) else None

        vals = {
            'reg_no': v_no,
            'date_time_c': dt_c,
            'date_time_d': dt_d,
            'date_time_e': dt_e,
            'date_time_f': dt_f,
        }
        _validate_string_lengths('vehicle_mileage_record', VehicleMileageRecord, vals, excel_row_no, 'Vehicle Mileage report')

        rec = VehicleMileageRecord(
            task_date=task_date,
            upload_date=today,
            **vals,
            mileage=_safe_float(row[mil_col]) if mil_col < len(row) else 0,
            ptop=_safe_float(row[ptop_col]) if ptop_col < len(row) else 0,
        )
        db.session.add(rec)
        count += 1
    return count


def _extract_activity_vehicle_no(ws):
    """Extract vehicle number from title like: Activity Report (GBC-22-039)."""
    title_candidates = []
    for r in range(1, 5):
        for c in range(1, 11):
            v = ws.cell(r, c).value
            if v is not None:
                title_candidates.append(str(v).strip())
    for text in title_candidates:
        m = _ACTIVITY_TITLE_RE.search(text)
        if m:
            return _normalize_vehicle_no(m.group(1))
    return ''


def _activity_cell_safe_str(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d %H:%M:%S')
    s = str(v).strip()
    return s if s else None


def _activity_cell_safe_float(v):
    if v in (None, ''):
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _parse_activity_report_single_file(f, task_date, upload_date, seen_rows):
    """Parse one Tracker Activity workbook; adds VehicleActivityRecord rows to session.
    Returns (rows_added, workbook_processed): workbook_processed False if file had no bytes."""
    import io
    import openpyxl

    raw = f.read()
    f.seek(0)
    if not raw:
        return 0, False
    if raw[:4] != b'PK\x03\x04':
        raise ValueError(f"Tracker Activity Report '{f.filename}' must be .xlsx format.")

    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    vehicle_no = _extract_activity_vehicle_no(ws)
    if not vehicle_no:
        wb.close()
        raise ValueError(f"Vehicle number not found in title for file '{f.filename}'. Expected: Activity Report (VEHICLE-NO).")

    header_row_idx = 9  # Row 10 in Excel (0-based index)
    all_rows = [list(r) for r in ws.iter_rows(min_row=1, values_only=True)]
    wb.close()
    if len(all_rows) <= header_row_idx:
        return 0, False

    headers = all_rows[header_row_idx][:11]
    norm_headers = [str(h).strip().lower().replace('\xa0', ' ') if h is not None else '' for h in headers]
    expected = ['group name', 'record date time', 'location', 'speed', 'direction', 'distance', 'travel time', 'stop time', 'reason']
    if not all(any(tok in (norm_headers[i] or '') for tok in expected[i].split()) for i in range(min(9, len(norm_headers)))):
        if not any('record' in h and 'time' in h for h in norm_headers):
            raise ValueError(f"Invalid heading row in file '{f.filename}'. Expected activity columns at row 10.")

    count_rows = 0
    for excel_row_no, row in enumerate(all_rows[header_row_idx + 1:], start=header_row_idx + 2):
        cells = list(row[:11]) + [None] * max(0, 11 - len(row))
        group_name = _activity_cell_safe_str(cells[0])
        record_date_time = _activity_cell_safe_str(cells[1])
        location = _activity_cell_safe_str(cells[2])
        direction = _activity_cell_safe_str(cells[4])
        travel_time = _activity_cell_safe_str(cells[6])
        stop_time = _activity_cell_safe_str(cells[7])
        reason = _activity_cell_safe_str(cells[8])
        speed = _activity_cell_safe_float(cells[3])
        distance = _activity_cell_safe_float(cells[5])
        latitude = _activity_cell_safe_float(cells[9]) if cells[9] not in (None, '') else None
        longitude = _activity_cell_safe_float(cells[10]) if cells[10] not in (None, '') else None

        if group_name and group_name.strip().lower() == 'total':
            continue
        if not any([group_name, record_date_time, location, direction, travel_time, stop_time, reason, speed, distance]):
            continue

        dedup_key = (vehicle_no, group_name or '', record_date_time or '', location or '', speed, direction or '', distance, travel_time or '', stop_time or '', reason or '', latitude, longitude)
        if dedup_key in seen_rows:
            continue
        seen_rows.add(dedup_key)

        vals = {
            'vehicle_no': vehicle_no,
            'group_name': group_name,
            'record_date_time': record_date_time,
            'location': location,
            'speed': speed,
            'direction': direction,
            'distance': distance,
            'travel_time': travel_time,
            'stop_time': stop_time,
            'reason': reason,
            'latitude': latitude,
            'longitude': longitude,
            'source_file': f.filename,
        }
        _validate_string_lengths('vehicle_activity_record', VehicleActivityRecord, vals, excel_row_no, 'Tracker Activity Report')

        db.session.add(VehicleActivityRecord(
            task_date=task_date,
            upload_date=upload_date,
            **vals,
        ))
        count_rows += 1

    return count_rows, True


def _parse_activity_report_excels(files, task_date):
    """Parse multiple Tracker Activity Report files (one per vehicle)."""
    valid_files = [f for f in (files or []) if f and getattr(f, 'filename', '').strip()]
    if not valid_files:
        return {'files': 0, 'rows': 0}

    VehicleActivityRecord.query.filter_by(task_date=task_date).delete()

    count_rows = 0
    count_files = 0
    today = pk_date()
    seen_rows = set()

    for f in valid_files:
        rows_i, processed = _parse_activity_report_single_file(f, task_date, today, seen_rows)
        count_rows += rows_i
        if processed:
            count_files += 1

    return {'files': count_files, 'rows': count_rows}


def _json_task_date():
    td = parse_date(request.form.get('task_date'))
    return td






def _norm_district_name_key(name):
    return (name or '').strip().lower()


# ────────────────────────────────────────────────
# Red Task Report
# ────────────────────────────────────────────────





# ────────────────────────────────────────────────
# Vehicle Move without Task Report
# ────────────────────────────────────────────────

def _ensure_unexecuted_task_table():
    UnexecutedTaskRecord.__table__.create(bind=db.session.get_bind(), checkfirst=True)


def _parse_emg_datetime(raw):
    s = (raw or '').strip()
    if not s:
        return None
    for fmt in (
        '%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M',
        '%d-%m-%Y %H:%M:%S', '%d-%m-%Y %H:%M',
        '%d %b %Y %H:%M:%S', '%d %b %Y %H:%M',
        '%d/%m/%Y %H:%M:%S', '%d/%m/%Y %H:%M',
        '%Y-%m-%d %I:%M:%S %p', '%Y-%m-%d %I:%M %p',
        '%d-%m-%Y %I:%M:%S %p', '%d-%m-%Y %I:%M %p',
        '%d/%m/%Y %I:%M:%S %p', '%d/%m/%Y %I:%M %p',
        '%d-%m-%y %H:%M:%S', '%d-%m-%y %H:%M',
        '%d/%m/%y %H:%M:%S', '%d/%m/%y %H:%M',
        '%d-%m-%y %I:%M:%S %p', '%d-%m-%y %I:%M %p',
        '%d/%m/%y %I:%M:%S %p', '%d/%m/%y %I:%M %p',
        '%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%d-%m-%y', '%d/%m/%y',
    ):
        try:
            return datetime.strptime(s, fmt)
        except (ValueError, TypeError):
            continue
    return _parse_activity_datetime(s)


def _shift_from_datetime(dt):
    if not dt:
        return '-'
    t = dt.time()
    return 'Day' if time(8, 0) <= t < time(20, 0) else 'Night'


def _fmt_duration(delta):
    if not delta:
        return '-'
    secs = int(delta.total_seconds())
    if secs < 0:
        secs = 0
    h = secs // 3600
    m = (secs % 3600) // 60
    return f'{h:02d}:{m:02d}'


def _unexecuted_task_rows(from_date, to_date, district_id=0, project_id=0, vehicle_id=0, category='', shift='',
                          check_type='', running_km_limit=None,
                          allowed_projects=None, allowed_districts=None, allowed_vehicles=None,
                          is_master_or_admin=True):
    allowed_projects = set(allowed_projects or [])
    allowed_districts = set(allowed_districts or [])
    allowed_vehicles = set(allowed_vehicles or [])
    selected_district = db.session.get(District, district_id) if district_id else None
    selected_district_name = (selected_district.name or '').strip().lower() if selected_district else ''

    def _norm_vno(vno):
        return (vno or '').strip().upper()

    emg_q = EmergencyTaskRecord.query.filter(
        EmergencyTaskRecord.task_date >= from_date,
        EmergencyTaskRecord.task_date <= to_date,
        EmergencyTaskRecord.category.in_(['Green', 'Yellow']),
        EmergencyTaskRecord.completed_date_time.isnot(None),
        EmergencyTaskRecord.excel_created_date.isnot(None),
    )
    if category in ('Green', 'Yellow'):
        emg_q = emg_q.filter(EmergencyTaskRecord.category == category)
    if vehicle_id:
        v = db.session.get(Vehicle, vehicle_id)
        emg_q = emg_q.filter(EmergencyTaskRecord.amb_reg_no == (v.vehicle_no if v else ''))

    emg_rows = emg_q.order_by(EmergencyTaskRecord.task_date.desc(), EmergencyTaskRecord.id.desc()).all()
    emg_vnos = [_norm_vno(r.amb_reg_no) for r in emg_rows if r.amb_reg_no]
    db_vehicles = Vehicle.query.filter(Vehicle.vehicle_no.in_(emg_vnos)).all() if emg_vnos else []
    vehicle_map = {_norm_vno(v.vehicle_no): v for v in db_vehicles}

    # Global eligibility guard (applies to all users): only include records where
    # district-project assignment exists and vehicle has district deployment in master data.
    assigned_project_pairs = set(
        (int(pid), int(did))
        for pid, did in db.session.query(project_district.c.project_id, project_district.c.district_id).all()
    )

    saved_map = {r.emergency_task_record_id: r for r in UnexecutedTaskRecord.query.filter(
        UnexecutedTaskRecord.emergency_task_record_id.in_([r.id for r in emg_rows])
    ).all()} if emg_rows else {}

    out_rows = []
    for r in emg_rows:
        assign_dt = _parse_emg_datetime(r.excel_created_date)
        close_dt = _parse_emg_datetime(r.completed_date_time)
        if not assign_dt or not close_dt:
            continue
        if close_dt < assign_dt:
            continue

        v = vehicle_map.get(_norm_vno(r.amb_reg_no))
        if not v or not v.project_id or not v.district_id:
            continue
        if (int(v.project_id), int(v.district_id)) not in assigned_project_pairs:
            continue

        # Apply user scope first (like Tracker Difference Report behavior).
        if not is_master_or_admin:
            if not v:
                continue
            # Vehicle scope is mandatory for non-admin users.
            if not allowed_vehicles or v.id not in allowed_vehicles:
                continue
            if allowed_districts and v.district_id not in allowed_districts:
                continue
            if allowed_projects and v.project_id not in allowed_projects:
                continue

        # Apply explicit filter values strictly. If vehicle mapping is missing, fallback to EMG district text
        # for district filter; for project/vehicle filters skip unmapped rows.
        if district_id:
            if v:
                if v.district_id != district_id:
                    continue
            else:
                emg_district_name = (r.district_name or '').strip().lower()
                if not selected_district_name or emg_district_name != selected_district_name:
                    continue
        if project_id and (not v or v.project_id != project_id):
            continue
        if vehicle_id and (not v or v.id != vehicle_id):
            continue

        activity_km = 0.0
        if v:
            acts = VehicleActivityRecord.query.filter(
                VehicleActivityRecord.vehicle_no == v.vehicle_no,
                VehicleActivityRecord.task_date >= assign_dt.date(),
                VehicleActivityRecord.task_date <= close_dt.date(),
            ).all()
            for a in acts:
                adt = _parse_activity_datetime(a.record_date_time)
                if adt and assign_dt <= adt <= close_dt:
                    activity_km += float(a.distance or 0)
        activity_km = round(activity_km, 2)

        if running_km_limit is not None:
            if check_type == 'above' and not (activity_km > running_km_limit):
                continue
            if check_type == 'below' and not (activity_km < running_km_limit):
                continue

        total_time = close_dt - assign_dt
        row_shift = _shift_from_datetime(assign_dt)
        if shift in ('day', 'night') and row_shift.lower() != shift:
            continue

        saved = saved_map.get(r.id)
        assigned_drivers = Driver.query.filter_by(vehicle_id=(v.id if v else None), status='Active').order_by(Driver.name).all() if v else []
        if saved and saved.driver_id:
            _saved_drv = db.session.get(Driver, saved.driver_id)
            if _saved_drv and all(d.id != _saved_drv.id for d in assigned_drivers):
                assigned_drivers.append(_saved_drv)

        out_rows.append({
            'emg': r,
            'vehicle': v,
            'district': v.district if v and v.district else None,
            'project': v.project if v and v.project else None,
            'assign_dt': assign_dt,
            'close_dt': close_dt,
            'total_time': _fmt_duration(total_time),
            'activity_km': activity_km,
            'shift': row_shift,
            'drivers': assigned_drivers,
            'saved': saved,
        })
    return out_rows


def _filter_unexecuted_rows_by_search(rows, table_search):
    s = (table_search or '').strip().lower()
    if not s:
        return rows
    out = []
    for row in rows:
        blob = ' '.join([
            row['district'].name if row.get('district') else (row['emg'].district_name or ''),
            row['project'].name if row.get('project') else '',
            row['vehicle'].vehicle_no if row.get('vehicle') else (row['emg'].amb_reg_no or ''),
            row['emg'].task_id_ext or '',
            row['assign_dt'].strftime('%d-%m-%Y %I:%M %p') if row.get('assign_dt') else '',
            row['close_dt'].strftime('%d-%m-%Y %I:%M %p') if row.get('close_dt') else '',
            row['total_time'] or '',
            row['emg'].category or '',
            f"{row['activity_km']:.2f}",
            row['shift'] or '',
            (row['saved'].driver.name if row.get('saved') and row['saved'].driver else ''),
            (row['saved'].fine if row.get('saved') else ''),
        ]).lower()
        if s in blob:
            out.append(row)
    return out







# ────────────────────────────────────────────────
# Penalty Record (Driver Status)
# ────────────────────────────────────────────────

def _penalty_record_query(from_date, to_date, district_id=0, project_id=0):
    """Shared query for penalty list/export/print."""
    query = PenaltyRecord.query.filter(
        PenaltyRecord.record_date >= from_date,
        PenaltyRecord.record_date <= to_date
    )
    if district_id:
        query = query.filter(PenaltyRecord.district_id == district_id)
    if project_id:
        query = query.filter(PenaltyRecord.project_id == project_id)
    return query.order_by(PenaltyRecord.record_date.desc(), PenaltyRecord.id.desc())






# ────────────────────────────────────────────────
# Party Name (Pump / Workshop / Spare parts shop)
# ────────────────────────────────────────────────












# ────────────────────────────────────────────────
# Products Name (used in Fueling / Oil / Maintenance)
# ────────────────────────────────────────────────









# ────────────────────────────────────────────────
# Fuel Expense
# ────────────────────────────────────────────────
_vehicle_maintenance_baseline_schema_ready = {'ok': False}


def _ensure_vehicle_maintenance_baseline_schema():
    if _vehicle_maintenance_baseline_schema_ready.get('ok'):
        return
    stmts = [
        """
        CREATE TABLE IF NOT EXISTS workspace_vehicle_maintenance_baseline (
            id SERIAL PRIMARY KEY,
            employee_id INTEGER NOT NULL,
            district_id INTEGER NULL,
            project_id INTEGER NULL,
            vehicle_id INTEGER NOT NULL,
            product_id INTEGER NOT NULL,
            job_category VARCHAR(120) NULL,
            interval_mode VARCHAR(20) NULL,
            interval_value INTEGER NULL,
            last_done_date DATE NULL,
            last_done_reading NUMERIC(12, 2) NULL,
            remarks TEXT NULL,
            created_by_user_id INTEGER NULL,
            created_at TIMESTAMP NULL DEFAULT NOW(),
            updated_at TIMESTAMP NULL DEFAULT NOW()
        )
        """,
        "ALTER TABLE workspace_vehicle_maintenance_baseline ADD COLUMN IF NOT EXISTS employee_id INTEGER",
        "ALTER TABLE workspace_vehicle_maintenance_baseline ADD COLUMN IF NOT EXISTS district_id INTEGER",
        "ALTER TABLE workspace_vehicle_maintenance_baseline ADD COLUMN IF NOT EXISTS project_id INTEGER",
        "ALTER TABLE workspace_vehicle_maintenance_baseline ADD COLUMN IF NOT EXISTS vehicle_id INTEGER",
        "ALTER TABLE workspace_vehicle_maintenance_baseline ADD COLUMN IF NOT EXISTS product_id INTEGER",
        "ALTER TABLE workspace_vehicle_maintenance_baseline ADD COLUMN IF NOT EXISTS job_category VARCHAR(120)",
        "ALTER TABLE workspace_vehicle_maintenance_baseline ADD COLUMN IF NOT EXISTS interval_mode VARCHAR(20)",
        "ALTER TABLE workspace_vehicle_maintenance_baseline ADD COLUMN IF NOT EXISTS interval_value INTEGER",
        "ALTER TABLE workspace_vehicle_maintenance_baseline ADD COLUMN IF NOT EXISTS last_done_date DATE",
        "ALTER TABLE workspace_vehicle_maintenance_baseline ADD COLUMN IF NOT EXISTS last_done_reading NUMERIC(12, 2)",
        "ALTER TABLE workspace_vehicle_maintenance_baseline ADD COLUMN IF NOT EXISTS remarks TEXT",
        "ALTER TABLE workspace_vehicle_maintenance_baseline ADD COLUMN IF NOT EXISTS created_by_user_id INTEGER",
        "ALTER TABLE workspace_vehicle_maintenance_baseline ADD COLUMN IF NOT EXISTS created_at TIMESTAMP",
        "ALTER TABLE workspace_vehicle_maintenance_baseline ADD COLUMN IF NOT EXISTS updated_at TIMESTAMP",
        "CREATE INDEX IF NOT EXISTS ix_ws_maint_baseline_employee_id ON workspace_vehicle_maintenance_baseline (employee_id)",
        "CREATE INDEX IF NOT EXISTS ix_ws_maint_baseline_vehicle_id ON workspace_vehicle_maintenance_baseline (vehicle_id)",
        "CREATE INDEX IF NOT EXISTS ix_ws_maint_baseline_product_id ON workspace_vehicle_maintenance_baseline (product_id)",
        "CREATE UNIQUE INDEX IF NOT EXISTS uq_ws_vehicle_maint_emp_vehicle_product ON workspace_vehicle_maintenance_baseline (employee_id, vehicle_id, product_id)",
    ]
    for stmt in stmts:
        try:
            db.session.execute(text(stmt))
            db.session.commit()
        except Exception:
            db.session.rollback()
    for stmt in [
        "ALTER TABLE workspace_vehicle_maintenance_baseline ALTER COLUMN product_id DROP NOT NULL",
        "DROP INDEX IF EXISTS uq_ws_vehicle_maint_emp_vehicle_product",
        "CREATE INDEX IF NOT EXISTS ix_ws_maint_baseline_emp_veh_jobcat ON workspace_vehicle_maintenance_baseline (employee_id, vehicle_id, job_category)",
    ]:
        try:
            db.session.execute(text(stmt))
            db.session.commit()
        except Exception:
            db.session.rollback()
    _vehicle_maintenance_baseline_schema_ready['ok'] = True


def _vehicle_latest_recorded_reading(vehicle_id):
    if not vehicle_id:
        return None
    latest_values = []
    last_fuel = FuelExpense.query.filter(
        FuelExpense.vehicle_id == vehicle_id,
        FuelExpense.current_reading.isnot(None)
    ).order_by(FuelExpense.fueling_date.desc(), FuelExpense.id.desc()).first()
    if last_fuel and last_fuel.current_reading is not None:
        latest_values.append(float(last_fuel.current_reading))
    last_oil = OilExpense.query.filter(
        OilExpense.vehicle_id == vehicle_id,
        OilExpense.current_reading.isnot(None)
    ).order_by(OilExpense.expense_date.desc(), OilExpense.id.desc()).first()
    if last_oil and last_oil.current_reading is not None:
        latest_values.append(float(last_oil.current_reading))
    last_maint = MaintenanceExpense.query.filter(
        MaintenanceExpense.vehicle_id == vehicle_id,
        MaintenanceExpense.current_reading.isnot(None)
    ).order_by(MaintenanceExpense.expense_date.desc(), MaintenanceExpense.id.desc()).first()
    if last_maint and last_maint.current_reading is not None:
        latest_values.append(float(last_maint.current_reading))
    return max(latest_values) if latest_values else None


def _job_category_config_by_name(name):
    if not name or not str(name).strip():
        return None
    n = (name or '').strip().lower()
    for j in _get_maintenance_job_categories() or []:
        if (j.get('name') or '').strip().lower() == n:
            return j
    return None


def _canonical_job_category_name(user_input):
    u = (user_input or '').strip()
    if not u:
        return None
    for j in _get_maintenance_job_categories() or []:
        jn = (j.get('name') or '').strip()
        if jn and jn.lower() == u.lower():
            return jn
    return None


def _merged_interval_for_baseline(baseline):
    """Row-stored interval (legacy) or from Add Maintenance job category settings."""
    mode = getattr(baseline, 'interval_mode', None) or None
    val = getattr(baseline, 'interval_value', None)
    try:
        val = int(val) if val is not None else None
    except (TypeError, ValueError):
        val = None
    if mode in ('interval_km', 'interval_day') and val and val > 0:
        return mode, val
    jc = getattr(baseline, 'job_category', None)
    cfg = _job_category_config_by_name(jc) if jc else None
    if not cfg:
        return None, None
    m = (cfg.get('interval_mode') or '').strip()
    if m not in ('interval_km', 'interval_day'):
        m = None
    try:
        v = int(float(cfg.get('interval_value')))
    except (TypeError, ValueError):
        v = None
    if m and v and v > 0:
        return m, v
    return None, None


def _format_baseline_reading_display(val):
    """Odometer / km: no trailing .00 (e.g. 2.00 -> 2; 19493.5 stays 19493.5)."""
    if val is None:
        return '-'
    try:
        x = float(val)
    except (TypeError, ValueError):
        return '-'
    r = round(x, 8)
    if abs(r - round(r)) < 1e-7:
        return str(int(round(r)))
    s = f'{r:.8f}'.rstrip('0').rstrip('.')
    return s


def _baseline_status(baseline, latest_reading=None):
    today = pk_date()
    status = 'No Interval'
    next_due_date = None
    next_due_reading = None
    remaining_days = None
    remaining_km = None
    mode, ival = _merged_interval_for_baseline(baseline)
    if mode == 'interval_day' and ival and baseline.last_done_date:
        next_due_date = baseline.last_done_date + timedelta(days=int(ival))
        remaining_days = (next_due_date - today).days
        if remaining_days < 0:
            status = 'Overdue'
        elif remaining_days <= 7:
            status = 'Due Soon'
        else:
            status = 'On Track'
    elif mode == 'interval_km' and ival and baseline.last_done_reading is not None:
        next_due_reading = float(baseline.last_done_reading) + float(ival)
        if latest_reading is None:
            status = 'Reading Needed'
        else:
            remaining_km = next_due_reading - float(latest_reading)
            if remaining_km < 0:
                status = 'Overdue'
            elif remaining_km <= 500:
                status = 'Due Soon'
            else:
                status = 'On Track'
    return {
        'status': status,
        'next_due_date': next_due_date,
        'next_due_date_label': next_due_date.strftime('%d-%m-%Y') if next_due_date else '-',
        'next_due_reading': next_due_reading,
        'next_due_reading_label': _format_baseline_reading_display(next_due_reading)
        if next_due_reading is not None
        else '-',
        'remaining_days': remaining_days,
        'remaining_days_label': str(remaining_days) if remaining_days is not None else '-',
        'remaining_km': remaining_km,
        'remaining_km_label': _format_baseline_reading_display(remaining_km) if remaining_km is not None else '-',
    }




def _vehicle_reading_setup_rows(employee_id, from_date=None, to_date=None, district_id=0, project_id=0, vehicle_id=0, search=''):
    q = WorkspaceVehicleReadingSetup.query.filter_by(employee_id=employee_id)
    if from_date:
        q = q.filter(WorkspaceVehicleReadingSetup.setup_date >= from_date)
    if to_date:
        q = q.filter(WorkspaceVehicleReadingSetup.setup_date <= to_date)
    if district_id:
        q = q.filter(WorkspaceVehicleReadingSetup.district_id == district_id)
    if project_id:
        q = q.filter(WorkspaceVehicleReadingSetup.project_id == project_id)
    if vehicle_id:
        q = q.filter(WorkspaceVehicleReadingSetup.vehicle_id == vehicle_id)
    if search:
        like = f"%{search}%"
        q = q.join(Vehicle, WorkspaceVehicleReadingSetup.vehicle_id == Vehicle.id).filter(
            or_(
                Vehicle.vehicle_no.ilike(like),
                WorkspaceVehicleReadingSetup.remarks.ilike(like),
            )
        )
    return q.order_by(
        WorkspaceVehicleReadingSetup.setup_date.desc(),
        WorkspaceVehicleReadingSetup.id.desc(),
    ).all()







def _fuel_expense_task_readings(vehicle_id, task_date):
    """Return (km_out_day_start, km_in_day_close) from VehicleDailyTask for given vehicle and date. None if no task."""
    task = VehicleDailyTask.query.filter_by(vehicle_id=vehicle_id, task_date=task_date).first()
    if not task:
        return None, None
    prev = VehicleDailyTask.query.filter(
        VehicleDailyTask.vehicle_id == vehicle_id,
        VehicleDailyTask.task_date < task_date
    ).order_by(VehicleDailyTask.task_date.desc()).first()
    km_out = float(prev.close_reading) if prev else None
    km_in = float(task.close_reading)
    return km_out, km_in


def _fuel_expense_previous_reading(vehicle_id, fueling_date=None, exclude_id=None, workspace_employee_id=None, current_reading=None):
    if not vehicle_id:
        return None
    q = FuelExpense.query.filter(FuelExpense.vehicle_id == vehicle_id)
    if exclude_id:
        q = q.filter(FuelExpense.id != exclude_id)
    if fueling_date and current_reading is not None:
        try:
            curr_val = float(current_reading)
        except (TypeError, ValueError):
            curr_val = None
        if curr_val is not None:
            same_day_prev = q.filter(
                FuelExpense.fueling_date == fueling_date,
                FuelExpense.current_reading.isnot(None),
                FuelExpense.current_reading < curr_val,
            ).order_by(
                FuelExpense.current_reading.desc(),
                FuelExpense.id.desc(),
            ).first()
            if same_day_prev and same_day_prev.current_reading is not None:
                return float(same_day_prev.current_reading)
    if fueling_date:
        if current_reading is not None:
            q = q.filter(FuelExpense.fueling_date < fueling_date)
        elif exclude_id:
            q = q.filter(
                db.or_(
                    FuelExpense.fueling_date < fueling_date,
                    db.and_(FuelExpense.fueling_date == fueling_date, FuelExpense.id < int(exclude_id)),
                )
            )
        else:
            q = q.filter(FuelExpense.fueling_date <= fueling_date)
    last_entry = q.order_by(FuelExpense.fueling_date.desc(), FuelExpense.id.desc()).first()
    if last_entry and last_entry.current_reading is not None:
        return float(last_entry.current_reading)
    return _fallback_vehicle_previous_reading(workspace_employee_id, vehicle_id, 'fuel')


def _resequence_vehicle_fuel_expenses(vehicle_id, workspace_employee_id=None):
    if not vehicle_id:
        return
    rows = FuelExpense.query.filter(
        FuelExpense.vehicle_id == vehicle_id
    ).order_by(
        FuelExpense.fueling_date.asc(),
        db.case((FuelExpense.current_reading.is_(None), 1), else_=0).asc(),
        FuelExpense.current_reading.asc(),
        FuelExpense.id.asc(),
    ).all()
    previous_current = None
    for idx, row in enumerate(rows):
        if idx == 0:
            if row.previous_reading is None:
                row.previous_reading = _fallback_vehicle_previous_reading(
                    workspace_employee_id or row.employee_id, vehicle_id, 'fuel'
                )
        else:
            row.previous_reading = previous_current

        prev_val = float(row.previous_reading) if row.previous_reading is not None else None
        curr_val = float(row.current_reading) if row.current_reading is not None else None
        row.km = (curr_val - prev_val) if (prev_val is not None and curr_val is not None) else None
        liters_val = float(row.liters) if row.liters is not None else None
        row.mpg = round(float(row.km) / liters_val, 2) if (row.km is not None and liters_val and liters_val > 0) else None

        previous_current = curr_val















def _workspace_expense_by_choices(employee_id):
    ensure_workspace_base_accounts(employee_id)
    rows = WorkspaceAccount.query.filter_by(employee_id=employee_id, is_active=True).order_by(WorkspaceAccount.code.asc()).all()
    driver_ids = sorted({
        int(a.entity_id) for a in rows
        if (a.entity_type or '').strip().lower() == 'driver' and a.entity_id
    })
    drivers_by_id = {}
    if driver_ids:
        for drv in Driver.query.filter(Driver.id.in_(driver_ids)).all():
            drivers_by_id[int(drv.id)] = drv
    vehicle_ids = sorted({
        int(getattr(drv, 'vehicle_id', 0) or 0)
        for drv in drivers_by_id.values()
        if getattr(drv, 'vehicle_id', None)
    })
    vehicles_by_id = {}
    if vehicle_ids:
        for veh in Vehicle.query.filter(Vehicle.id.in_(vehicle_ids)).all():
            vehicles_by_id[int(veh.id)] = veh

    choices = [('', '-- Default (Auto from Workspace COA) --')]

    def _balance_side(account_type, bal):
        if account_type in ('Asset', 'Expense'):
            return 'Dr' if bal >= 0 else 'Cr'
        return 'Cr' if bal >= 0 else 'Dr'

    for a in rows:
        # Show only likely payment/counterparty heads for cleaner dropdown.
        if a.account_type == 'Expense' or a.code in ('1000', '5000', '5100'):
            continue
        label = f"{a.code} - {a.name} ({a.account_type})"
        if (a.entity_type or '').strip().lower() == 'driver' and a.entity_id:
            drv = drivers_by_id.get(int(a.entity_id))
            veh = vehicles_by_id.get(int(drv.vehicle_id)) if drv and getattr(drv, 'vehicle_id', None) else None
            vehicle_no = (veh.vehicle_no if veh and getattr(veh, 'vehicle_no', None) else None) or ''
            if vehicle_no:
                label = f"{label} | Vehicle: {vehicle_no}"
        bal = Decimal(str(a.current_balance or 0))
        side = _balance_side(a.account_type, bal)
        sign = '+' if bal > 0 else ('-' if bal < 0 else '')
        label = f"{label} | Bal: {sign}{abs(bal):,.2f} {side}"
        choices.append((f'acct-{a.id}', label))
    return choices


def _workspace_default_cash_expense_by(employee_id):
    if not employee_id:
        return ''
    ensure_workspace_base_accounts(employee_id)
    cash_account = WorkspaceAccount.query.filter_by(
        employee_id=employee_id,
        code='1100',
        is_active=True,
    ).first()
    return f'acct-{cash_account.id}' if cash_account else ''


def _workspace_default_hbl_expense_by(employee_id):
    if not employee_id:
        return ''
    ensure_workspace_base_accounts(employee_id)
    hbl_account = WorkspaceAccount.query.filter_by(
        employee_id=employee_id,
        code='1110',
        is_active=True,
    ).first()
    return f'acct-{hbl_account.id}' if hbl_account else ''


def _workspace_account_id_from_expense_by(expense_by_val, employee_id):
    if not expense_by_val:
        return None
    parts = str(expense_by_val).split('-', 1)
    if len(parts) != 2 or parts[0] != 'acct' or not parts[1].isdigit():
        return None
    acct_id = int(parts[1])
    acct = WorkspaceAccount.query.filter_by(id=acct_id, employee_id=employee_id, is_active=True).first()
    return acct.id if acct else None


def _workspace_journal_expense_by_account_id(journal_entry, lines_by_je, exclude_account_ids, accounts_by_id):
    """Credit-side workspace account used to pay/settle an expense (Expense By)."""
    lines = lines_by_je.get(journal_entry.id, [])
    credit_lines = sorted(
        [
            ln for ln in lines
            if Decimal(str(ln.credit or 0)) > 0 and ln.account_id not in exclude_account_ids
        ],
        key=lambda x: Decimal(str(x.credit or 0)),
        reverse=True,
    )
    for ln in credit_lines:
        acct = accounts_by_id.get(ln.account_id)
        if acct and acct.code not in ('5100', '5000'):
            return acct.id
    return None


def _workspace_expense_by_account_id_for_reference(employee_id, reference_type, reference_id):
    if not employee_id or not reference_type or not reference_id:
        return None
    rows = WorkspaceJournalEntry.query.filter_by(
        employee_id=employee_id,
        reference_type=reference_type,
        reference_id=reference_id,
    ).order_by(WorkspaceJournalEntry.id.desc()).all()
    if not rows:
        return None
    lines_by_je = {}
    line_rows = WorkspaceJournalEntryLine.query.filter(
        WorkspaceJournalEntryLine.journal_entry_id.in_([je.id for je in rows])
    ).all()
    for ln in line_rows:
        lines_by_je.setdefault(ln.journal_entry_id, []).append(ln)
    ensure_workspace_base_accounts(employee_id)
    expense_head = WorkspaceAccount.query.filter_by(employee_id=employee_id, code='5100').first()
    exclude_ids = {expense_head.id} if expense_head else set()
    acct_ids = {ln.account_id for ln in line_rows}
    accounts_by_id = {}
    if acct_ids:
        for acct in WorkspaceAccount.query.filter(
            WorkspaceAccount.id.in_(list(acct_ids)),
            WorkspaceAccount.employee_id == employee_id,
            WorkspaceAccount.is_active == True,
        ).all():
            accounts_by_id[acct.id] = acct
    transfer_rows = [je for je in rows if (je.entry_type or '').strip().lower() == 'transfer']
    expense_rows = [je for je in rows if (je.entry_type or '').strip().lower() == 'expense']
    for group in (transfer_rows, expense_rows):
        for je in group:
            acct_id = _workspace_journal_expense_by_account_id(je, lines_by_je, exclude_ids, accounts_by_id)
            if acct_id:
                return acct_id
    return None


def _workspace_expense_by_label_for_reference(employee_id, reference_type, reference_id):
    acct_id = _workspace_expense_by_account_id_for_reference(employee_id, reference_type, reference_id)
    if not acct_id:
        return ''
    acct = WorkspaceAccount.query.filter_by(id=acct_id, employee_id=employee_id, is_active=True).first()
    if not acct:
        return ''
    return f"{acct.code} - {acct.name}"


def _workspace_expense_by_labels_for_references(employee_id, reference_type, reference_ids):
    ids = sorted({int(x) for x in reference_ids if x})
    if not employee_id or not reference_type or not ids:
        return {}
    entries = WorkspaceJournalEntry.query.filter(
        WorkspaceJournalEntry.employee_id == employee_id,
        WorkspaceJournalEntry.reference_type == reference_type,
        WorkspaceJournalEntry.reference_id.in_(ids),
    ).order_by(
        WorkspaceJournalEntry.reference_id.asc(),
        WorkspaceJournalEntry.id.desc(),
    ).all()
    if not entries:
        return {ref_id: '' for ref_id in ids}
    lines_by_je = {}
    line_rows = WorkspaceJournalEntryLine.query.filter(
        WorkspaceJournalEntryLine.journal_entry_id.in_([je.id for je in entries])
    ).all()
    for ln in line_rows:
        lines_by_je.setdefault(ln.journal_entry_id, []).append(ln)
    ensure_workspace_base_accounts(employee_id)
    expense_head = WorkspaceAccount.query.filter_by(employee_id=employee_id, code='5100').first()
    exclude_ids = {expense_head.id} if expense_head else set()
    acct_ids = {ln.account_id for ln in line_rows}
    accounts_by_id = {}
    if acct_ids:
        for acct in WorkspaceAccount.query.filter(
            WorkspaceAccount.id.in_(list(acct_ids)),
            WorkspaceAccount.employee_id == employee_id,
            WorkspaceAccount.is_active == True,
        ).all():
            accounts_by_id[acct.id] = acct
    by_ref = {}
    for je in entries:
        by_ref.setdefault(je.reference_id, []).append(je)
    labels = {}
    for ref_id in ids:
        acct_id = None
        je_list = by_ref.get(ref_id, [])
        transfer_rows = [je for je in je_list if (je.entry_type or '').strip().lower() == 'transfer']
        expense_rows = [je for je in je_list if (je.entry_type or '').strip().lower() == 'expense']
        for group in (transfer_rows, expense_rows):
            for je in group:
                acct_id = _workspace_journal_expense_by_account_id(je, lines_by_je, exclude_ids, accounts_by_id)
                if acct_id:
                    break
            if acct_id:
                break
        acct = accounts_by_id.get(acct_id) if acct_id else None
        labels[ref_id] = f"{acct.code} - {acct.name}" if acct else ''
    return labels


def _workspace_expense_by_for_reference(employee_id, reference_type, reference_id):
    acct_id = _workspace_expense_by_account_id_for_reference(employee_id, reference_type, reference_id)
    return f'acct-{acct_id}' if acct_id else ''


def _fuel_expense_ids_matching_expense_by_search(employee_id, search_token):
    """Fuel expense IDs whose Workspace COA payer account matches a single search token."""
    if not employee_id or not search_token:
        return None
    like_q = f"%{search_token.strip()}%"
    driver_join = db.and_(
        func.lower(WorkspaceAccount.entity_type) == 'driver',
        WorkspaceAccount.entity_id == Driver.id,
    )
    return (
        db.session.query(WorkspaceJournalEntry.reference_id)
        .join(
            WorkspaceJournalEntryLine,
            WorkspaceJournalEntryLine.journal_entry_id == WorkspaceJournalEntry.id,
        )
        .join(
            WorkspaceAccount,
            WorkspaceAccount.id == WorkspaceJournalEntryLine.account_id,
        )
        .outerjoin(Driver, driver_join)
        .filter(
            WorkspaceJournalEntry.employee_id == employee_id,
            WorkspaceJournalEntry.reference_type == 'FuelExpense',
            WorkspaceJournalEntry.reference_id.isnot(None),
            WorkspaceJournalEntry.entry_type.in_(['Expense', 'Transfer']),
            WorkspaceJournalEntryLine.credit > 0,
            WorkspaceAccount.employee_id == employee_id,
            ~WorkspaceAccount.code.in_(['5100', '5000']),
            or_(
                WorkspaceAccount.name.ilike(like_q),
                WorkspaceAccount.code.ilike(like_q),
                (WorkspaceAccount.code + ' - ' + WorkspaceAccount.name).ilike(like_q),
                Driver.name.ilike(like_q),
            ),
        )
        .distinct()
    )


def _fuel_expense_list_search_filter(search_q, workspace_employee_id):
    """Multi-word AND search: each token must match at least one searchable field."""
    tokens = [t for t in (search_q or '').split() if t]
    if not tokens:
        return None
    search_columns = [
        FuelExpense.slip_no,
        FuelExpense.payment_type,
        FuelExpense.fuel_type,
        cast(FuelExpense.id, SAString),
        cast(FuelExpense.current_reading, SAString),
        cast(FuelExpense.previous_reading, SAString),
        cast(FuelExpense.amount, SAString),
        cast(FuelExpense.liters, SAString),
        cast(FuelExpense.km, SAString),
        cast(FuelExpense.fueling_date, SAString),
        cast(FuelExpense.card_swipe_date, SAString),
    ]
    token_clauses = []
    for tok in tokens:
        like_tok = f'%{tok}%'
        token_or_parts = [col.ilike(like_tok) for col in search_columns]
        token_or_parts.extend([
            FuelExpense.vehicle_id.in_(db.session.query(Vehicle.id).filter(Vehicle.vehicle_no.ilike(like_tok))),
            FuelExpense.project_id.in_(db.session.query(Project.id).filter(Project.name.ilike(like_tok))),
            FuelExpense.district_id.in_(db.session.query(District.id).filter(District.name.ilike(like_tok))),
            FuelExpense.workspace_pump_id.in_(db.session.query(WorkspaceParty.id).filter(WorkspaceParty.name.ilike(like_tok))),
            FuelExpense.fuel_pump_id.in_(db.session.query(Party.id).filter(Party.name.ilike(like_tok))),
        ])
        expense_by_ids = _fuel_expense_ids_matching_expense_by_search(workspace_employee_id, tok)
        if expense_by_ids is not None:
            token_or_parts.append(FuelExpense.id.in_(expense_by_ids))
        token_clauses.append(or_(*token_or_parts))
    return and_(*token_clauses)


def _require_workspace_employee_for_expense_management():
    """Expense Management is now part of Employee Workspace."""
    if not session.get('workspace_employee_id'):
        flash('Employee Workspace select karna zaroori hai.', 'warning')
        return redirect(url_for('workspace_dashboard'))
    return None


def _workspace_employee_id_for_expenses():
    return session.get('workspace_employee_id')


def _safe_internal_path(return_to, default_path):
    """App-relative path only: prevents open redirects from return_to query param."""
    from urllib.parse import unquote
    if not return_to or not isinstance(return_to, str):
        return default_path
    raw = unquote(return_to.strip())
    if not raw or not raw.startswith('/') or raw.startswith('//'):
        return default_path
    if any(c in raw for c in '\n\r\x00'):
        return default_path
    if len(raw) > 2000:
        return default_path
    return raw


def _workspace_employee_default_district_id(employee_id):
    if not employee_id:
        return None
    emp = db.session.get(Employee, employee_id)
    if not emp:
        return None
    try:
        first_district = emp.districts.order_by(District.name.asc()).first()
    except Exception:
        first_district = None
    return first_district.id if first_district else None


def _workspace_employee_default_project_id(employee_id, preferred_district_id=None):
    if not employee_id:
        return None
    emp = db.session.get(Employee, employee_id)
    if not emp:
        return None
    preferred_district_id = int(preferred_district_id) if preferred_district_id else None
    try:
        emp_projects = list(emp.projects)
    except Exception:
        emp_projects = []
    if not emp_projects:
        return None
    emp_projects.sort(key=lambda p: (p.name or '').lower())
    if preferred_district_id:
        for p in emp_projects:
            try:
                linked = p.districts.filter(District.id == preferred_district_id).first() if hasattr(p.districts, 'filter') else None
            except Exception:
                linked = None
            if linked:
                return p.id
    return emp_projects[0].id


def _fallback_vehicle_previous_reading(employee_id, vehicle_id, mode):
    if not vehicle_id:
        return None
    setup = None
    if employee_id:
        setup = WorkspaceVehicleReadingSetup.query.filter_by(
            employee_id=employee_id,
            vehicle_id=vehicle_id,
        ).first()
    if not setup:
        setup = WorkspaceVehicleReadingSetup.query.filter_by(
            vehicle_id=vehicle_id,
        ).order_by(WorkspaceVehicleReadingSetup.id.desc()).first()
    if not setup:
        return None
    if mode == 'fuel':
        raw = setup.fuel_previous_reading
    elif mode == 'oil':
        raw = setup.oil_previous_reading
    elif mode == 'maintenance':
        _ensure_vehicle_maintenance_baseline_schema()
        baseline_q = WorkspaceVehicleMaintenanceBaseline.query.filter(
            WorkspaceVehicleMaintenanceBaseline.vehicle_id == vehicle_id
        )
        if employee_id:
            baseline_q = baseline_q.filter(WorkspaceVehicleMaintenanceBaseline.employee_id == employee_id)
        baseline = baseline_q.order_by(
            db.case((WorkspaceVehicleMaintenanceBaseline.last_done_date.is_(None), 1), else_=0).asc(),
            WorkspaceVehicleMaintenanceBaseline.last_done_date.desc(),
            db.case((WorkspaceVehicleMaintenanceBaseline.last_done_reading.is_(None), 1), else_=0).asc(),
            WorkspaceVehicleMaintenanceBaseline.last_done_reading.desc(),
            WorkspaceVehicleMaintenanceBaseline.id.desc(),
        ).first()
        if baseline and baseline.last_done_reading is not None:
            return float(baseline.last_done_reading)
        raw = None
    else:
        raw = None
    return float(raw) if raw is not None else None


def _workspace_reverse_expense_journals(reference_type, reference_id, employee_id):
    if not employee_id:
        return
    rows = WorkspaceJournalEntry.query.filter_by(
        employee_id=employee_id,
        reference_type=reference_type,
        reference_id=reference_id
    ).all()
    for je in rows:
        workspace_reverse_journal_entry(je.id)


_WORKSPACE_REGULAR_EXPENSE_SYNC_AVAILABLE = None


def _workspace_regular_expense_sync_available():
    global _WORKSPACE_REGULAR_EXPENSE_SYNC_AVAILABLE
    if _WORKSPACE_REGULAR_EXPENSE_SYNC_AVAILABLE is not None:
        return _WORKSPACE_REGULAR_EXPENSE_SYNC_AVAILABLE
    try:
        insp = inspect(db.engine)
        if not insp.has_table('workspace_expense'):
            _WORKSPACE_REGULAR_EXPENSE_SYNC_AVAILABLE = False
            return False
        cols = {c.get('name') for c in insp.get_columns('workspace_expense')}
        required = {
            'employee_id', 'expense_number', 'expense_date', 'expense_type',
            'description', 'amount', 'payment_mode', 'category',
            'workspace_party_id', 'journal_entry_id',
        }
        _WORKSPACE_REGULAR_EXPENSE_SYNC_AVAILABLE = required.issubset(cols)
    except Exception:
        _WORKSPACE_REGULAR_EXPENSE_SYNC_AVAILABLE = False
    return _WORKSPACE_REGULAR_EXPENSE_SYNC_AVAILABLE


def _workspace_regular_expense_number(reference_type, reference_id):
    if not reference_type or not reference_id:
        return ''
    key = str(reference_type).strip()
    try:
        rid = int(reference_id)
    except (TypeError, ValueError):
        return ''
    mapping = {
        'FuelExpense': 'FUEL',
        'OilExpense': 'OIL',
        'MaintenanceExpense': 'MAINT',
    }
    prefix = mapping.get(key, key.upper()[:12])
    return f'{prefix}-{rid}'


def _workspace_sync_regular_expense(employee_id, reference_type, reference_id, expense_date, amount,
                                    description, expense_type, payment_mode, category,
                                    workspace_party_id=None, journal_entry_id=None):
    if not _workspace_regular_expense_sync_available():
        return None
    if not employee_id:
        return None
    exp_no = _workspace_regular_expense_number(reference_type, reference_id)
    if not exp_no:
        return None
    amount_val = Decimal(str(amount or 0))
    existing = WorkspaceExpense.query.filter_by(
        employee_id=employee_id,
        expense_number=exp_no,
    ).first()
    if amount_val <= Decimal('0'):
        if existing:
            db.session.delete(existing)
        return None
    if not existing:
        existing = WorkspaceExpense(
            employee_id=employee_id,
            expense_number=exp_no,
            created_by_user_id=session.get('user_id'),
        )
        db.session.add(existing)
    existing.expense_date = expense_date or pk_date()
    existing.expense_type = expense_type or reference_type or 'Expense'
    existing.workspace_party_id = workspace_party_id
    existing.workspace_product_id = None
    existing.to_driver_id = None
    existing.description = description or existing.expense_type
    existing.amount = amount_val
    existing.payment_mode = payment_mode or 'Cash'
    existing.category = category or None
    existing.journal_entry_id = journal_entry_id
    return existing


def _workspace_delete_regular_expense(employee_id, reference_type, reference_id):
    if not _workspace_regular_expense_sync_available():
        return
    if not employee_id:
        return
    exp_no = _workspace_regular_expense_number(reference_type, reference_id)
    if not exp_no:
        return
    row = WorkspaceExpense.query.filter_by(employee_id=employee_id, expense_number=exp_no).first()
    if row:
        db.session.delete(row)


def _workspace_post_expense_journal(employee_id, reference_type, reference_id, expense_date, amount, description, category_code, workspace_party_id=None, credit_account_id=None):
    if not employee_id:
        return None
    amount_val = Decimal(str(amount or 0))
    if amount_val <= Decimal("0"):
        return None

    ensure_workspace_base_accounts(employee_id)
    expense_head = WorkspaceAccount.query.filter_by(employee_id=employee_id, code='5100').first()
    cash_head = WorkspaceAccount.query.filter_by(employee_id=employee_id, code='1100').first()
    if not (expense_head and cash_head):
        return None

    selected_credit_id = credit_account_id
    credit_account_id = cash_head.id
    if selected_credit_id:
        user_credit = WorkspaceAccount.query.filter_by(id=selected_credit_id, employee_id=employee_id, is_active=True).first()
        if user_credit:
            credit_account_id = user_credit.id
    if workspace_party_id:
        try:
            party_acct = ensure_workspace_counterparty_account(employee_id, party_id=workspace_party_id)
            if party_acct:
                # Explicit selected account takes priority; otherwise party account.
                if not credit_account_id or credit_account_id == cash_head.id:
                    credit_account_id = party_acct.id
        except Exception:
            pass

    line_desc = description or 'Expense posted'
    return workspace_create_journal_entry(
        employee_id=employee_id,
        entry_type='Expense',
        entry_date=expense_date or pk_date(),
        description=description or reference_type,
        lines=[
            {'account_id': expense_head.id, 'debit': amount_val, 'credit': 0, 'description': line_desc},
            {'account_id': credit_account_id, 'debit': 0, 'credit': amount_val, 'description': line_desc},
        ],
        reference_type=reference_type,
        reference_id=reference_id,
        created_by_user_id=session.get('user_id'),
        category=category_code,
    )


def _workspace_post_credit_settlement_journal(employee_id, reference_type, reference_id, expense_date, amount,
                                              category_code, workspace_party_id, credit_account_id, description):
    if not employee_id or not workspace_party_id or not credit_account_id:
        return None
    settle_amount = Decimal(str(amount or 0))
    if settle_amount <= 0:
        return None
    try:
        party_acct = ensure_workspace_counterparty_account(employee_id, party_id=workspace_party_id)
    except Exception:
        party_acct = None
    if not party_acct or int(credit_account_id) == int(party_acct.id):
        return None
    settle_desc = description or 'Credit settlement'
    return workspace_create_journal_entry(
        employee_id=employee_id,
        entry_type='Transfer',
        entry_date=expense_date or pk_date(),
        description=settle_desc,
        lines=[
            {
                'account_id': party_acct.id,
                'debit': settle_amount,
                'credit': 0,
                'description': settle_desc,
            },
            {
                'account_id': credit_account_id,
                'debit': 0,
                'credit': settle_amount,
                'description': settle_desc,
            },
        ],
        reference_type=reference_type,
        reference_id=reference_id,
        created_by_user_id=session.get('user_id'),
        category=category_code,
    )









# ────────────────────────────────────────────────
# Oil Expense
# ────────────────────────────────────────────────

def _oil_expense_previous_reading(vehicle_id, expense_date=None, exclude_id=None, workspace_employee_id=None, current_reading=None):
    if not vehicle_id:
        return None
    q = OilExpense.query.filter(OilExpense.vehicle_id == vehicle_id)
    if exclude_id:
        q = q.filter(OilExpense.id != exclude_id)
    if expense_date and current_reading is not None:
        try:
            curr_val = float(current_reading)
        except (TypeError, ValueError):
            curr_val = None
        if curr_val is not None:
            same_day_prev = q.filter(
                OilExpense.expense_date == expense_date,
                OilExpense.current_reading.isnot(None),
                OilExpense.current_reading < curr_val,
            ).order_by(
                OilExpense.current_reading.desc(),
                OilExpense.id.desc(),
            ).first()
            if same_day_prev and same_day_prev.current_reading is not None:
                return float(same_day_prev.current_reading)
    if expense_date:
        if current_reading is not None:
            q = q.filter(OilExpense.expense_date < expense_date)
        elif exclude_id:
            q = q.filter(
                db.or_(
                    OilExpense.expense_date < expense_date,
                    db.and_(OilExpense.expense_date == expense_date, OilExpense.id < int(exclude_id)),
                )
            )
        else:
            q = q.filter(OilExpense.expense_date <= expense_date)
    last_entry = q.order_by(OilExpense.expense_date.desc(), OilExpense.id.desc()).first()
    if last_entry and last_entry.current_reading is not None:
        return float(last_entry.current_reading)
    return _fallback_vehicle_previous_reading(workspace_employee_id, vehicle_id, 'oil')


def _resequence_vehicle_oil_expenses(vehicle_id, workspace_employee_id=None):
    if not vehicle_id:
        return
    rows = OilExpense.query.filter(
        OilExpense.vehicle_id == vehicle_id
    ).order_by(
        OilExpense.expense_date.asc(),
        db.case((OilExpense.current_reading.is_(None), 1), else_=0).asc(),
        OilExpense.current_reading.asc(),
        OilExpense.id.asc(),
    ).all()
    previous_current = None
    for idx, row in enumerate(rows):
        if idx == 0:
            if row.previous_reading is None:
                row.previous_reading = _fallback_vehicle_previous_reading(
                    workspace_employee_id or row.employee_id, vehicle_id, 'oil'
                )
        else:
            row.previous_reading = previous_current

        prev_val = float(row.previous_reading) if row.previous_reading is not None else None
        curr_val = float(row.current_reading) if row.current_reading is not None else None
        row.km = (curr_val - prev_val) if (prev_val is not None and curr_val is not None) else None
        previous_current = curr_val




def _expense_history_num_label(val):
    try:
        n = float(val)
    except (TypeError, ValueError):
        return '-'
    return f'{n:.2f}'


def _build_oil_product_price_history(product_id, workspace_party_id=None, current_id=None, workspace_employee_id=None, expense_date=None):
    q = db.session.query(OilExpenseItem, OilExpense, WorkspaceParty).join(
        OilExpense, OilExpense.id == OilExpenseItem.oil_expense_id
    ).outerjoin(
        WorkspaceParty, WorkspaceParty.id == OilExpense.workspace_party_id
    ).filter(
        OilExpenseItem.product_id == product_id
    )
    if workspace_employee_id:
        q = q.filter(OilExpense.employee_id == workspace_employee_id)
    if current_id:
        q = q.filter(OilExpense.id != current_id)
    rows = q.order_by(OilExpense.expense_date.desc(), OilExpense.id.desc(), OilExpenseItem.id.desc()).limit(120).all()

    same_party = []
    other_parties = []
    match_same_date_same_party = None
    match_same_date = None
    match_same_party = None
    match_any = None
    for it, rec, party in rows:
        qty = float(it.purchase_qty if it.purchase_qty is not None else (it.qty if it.qty is not None else 0))
        price = float(it.price or 0)
        if qty <= 0 and price <= 0:
            continue
        total = float(it.amount) if it.amount is not None else (qty * price)
        row = {
            'date_label': rec.expense_date.strftime('%d-%m-%Y') if rec.expense_date else '-',
            'invoice_no': f'OIL-{rec.id}',
            'party_name': (party.name if party else '-'),
            'qty_label': _expense_history_num_label(qty),
            'price_label': _expense_history_num_label(price),
            'total_label': _expense_history_num_label(total),
        }
        if workspace_party_id and rec.workspace_party_id == workspace_party_id:
            same_party.append(row)
        elif rec.workspace_party_id and rec.workspace_party_id != workspace_party_id:
            other_parties.append(row)
        elif not workspace_party_id:
            other_parties.append(row)
        if price <= 0:
            continue
        rec_date = rec.expense_date
        if match_any is None:
            match_any = price
        if workspace_party_id and rec.workspace_party_id == workspace_party_id and match_same_party is None:
            match_same_party = price
        if expense_date and rec_date == expense_date and match_same_date is None:
            match_same_date = price
        if expense_date and workspace_party_id and rec_date == expense_date and rec.workspace_party_id == workspace_party_id and match_same_date_same_party is None:
            match_same_date_same_party = price
    suggested_price = match_same_date_same_party
    if suggested_price is None:
        suggested_price = match_same_date
    if suggested_price is None:
        suggested_price = match_same_party
    if suggested_price is None:
        suggested_price = match_any
    return same_party[:8], other_parties[:8], suggested_price


def _build_maintenance_product_price_history(product_id, workspace_party_id=None, current_id=None, workspace_employee_id=None, expense_date=None):
    q = db.session.query(MaintenanceExpenseItem, MaintenanceExpense, WorkspaceParty).join(
        MaintenanceExpense, MaintenanceExpense.id == MaintenanceExpenseItem.maintenance_expense_id
    ).outerjoin(
        WorkspaceParty, WorkspaceParty.id == MaintenanceExpense.workspace_party_id
    ).filter(
        MaintenanceExpenseItem.product_id == product_id
    )
    if workspace_employee_id:
        q = q.filter(MaintenanceExpense.employee_id == workspace_employee_id)
    if current_id:
        q = q.filter(MaintenanceExpense.id != current_id)
    rows = q.order_by(MaintenanceExpense.expense_date.desc(), MaintenanceExpense.id.desc(), MaintenanceExpenseItem.id.desc()).limit(120).all()

    same_party = []
    other_parties = []
    match_same_date_same_party = None
    match_same_date = None
    match_same_party = None
    match_any = None
    for it, rec, party in rows:
        qty = float(it.qty or 0)
        price = float(it.price or 0)
        if qty <= 0 and price <= 0:
            continue
        total = float(it.amount) if it.amount is not None else (qty * price)
        row = {
            'date_label': rec.expense_date.strftime('%d-%m-%Y') if rec.expense_date else '-',
            'invoice_no': f'MAINT-{rec.id}',
            'party_name': (party.name if party else '-'),
            'qty_label': _expense_history_num_label(qty),
            'price_label': _expense_history_num_label(price),
            'total_label': _expense_history_num_label(total),
        }
        if workspace_party_id and rec.workspace_party_id == workspace_party_id:
            same_party.append(row)
        elif rec.workspace_party_id and rec.workspace_party_id != workspace_party_id:
            other_parties.append(row)
        elif not workspace_party_id:
            other_parties.append(row)
        if price <= 0:
            continue
        rec_date = rec.expense_date
        if match_any is None:
            match_any = price
        if workspace_party_id and rec.workspace_party_id == workspace_party_id and match_same_party is None:
            match_same_party = price
        if expense_date and rec_date == expense_date and match_same_date is None:
            match_same_date = price
        if expense_date and workspace_party_id and rec_date == expense_date and rec.workspace_party_id == workspace_party_id and match_same_date_same_party is None:
            match_same_date_same_party = price
    suggested_price = match_same_date_same_party
    if suggested_price is None:
        suggested_price = match_same_date
    if suggested_price is None:
        suggested_price = match_same_party
    if suggested_price is None:
        suggested_price = match_any
    return same_party[:8], other_parties[:8], suggested_price



def _get_or_create_product_balance(product_id):
    bal = ProductBalance.query.filter_by(product_id=product_id).first()
    if not bal:
        bal = ProductBalance(product_id=product_id, balance_qty=0)
        db.session.add(bal)
    return bal


def _apply_oil_expense_items_balance(items, reverse=False):
    for item in items:
        bal = _get_or_create_product_balance(item.product_id)
        purchase_qty = float(item.purchase_qty or 0)
        used_qty = float(item.used_qty or 0)
        delta = (purchase_qty - used_qty) if not reverse else (used_qty - purchase_qty)
        if delta:
            from decimal import Decimal
            bal.balance_qty = Decimal(str(bal.balance_qty or 0)) + Decimal(str(delta))
    db.session.flush()


def _workspace_products_for_expense_form(employee_id, form_token):
    """Return Product rows for workspace products, creating master Product mirrors when needed.

    Oil/Maintenance item tables reference master Product FK, so workspace products are mapped by
    name and auto-synced on demand.
    """
    token = (form_token or '').strip()
    if not employee_id:
        rows = Product.query.filter(
            db.or_(
                Product.used_in_forms.is_(None),
                Product.used_in_forms == '',
                Product.used_in_forms.like(f'%{token}%'),
            )
        ).order_by(Product.name).all()
        for p in rows:
            # Product table may not have default_price; keep template/API access safe.
            if not hasattr(p, 'default_price'):
                setattr(p, 'default_price', None)
        return rows

    ws_query = WorkspaceProduct.query.filter(
        WorkspaceProduct.employee_id == employee_id,
        WorkspaceProduct.is_active.is_(True),
    )
    if token:
        ws_query = ws_query.filter(
            db.or_(
                WorkspaceProduct.used_in_forms.is_(None),
                WorkspaceProduct.used_in_forms == '',
                WorkspaceProduct.used_in_forms.ilike(f'%{token}%'),
            )
        )
    ws_rows = ws_query.order_by(WorkspaceProduct.name.asc()).all()
    ws_default_price_by_name = {}
    for ws in ws_rows:
        key = (ws.name or '').strip().lower()
        if not key:
            continue
        # Prefer latest non-null configured default price per workspace product name.
        if key not in ws_default_price_by_name or ws.default_price is not None:
            ws_default_price_by_name[key] = ws.default_price

    mapped = []
    changed = False
    for ws in ws_rows:
        name = (ws.name or '').strip()
        if not name:
            continue
        prod = Product.query.filter(func.lower(Product.name) == name.lower()).first()
        if not prod:
            prod = Product(name=name, used_in_forms=token or None, remarks=ws.remarks)
            db.session.add(prod)
            try:
                db.session.flush()
                changed = True
            except IntegrityError:
                db.session.rollback()
                prod = Product.query.filter(func.lower(Product.name) == name.lower()).first()
                if not prod:
                    continue
        elif token:
            existing_tokens = [x.strip() for x in (prod.used_in_forms or '').split(',') if x.strip()]
            if token not in existing_tokens:
                existing_tokens.append(token)
                prod.used_in_forms = ','.join(existing_tokens)
                changed = True
        default_price = ws_default_price_by_name.get(name.lower())
        if not hasattr(prod, 'default_price'):
            setattr(prod, 'default_price', default_price)
        else:
            prod.default_price = default_price
        mapped.append(prod)

    if changed:
        db.session.commit()
    # Keep deterministic ordering and avoid duplicates if names collide by case.
    uniq = {}
    for p in mapped:
        key = (p.name or '').strip().lower()
        if not hasattr(p, 'default_price'):
            setattr(p, 'default_price', ws_default_price_by_name.get(key))
        uniq[p.id] = p
    return sorted(uniq.values(), key=lambda x: (x.name or '').lower())









# ────────────────────────────────────────────────
# Maintenance Expense
# ────────────────────────────────────────────────



def _maintenance_approval_qty_label(qty_val):
    try:
        qv = Decimal(str(qty_val))
    except Exception:
        qv = Decimal('0')
    qv = qv.quantize(Decimal('0.01'))
    s = format(qv, 'f')
    if '.' in s:
        s = s.rstrip('0').rstrip('.')
    return s or '0'


def _maintenance_approval_detail_line(p_name, qty, amount):
    amt = int(round(float(amount or 0)))
    if 'labour' in (p_name or '').lower():
        return f'* {p_name} = Rs. {amt:,}'
    qty_lbl = _maintenance_approval_qty_label(qty)
    return f'* {p_name} ×{qty_lbl} = Rs. {amt:,}'


def _maintenance_approval_total_line(total):
    amt = int(round(float(total or 0)))
    return f'💰 Total Amount: Rs. {amt:,}/-'






def _maintenance_expense_previous_reading(vehicle_id, expense_date=None, exclude_id=None, workspace_employee_id=None, current_reading=None):
    if not vehicle_id:
        return None
    q = MaintenanceExpense.query.filter(MaintenanceExpense.vehicle_id == vehicle_id)
    if exclude_id:
        q = q.filter(MaintenanceExpense.id != exclude_id)
    if expense_date and current_reading is not None:
        try:
            curr_val = float(current_reading)
        except (TypeError, ValueError):
            curr_val = None
        if curr_val is not None:
            same_day_prev = q.filter(
                MaintenanceExpense.expense_date == expense_date,
                MaintenanceExpense.current_reading.isnot(None),
                MaintenanceExpense.current_reading < curr_val,
            ).order_by(
                MaintenanceExpense.current_reading.desc(),
                MaintenanceExpense.id.desc(),
            ).first()
            if same_day_prev and same_day_prev.current_reading is not None:
                return float(same_day_prev.current_reading)
    if expense_date:
        if current_reading is not None:
            q = q.filter(MaintenanceExpense.expense_date < expense_date)
        elif exclude_id:
            q = q.filter(
                db.or_(
                    MaintenanceExpense.expense_date < expense_date,
                    db.and_(MaintenanceExpense.expense_date == expense_date, MaintenanceExpense.id < int(exclude_id)),
                )
            )
        else:
            q = q.filter(MaintenanceExpense.expense_date <= expense_date)
    last_entry = q.order_by(MaintenanceExpense.expense_date.desc(), MaintenanceExpense.id.desc()).first()
    if last_entry and last_entry.current_reading is not None:
        return float(last_entry.current_reading)
    fallback = _fallback_vehicle_previous_reading(workspace_employee_id, vehicle_id, 'maintenance')
    return float(fallback) if fallback is not None else None


def _resequence_vehicle_maintenance_expenses(vehicle_id, workspace_employee_id=None):
    if not vehicle_id:
        return
    rows = MaintenanceExpense.query.filter(
        MaintenanceExpense.vehicle_id == vehicle_id
    ).order_by(
        MaintenanceExpense.expense_date.asc(),
        db.case((MaintenanceExpense.current_reading.is_(None), 1), else_=0).asc(),
        MaintenanceExpense.current_reading.asc(),
        MaintenanceExpense.id.asc(),
    ).all()
    previous_current = None
    for idx, row in enumerate(rows):
        if idx == 0:
            if row.previous_reading is None:
                row.previous_reading = _maintenance_expense_previous_reading(
                    vehicle_id=vehicle_id,
                    expense_date=row.expense_date,
                    exclude_id=row.id,
                    workspace_employee_id=(workspace_employee_id or row.employee_id),
                )
        else:
            row.previous_reading = previous_current

        prev_val = float(row.previous_reading) if row.previous_reading is not None else None
        curr_val = float(row.current_reading) if row.current_reading is not None else None
        row.km = (curr_val - prev_val) if (prev_val is not None and curr_val is not None) else None
        previous_current = curr_val







def _next_maintenance_work_order_no(opened_on=None):
    dt = opened_on or pk_date()
    yymm = dt.strftime('%y%m')
    prefix = f"MWO-{yymm}-"
    latest = MaintenanceWorkOrder.query.filter(
        MaintenanceWorkOrder.work_order_no.like(f"{prefix}%")
    ).order_by(MaintenanceWorkOrder.id.desc()).first()
    serial = 1
    if latest and latest.work_order_no:
        try:
            serial = int(str(latest.work_order_no).split('-')[-1]) + 1
        except Exception:
            serial = 1
    return f"{prefix}{serial:04d}"


_maintenance_work_order_schema_ready = {'ok': False}


def _ensure_maintenance_work_order_schema():
    """Safety net for environments where migration is delayed/missed.

    Keeps maintenance pages functional by creating required table/columns if absent.

    NOTE: SQLite does NOT support `ALTER TABLE ... ADD COLUMN IF NOT EXISTS`
    (raises "near EXISTS: syntax error") nor `SERIAL`. So we use SQLAlchemy's
    inspector to check existing columns and only add genuinely missing ones,
    which is valid on both SQLite and PostgreSQL.
    """
    if _maintenance_work_order_schema_ready.get('ok'):
        return
    is_sqlite = db.engine.dialect.name == 'sqlite'
    # CREATE TABLE IF NOT EXISTS works on both dialects; SERIAL is PG-only.
    _pk_def = 'INTEGER PRIMARY KEY AUTOINCREMENT' if is_sqlite else 'SERIAL PRIMARY KEY'
    create_stmts = [
        f"""
        CREATE TABLE IF NOT EXISTS maintenance_work_order (
            id {_pk_def},
            work_order_no VARCHAR(40),
            district_id INTEGER NULL,
            project_id INTEGER NULL,
            employee_id INTEGER NULL,
            vehicle_id INTEGER NULL,
            opened_on DATE NULL,
            closed_on DATE NULL,
            work_type VARCHAR(120) NULL,
            title VARCHAR(180) NULL,
            status VARCHAR(20) NULL DEFAULT 'open',
            remarks TEXT NULL,
            created_at TIMESTAMP NULL
        )
        """,
        "CREATE INDEX IF NOT EXISTS ix_maintenance_work_order_work_order_no ON maintenance_work_order (work_order_no)",
        f"""
        CREATE TABLE IF NOT EXISTS maintenance_work_order_attachment (
            id {_pk_def},
            work_order_id INTEGER NOT NULL,
            file_path VARCHAR(2048) NOT NULL,
            file_type VARCHAR(20),
            original_name VARCHAR(255),
            created_at TIMESTAMP
        )
        """,
        "CREATE INDEX IF NOT EXISTS ix_mwo_attachment_work_order_id ON maintenance_work_order_attachment (work_order_id)",
        "CREATE INDEX IF NOT EXISTS ix_maintenance_expense_work_order_id ON maintenance_expense (work_order_id)",
    ]
    # Columns that should exist (table, column, SQL type). Added only if missing.
    required_columns = [
        ('maintenance_work_order', 'work_order_no', 'VARCHAR(40)'),
        ('maintenance_work_order', 'district_id', 'INTEGER'),
        ('maintenance_work_order', 'project_id', 'INTEGER'),
        ('maintenance_work_order', 'employee_id', 'INTEGER'),
        ('maintenance_work_order', 'vehicle_id', 'INTEGER'),
        ('maintenance_work_order', 'opened_on', 'DATE'),
        ('maintenance_work_order', 'closed_on', 'DATE'),
        ('maintenance_work_order', 'work_type', 'VARCHAR(120)'),
        ('maintenance_work_order', 'title', 'VARCHAR(180)'),
        ('maintenance_work_order', 'status', 'VARCHAR(20)'),
        ('maintenance_work_order', 'remarks', 'TEXT'),
        ('maintenance_work_order', 'created_at', 'TIMESTAMP'),
        ('maintenance_work_order', 'upload_status', 'VARCHAR(20)'),
        ('maintenance_work_order', 'upload_total', 'INTEGER DEFAULT 0'),
        ('maintenance_work_order', 'upload_done', 'INTEGER DEFAULT 0'),
        ('maintenance_work_order', 'upload_failed', 'INTEGER DEFAULT 0'),
        ('maintenance_work_order', 'upload_error', 'TEXT'),
        ('maintenance_work_order', 'upload_manifest_json', 'TEXT'),
        ('maintenance_work_order', 'upload_started_at', 'TIMESTAMP'),
        ('maintenance_work_order', 'upload_finished_at', 'TIMESTAMP'),
        ('maintenance_expense', 'work_order_id', 'INTEGER'),
    ]
    try:
        existing_tables = set(inspect(db.engine).get_table_names())
        # 1) CREATE TABLE / INDEX statements (idempotent on both dialects)
        for stmt in create_stmts:
            db.session.execute(text(stmt))
        db.session.commit()
        # 2) Add genuinely missing columns only (avoids SQLite IF NOT EXISTS gap)
        for _tbl, _col, _coltype in required_columns:
            if _tbl not in existing_tables:
                continue  # table just created above with all columns
            _existing_cols = {c['name'] for c in inspect(db.engine).get_columns(_tbl)}
            if _col not in _existing_cols:
                db.session.execute(text(f'ALTER TABLE {_tbl} ADD COLUMN {_col} {_coltype}'))
        db.session.commit()
        _maintenance_work_order_schema_ready['ok'] = True
    except Exception:
        db.session.rollback()
        app.logger.exception('Maintenance work-order schema safety sync failed')













def _mwo_unified_gallery_build_items(wo):
    """Build combined media_items for bill + WO attachments (same shape as maintenance_expense_media)."""

    def _human_size(n):
        if n is None:
            return ''
        size = float(n)
        units = ['B', 'KB', 'MB', 'GB', 'TB']
        idx = 0
        while size >= 1024 and idx < len(units) - 1:
            size /= 1024.0
            idx += 1
        if idx == 0:
            return f"{int(size)} {units[idx]}"
        return f"{size:.1f} {units[idx]}"

    def _local_sz(stored_path):
        if not stored_path:
            return None
        p = str(stored_path).strip()
        if p.startswith('http://') or p.startswith('https://'):
            return None
        upload_root = os.path.abspath(app.config.get('UPLOAD_FOLDER', ''))
        if not upload_root:
            return None
        rel = p.replace('\\', '/').lstrip('/')
        if rel.startswith('uploads/'):
            rel = rel[len('uploads/'):]
        full = os.path.abspath(os.path.join(upload_root, rel.replace('/', os.sep)))
        if not full.startswith(upload_root):
            return None
        try:
            return os.path.getsize(full)
        except OSError:
            return None

    def _local_fp(stored_path):
        if not stored_path:
            return None
        p = str(stored_path).strip()
        if p.startswith('http://') or p.startswith('https://'):
            return None
        upload_root = os.path.abspath(app.config.get('UPLOAD_FOLDER', ''))
        if not upload_root:
            return None
        rel = p.replace('\\', '/').lstrip('/')
        if rel.startswith('uploads/'):
            rel = rel[len('uploads/'):]
        full = os.path.abspath(os.path.join(upload_root, rel.replace('/', os.sep)))
        if not full.startswith(upload_root):
            return None
        return full if os.path.isfile(full) else None

    media_items = []
    expenses = wo.expenses.order_by(MaintenanceExpense.expense_date.asc(), MaintenanceExpense.id.asc()).all()
    for exp in expenses:
        for att in exp.attachments.order_by(MaintenanceExpenseAttachment.created_at.asc(), MaintenanceExpenseAttachment.id.asc()).all():
            url = media_url_filter(att.file_path or '')
            if not url:
                continue
            ftype = (att.file_type or '').strip().lower()
            if ftype not in ('image', 'video'):
                path = (att.file_path or '').lower()
                if any(path.endswith(x) for x in ('.mp4', '.webm', '.mov')):
                    ftype = 'video'
                else:
                    ftype = 'image'
            size_bytes = _local_sz(att.file_path or '')
            created_at = att.created_at
            label = f"Bill MAINT-{exp.id}"
            media_items.append({
                'url': url,
                'type': ftype,
                'name': f"{label} — {att.original_name or os.path.basename(att.file_path or '') or 'Attachment'}",
                'created_at': created_at.strftime('%d-%m-%Y %I:%M %p') if created_at else '',
                'created_at_iso': created_at.isoformat() if created_at else '',
                'size_bytes': size_bytes,
                'size_label': _human_size(size_bytes),
                'download_url': url_for('maintenance_expense_media_download', pk=exp.id, att_id=att.id),
                'is_local_file': bool(_local_fp(att.file_path or '')),
            })
    wo_start = len(media_items)
    for att in wo.attachments.order_by(MaintenanceWorkOrderAttachment.created_at.asc(), MaintenanceWorkOrderAttachment.id.asc()).all():
        url = media_url_filter(att.file_path or '')
        if not url:
            continue
        ftype = (att.file_type or '').strip().lower()
        if ftype not in ('image', 'video'):
            path = (att.file_path or '').lower()
            if any(path.endswith(x) for x in ('.mp4', '.webm', '.mov')):
                ftype = 'video'
            else:
                ftype = 'image'
        size_bytes = _local_sz(att.file_path or '')
        created_at = att.created_at
        media_items.append({
            'url': url,
            'type': ftype,
            'name': f"WO {wo.work_order_no or ''} — {att.original_name or os.path.basename(att.file_path or '') or 'Attachment'}",
            'created_at': created_at.strftime('%d-%m-%Y %I:%M %p') if created_at else '',
            'created_at_iso': created_at.isoformat() if created_at else '',
            'size_bytes': size_bytes,
            'size_label': _human_size(size_bytes),
            'download_url': url_for('maintenance_work_order_media_download', pk=wo.id, att_id=att.id),
            'is_local_file': bool(_local_fp(att.file_path or '')),
        })
    return media_items, wo_start


def _mwo_job_gallery_build_items(wo):
    def _human_size(n):
        if n is None:
            return ''
        size = float(n)
        units = ['B', 'KB', 'MB', 'GB', 'TB']
        idx = 0
        while size >= 1024 and idx < len(units) - 1:
            size /= 1024.0
            idx += 1
        if idx == 0:
            return f"{int(size)} {units[idx]}"
        return f"{size:.1f} {units[idx]}"

    def _local_sz(stored_path):
        if not stored_path:
            return None
        p = str(stored_path).strip()
        if p.startswith('http://') or p.startswith('https://'):
            return None
        upload_root = os.path.abspath(app.config.get('UPLOAD_FOLDER', ''))
        if not upload_root:
            return None
        rel = p.replace('\\', '/').lstrip('/')
        if rel.startswith('uploads/'):
            rel = rel[len('uploads/'):]
        full = os.path.abspath(os.path.join(upload_root, rel.replace('/', os.sep)))
        if not full.startswith(upload_root):
            return None
        try:
            return os.path.getsize(full)
        except OSError:
            return None

    def _local_fp(stored_path):
        if not stored_path:
            return None
        p = str(stored_path).strip()
        if p.startswith('http://') or p.startswith('https://'):
            return None
        upload_root = os.path.abspath(app.config.get('UPLOAD_FOLDER', ''))
        if not upload_root:
            return None
        rel = p.replace('\\', '/').lstrip('/')
        if rel.startswith('uploads/'):
            rel = rel[len('uploads/'):]
        full = os.path.abspath(os.path.join(upload_root, rel.replace('/', os.sep)))
        if not full.startswith(upload_root):
            return None
        return full if os.path.isfile(full) else None

    media_items = []
    for att in wo.attachments.order_by(MaintenanceWorkOrderAttachment.created_at.asc(), MaintenanceWorkOrderAttachment.id.asc()).all():
        url = media_url_filter(att.file_path or '')
        if not url:
            continue
        ftype = (att.file_type or '').strip().lower()
        if ftype not in ('image', 'video'):
            path = (att.file_path or '').lower()
            if any(path.endswith(x) for x in ('.mp4', '.webm', '.mov')):
                ftype = 'video'
            else:
                ftype = 'image'
        size_bytes = _local_sz(att.file_path or '')
        created_at = att.created_at
        media_items.append({
            'url': url,
            'type': ftype,
            'name': att.original_name or os.path.basename(att.file_path or '') or 'Attachment',
            'created_at': created_at.strftime('%d-%m-%Y %I:%M %p') if created_at else '',
            'created_at_iso': created_at.isoformat() if created_at else '',
            'size_bytes': size_bytes,
            'size_label': _human_size(size_bytes),
            'download_url': url_for('maintenance_work_order_media_download', pk=wo.id, att_id=att.id),
            'is_local_file': bool(_local_fp(att.file_path or '')),
        })
    return media_items












def _maintenance_attachment_download_name(att):
    name = secure_filename((att.original_name or '').strip())
    if not name:
        base = secure_filename(os.path.basename(att.file_path or '').strip()) or f'attachment_{att.id}'
        name = base
    if '.' not in name:
        if (att.file_type or '').lower() == 'video':
            name += '.mp4'
        else:
            name += '.jpg'
    return name


def _maintenance_attachment_local_full_path(stored_path):
    if not stored_path:
        return None
    p = str(stored_path).strip()
    if p.startswith('http://') or p.startswith('https://'):
        return None
    upload_root = os.path.abspath(app.config.get('UPLOAD_FOLDER', ''))
    if not upload_root:
        return None
    rel = p.replace('\\', '/').lstrip('/')
    if rel.startswith('uploads/'):
        rel = rel[len('uploads/'):]
    full = os.path.abspath(os.path.join(upload_root, rel.replace('/', os.sep)))
    if not full.startswith(upload_root):
        return None
    return full if os.path.isfile(full) else None


def _maintenance_attachment_read_bytes(stored_path):
    full = _maintenance_attachment_local_full_path(stored_path)
    if full:
        with open(full, 'rb') as fh:
            data = fh.read()
        mt, _ = mimetypes.guess_type(full)
        return data, (mt or 'application/octet-stream')

    url = media_url_filter(stored_path or '')
    if not url:
        return b'', 'application/octet-stream'
    req = Request(url, headers={'User-Agent': 'fleet-manager/maintenance-media-download'})
    with urlopen(req, timeout=90) as resp:
        data = resp.read()
        ct = (resp.headers.get('Content-Type') or 'application/octet-stream').split(';')[0].strip()
    return data, (ct or 'application/octet-stream')










# ────────────────────────────────────────────────
# Employee Expense - routes now in routes_finance.py
# ────────────────────────────────────────────────


# ────────────────────────────────────────────────
# Reports Index & Multiple Report Types
# ────────────────────────────────────────────────
def _linked_driver_id_for_current_user():
    """If logged-in user matches a driver by CNIC (username), return driver.id for 'my profile' links."""
    try:
        uid = session.get('user_id')
        if not uid:
            return None
        user = db.session.get(User, uid)
        if not user:
            return None
        uname = (user.username or '').strip()
        if not uname:
            return None
        cnic_variants = [uname]
        digits = re.sub(r'\D', '', uname)
        if len(digits) == 13:
            cnic_variants.append(digits[:5] + '-' + digits[5:12] + '-' + digits[12:])
        for c in cnic_variants:
            drv = Driver.query.filter(func.lower(Driver.cnic_no) == c.lower()).first()
            if drv:
                return drv.id
    except Exception:
        return None
    return None


def _report_centre_badge_counts(can_page, is_master):
    """Actionable counts for Report Centre mobile tiles (permission-gated, best-effort)."""
    badges = {}

    def _set(key, val):
        n = int(val or 0)
        if n > 0:
            badges[key] = n if n < 100 else 99

    try:
        if can_page('missing_documents_report'):
            miss_q = Driver.query.filter(Driver.status != 'Left').filter(
                db.or_(
                    Driver.photo_path.is_(None), Driver.photo_path == '',
                    Driver.cnic_front_path.is_(None), Driver.cnic_front_path == '',
                    Driver.cnic_back_path.is_(None), Driver.cnic_back_path == '',
                    Driver.license_front_path.is_(None), Driver.license_front_path == '',
                    Driver.license_back_path.is_(None), Driver.license_back_path == '',
                    Driver.verify_license_photo_path.is_(None), Driver.verify_license_photo_path == '',
                    Driver.document_path.is_(None), Driver.document_path == '',
                )
            )
            _set('missing-docs', miss_q.count())
    except Exception:
        pass

    try:
        if can_page('book_pending_returns'):
            from models import BookAssignment
            _set('book-return', BookAssignment.query.filter_by(status='Active').count())
    except Exception:
        pass

    try:
        if can_page('oil_change_alert_report'):
            from auth_utils import get_user_context
            uid = session.get('user_id')
            ctx = get_user_context(uid) if uid else {}
            rows = _oil_change_alert_rows(
                statuses=['near', 'crossed'],
                allowed_projects=ctx.get('allowed_projects'),
                allowed_districts=ctx.get('allowed_districts'),
                allowed_vehicles=ctx.get('allowed_vehicles'),
                is_master_or_admin=ctx.get('is_master_or_admin', is_master),
            )
            _set('oil-alert', len(rows))
    except Exception:
        pass

    try:
        if can_page('red_task_list'):
            from models import RedTask
            today = pk_date()
            _set('red-task', RedTask.query.filter(RedTask.task_date == today).count())
    except Exception:
        pass

    try:
        if can_page('report_expiry'):
            from datetime import timedelta
            from auth_utils import get_user_context
            uid = session.get('user_id')
            ctx = get_user_context(uid) if uid else {}
            today = pk_date()
            horizon = today + timedelta(days=30)
            exp_q = Driver.query.filter(
                Driver.status != 'Left',
                Driver.vehicle_id.isnot(None),
            )
            if not ctx.get('is_master_or_admin', is_master):
                av = ctx.get('allowed_vehicles') or set()
                if av:
                    exp_q = exp_q.filter(Driver.vehicle_id.in_(list(av)))
            exp_q = exp_q.filter(
                db.or_(
                    db.and_(Driver.license_expiry_date.isnot(None), Driver.license_expiry_date <= horizon),
                    db.and_(Driver.cnic_expiry_date.isnot(None), Driver.cnic_expiry_date <= horizon),
                )
            )
            _set('expiry', exp_q.count())
    except Exception:
        pass

    return badges


def _report_centre_visibility(linked_driver_id=None):
    """Which Report Centre tabs/columns have at least one permitted link (same rules as can_see_page in template)."""
    from permissions_config import can_see_page as _can_pg

    perms = session.get('permissions') or []
    is_master = session.get('is_master', False)
    linked_id = linked_driver_id if linked_driver_id is not None else _linked_driver_id_for_current_user()

    def c(page_key):
        if is_master:
            return True
        return _can_pg(perms, page_key)

    fleet_vehicle = (
        c('report_vehicle_summary') or c('vehicles_list') or c('report_vehicle_profile')
        or c('report_parking_utilization')
    )
    fleet_project = c('report_project_summary') or c('report_district_summary') or c('report_company_profile')
    fleet_expense = (
        c('fuel_expense_list') or c('oil_expense_list') or c('maintenance_expense_list')
        or c('oil_change_alert_report')
    )
    show_fleet = bool(fleet_vehicle or fleet_project or fleet_expense)

    task_daily = (
        c('task_report_list') or c('task_report_new') or c('task_report_entry') or c('red_task_list') or c('red_task_summary') or c('red_task_summary_detail') or c('without_task_list')
        or c('speed_monitoring_report') or c('mileage_report') or c('tracker_difference_report')
        or c('unauthorized_movement_report') or c('task_start_delay_report') or c('task_turnaround_report')
        or c('unexecuted_task_report')
    )
    task_logbook = c('task_report_logbook_cover')
    task_upload = c('task_report_upload')
    show_task = bool(task_daily or task_logbook or task_upload)

    hr_driver = (
        (bool(linked_id) and c('report_driver_profile'))
        or c('active_drivers_report') or c('driver_seat_available_report') or c('missing_documents_report')
        or c('penalty_record_list') or c('driver_salary_slip')
    )
    hr_att = c('driver_attendance_report') or c('driver_attendance_tra_report') or c('report_expiry')
    hr_workforce = c('driver_job_left_list') or c('driver_rejoin_list')
    show_hr = bool(hr_driver or hr_att or hr_workforce)

    finance_books = (
        c('payment_vouchers_list') or c('receipt_vouchers_list') or c('bank_entries_list')
        or c('journal_vouchers_list')
    )
    finance_ledger = c('accounts_account_ledger') or c('wallet_dashboard') or c('fund_transfers_list')
    finance_final = c('accounts_balance_sheet') or c('chart_of_accounts_list')
    show_finance = bool(finance_books or finance_ledger or finance_final)

    admin_activity = c('activity_log_report')
    admin_ai = c('report_ai')
    admin_books = c('book_inventory_list') or c('book_assignment_list') or c('book_pending_returns')
    show_admin = bool(admin_activity or admin_ai or admin_books)

    order = [
        ('fleet', show_fleet),
        ('task', show_task),
        ('hr', show_hr),
        ('finance', show_finance),
        ('admin', show_admin),
    ]
    first_tab = 'fleet'
    for tid, ok in order:
        if ok:
            first_tab = tid
            break

    any_tab = bool(show_fleet or show_task or show_hr or show_finance or show_admin)

    badges = _report_centre_badge_counts(c, is_master)

    return {
        'first_tab': first_tab,
        'any_tab': any_tab,
        'badges': badges,
        'show_fleet': show_fleet,
        'show_task': show_task,
        'show_hr': show_hr,
        'show_finance': show_finance,
        'show_admin': show_admin,
        'fleet_vehicle_col': fleet_vehicle,
        'fleet_project_col': fleet_project,
        'fleet_expense_col': fleet_expense,
        'task_daily_col': task_daily,
        'task_logbook_col': task_logbook,
        'task_upload_col': task_upload,
        'hr_driver_col': hr_driver,
        'hr_att_col': hr_att,
        'hr_workforce_col': hr_workforce,
        'finance_books_col': finance_books,
        'finance_ledger_col': finance_ledger,
        'finance_final_col': finance_final,
        'admin_activity_col': admin_activity,
        'admin_ai_col': admin_ai,
        'admin_books_col': admin_books,
    }




_ACTIVITY_LOG_MAX_DAYS = 10


def _activity_log_datetime_in_range(column, date_from, date_to):
    """Calendar-date filter safe for SQLite ISO strings (T) from sync_master."""
    return and_(func.date(column) >= date_from, func.date(column) <= date_to)





def _render_ai_report_table(headers, rows):
    from markupsafe import escape as _esc
    if not rows:
        return '<p class="text-muted">No data found.</p>'
    keys = list(rows[0].keys()) if rows else []
    h = '<thead class="table-light"><tr>' + ''.join(f'<th>{_esc(x)}</th>' for x in headers) + '</tr></thead>'
    body = '<tbody>'
    for r in rows:
        body += '<tr>' + ''.join(f'<td>{_esc(r.get(k, ""))}</td>' for k in keys) + '</tr>'
    body += '</tbody>'
    return '<table class="table table-bordered table-sm">' + h + body + '</table>'





def _driver_profile_view_core(driver_id):
    """Shared template context for driver profile (authenticated + public share link)."""
    driver = Driver.query.get_or_404(driver_id)
    transfers = DriverTransfer.query.filter_by(driver_id=driver_id).order_by(DriverTransfer.transfer_date.asc()).all()
    status_changes = DriverStatusChange.query.filter_by(driver_id=driver_id).order_by(DriverStatusChange.change_date.asc()).all()

    job_history = []
    if driver.assign_date:
        if transfers:
            first_t = transfers[0]
            _assign_snap = type('Snap', (), {
                'vehicle': first_t.old_vehicle,
                'project': first_t.old_project,
                'district': first_t.old_district,
                'shift': first_t.old_shift,
                'assign_remarks': driver.assign_remarks,
            })()
        else:
            _assign_snap = driver
        job_history.append({'date': driver.assign_date, 'type': 'assignment', 'data': _assign_snap})
    for t in transfers:
        job_history.append({'date': t.transfer_date, 'type': 'transfer', 'data': t})
    for s in status_changes:
        job_history.append({'date': s.change_date, 'type': 'status', 'data': s})
    job_history.sort(key=lambda x: x['date'])

    from datetime import date as _date
    _today_d = _date.today()
    for i, h in enumerate(job_history):
        next_date = job_history[i + 1]['date'] if i + 1 < len(job_history) else _today_d
        h['duration_days'] = (next_date - h['date']).days

    total_actions = len(job_history)
    last_action = job_history[-1]['date'] if job_history else None
    last_action_type = None
    if job_history:
        _la = job_history[-1]
        if _la['type'] == 'assignment':
            last_action_type = 'Assigned'
        elif _la['type'] == 'transfer':
            last_action_type = 'Transferred'
        elif _la['type'] == 'status':
            last_action_type = 'Left' if _la['data'].action_type == 'left' else 'Rejoined'

    _today = pk_date()
    service_days = (_today - driver.application_date).days if driver.application_date else None
    driver_age = None
    if driver.dob:
        driver_age = _today.year - driver.dob.year - ((_today.month, _today.day) < (driver.dob.month, driver.dob.day))
    doc_fields = [driver.photo_path, driver.cnic_front_path, driver.cnic_back_path,
                  driver.license_front_path, driver.license_back_path,
                  driver.verify_license_photo_path, driver.document_path]
    doc_uploaded = sum(1 for d in doc_fields if d)
    doc_total = len(doc_fields)
    jh_counts = {'assignment': 0, 'transfer': 0, 'shift_change': 0, 'left': 0, 'rejoin': 0}
    for h in job_history:
        if h['type'] == 'assignment':
            jh_counts['assignment'] += 1
        elif h['type'] == 'transfer':
            if h['data'].is_shift_only:
                jh_counts['shift_change'] += 1
            else:
                jh_counts['transfer'] += 1
        elif h['type'] == 'status':
            if h['data'].action_type == 'left':
                jh_counts['left'] += 1
            else:
                jh_counts['rejoin'] += 1

    return {
        'driver': driver,
        'job_history': job_history,
        'total_actions': total_actions,
        'last_action': last_action,
        'last_action_type': last_action_type,
        'today': _today,
        'service_days': service_days,
        'driver_age': driver_age,
        'doc_uploaded': doc_uploaded,
        'doc_total': doc_total,
        'jh_counts': jh_counts,
    }


def _driver_update_vehicle_choices(driver=None):
    """Vehicle numbers from Vehicle Parking Inventory (permission-scoped)."""
    from auth_utils import get_user_context

    user_id = session.get('user_id')
    user_context = get_user_context(user_id) if user_id else {}
    allowed_projects = user_context.get('allowed_projects', set())
    allowed_districts = user_context.get('allowed_districts', set())
    allowed_vehicles = user_context.get('allowed_vehicles', set())
    is_master_or_admin = user_context.get('is_master_or_admin', False)

    q = Vehicle.query.filter(
        Vehicle.parking_station_id.isnot(None),
        Vehicle.vehicle_no.isnot(None),
        Vehicle.vehicle_no != '',
    )
    if not is_master_or_admin:
        if allowed_projects:
            q = q.filter(Vehicle.project_id.in_(list(allowed_projects)))
        if allowed_districts:
            q = q.filter(Vehicle.district_id.in_(list(allowed_districts)))
        if allowed_vehicles:
            q = q.filter(Vehicle.id.in_(list(allowed_vehicles)))
    vehicles = q.order_by(*vehicle_order_by()).all()

    choices = []
    seen = set()
    for v in vehicles:
        vno = (v.vehicle_no or '').strip()
        key = vno.lower()
        if vno and key not in seen:
            seen.add(key)
            choices.append(vno)

    if driver:
        default = (driver.vehicle.vehicle_no if getattr(driver, 'vehicle', None) else '') or ''
        default = default.strip()
        if default and default.lower() not in seen:
            choices.insert(0, default)
    return choices






# ════════════════════════════════════════════════════════════════════════════════

import collections as _sh_coll
import time as _sh_time

_health_cache      = {'data': None, 'ts': None}
_HEALTH_CACHE_TTL  = 900
_api_latency_ms    = _sh_coll.deque(maxlen=100)
_latency_history   = _sh_coll.deque(maxlen=24)
_session_history   = _sh_coll.deque(maxlen=24)
_route_perf_log    = _sh_coll.deque(maxlen=2500)
_last_backup_ts    = {'ts': None}
_health_alert_sent = {'db': False, 'r2': False}
_app_start_time    = _sh_time.time()


@app.before_request
def _sh_before_request():
    from flask import g
    g._sh_t0 = _sh_time.time()


@app.after_request
def _sh_after_request(response):
    from flask import g
    if hasattr(g, '_sh_t0'):
        ms = round((_sh_time.time() - g._sh_t0) * 1000, 1)
        if request.path.startswith('/api/'):
            _api_latency_ms.append(ms)
        try:
            skip_perf_path = (
                request.path.startswith('/static/')
                or request.path.startswith('/assets/')
                or request.path.startswith('/admin/system-health')
                or request.path.startswith('/network-probe')
            )
            if not skip_perf_path:
                payload_bytes = 0
                try:
                    payload_bytes = int(response.calculate_content_length() or 0)
                except Exception:
                    try:
                        payload_bytes = int(response.headers.get('Content-Length') or 0)
                    except Exception:
                        payload_bytes = 0
                _route_perf_log.append({
                    'ts': int(_sh_time.time()),
                    'method': request.method,
                    'path': request.path,
                    'endpoint': request.endpoint or '',
                    'ms': ms,
                    'status': int(getattr(response, 'status_code', 0) or 0),
                    'payload_bytes': payload_bytes,
                })
                uid = session.get('user_id')
                if uid and request.endpoint != 'api_client_diagnostics':
                    status_code = int(getattr(response, 'status_code', 0) or 0)
                    slow_html = (
                        request.method == 'GET'
                        and ms >= 8000
                        and not request.path.startswith('/api/')
                    )
                    if status_code >= 400 or slow_html:
                        _persist_client_diagnostic(
                            'http_error' if status_code >= 400 else 'slow_server',
                            page_path=request.path,
                            message=f'{request.method} {request.path} → {status_code}',
                            duration_ms=int(ms),
                            status_code=status_code,
                        )
        except Exception:
            pass
    return response


def _maybe_send_health_alert(data):
    """Legacy health notifications disabled (notifications v2)."""
    return
    try:
        if data.get('db_critical') and not _health_alert_sent['db']:
            _health_alert_sent['db'] = True
            _db_title = 'Database Storage Critical'
            _db_msg = (
                f"Database is at {data['db_pct']}% "
                f"({data['db_size_mb']} MB / {data['db_size_limit_mb']} MB). "
                "Upgrade plan or clean up data immediately."
            )
            n = Notification(
                title=_db_title, message=_db_msg,
                notification_type='danger', created_by_user_id=None,
                required_permission='backup,users_manage',
            )
            db.session.add(n)
            db.session.commit()
            try:
                from push_notifications import send_push_to_permitted
                send_push_to_permitted(['backup', 'users_manage'], _db_title, _db_msg)
            except Exception:
                pass
    except Exception:
        try:
            db.session.rollback()
        except Exception:
            pass
    try:
        if data.get('r2_critical') and not _health_alert_sent['r2']:
            _health_alert_sent['r2'] = True
            _r2_title = 'R2 Bucket Storage Critical'
            _r2_limit_gb = round(data.get('r2_size_limit_mb', 10240) / 1024)
            _r2_msg = (
                f"Cloudflare R2 is at {data['r2_pct']}% of its {_r2_limit_gb} GB limit. "
                "Delete unused files or upgrade your R2 plan."
            )
            n2 = Notification(
                title=_r2_title, message=_r2_msg,
                notification_type='danger', created_by_user_id=None,
                required_permission='backup,users_manage',
            )
            db.session.add(n2)
            db.session.commit()
            try:
                from push_notifications import send_push_to_permitted
                send_push_to_permitted(['backup', 'users_manage'], _r2_title, _r2_msg)
            except Exception:
                pass
    except Exception:
        try:
            db.session.rollback()
        except Exception:
            pass


def _build_route_diagnostics(window_minutes=15):
    now_ts = int(_sh_time.time())
    cutoff = now_ts - (int(window_minutes) * 60)
    all_perf = list(_route_perf_log)
    recent = [x for x in all_perf if int(x.get('ts', 0)) >= cutoff]
    history_route_times = {}
    for row in all_perf:
        method = str(row.get('method') or 'GET').upper()
        endpoint_or_path = row.get('endpoint') or row.get('path') or 'unknown'
        hist_key = f'{method} {endpoint_or_path}'
        t = float(row.get('ms') or 0)
        if t <= 0:
            continue
        history_route_times.setdefault(hist_key, []).append(t)

    def _fmt_size(bytes_val):
        try:
            n = float(bytes_val or 0)
        except Exception:
            n = 0.0
        if n >= (1024 * 1024):
            return f'{round(n / (1024 * 1024), 2)} MB'
        if n >= 1024:
            return f'{round(n / 1024, 1)} KB'
        return f'{int(n)} B'

    route_buckets = {}
    for row in recent:
        method = str(row.get('method') or 'GET').upper()
        endpoint_or_path = row.get('endpoint') or row.get('path') or 'unknown'
        key = f'{method} {endpoint_or_path}'
        b = route_buckets.setdefault(key, {'times': [], 'errors': 0, 'path': row.get('path') or '', 'payloads': []})
        t = float(row.get('ms') or 0)
        if t > 0:
            b['times'].append(t)
        payload_bytes = int(row.get('payload_bytes') or 0)
        if payload_bytes > 0:
            b['payloads'].append(payload_bytes)
        st = int(row.get('status') or 0)
        if st >= 500:
            b['errors'] += 1

    top_slow = []
    for route_key, b in route_buckets.items():
        times = sorted(b['times'])
        if not times:
            continue
        hits = len(times)
        avg = round(sum(times) / hits, 1)
        p95 = round(times[max(0, int(hits * 0.95) - 1)], 1)
        mx = round(times[-1], 1)
        payloads = b.get('payloads') or []
        avg_payload_bytes = int(round(sum(payloads) / len(payloads))) if payloads else 0
        top_slow.append({
            'route': route_key,
            'path': b['path'],
            'hits': hits,
            'avg_ms': avg,
            'p95_ms': p95,
            'max_ms': mx,
            'error_count': int(b['errors']),
            'avg_payload_bytes': avg_payload_bytes,
            'avg_payload_size': _fmt_size(avg_payload_bytes),
        })
    top_slow.sort(key=lambda x: (x['p95_ms'], x['avg_ms']), reverse=True)
    top_slow = top_slow[:8]

    recent_errors = []
    for row in sorted(recent, key=lambda x: x.get('ts', 0), reverse=True):
        st = int(row.get('status') or 0)
        if st < 500:
            continue
        recent_errors.append({
            'time': datetime.utcfromtimestamp(int(row.get('ts', now_ts))).strftime('%H:%M:%S'),
            'route': row.get('endpoint') or row.get('path') or 'unknown',
            'status': st,
            'ms': round(float(row.get('ms') or 0), 1),
        })
        if len(recent_errors) >= 8:
            break

    overall_times = sorted([float(r.get('ms') or 0) for r in recent if float(r.get('ms') or 0) > 0])
    req_count = len(overall_times)
    avg_ms = round(sum(overall_times) / req_count, 1) if req_count else None
    p95_ms = round(overall_times[max(0, int(req_count * 0.95) - 1)], 1) if req_count else None
    err_count = sum(1 for r in recent if int(r.get('status') or 0) >= 500)
    slow_count = sum(1 for t in overall_times if t >= 1000)
    slow_rate_pct = round((slow_count * 100.0) / req_count, 1) if req_count else 0.0

    analysis = []
    if err_count > 0:
        analysis.append({
            'level': 'danger',
            'text': f'Found {err_count} server error response(s) in the last {window_minutes} minutes.',
        })
    if p95_ms is not None and p95_ms >= 1800:
        analysis.append({
            'level': 'danger',
            'text': f'High backend response time detected (p95 {p95_ms} ms). Software/database slowdown likely.',
        })
    elif p95_ms is not None and p95_ms >= 900:
        analysis.append({
            'level': 'warning',
            'text': f'Moderate slowdown detected (p95 {p95_ms} ms). Check heavy routes below.',
        })
    if req_count < 20:
        analysis.append({
            'level': 'info',
            'text': f'Low sample volume in the last {window_minutes} minutes. Trigger normal user actions for stronger diagnosis.',
        })
    smart_candidates = []
    for r in top_slow:
        hist = history_route_times.get(r['route']) or []
        if len(hist) < 40 or int(r.get('hits') or 0) < 5:
            continue
        baseline_avg = sum(hist) / len(hist)
        curr_avg = float(r.get('avg_ms') or 0)
        if baseline_avg <= 0 or curr_avg < 300:
            continue
        slow_pct = ((curr_avg - baseline_avg) / baseline_avg) * 100.0
        if slow_pct >= 20:
            smart_candidates.append((slow_pct, r['route']))
    if smart_candidates:
        smart_candidates.sort(reverse=True, key=lambda x: x[0])
        slow_pct, route_name = smart_candidates[0]
        analysis.append({
            'level': 'warning',
            'text': f"Warning: '{route_name}' is performing {round(slow_pct, 1)}% slower than usual.",
        })
    if top_slow:
        top_max = max(top_slow, key=lambda x: float(x.get('max_ms') or 0))
        top_max_ms = float(top_max.get('max_ms') or 0)
        if top_max_ms >= 2500:
            analysis.append({
                'level': 'warning',
                'text': f'Outlier spike detected on {top_max.get("route")} (max {round(top_max_ms, 1)} ms, hits {top_max.get("hits")}).',
            })
    if slow_rate_pct >= 15:
        analysis.append({
            'level': 'warning',
            'text': f'High slow-request ratio detected: {slow_rate_pct}% requests are >= 1000 ms.',
        })
    if not analysis:
        analysis.append({
            'level': 'success',
            'text': 'No major software bottleneck detected in recent server timings.',
        })

    return {
        'window_minutes': int(window_minutes),
        'request_count': req_count,
        'avg_ms': avg_ms,
        'p95_ms': p95_ms,
        'error_count': err_count,
        'slow_count': slow_count,
        'slow_rate_pct': slow_rate_pct,
        'top_slow_routes': top_slow,
        'recent_errors': recent_errors,
        'analysis': analysis,
    }


def _build_health_data():
    """Fetch live infrastructure metrics from Render API, PostgreSQL, R2, and internal sources."""
    import json
    import urllib.request
    import platform
    import flask as _flask_mod

    _db_limit = int(os.environ.get('DB_SIZE_LIMIT_MB', '1024'))
    _r2_limit = int(os.environ.get('R2_SIZE_LIMIT_GB', '10')) * 1024

    result = {
        'service':          None,
        'last_deploy':      None,
        'recent_deploys':   [],
        'db_size_mb':       None,
        'db_size_limit_mb': _db_limit,
        'r2_size_mb':       None,
        'r2_total_objects': 0,
        'r2_size_limit_mb': _r2_limit,
        'active_sessions':  None,
        'api_avg_ms':       None,
        'api_p95_ms':       None,
        'api_sample_count': 0,
        'last_backup_ts':   _last_backup_ts.get('ts'),
        'checks':           {},
        'errors':           [],
        'fetched_at':       pk_now().strftime('%d-%m-%Y %H:%M UTC'),
        'ram_mb':           None,
        'ram_limit_mb':     int(os.environ.get('RAM_LIMIT_MB', '512')),
        'ram_pct':          None,
        'upload_size_mb':   None,
        'upload_file_count': 0,
        'db_table_stats':   [],
        'fcm_total':        0,
        'fcm_active':       0,
        'fcm_inactive':     0,
        'sys_python':       platform.python_version(),
        'sys_flask':        _flask_mod.__version__,
        'sys_os':           f'{platform.system()} {platform.release()}',
        'sys_timezone':     app.config.get('APP_TIMEZONE', 'UTC'),
        'sys_server_time':  pk_now().strftime('%d-%m-%Y %H:%M:%S'),
        'sys_uptime_sec':   int(_sh_time.time() - _app_start_time),
        'latency_history':  list(_latency_history),
        'session_history':  list(_session_history),
        'backup_schedule_enabled': app.config.get('BACKUP_SCHEDULE_ENABLED', False),
        'backup_schedule_time':    app.config.get('BACKUP_SCHEDULE_TIME', '02:00'),
        'backup_email_to':         app.config.get('BACKUP_EMAIL_TO', ''),
        'diagnostics':             {},
    }

    render_key = os.environ.get('RENDER_API_KEY', '').strip()
    service_id = os.environ.get('RENDER_SERVICE_ID', '').strip()

    # 1. Render Service + Deploy History
    if render_key and service_id:
        try:
            hdr = {'Authorization': f'Bearer {render_key}', 'Accept': 'application/json'}
            req = urllib.request.Request(
                f'https://api.render.com/v1/services/{service_id}', headers=hdr)
            with urllib.request.urlopen(req, timeout=8) as r:
                result['service'] = json.loads(r.read())
            req2 = urllib.request.Request(
                f'https://api.render.com/v1/services/{service_id}/deploys?limit=5', headers=hdr)
            with urllib.request.urlopen(req2, timeout=8) as r2:
                deploys = json.loads(r2.read())
                parsed = [d.get('deploy', d) for d in deploys] if deploys else []
                result['recent_deploys'] = parsed
                if parsed:
                    result['last_deploy'] = parsed[0]
            svc_name   = (result['service'] or {}).get('name', service_id)
            svc_status = (result['service'] or {}).get('status', '?')
            result['checks']['render_api'] = {'status': 'ok', 'msg': f'{svc_name} is {svc_status}'}
        except Exception as e:
            msg = str(e)[:120]
            result['checks']['render_api'] = {'status': 'error', 'msg': msg}
            result['errors'].append(f'Render API: {msg}')
    else:
        result['checks']['render_api'] = {
            'status': 'skip', 'msg': 'RENDER_API_KEY or RENDER_SERVICE_ID not configured'}

    # 2. Database Size
    try:
        db_uri = app.config.get('SQLALCHEMY_DATABASE_URI', '')
        if 'postgresql' in db_uri:
            row = db.session.execute(text('SELECT pg_database_size(current_database()) AS sz')).fetchone()
            if row:
                result['db_size_mb'] = round(row.sz / (1024 * 1024), 1)
        else:
            db_path = db_uri.replace('sqlite:///', '')
            if db_path and os.path.isfile(db_path):
                result['db_size_mb'] = round(os.path.getsize(db_path) / (1024 * 1024), 1)
                result['db_size_limit_mb'] = 512
        result['checks']['db_size'] = {'status': 'ok', 'msg': f'{result["db_size_mb"]} MB used'}
    except Exception as e:
        msg = str(e)[:120]
        result['checks']['db_size'] = {'status': 'error', 'msg': msg}
        result['errors'].append(f'DB size: {msg}')

    # 3. Cloudflare R2 Bucket
    try:
        from r2_storage import _get_s3_client, R2_BUCKET_NAME
        client = _get_s3_client()
        paginator = client.get_paginator('list_objects_v2')
        total_bytes, total_objs = 0, 0
        for page in paginator.paginate(Bucket=R2_BUCKET_NAME):
            for obj in page.get('Contents', []):
                total_bytes += obj.get('Size', 0)
                total_objs += 1
        result['r2_size_mb']       = round(total_bytes / (1024 * 1024), 1)
        result['r2_total_objects'] = total_objs
        result['checks']['r2_storage'] = {
            'status': 'ok', 'msg': f'{result["r2_size_mb"]} MB, {total_objs} objects'}
    except Exception as e:
        msg = str(e)[:120]
        result['checks']['r2_storage'] = {'status': 'error', 'msg': msg}
        result['errors'].append(f'R2 storage: {msg}')

    # 4. Active Sessions
    try:
        cutoff = pk_now() - timedelta(minutes=30)
        active = db.session.query(ActivityLog.user_id).filter(
            ActivityLog.created_at >= cutoff).distinct().count()
        result['active_sessions'] = active
        result['checks']['sessions'] = {'status': 'ok', 'msg': f'{active} user(s) active in last 30 min'}
        _session_history.append(active)
    except Exception as e:
        result['checks']['sessions'] = {'status': 'error', 'msg': str(e)[:80]}

    # 5. API Latency
    if _api_latency_ms:
        samples = list(_api_latency_ms)
        avg = round(sum(samples) / len(samples), 1)
        result['api_avg_ms']       = avg
        result['api_sample_count'] = len(samples)
        p95_idx                    = max(0, int(len(samples) * 0.95) - 1)
        result['api_p95_ms']       = sorted(samples)[p95_idx]
        result['checks']['api_latency'] = {
            'status': 'ok',
            'msg': f'avg {avg}ms, p95 {result["api_p95_ms"]}ms ({len(samples)} samples)'}
        _latency_history.append(avg)
    else:
        result['checks']['api_latency'] = {
            'status': 'na', 'msg': 'No /api/* calls recorded yet in this session'}

    # 6. RAM / Memory Usage
    try:
        import resource
        rss_bytes = resource.getrusage(resource.RUSAGE_SELF).ru_maxrss * 1024
        result['ram_mb'] = round(rss_bytes / (1024 * 1024), 1)
    except Exception:
        try:
            import psutil
            result['ram_mb'] = round(psutil.Process().memory_info().rss / (1024 * 1024), 1)
        except Exception:
            pass
    if result['ram_mb'] is not None and result['ram_limit_mb']:
        result['ram_pct'] = round(result['ram_mb'] / result['ram_limit_mb'] * 100, 1)

    # 7. Upload Folder Size
    try:
        upload_dir = app.config.get('UPLOAD_FOLDER', '')
        if upload_dir and os.path.isdir(upload_dir):
            total_sz, total_cnt = 0, 0
            for dirpath, _dirnames, filenames in os.walk(upload_dir):
                for f in filenames:
                    fp = os.path.join(dirpath, f)
                    try:
                        total_sz += os.path.getsize(fp)
                        total_cnt += 1
                    except OSError:
                        pass
            result['upload_size_mb'] = round(total_sz / (1024 * 1024), 1)
            result['upload_file_count'] = total_cnt
    except Exception:
        pass

    # 8. Database Table Stats (PostgreSQL only)
    try:
        db_uri = app.config.get('SQLALCHEMY_DATABASE_URI', '')
        if 'postgresql' in db_uri:
            tbl_q = text("""
                SELECT relname AS name,
                       n_live_tup AS row_count,
                       pg_total_relation_size(relid) AS size_bytes
                FROM pg_stat_user_tables
                ORDER BY pg_total_relation_size(relid) DESC
                LIMIT 10
            """)
            rows = db.session.execute(tbl_q).fetchall()
            result['db_table_stats'] = [
                {'name': r.name, 'rows': r.row_count,
                 'size_mb': round(r.size_bytes / (1024 * 1024), 2)}
                for r in rows
            ]
    except Exception:
        pass

    # 9. FCM / Notification Stats
    try:
        result['fcm_total']    = DeviceFCMToken.query.count()
        result['fcm_active']   = DeviceFCMToken.query.filter_by(is_active=True).count()
        result['fcm_inactive'] = result['fcm_total'] - result['fcm_active']
    except Exception:
        pass

    # 10. Security — Login Attempts (last 24h)
    try:
        _sec_cutoff = pk_now() - timedelta(hours=24)
        result['failed_logins_24h'] = LoginAttempt.query.filter(
            LoginAttempt.success == False,
            LoginAttempt.created_at >= _sec_cutoff
        ).count()
        result['success_logins_24h'] = LoginAttempt.query.filter(
            LoginAttempt.success == True,
            LoginAttempt.created_at >= _sec_cutoff
        ).count()
        _top_ips_q = db.session.query(
            LoginAttempt.ip_address, func.count(LoginAttempt.id).label('cnt')
        ).filter(
            LoginAttempt.success == False,
            LoginAttempt.created_at >= _sec_cutoff
        ).group_by(LoginAttempt.ip_address).order_by(func.count(LoginAttempt.id).desc()).limit(5).all()
        result['top_fail_ips'] = [{'ip': r[0] or 'unknown', 'count': r[1]} for r in _top_ips_q]
        _active_users = db.session.query(
            User.full_name, ActivityLog.created_at
        ).join(User, User.id == ActivityLog.user_id).filter(
            ActivityLog.created_at >= pk_now() - timedelta(minutes=30)
        ).group_by(User.id, User.full_name, ActivityLog.created_at).order_by(
            ActivityLog.created_at.desc()
        ).limit(20).all()
        _seen = set()
        _active_list = []
        for name, ts in _active_users:
            if name not in _seen:
                _seen.add(name)
                _active_list.append({'name': name, 'last_seen': ts.strftime('%H:%M') if ts else '?'})
        result['active_user_list'] = _active_list[:10]
    except Exception:
        result['failed_logins_24h'] = 0
        result['success_logins_24h'] = 0
        result['top_fail_ips'] = []
        result['active_user_list'] = []

    # 11. Row counts for cleanup targets
    try:
        result['row_count_activity'] = ActivityLog.query.count()
        result['row_count_login'] = LoginLog.query.count()
        result['row_count_notifications'] = Notification.query.count()
        result['row_count_login_attempts'] = LoginAttempt.query.count()
    except Exception:
        result['row_count_activity'] = 0
        result['row_count_login'] = 0
        result['row_count_notifications'] = 0
        result['row_count_login_attempts'] = 0

    # 12. Persistent backup info from SystemSetting
    try:
        result['last_backup_ts_persistent'] = SystemSetting.get('last_backup_ts')
        result['last_backup_result'] = SystemSetting.get('last_backup_result', 'unknown')
        result['last_backup_size'] = SystemSetting.get('last_backup_size')
    except Exception:
        result['last_backup_ts_persistent'] = None
        result['last_backup_result'] = 'unknown'
        result['last_backup_size'] = None

    # Percentages & Critical Flags
    result['db_pct']       = round(result['db_size_mb'] / result['db_size_limit_mb'] * 100, 1) if result['db_size_mb'] is not None else None
    result['r2_pct']       = round(result['r2_size_mb'] / result['r2_size_limit_mb'] * 100, 1) if result['r2_size_mb'] is not None else None
    result['db_critical']  = bool(result['db_pct'] is not None and result['db_pct'] >= 80)
    result['r2_critical']  = bool(result['r2_pct'] is not None and result['r2_pct'] >= 80)
    result['any_critical'] = result['db_critical'] or result['r2_critical']

    # 13. Software diagnostics from route-level timings
    try:
        result['diagnostics'] = _build_route_diagnostics(window_minutes=15)
    except Exception as e:
        result['diagnostics'] = {
            'window_minutes': 15,
            'request_count': 0,
            'avg_ms': None,
            'p95_ms': None,
            'error_count': 0,
            'top_slow_routes': [],
            'recent_errors': [],
            'analysis': [{'level': 'danger', 'text': f'Diagnostics engine error: {str(e)[:120]}'}],
        }

    _maybe_send_health_alert(result)
    return result


def _fetch_system_health(force=False):
    import time as _t
    now = _t.time()
    if (not force) and _health_cache['data'] and _health_cache['ts'] and (now - _health_cache['ts']) < _HEALTH_CACHE_TTL:
        return _health_cache['data']
    data = _build_health_data()
    _health_cache['data'] = data
    _health_cache['ts']   = now
    return data








@app.cli.command('fix-driver-status')
def fix_driver_status():
    """One-time backfill: calculate cnic_status and license_status for all drivers where it is blank."""
    today = pk_date()
    drivers = Driver.query.all()
    updated = 0
    for d in drivers:
        changed = False
        if d.cnic_expiry_date:
            correct = 'Valid' if d.cnic_expiry_date >= today else 'Expired'
            if d.cnic_status != correct:
                d.cnic_status = correct
                changed = True
        if d.license_expiry_date:
            correct = 'Valid' if d.license_expiry_date >= today else 'Expired'
            if d.license_status != correct:
                d.license_status = correct
                changed = True
        if changed:
            updated += 1
    db.session.commit()
    app.logger.info('Driver photo path fix: %d/%d updated.', updated, len(drivers))


# ══════════════════════════════════════════════════════════════════════════════
# Tracker Automation — TrackingWorld portal robot
# ══════════════════════════════════════════════════════════════════════════════






# ════════════════════════════════════════════════════════════════════════════════
# DRIVER DOCUMENT UPDATE PORTAL – LIST + PORTAL
# ════════════════════════════════════════════════════════════════════════════════





